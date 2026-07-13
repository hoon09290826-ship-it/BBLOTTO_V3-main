from fastapi import FastAPI, HTTPException, Request, Header
from fastapi.responses import FileResponse, StreamingResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from pathlib import Path
import sqlite3, json, random, re, itertools, datetime, secrets, hashlib, hmac, io, csv, collections, shutil, os, urllib.request, urllib.parse, time, threading

BASE = Path(__file__).resolve().parents[1]

# PHASE21: 데이터 초기화 방지 / 영구저장 경로 고정
# 우선순위
# 1) BBLOTTO_DB_DIR 환경변수
# 2) Render Disk 권장 마운트 경로 /data
# 3) 로컬 개발용 프로젝트 database 폴더
def resolve_db_dir():
    env_dir = os.getenv('BBLOTTO_DB_DIR', '').strip()
    if env_dir:
        return Path(env_dir)
    render_disk = Path('/data')
    if render_disk.exists() and os.access(str(render_disk), os.W_OK):
        return render_disk / 'bblotto_database'
    return BASE / 'database'

DB_DIR = resolve_db_dir(); DB_DIR.mkdir(parents=True, exist_ok=True)
EXPORT_DIR = Path(os.getenv('BBLOTTO_EXPORT_DIR', str(DB_DIR / 'exports'))); EXPORT_DIR.mkdir(parents=True, exist_ok=True)
DB = DB_DIR / 'bblotto_v34.db'
FRONT = BASE / 'frontend'

RC_VERSION = 'STABLE_CORE_CLEAN_BUILD'
APP_VERSION = 'BBLOTTO V3 STABLE'
app = FastAPI(title=f'{APP_VERSION} {RC_VERSION}', docs_url=None, redoc_url=None, openapi_url=None)
RC3_8_VERSION = 'V2_STABLE_RC3_15'
RC3_9_VERSION = 'V2_STABLE_RC3_15'
RC3_10_VERSION = 'V2_STABLE_RC3_15'
app.mount('/static', StaticFiles(directory=str(FRONT)), name='static')

# RC11.3: 운영 보안 보조 함수
_RC11_MAX_BODY_BYTES = 2 * 1024 * 1024
_RC11_ALLOWED_METHODS_WITH_BODY = {'POST', 'PUT', 'PATCH'}

def _rc113_client_ip(request: Request) -> str:
    # Railway/Render 같은 역방향 프록시 환경에서는 Forwarded 헤더를 우선 사용합니다.
    forwarded = str(request.headers.get('x-forwarded-for', '') or '').split(',')[0].strip()
    if forwarded and len(forwarded) <= 64 and re.fullmatch(r'[0-9a-fA-F:.]+', forwarded):
        return forwarded
    return request.client.host if request.client else 'unknown'

def _rc113_validate_origin(request: Request):
    if request.method not in {'POST', 'PUT', 'PATCH', 'DELETE'} or not request.url.path.startswith('/api/'):
        return
    origin = str(request.headers.get('origin', '') or '').strip()
    if not origin:
        return
    expected = f'{request.url.scheme}://{request.headers.get("host", "")}'.rstrip('/')
    forwarded_proto = str(request.headers.get('x-forwarded-proto', '') or '').split(',')[0].strip()
    if forwarded_proto:
        expected = f'{forwarded_proto}://{request.headers.get("host", "")}'.rstrip('/')
    if origin.rstrip('/') != expected:
        raise HTTPException(403, '허용되지 않은 요청 출처입니다.')

def validate_admin_username(username: str) -> str:
    value = str(username or '').strip()
    if not re.fullmatch(r'[A-Za-z0-9._-]{4,40}', value):
        raise HTTPException(400, '관리자 아이디는 영문, 숫자, 점, 밑줄, 하이픈을 사용해 4~40자로 입력해주세요.')
    return value

# RC11: 기본 보안 헤더. 외부 리소스 없이 자체 파일만 사용하는 현재 구조에 맞춘 정책입니다.
@app.middleware('http')
async def rc11_security_headers(request: Request, call_next):
    try:
        _rc113_validate_origin(request)
        if request.method in _RC11_ALLOWED_METHODS_WITH_BODY:
            length = request.headers.get('content-length')
            if length:
                try:
                    if int(length) > _RC11_MAX_BODY_BYTES:
                        raise HTTPException(413, '요청 데이터가 너무 큽니다.')
                except ValueError:
                    raise HTTPException(400, '잘못된 Content-Length입니다.')
    except HTTPException as exc:
        return JSONResponse(status_code=exc.status_code, content={'ok':False,'error':{'type':'SecurityValidation','status_code':exc.status_code,'message':exc.detail},'path':str(request.url.path),'time':datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')})
    response = await call_next(request)
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'DENY'
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    response.headers['Permissions-Policy'] = 'camera=(), microphone=(), geolocation=()'
    response.headers['Cross-Origin-Opener-Policy'] = 'same-origin'
    response.headers['Cross-Origin-Resource-Policy'] = 'same-origin'
    response.headers['X-Permitted-Cross-Domain-Policies'] = 'none'
    response.headers['X-Robots-Tag'] = 'noindex, nofollow, noarchive'
    if request.url.scheme == 'https' or str(request.headers.get('x-forwarded-proto','')).lower().startswith('https'):
        response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    response.headers['Content-Security-Policy'] = (
        "default-src 'self'; img-src 'self' data:; style-src 'self' 'unsafe-inline'; "
        "script-src 'self'; connect-src 'self'; frame-ancestors 'none'; base-uri 'self'; form-action 'self'"
    )
    if request.url.path.startswith('/api/') or request.url.path in {'/','/dashboard','/app.js','/login.js','/style.css'}:
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
    return response

# RC11: 로그인 무차별 대입 방지(서버 프로세스 단위). DB 로그와 함께 동작합니다.
_RC11_LOGIN_LOCK = threading.Lock()
_RC11_LOGIN_ATTEMPTS = collections.defaultdict(list)
_RC11_LOGIN_WINDOW_SECONDS = 15 * 60
_RC11_LOGIN_MAX_FAILURES = 7

def _rc11_login_key(request: Request, username: str) -> str:
    ip = _rc113_client_ip(request)
    return f'{ip}|{str(username or "").strip().lower()[:80]}'

def _rc11_check_login_limit(request: Request, username: str):
    key = _rc11_login_key(request, username)
    current = time.time()
    with _RC11_LOGIN_LOCK:
        recent = [t for t in _RC11_LOGIN_ATTEMPTS.get(key, []) if current - t < _RC11_LOGIN_WINDOW_SECONDS]
        _RC11_LOGIN_ATTEMPTS[key] = recent
        if len(recent) >= _RC11_LOGIN_MAX_FAILURES:
            retry = max(1, int(_RC11_LOGIN_WINDOW_SECONDS - (current - recent[0])))
            raise HTTPException(429, f'로그인 시도가 너무 많습니다. 약 {max(1, retry // 60)}분 후 다시 시도해주세요.')

def _rc11_record_login_failure(request: Request, username: str):
    key = _rc11_login_key(request, username)
    with _RC11_LOGIN_LOCK:
        _RC11_LOGIN_ATTEMPTS[key].append(time.time())

def _rc11_clear_login_failures(request: Request, username: str):
    key = _rc11_login_key(request, username)
    with _RC11_LOGIN_LOCK:
        _RC11_LOGIN_ATTEMPTS.pop(key, None)


# RC2 Sprint 6: 표준 오류 응답 핸들러
@app.exception_handler(HTTPException)
def rc2_http_exception_handler(request: Request, exc: HTTPException):
    return JSONResponse(status_code=exc.status_code, content={
        'ok': False,
        'error': {'type': 'HTTPException', 'status_code': exc.status_code, 'message': exc.detail},
        'path': str(request.url.path),
        'time': datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    })

@app.exception_handler(Exception)
def rc2_unhandled_exception_handler(request: Request, exc: Exception):
    return JSONResponse(status_code=500, content={
        'ok': False,
        'error': {'type': exc.__class__.__name__, 'message': '서버 처리 중 오류가 발생했습니다.'},
        'path': str(request.url.path),
        'time': datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    })

DEFAULT_DRAWS = [
 (1230,'2026.06.27',[3,8,9,22,28,42],45),(1229,'2026.06.20',[12,13,29,34,37,42],16),(1228,'2026.06.13',[24,29,30,31,35,44],1),
 (1227,'2026.06.06',[1,14,16,34,41,44],13),(1226,'2026.05.30',[4,6,13,17,26,28],41),(1225,'2026.05.23',[8,9,19,25,41,42],33),
 (1224,'2026.05.16',[9,18,21,27,44,45],28),(1223,'2026.05.09',[16,18,20,32,33,39],26),(1222,'2026.05.02',[4,11,17,22,32,41],34),
 (1221,'2026.04.25',[6,13,18,28,30,36],9),(1220,'2026.04.18',[2,22,25,28,34,43],16),(1219,'2026.04.11',[1,2,15,28,39,45],31),
]
try:
    from .data import DRAWS as FULL_DRAWS
    DEFAULT_DRAWS = [(int(x['r']), x.get('d',''), list(x.get('n',[])), int(x.get('b',0))) for x in FULL_DRAWS]
except Exception:
    pass

PRIZE_TABLE = {'1등': 2000000000, '2등': 50000000, '3등': 1500000, '4등': 50000, '5등': 5000, '낙첨': 0}
COST_PER_COMBO = 1000

# PHASE30: PostgreSQL 영구저장 지원
# Render Free에서 SQLite 파일이 초기화되는 문제를 해결하기 위해 DATABASE_URL이 있으면 PostgreSQL을 사용합니다.
# DATABASE_URL이 없으면 기존처럼 SQLite를 사용하므로 로컬 실행도 그대로 가능합니다.
def _normalize_database_url(url: str) -> str:
    url = (url or '').strip()
    if not url or url.startswith('${{'):
        return ''
    if url.startswith('postgres://'):
        return 'postgresql://' + url[len('postgres://'):]
    return url

def _build_database_url_from_pg_vars() -> str:
    host = os.getenv('PGHOST', '').strip()
    port = os.getenv('PGPORT', '5432').strip() or '5432'
    user = os.getenv('PGUSER', '').strip() or os.getenv('POSTGRES_USER', '').strip()
    password = os.getenv('PGPASSWORD', '').strip() or os.getenv('POSTGRES_PASSWORD', '').strip()
    dbname = os.getenv('PGDATABASE', '').strip() or os.getenv('POSTGRES_DB', '').strip()
    if host and user and dbname:
        return f"postgresql://{urllib.parse.quote(user)}:{urllib.parse.quote(password)}@{host}:{port}/{dbname}"
    return ''

# RC3 Database Migration: Railway/PostgreSQL 자동 감지
# 우선순위: DATABASE_URL > POSTGRES_URL > PGHOST/PGUSER/PGDATABASE 조합 > SQLite
DATABASE_URL = _normalize_database_url(
    os.getenv('DATABASE_URL', '')
    or os.getenv('POSTGRES_URL', '')
    or _build_database_url_from_pg_vars()
)
DB_ENGINE = 'postgresql' if DATABASE_URL else 'sqlite'

class CompatRow:
    def __init__(self, columns, values):
        self._columns = list(columns or [])
        self._values = list(values or [])
        self._map = {str(k): self._values[i] for i, k in enumerate(self._columns) if i < len(self._values)}
    def __getitem__(self, key):
        if isinstance(key, int): return self._values[key]
        return self._map.get(key)
    def get(self, key, default=None): return self._map.get(key, default)
    def keys(self): return self._map.keys()
    def items(self): return self._map.items()
    def __iter__(self): return iter(self._map)
    def __len__(self): return len(self._values)
    def __dict__(self): return dict(self._map)

class PgCursorCompat:
    def __init__(self, conn):
        self.conn = conn
        self.cur = conn._conn.cursor()
        self.rowcount = -1
        self.lastrowid = None
        self._last_rows = None
        self._last_columns = None
    def _convert_sql(self, sql):
        s = str(sql)
        s = re.sub(r'INTEGER\s+PRIMARY\s+KEY\s+AUTOINCREMENT', 'SERIAL PRIMARY KEY', s, flags=re.I)
        s = s.replace('INSERT OR IGNORE INTO', 'INSERT INTO')
        s = s.replace('INSERT OR REPLACE INTO', 'INSERT INTO')
        s = s.replace('DEFAULT ""', "DEFAULT ''")
        s = s.replace('DEFAULT "관리자"', "DEFAULT '관리자'")
        s = s.replace('DEFAULT "전체권한"', "DEFAULT '전체권한'")
        s = s.replace('DEFAULT "일반"', "DEFAULT '일반'")
        s = s.replace('DEFAULT "활성"', "DEFAULT '활성'")
        s = s.replace('DEFAULT "보통"', "DEFAULT '보통'")
        s = s.replace('DEFAULT "직접등록"', "DEFAULT '직접등록'")
        s = s.replace('DEFAULT "balanced"', "DEFAULT 'balanced'")
        s = s.replace('DEFAULT "[]"', "DEFAULT '[]'")
        s = s.replace('DEFAULT "{}"', "DEFAULT '{}'")
        s = s.replace('DEFAULT "mock"', "DEFAULT 'mock'")
        s = s.replace('DEFAULT "saved"', "DEFAULT 'saved'")
        s = s.replace('COALESCE(NULLIF(name,""),"관리자")', "COALESCE(NULLIF(name,''),'관리자')")
        s = s.replace('COALESCE(status,"활성")', "COALESCE(status,'활성')")
        s = s.replace('COALESCE(priority,"보통")', "COALESCE(priority,'보통')")
        s = s.replace('COALESCE(grade,"일반")', "COALESCE(grade,'일반')")
        s = s.replace('COALESCE(last_contact_at,"")', "COALESCE(last_contact_at,'')")
        s = s.replace('COALESCE(source,"직접등록")', "COALESCE(source,'직접등록')")
        for lit in ['활성','VIP','다이아','높음','최우선','보통','일반','직접등록','manual','manual_auto','starter100','balanced']:
            s = s.replace(f'"{lit}"', f"'{lit}'")
        s = s.replace('COALESCE(source,"")', "COALESCE(source,'')")
        s = s.replace('COALESCE(priority,"")', "COALESCE(priority,'')")
        s = re.sub(r'\s+COLLATE\s+NOCASE', '', s, flags=re.I)
        s = s.replace('COALESCE(last_contact_at,"")=""', "COALESCE(last_contact_at,'')=''")
        # sqlite의 ? 파라미터를 psycopg2의 %s로 변환한다.
        # 단, SQL 안의 LIKE '%super%' 같은 리터럴에 들어있는 %s를
        # 파라미터로 착각하지 않도록 ?를 먼저 안전 토큰으로 바꾼다.
        pg_param_token = '__BBLOTTO_PG_PARAM__'
        out=[]; in_str=False; quote=''
        for ch in s:
            if ch in ("'", '"'):
                if not in_str:
                    in_str=True; quote=ch
                elif quote==ch:
                    in_str=False; quote=''
            if ch=='?' and not in_str:
                out.append(pg_param_token)
            else:
                out.append(ch)
        s=''.join(out)
        # psycopg2는 SQL 문자열 안의 %도 포맷 기호로 해석한다.
        # 실제 파라미터 토큰은 보호하고, 나머지 %만 %%로 이스케이프한다.
        if '%' in s:
            s = s.replace('%', '%%')
        s = s.replace(pg_param_token, '%s')
        # PostgreSQL upsert/ignore 호환
        if re.match(r'\s*INSERT\s+INTO\s+draws\s*\(', s, re.I) and 'ON CONFLICT' not in s:
            s += ' ON CONFLICT (round_no) DO UPDATE SET draw_date=EXCLUDED.draw_date, numbers=EXCLUDED.numbers, bonus=EXCLUDED.bonus, source=EXCLUDED.source, updated_at=EXCLUDED.updated_at'
        elif re.match(r'\s*INSERT\s+INTO\s+settings\s*\(', s, re.I) and 'ON CONFLICT' not in s:
            s += ' ON CONFLICT (key) DO UPDATE SET value=EXCLUDED.value, updated_at=EXCLUDED.updated_at'
        elif re.match(r'\s*INSERT\s+INTO\s+admins\s*\(', s, re.I) and 'ON CONFLICT' not in s and 'RETURNING' not in s:
            # 기본 관리자 생성 또는 관리자 추가에서 username 중복 방지
            if 'password_hash' in s and 'VALUES' in s:
                s += ' ON CONFLICT (username) DO NOTHING'
        return s
    def _table_info_rows(self, table):
        q = "SELECT column_name FROM information_schema.columns WHERE table_schema='public' AND table_name=%s ORDER BY ordinal_position"
        self.cur.execute(q, (table,))
        rows = self.cur.fetchall()
        # sqlite PRAGMA table_info 호환: r[1]이 컬럼명
        self._last_columns = ['cid','name','type','notnull','dflt_value','pk']
        self._last_rows = [CompatRow(self._last_columns, [i, r[0], '', 0, None, 0]) for i, r in enumerate(rows)]
        return self
    def execute(self, sql, params=None):
        params = tuple(params or ())
        raw = str(sql).strip()
        m = re.match(r'PRAGMA\s+table_info\(([^)]+)\)', raw, re.I)
        if m:
            return self._table_info_rows(m.group(1).strip().strip('"'))
        s = self._convert_sql(sql)
        # lastrowid가 필요한 주요 INSERT는 RETURNING id 추가
        if re.match(r'\s*INSERT\s+INTO\s+(members|recommendations|sms_logs)\s*\(', s, re.I) and 'RETURNING' not in s and 'ON CONFLICT' not in s:
            s += ' RETURNING id'
        try:
            self.cur.execute(s, params)
        except Exception as e:
            # ALTER TABLE ADD COLUMN 중복은 안전하게 무시
            if raw.upper().startswith('ALTER TABLE') and ('already exists' in str(e).lower() or 'duplicate column' in str(e).lower()):
                self.conn._conn.rollback(); return self
            raise
        self.rowcount = self.cur.rowcount
        self._last_rows = None; self._last_columns = None; self.lastrowid = None
        if self.cur.description:
            cols = [d[0] for d in self.cur.description]
            rows = self.cur.fetchall()
            self._last_columns = cols
            self._last_rows = [CompatRow(cols, r) for r in rows]
            if len(rows)==1 and 'id' in cols:
                try: self.lastrowid = self._last_rows[0]['id']
                except Exception: pass
        return self
    def fetchone(self):
        if not self._last_rows: return None
        return self._last_rows[0]
    def fetchall(self): return self._last_rows or []

class PgConnCompat:
    def __init__(self):
        import psycopg2
        self._conn = psycopg2.connect(DATABASE_URL)
        self._conn.autocommit = False
    def cursor(self): return PgCursorCompat(self)
    def execute(self, sql, params=None):
        cur = self.cursor(); return cur.execute(sql, params)
    def commit(self): self._conn.commit()
    def rollback(self): self._conn.rollback()
    def close(self): self._conn.close()
    def __enter__(self): return self
    def __exit__(self, exc_type, exc, tb):
        if exc_type: self.rollback()
        else: self.commit()
        self.close()

def con():
    if DB_ENGINE == 'postgresql':
        return PgConnCompat()
    c = sqlite3.connect(DB); c.row_factory = sqlite3.Row; return c

def now(): return datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')

def normalize_date_text(value, fallback=''):
    text = str(value or '').strip()
    if not text:
        return fallback
    # HTML date input(YYYY-MM-DD) 또는 기존 created_at(YYYY-MM-DD HH:MM:SS) 모두 허용
    m = re.match(r'^(\d{4}-\d{2}-\d{2})', text)
    if m:
        return m.group(1)
    m = re.match(r'^(\d{4})[./](\d{1,2})[./](\d{1,2})', text)
    if m:
        return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    return fallback

def add_months_date(date_text, months=12):
    base = normalize_date_text(date_text, datetime.datetime.now().strftime('%Y-%m-%d'))
    try:
        d = datetime.datetime.strptime(base, '%Y-%m-%d').date()
        months = int(months or 12)
        if months not in (6, 12, 24, 36):
            months = 12
        y = d.year + ((d.month - 1 + months) // 12)
        m = ((d.month - 1 + months) % 12) + 1
        # 대상 월의 마지막 일자보다 큰 날은 말일로 보정합니다.
        import calendar
        day = min(d.day, calendar.monthrange(y, m)[1])
        return datetime.date(y, m, day).strftime('%Y-%m-%d')
    except Exception:
        return ''

def add_one_year_date(date_text):
    return add_months_date(date_text, 12)


def clean_template_text(value):
    if value is None:
        return ''
    if isinstance(value, (dict, list)):
        obj = value
    else:
        text = str(value)
        for _ in range(3):
            t = text.strip()
            if not ((t.startswith('{') and t.endswith('}')) or (t.startswith('[') and t.endswith(']'))):
                break
            try:
                obj = json.loads(t)
            except Exception:
                break
            extracted = clean_template_text(obj)
            if not extracted or extracted == text:
                break
            text = extracted
        return text.replace('\\n', '\n').replace('\\t', '\t')
    if isinstance(obj, list):
        return '\n'.join(clean_template_text(x) for x in obj if clean_template_text(x))
    for key in ('body','value','text','message','sms_template','template','content'):
        if key in obj and obj[key] is not None:
            return clean_template_text(obj[key])
    return ''

def hash_password(password: str, salt: str|None=None, iterations: int=260000) -> str:
    salt = salt or secrets.token_hex(16)
    iterations = max(120000, int(iterations or 260000))
    dk = hashlib.pbkdf2_hmac('sha256', password.encode('utf-8'), salt.encode('utf-8'), iterations)
    return f'pbkdf2_sha256${iterations}${salt}${dk.hex()}'

def verify_password(password: str, stored: str) -> bool:
    try:
        stored = str(stored or '')
        # RC11 방식: 알고리즘$반복횟수$salt$digest
        if stored.startswith('pbkdf2_sha256$'):
            parts = stored.split('$')
            if len(parts) == 4:
                _, iterations, salt, digest = parts
                calc = hash_password(password, salt, int(iterations)).split('$', 3)[3]
                return hmac.compare_digest(calc, digest)
            # RC10 이하 방식: 알고리즘$salt$digest (120,000회)
            if len(parts) == 3:
                _, salt, digest = parts
                dk = hashlib.pbkdf2_hmac('sha256', password.encode('utf-8'), salt.encode('utf-8'), 120000).hex()
                return hmac.compare_digest(dk, digest)
        # 구버전 호환: salt$digest
        if '$' in stored:
            salt, digest = stored.split('$', 1)
            dk = hashlib.pbkdf2_hmac('sha256', password.encode('utf-8'), salt.encode('utf-8'), 120000).hex()
            return hmac.compare_digest(dk, digest)
        # 아주 초기 DB만 1회 로그인 허용하며 성공 즉시 RC11 해시로 교체합니다.
        return bool(stored) and hmac.compare_digest(stored, password)
    except Exception:
        return False

def password_needs_rehash(stored: str) -> bool:
    try:
        parts = str(stored or '').split('$')
        return not (len(parts) == 4 and parts[0] == 'pbkdf2_sha256' and int(parts[1]) >= 260000)
    except Exception:
        return True

def validate_password_strength(password: str):
    value = str(password or '')
    if len(value) < 10:
        raise HTTPException(400, '비밀번호는 10자 이상으로 설정해주세요.')
    groups = sum(bool(re.search(pattern, value)) for pattern in (r'[A-Z]', r'[a-z]', r'\d', r'[^A-Za-z0-9]'))
    if groups < 3:
        raise HTTPException(400, '비밀번호에는 영문 대·소문자, 숫자, 특수문자 중 3종류 이상을 사용해주세요.')

def parse_nums(value):
    if value is None: return []
    if isinstance(value, (list, tuple)):
        nums = [int(x) for x in value if 1 <= int(x) <= 45]
    else:
        text = str(value)
        try:
            data = json.loads(text)
            if isinstance(data, list):
                nums = [int(x) for x in data if 1 <= int(x) <= 45]
            else:
                nums = [int(x) for x in re.findall(r'\d+', text) if 1 <= int(x) <= 45]
        except Exception:
            nums = [int(x) for x in re.findall(r'\d+', text) if 1 <= int(x) <= 45]
    return sorted(list(dict.fromkeys(nums)))[:6]

def table_cols(c, table):
    try: return [r[1] for r in c.execute(f'PRAGMA table_info({table})').fetchall()]
    except Exception: return []


# === RC3-15: 당첨번호 회차 무결성 보조 ===
def _rc315_expected_round_and_completed(now_dt=None):
    """현재 한국시간 기준 추천 회차와 실제 추첨 완료 회차를 분리합니다.
    예: 1232회 추천 기간이어도 추첨 전이면 완료 회차는 1231회입니다.
    """
    try:
        dt = now_dt or (datetime.datetime.utcnow() + datetime.timedelta(hours=9))
        first = datetime.date(2002, 12, 7)
        today = dt.date() if isinstance(dt, datetime.datetime) else dt
        expected = int(((today - first).days // 7) + 1) if today >= first else 1
        draw_dt = datetime.datetime.combine(first + datetime.timedelta(days=(expected-1)*7), datetime.time(20, 35))
        completed = expected if dt >= draw_dt else max(1, expected - 1)
        return expected, completed
    except Exception:
        return 0, 0


def _rc315_clean_future_draws_in_conn(c):
    """추첨 전/미래 회차에 잘못 저장된 당첨번호를 제거합니다.
    RC3-15 핵심: 추천 회차(다음 회차)와 당첨 확인 회차(최근 추첨 완료 회차)를 분리합니다.
    """
    expected, completed = _rc315_expected_round_and_completed()
    removed = 0
    suspicious = []
    try:
        rows = c.execute('SELECT round_no,numbers,bonus,source,updated_at FROM draws WHERE round_no>? ORDER BY round_no DESC', (completed,)).fetchall()
        for row in rows:
            suspicious.append(dict(row) if hasattr(row, 'keys') else {'round_no': row[0]})
        c.execute('DELETE FROM draws WHERE round_no>?', (completed,))
        removed = int(getattr(c, 'rowcount', 0) or 0)
    except Exception:
        removed = 0
    return {'expected_round': expected, 'completed_round': completed, 'removed': removed, 'suspicious': suspicious}


def _rc315_validate_draw_payload(round_no, numbers, bonus, allow_completed_only=True):
    nums = parse_nums(numbers)
    if len(nums) != 6:
        raise HTTPException(400, '당첨번호 6개가 필요합니다.')
    if len(set(nums)) != 6 or any(n < 1 or n > 45 for n in nums):
        raise HTTPException(400, '당첨번호는 1~45 사이의 서로 다른 6개 숫자여야 합니다.')
    try:
        b = int(bonus)
    except Exception:
        b = 0
    if not (1 <= b <= 45):
        raise HTTPException(400, '보너스 번호는 1~45 사이여야 합니다.')
    if b in nums:
        raise HTTPException(400, '보너스 번호는 당첨번호 6개와 달라야 합니다.')
    expected, completed = _rc315_expected_round_and_completed()
    r = int(round_no or 0)
    if allow_completed_only and completed and r > completed:
        raise HTTPException(400, f'{r}회는 아직 추첨 완료 전입니다. 현재 저장 가능한 최신 당첨번호는 {completed}회입니다.')
    return r, nums, b, expected, completed

def init_db():
    with con() as c:
        c.execute('CREATE TABLE IF NOT EXISTS admins(id INTEGER PRIMARY KEY AUTOINCREMENT, username TEXT UNIQUE NOT NULL, name TEXT DEFAULT "관리자", password_hash TEXT NOT NULL, is_active INTEGER DEFAULT 1, created_at TEXT, last_login_at TEXT)')
        # V34 Priority4: 관리자 운영 컬럼을 기존 DB에 안전하게 추가
        admin_cols = table_cols(c, 'admins')
        for col, ddl in {
            'name':'TEXT DEFAULT "관리자"',
            'is_active':'INTEGER DEFAULT 1',
            'last_login_at':'TEXT DEFAULT ""',
            'role':'TEXT DEFAULT "전체권한"',
            'memo':'TEXT DEFAULT ""',
            'updated_at':'TEXT DEFAULT ""',
            'last_ip':'TEXT DEFAULT ""',
            'phone':'TEXT DEFAULT ""'
        }.items():
            if col not in admin_cols:
                c.execute(f'ALTER TABLE admins ADD COLUMN {col} {ddl}')
        c.execute('UPDATE admins SET name=COALESCE(NULLIF(name,""),"관리자"), is_active=COALESCE(is_active,1)')
        c.execute('CREATE TABLE IF NOT EXISTS sessions(token TEXT PRIMARY KEY, admin_id INTEGER, created_at TEXT, expires_at TEXT)')
        session_cols = table_cols(c, 'sessions')
        for col, ddl in {
            'last_seen_at':'TEXT DEFAULT ""',
            'ip':'TEXT DEFAULT ""',
            'user_agent':'TEXT DEFAULT ""'
        }.items():
            if col not in session_cols:
                c.execute(f'ALTER TABLE sessions ADD COLUMN {col} {ddl}')
        c.execute('CREATE TABLE IF NOT EXISTS admin_logs(id INTEGER PRIMARY KEY AUTOINCREMENT, admin_id INTEGER, username TEXT, action TEXT, detail TEXT, ip TEXT, created_at TEXT)')
        # RC3-3: 로그인 이력은 관리자 활동 로그와 별도 테이블로 영구 저장합니다.
        c.execute('CREATE TABLE IF NOT EXISTS login_logs(id INTEGER PRIMARY KEY AUTOINCREMENT, admin_id INTEGER DEFAULT 0, username TEXT DEFAULT "", success INTEGER DEFAULT 0, ip TEXT DEFAULT "", user_agent TEXT DEFAULT "", message TEXT DEFAULT "", created_at TEXT)')
        c.execute('CREATE TABLE IF NOT EXISTS backup_history(id INTEGER PRIMARY KEY AUTOINCREMENT, filename TEXT, reason TEXT DEFAULT "manual", size_bytes INTEGER DEFAULT 0, created_by INTEGER, created_at TEXT)')
        c.execute('CREATE TABLE IF NOT EXISTS members(id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT NOT NULL, phone TEXT DEFAULT "", grade TEXT DEFAULT "일반", memo TEXT DEFAULT "", created_by INTEGER, created_at TEXT)')
        c.execute("CREATE TABLE IF NOT EXISTS member_notes(id INTEGER PRIMARY KEY AUTOINCREMENT, member_id INTEGER, note TEXT DEFAULT '', note_type TEXT DEFAULT '상담', created_by INTEGER DEFAULT 0, created_by_name TEXT DEFAULT '', created_at TEXT)")
        # V34 Final Member Phase3: 기존 DB와 충돌 없이 필요한 컬럼만 추가
        member_cols = table_cols(c, 'members')
        if 'status' not in member_cols:
            c.execute('ALTER TABLE members ADD COLUMN status TEXT DEFAULT "활성"')
        if 'last_contact_at' not in member_cols:
            c.execute('ALTER TABLE members ADD COLUMN last_contact_at TEXT DEFAULT ""')
        if 'updated_at' not in member_cols:
            c.execute('ALTER TABLE members ADD COLUMN updated_at TEXT DEFAULT ""')
        if 'priority' not in member_cols:
            c.execute('ALTER TABLE members ADD COLUMN priority TEXT DEFAULT "보통"')
        if 'source' not in member_cols:
            c.execute('ALTER TABLE members ADD COLUMN source TEXT DEFAULT "직접등록"')
        if 'created_by' not in member_cols:
            c.execute('ALTER TABLE members ADD COLUMN created_by INTEGER DEFAULT 0')
        if 'preferred_count' not in member_cols:
            c.execute('ALTER TABLE members ADD COLUMN preferred_count INTEGER DEFAULT 10')
        if 'contract_end_at' not in member_cols:
            c.execute('ALTER TABLE members ADD COLUMN contract_end_at TEXT DEFAULT ""')
        if 'contract_months' not in member_cols:
            c.execute('ALTER TABLE members ADD COLUMN contract_months INTEGER DEFAULT 12')
        c.execute('CREATE TABLE IF NOT EXISTS recommendations(id INTEGER PRIMARY KEY AUTOINCREMENT, member_id INTEGER, member_name TEXT, round_no INTEGER, mode TEXT, count INTEGER, numbers TEXT, analysis TEXT, sms TEXT, created_by INTEGER, created_at TEXT)')
        # V40 Phase1: 기존 Render SQLite가 오래된 스키마여도 자동 복구합니다.
        rec_cols = table_cols(c, 'recommendations')
        for col, ddl in {
            'member_id':'INTEGER',
            'member_name':'TEXT DEFAULT ""',
            'round_no':'INTEGER DEFAULT 0',
            'mode':'TEXT DEFAULT "balanced"',
            'count':'INTEGER DEFAULT 0',
            'numbers':'TEXT DEFAULT "[]"',
            'analysis':'TEXT DEFAULT ""',
            'sms':'TEXT DEFAULT ""',
            'created_by':'INTEGER DEFAULT 0',
            'created_at':'TEXT DEFAULT ""',
            'avg_score':'REAL DEFAULT 0',
            'engine_json':'TEXT DEFAULT "{}"',
            'details_json':'TEXT DEFAULT "[]"',
            'explicit_saved':'INTEGER DEFAULT 0'
        }.items():
            if col not in rec_cols:
                c.execute(f'ALTER TABLE recommendations ADD COLUMN {col} {ddl}')
        c.execute('CREATE TABLE IF NOT EXISTS sms_logs(id INTEGER PRIMARY KEY AUTOINCREMENT, member_id INTEGER, member_name TEXT, round_no INTEGER, body TEXT, combos TEXT DEFAULT "[]", created_by INTEGER, created_at TEXT)')
        sms_cols = table_cols(c, 'sms_logs')
        for col, ddl in {
            'phone':'TEXT DEFAULT ""',
            'provider':'TEXT DEFAULT "mock"',
            'status':'TEXT DEFAULT "saved"',
            'result_message':'TEXT DEFAULT ""',
            'sent_at':'TEXT DEFAULT ""'
        }.items():
            if col not in sms_cols:
                c.execute(f'ALTER TABLE sms_logs ADD COLUMN {col} {ddl}')
        c.execute('CREATE TABLE IF NOT EXISTS winning_checks(id INTEGER PRIMARY KEY AUTOINCREMENT, member_id INTEGER, member_name TEXT, round_no INTEGER, target_numbers TEXT, win_numbers TEXT, bonus INTEGER, match_count INTEGER, bonus_match INTEGER, rank TEXT, prize INTEGER DEFAULT 0, cost INTEGER DEFAULT 1000, profit INTEGER DEFAULT 0, roi REAL DEFAULT 0, created_by INTEGER, created_at TEXT)')
        c.execute('CREATE TABLE IF NOT EXISTS draws(round_no INTEGER PRIMARY KEY, draw_date TEXT DEFAULT "", numbers TEXT, bonus INTEGER, source TEXT DEFAULT "manual", updated_at TEXT)')
        draw_cols = table_cols(c, 'draws')
        if 'source' not in draw_cols:
            c.execute('ALTER TABLE draws ADD COLUMN source TEXT DEFAULT "manual"')
        if 'updated_at' not in draw_cols:
            c.execute('ALTER TABLE draws ADD COLUMN updated_at TEXT DEFAULT ""')
        c.execute('CREATE TABLE IF NOT EXISTS settings(key TEXT PRIMARY KEY, value TEXT, updated_at TEXT)')
        if c.execute('SELECT COUNT(*) FROM admins').fetchone()[0] == 0:
            init_username = os.getenv('BBLOTTO_ADMIN_USERNAME', 'admin').strip() or 'admin'
            init_password = os.getenv('BBLOTTO_ADMIN_PASSWORD', '').strip()
            if not init_password:
                init_password = secrets.token_urlsafe(18)
                print(f'[BBLOTTO][SECURITY] 초기 관리자 비밀번호가 자동 생성되었습니다. Railway/서버 로그에서 확인 후 즉시 변경하세요. username={init_username} password={init_password}')
            c.execute('INSERT INTO admins(username,name,password_hash,is_active,created_at) VALUES(?,?,?,?,?)', (init_username,'대표 관리자',hash_password(init_password),1,now()))
        # 기본 당첨번호 DB는 비어 있을 때만 넣는 방식이 아니라, 누락 회차를 항상 보강합니다.
        # 그래서 기존 DB에 20회차만 남아 있어도 최근 100회 통계가 바로 동작합니다.
        for r,d,n,b in DEFAULT_DRAWS:
            c.execute('INSERT OR IGNORE INTO draws(round_no,draw_date,numbers,bonus,source,updated_at) VALUES(?,?,?,?,?,?)', (r,d,json.dumps(n),b,'starter100',now()))
        default_sms = '안녕하세요 {회원명}님, BBLOTTO입니다.\n{회차}회차 추천번호 및 이번회차 분석입니다.\n\n[이번회차 핵심 분석]\n{분석}\n\n[추천번호]\n{추천번호}\n\n좋은 결과 있으시길 바랍니다.'
        c.execute('INSERT OR IGNORE INTO settings(key,value,updated_at) VALUES(?,?,?)', ('sms_template', default_sms, now()))
        current_sms = c.execute('SELECT value FROM settings WHERE key=?', ('sms_template',)).fetchone()
        if current_sms:
            cleaned = clean_template_text(current_sms['value'])
            if cleaned and cleaned != current_sms['value']:
                c.execute('UPDATE settings SET value=?, updated_at=? WHERE key=?', (cleaned, now(), 'sms_template'))
        c.execute('INSERT OR IGNORE INTO settings(key,value,updated_at) VALUES(?,?,?)', ('sms_provider', 'mock', now()))
        c.execute('INSERT OR IGNORE INTO settings(key,value,updated_at) VALUES(?,?,?)', ('sms_sender', '', now()))
        c.execute('INSERT OR IGNORE INTO settings(key,value,updated_at) VALUES(?,?,?)', ('sms_api_url', '', now()))
        c.execute('INSERT OR IGNORE INTO settings(key,value,updated_at) VALUES(?,?,?)', ('sms_api_key', '', now()))
        c.execute('INSERT OR IGNORE INTO settings(key,value,updated_at) VALUES(?,?,?)', ('session_timeout_minutes', '600', now()))
        # PHASE25: 기존 DB에 480분 이하로 남아 있던 자동 로그아웃 값을 10시간(600분)으로 보정
        try:
            old_timeout = c.execute('SELECT value FROM settings WHERE key=?', ('session_timeout_minutes',)).fetchone()
            old_val = int((old_timeout or {'value':'600'})['value'] or 600)
            if old_val < 600:
                c.execute('UPDATE settings SET value=?, updated_at=? WHERE key=?', ('600', now(), 'session_timeout_minutes'))
        except Exception:
            c.execute('UPDATE settings SET value=?, updated_at=? WHERE key=?', ('600', now(), 'session_timeout_minutes'))
        c.execute('INSERT OR IGNORE INTO settings(key,value,updated_at) VALUES(?,?,?)', ('auto_logout_warning_minutes', '5', now()))
        # RC3-15: 운영 DB에 1232회처럼 추첨 전 회차가 잘못 저장되어 있으면 시작 시 자동 정리
        _rc315_clean_future_draws_in_conn(c)
        c.commit()

def log_action(admin, action, detail='', request: Request|None=None):
    ip = _rc113_client_ip(request) if request else ''
    with con() as c:
        c.execute('INSERT INTO admin_logs(admin_id,username,action,detail,ip,created_at) VALUES(?,?,?,?,?,?)', (admin.get('id') if admin else None, admin.get('username') if admin else '', action, detail, ip, now()))
        c.commit()

def log_login_event(admin_id=0, username='', success=0, message='', request: Request|None=None):
    """RC3-3: 로그인 성공/실패 이력을 PostgreSQL/SQLite에 공통 저장합니다."""
    ip = _rc113_client_ip(request) if request else ''
    ua = request.headers.get('user-agent','')[:240] if request else ''
    try:
        with con() as c:
            c.execute('INSERT INTO login_logs(admin_id,username,success,ip,user_agent,message,created_at) VALUES(?,?,?,?,?,?,?)', (int(admin_id or 0), username or '', 1 if success else 0, ip, ua, message or '', now()))
            c.commit()
    except Exception:
        # 로그인 자체가 실패하지 않도록 로그 저장 오류는 무시합니다.
        pass


# RC3-4: PostgreSQL/SQLite 공통 운영 백업 테이블 목록
RC3_BACKUP_TABLES = [
    'admins', 'members', 'settings', 'draws', 'recommendations', 'winning_checks',
    'sms_logs', 'admin_logs', 'login_logs', 'backup_history', 'sessions',
    'engine_runs', 'dashboard_snapshots'
]


def _safe_backup_filename(ext='json', label='BBLOTTO_RC3_BACKUP'):
    stamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    ext = str(ext or 'json').lstrip('.')
    return f'{label}_{stamp}.{ext}'


def _json_default(value):
    try:
        if isinstance(value, (datetime.datetime, datetime.date)):
            return value.isoformat()
    except Exception:
        pass
    return str(value)


def _backup_export_json(reason='manual', admin=None):
    """PostgreSQL/SQLite 공통 JSON 백업 생성. DB 엔진과 무관하게 모든 운영 테이블을 보관합니다."""
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    filename = _safe_backup_filename('json')
    dest = EXPORT_DIR / filename
    payload = {
        'app': 'BBLOTTO PRO',
        'version': 'RC3-4',
        'engine': DB_ENGINE,
        'created_at': now(),
        'reason': reason,
        'tables': {},
    }
    with con() as c:
        for table in RC3_BACKUP_TABLES:
            try:
                rows = c.execute(f'SELECT * FROM {table}').fetchall()
                payload['tables'][table] = [dict(r) for r in rows]
            except Exception:
                payload['tables'][table] = []
    dest.write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=_json_default), encoding='utf-8')
    size = dest.stat().st_size if dest.exists() else 0
    try:
        with con() as c:
            c.execute('INSERT INTO backup_history(filename,reason,size_bytes,created_by,created_at) VALUES(?,?,?,?,?)', (filename, reason, size, admin.get('id') if admin else None, now()))
            c.commit()
    except Exception:
        pass
    return {'filename': filename, 'format': 'json', 'engine': DB_ENGINE, 'reason': reason, 'size_bytes': size, 'created_at': now()}


def create_db_backup(reason='manual', admin=None):
    """RC3-4 운영 백업.
    - PostgreSQL: JSON 덤프 파일 생성
    - SQLite: 기존 .db 복사 + JSON 백업 생성 가능 구조 유지
    """
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    if DB_ENGINE == 'postgresql':
        return _backup_export_json(reason, admin)
    stamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'BBLOTTO_V34_BACKUP_{stamp}.db'
    dest = EXPORT_DIR / filename
    shutil.copy2(DB, dest)
    size = dest.stat().st_size if dest.exists() else 0
    try:
        with con() as c:
            c.execute('INSERT INTO backup_history(filename,reason,size_bytes,created_by,created_at) VALUES(?,?,?,?,?)', (filename, reason, size, admin.get('id') if admin else None, now()))
            c.commit()
    except Exception:
        pass
    return {'filename': filename, 'format': 'sqlite_db', 'engine': DB_ENGINE, 'reason': reason, 'size_bytes': size, 'created_at': now()}


def _validate_backup_json(path: Path):
    if not path.exists() or path.suffix.lower() != '.json':
        raise HTTPException(400, 'JSON 백업 파일만 복원할 수 있습니다.')
    try:
        data = json.loads(path.read_text(encoding='utf-8'))
    except Exception:
        raise HTTPException(400, '백업 JSON 파일을 읽을 수 없습니다.')
    if not isinstance(data, dict) or not isinstance(data.get('tables'), dict):
        raise HTTPException(400, 'BBLOTTO 백업 형식이 아닙니다.')
    return data


def _restore_json_backup(path: Path, admin=None, request: Request|None=None):
    data = _validate_backup_json(path)
    tables = data.get('tables') or {}
    restore_order = [t for t in reversed(RC3_BACKUP_TABLES) if t in tables]
    inserted = {}
    with con() as c:
        for table in restore_order:
            try:
                c.execute(f'DELETE FROM {table}')
            except Exception:
                pass
        for table in RC3_BACKUP_TABLES:
            rows = tables.get(table) or []
            if not rows:
                inserted[table] = 0
                continue
            cols = table_cols(c, table)
            allowed = [col for col in rows[0].keys() if col in cols]
            if not allowed:
                inserted[table] = 0
                continue
            marks = ','.join(['?'] * len(allowed))
            col_sql = ','.join(allowed)
            count = 0
            for row in rows:
                vals = [row.get(col) for col in allowed]
                try:
                    c.execute(f'INSERT INTO {table}({col_sql}) VALUES({marks})', vals)
                    count += 1
                except Exception:
                    pass
            inserted[table] = count
        c.commit()
    try:
        log_action(admin or {}, 'RESTORE_BACKUP', f'복원 완료: {path.name}', request)
    except Exception:
        pass
    return {'ok': True, 'restored_from': path.name, 'engine': DB_ENGINE, 'inserted': inserted, 'restored_at': now()}

def ensure_daily_backup():
    """하루 1회 자동 백업. 실행 시 충돌 없이 보관만 합니다."""
    today = datetime.datetime.now().strftime('%Y-%m-%d')
    try:
        with con() as c:
            r = c.execute('SELECT COUNT(*) c FROM backup_history WHERE reason=? AND created_at LIKE ?', ('auto_daily', today+'%')).fetchone()
        if not r or r['c'] == 0:
            create_db_backup('auto_daily', None)
    except Exception:
        pass

init_db()
ensure_daily_backup()

def current_admin(authorization: str|None):
    if not authorization or not authorization.lower().startswith('bearer '):
        raise HTTPException(401, '로그인이 필요합니다.')
    token = authorization.split(' ',1)[1].strip()
    with con() as c:
        row = c.execute('SELECT s.token,s.expires_at,a.* FROM sessions s JOIN admins a ON a.id=s.admin_id WHERE s.token=? AND a.is_active=1', (token,)).fetchone()
        if not row: raise HTTPException(401, '세션이 만료되었습니다.')
        if row['expires_at'] < now():
            c.execute('DELETE FROM sessions WHERE token=?', (token,)); c.commit()
            raise HTTPException(401, '세션이 만료되었습니다.')
        try:
            c.execute('UPDATE sessions SET last_seen_at=? WHERE token=?', (now(), token)); c.commit()
        except Exception:
            pass
        return dict(row)

def require_admin(authorization: str|None = Header(default=None)):
    return current_admin(authorization)

def require_admin_any(authorization: str|None = None, token: str|None = None):
    if authorization:
        return current_admin(authorization)
    if token:
        return current_admin('Bearer ' + token)
    raise HTTPException(401, '로그인이 필요합니다.')

def is_super_admin(admin: dict|sqlite3.Row|None):
    if not admin:
        return False
    if hasattr(admin, 'get'):
        username = str(admin.get('username','') or '').strip().lower()
        role = str(admin.get('role','') or '').replace(' ', '').lower()
        name = str(admin.get('name','') or '').replace(' ', '').lower()
    else:
        keys = admin.keys()
        username = str(admin['username'] if 'username' in keys else '').strip().lower()
        role = str(admin['role'] if 'role' in keys else '').replace(' ', '').lower()
        name = str(admin['name'] if 'name' in keys else '').replace(' ', '').lower()
    return (username == 'admin' or '최고관리자' in role or '대표관리자' in role or '전체권한' in role or role in {'대표','최고','전체','전체권한'} or 'super' in role or 'owner' in role or '최고관리자' in name or '대표관리자' in name or '전체권한' in name or name in {'대표','최고','전체','전체권한'})

def require_super_admin(admin):
    if not is_super_admin(admin):
        raise HTTPException(403, '최고 관리자만 처리할 수 있습니다.')
    return admin


def member_scope_condition(admin, table_alias: str = ''):
    """RC4-4: 최고관리자는 전체 회원, 일반관리자는 본인이 등록한 회원만 조회/관리합니다."""
    if is_super_admin(admin):
        return '', []
    prefix = (table_alias + '.') if table_alias else ''
    return f'COALESCE({prefix}created_by,0)=?', [int(admin.get('id') or 0)]

def assert_member_access(c, admin, member_id: int):
    """회원 단건 작업 권한 확인."""
    if is_super_admin(admin):
        row = c.execute('SELECT id,name,created_by FROM members WHERE id=?', (member_id,)).fetchone()
    else:
        row = c.execute('SELECT id,name,created_by FROM members WHERE id=? AND COALESCE(created_by,0)=?', (member_id, int(admin.get('id') or 0))).fetchone()
    if not row:
        raise HTTPException(404, '관리 가능한 회원을 찾을 수 없습니다.')
    return row


# ===== PHASE19: 회차 자동 확인 / 속도 최적화 보조 =====
LOTTO_FIRST_DRAW_DATE = datetime.date(2002, 12, 7)  # 1회 추첨일
LOTTO_DRAW_WEEKDAY = 5  # Saturday
LOTTO_DRAW_HOUR = 20
LOTTO_DRAW_MINUTE = 35

def kst_now():
    """서버 위치와 무관하게 한국 시간 기준으로 회차를 계산합니다."""
    return datetime.datetime.utcnow() + datetime.timedelta(hours=9)

def expected_lotto_round(dt=None):
    """오늘 기준으로 관리해야 할 로또 회차를 계산합니다.
    토요일 추첨 전에도 오늘 추첨될 회차를 반환하므로 추천번호/당첨확인 회차가 밀리지 않습니다.
    """
    dt = dt or kst_now()
    today = dt.date() if isinstance(dt, datetime.datetime) else dt
    if today < LOTTO_FIRST_DRAW_DATE:
        return 1
    return int(((today - LOTTO_FIRST_DRAW_DATE).days // 7) + 1)

def draw_date_for_round(round_no:int):
    try:
        d = LOTTO_FIRST_DRAW_DATE + datetime.timedelta(days=(int(round_no)-1)*7)
        return d.strftime('%Y.%m.%d')
    except Exception:
        return ''

def draw_status_for_round(round_no:int):
    """scheduled: 추첨 전, pending: 추첨 후 번호 미저장, saved: DB 저장 완료"""
    now_kst = kst_now()
    r = int(round_no or 0)
    expected = expected_lotto_round(now_kst)
    draw_dt = datetime.datetime.combine(LOTTO_FIRST_DRAW_DATE + datetime.timedelta(days=(r-1)*7), datetime.time(LOTTO_DRAW_HOUR, LOTTO_DRAW_MINUTE)) if r > 0 else now_kst
    if r > expected:
        return 'future'
    if r == expected and now_kst < draw_dt:
        return 'scheduled'
    return 'pending'

# RC3-10: 동행복권 API가 일시 차단/지연될 때 운영이 멈추지 않도록
# 검증된 최신 회차를 보조 캐시로 둡니다. 최신 회차가 추가되면 여기만 보강해도 자동확인이 복구됩니다.
OFFICIAL_DRAW_FALLBACKS = {
    1231: {'round_no': 1231, 'draw_date': '2026.07.04', 'numbers': [4, 13, 14, 18, 31, 38], 'bonus': 15, 'source': 'official_cache'},
}


def _normalize_official_payload(data, round_no:int, source:str):
    if not data or data.get('returnValue') != 'success':
        return None
    nums = [int(data.get(f'drwtNo{i}', 0) or 0) for i in range(1, 7)]
    bonus = int(data.get('bnusNo', 0) or 0)
    if len(set(nums)) == 6 and all(1 <= n <= 45 for n in nums) and 1 <= bonus <= 45 and bonus not in nums:
        return {
            'round_no': int(data.get('drwNo') or round_no),
            'draw_date': str(data.get('drwNoDate') or draw_date_for_round(round_no)).replace('-', '.'),
            'numbers': sorted(nums),
            'bonus': bonus,
            'source': source,
        }
    return None


def _fallback_draw(round_no:int):
    cached = OFFICIAL_DRAW_FALLBACKS.get(int(round_no))
    if not cached:
        return None
    # 추첨 전 회차에 캐시가 잘못 적용되는 것을 방지합니다.
    if draw_status_for_round(int(round_no)) == 'scheduled':
        return None
    return dict(cached)


def fetch_official_lotto(round_no:int):
    """동행복권 공개 조회 JSON을 안전하게 시도합니다.
    RC3-10: 공식 API 실패 시 검증 캐시로 한 번 더 복구합니다.
    """
    r = int(round_no)
    urls = [
        'https://www.dhlottery.co.kr/common.do?method=getLottoNumber&drwNo=' + urllib.parse.quote(str(r)),
        'http://www.dhlottery.co.kr/common.do?method=getLottoNumber&drwNo=' + urllib.parse.quote(str(r)),
    ]
    last_error = ''
    for idx, url in enumerate(urls):
        try:
            req = urllib.request.Request(url, headers={
                'User-Agent': 'Mozilla/5.0 (compatible; BBLOTTO-RC3-10/1.0)',
                'Accept': 'application/json,text/plain,*/*',
                'Referer': 'https://www.dhlottery.co.kr/',
            })
            with urllib.request.urlopen(req, timeout=7) as resp:
                data = json.loads(resp.read().decode('utf-8', errors='ignore'))
            normalized = _normalize_official_payload(data, r, 'dhlottery' if idx == 0 else 'dhlottery_http')
            if normalized:
                return normalized
            last_error = '공식 API 응답에 당첨번호가 없습니다.'
        except Exception as e:
            last_error = str(e)[:180]
            continue
    cached = _fallback_draw(r)
    if cached:
        cached['fetch_error'] = last_error
        return cached
    return None

def save_draw_if_missing(draw):
    if not draw:
        return None
    try:
        r, nums, bonus, expected, completed = _rc315_validate_draw_payload(draw.get('round_no'), draw.get('numbers', []), draw.get('bonus', 0), allow_completed_only=True)
        draw['round_no'] = r
        draw['numbers'] = nums
        draw['bonus'] = bonus
        with con() as c:
            old = c.execute('SELECT round_no FROM draws WHERE round_no=?', (r,)).fetchone()
            if not old:
                c.execute('INSERT OR REPLACE INTO draws(round_no,draw_date,numbers,bonus,source,updated_at) VALUES(?,?,?,?,?,?)', (r, draw.get('draw_date',''), json.dumps(nums), bonus, draw.get('source','official'), now()))
                c.commit()
        return draw
    except HTTPException:
        return None
    except Exception:
        return None

def get_draw(round_no:int):
    with con() as c:
        r = c.execute('SELECT * FROM draws WHERE round_no=?', (round_no,)).fetchone()
    if not r: return None
    return {'round_no': r['round_no'], 'draw_date': r['draw_date'], 'numbers': parse_nums(r['numbers']), 'bonus': r['bonus']}

def resolve_draw_for_check(round_no:int|None=None, allow_fetch:bool=True):
    """당첨확인 화면용 회차/당첨번호를 자동 판정합니다."""
    expected = expected_lotto_round()
    r = int(round_no or expected)
    if r <= 0:
        r = expected
    status = draw_status_for_round(r)
    draw = get_draw(r)
    auto_fetched = False
    if not draw and allow_fetch and status == 'pending':
        fetched = fetch_official_lotto(r)
        if fetched:
            draw = save_draw_if_missing(fetched)
            auto_fetched = True
    if draw:
        return {
            'round_no': int(r), 'expected_round': int(expected), 'draw_date': draw.get('draw_date') or draw_date_for_round(r),
            'status': 'saved', 'can_check': True, 'numbers': draw.get('numbers') or [], 'bonus': int(draw.get('bonus') or 0),
            'source': draw.get('source','db'), 'auto_fetched': auto_fetched,
            'message': f'{r}회 당첨번호가 자동 확인되었습니다.' + ((' 공식 API/보조 캐시로 저장했습니다.' if draw.get('source') == 'official_cache' else ' 동행복권 공개 데이터로 저장했습니다.') if auto_fetched else '')
        }
    if status == 'scheduled':
        return {
            'round_no': int(r), 'expected_round': int(expected), 'draw_date': draw_date_for_round(r),
            'status': 'scheduled', 'can_check': False, 'numbers': [], 'bonus': 0, 'source': '', 'auto_fetched': False,
            'message': f'{r}회는 오늘 추첨 예정 회차입니다. 추첨 전이라 당첨번호는 아직 자동 입력하지 않습니다.'
        }
    if status == 'future':
        return {
            'round_no': int(r), 'expected_round': int(expected), 'draw_date': draw_date_for_round(r),
            'status': 'future', 'can_check': False, 'numbers': [], 'bonus': 0, 'source': '', 'auto_fetched': False,
            'message': f'{r}회는 아직 추첨 예정 전 회차입니다.'
        }
    return {
        'round_no': int(r), 'expected_round': int(expected), 'draw_date': draw_date_for_round(r),
        'status': 'pending', 'can_check': False, 'numbers': [], 'bonus': 0, 'source': '', 'auto_fetched': False,
        'message': f'{r}회 추첨 시간이 지났지만 당첨번호가 아직 공개/저장되지 않았습니다. 잠시 후 다시 확인하거나 수동 입력하세요.'
    }

def rank_result(combo, wins, bonus):
    combo = parse_nums(combo); wins = parse_nums(wins)
    matched = sorted(set(combo)&set(wins)); m=len(matched); bm = int(bonus in combo)
    rank = '낙첨'
    if m == 6: rank='1등'
    elif m == 5 and bm: rank='2등'
    elif m == 5: rank='3등'
    elif m == 4: rank='4등'
    elif m == 3: rank='5등'
    prize = PRIZE_TABLE[rank]; cost = COST_PER_COMBO; profit = prize-cost; roi = round((profit/cost)*100, 2) if cost else 0
    return {'combo':combo,'matched':matched,'match_count':m,'bonus_match':bool(bm),'rank':rank,'prize':prize,'cost':cost,'profit':profit,'roi':roi}

def all_draw_nums(limit=100):
    with con() as c:
        rows = c.execute('SELECT * FROM draws ORDER BY round_no DESC LIMIT ?', (limit,)).fetchall()
    nums=[]
    for r in rows: nums.extend(parse_nums(r['numbers']))
    return nums

def make_combos(count=10, fixed='', excluded='', mode='balanced'):
    nums = all_draw_nums(100); freq={n:nums.count(n) for n in range(1,46)}
    hot = sorted(range(1,46), key=lambda n:(-freq[n],n))[:15]
    cold = sorted(range(1,46), key=lambda n:(freq[n],n))[:15]
    fixed_set=set(parse_nums(fixed)); excluded_set=set(parse_nums(excluded))
    pool=[n for n in range(1,46) if n not in excluded_set and n not in fixed_set]
    combos=[]; tries=0
    while len(combos)<max(1,min(50,count)) and tries<5000:
        tries+=1
        combo=set(fixed_set)
        if mode=='aggressive': base=random.sample(hot, min(3,len(hot))) + random.sample(pool, min(10,len(pool)))
        elif mode=='conservative': base=random.sample(cold, min(2,len(cold))) + random.sample(pool, min(12,len(pool)))
        else: base=random.sample(hot, min(2,len(hot))) + random.sample(cold, min(2,len(cold))) + random.sample(pool, min(12,len(pool)))
        for n in base:
            if n not in excluded_set and len(combo)<6: combo.add(n)
        while len(combo)<6: combo.add(random.choice(pool))
        arr=sorted(combo)
        odd=sum(n%2 for n in arr)
        sections=[sum(n<=15 for n in arr),sum(16<=n<=30 for n in arr),sum(n>=31 for n in arr)]
        if odd in [2,3,4] and max(sections)<=4 and arr not in combos: combos.append(arr)
    return combos



def latest_stats(limit=100):
    """V40 Phase1 통계 엔진: 최근 10/30/50/100회 흐름, 미출현, 동반출현, 끝수/구간/홀짝/합계를 계산합니다."""
    with con() as c:
        rows = c.execute('SELECT * FROM draws ORDER BY round_no DESC LIMIT ?', (max(10, int(limit or 100)),)).fetchall()
    draws=[]
    for r in rows:
        nums=parse_nums(r['numbers'])
        if len(nums)==6:
            draws.append({'round_no':int(r['round_no']), 'draw_date':r['draw_date'] or '', 'numbers':nums, 'bonus':int(r['bonus'] or 0)})
    if not draws:
        draws=[{'round_no':r,'draw_date':d,'numbers':n,'bonus':b} for r,d,n,b in DEFAULT_DRAWS if len(n)==6]
    windows={10:draws[:10], 30:draws[:30], 50:draws[:50], 100:draws[:100]}
    freq_by_window={}
    for w,ds in windows.items():
        f={n:0 for n in range(1,46)}
        for d in ds:
            for n in d['numbers']:
                f[n]+=1
        freq_by_window[w]=f
    freq=freq_by_window[100]
    last_seen={n:999 for n in range(1,46)}
    for idx,d in enumerate(draws[:100]):
        for n in d['numbers']:
            if last_seen[n] == 999:
                last_seen[n]=idx
    pair_counts=collections.Counter()
    for d in draws[:100]:
        for a,b in itertools.combinations(d['numbers'],2):
            pair_counts[tuple(sorted((a,b)))] += 1
    end_counts={i:0 for i in range(10)}
    zone_counts={'1~15':0,'16~30':0,'31~45':0}
    odd_total=0; sums=[]
    for d in draws[:100]:
        sums.append(sum(d['numbers']))
        for n in d['numbers']:
            end_counts[n%10]+=1
            odd_total += n%2
            if n<=15: zone_counts['1~15']+=1
            elif n<=30: zone_counts['16~30']+=1
            else: zone_counts['31~45']+=1
    avg100=sum(freq.values())/45 if freq else 0
    hot=sorted(range(1,46), key=lambda n:(-(freq_by_window[30][n]*1.7 + freq_by_window[100][n]*0.7 + freq_by_window[10][n]*2.0), last_seen[n], n))[:12]
    cold=sorted(range(1,46), key=lambda n:(freq_by_window[100][n], -last_seen[n], n))[:12]
    overdue=sorted(range(1,46), key=lambda n:(last_seen[n], -freq_by_window[100][n], n), reverse=True)[:12]
    mid=sorted(range(1,46), key=lambda n:(abs(freq_by_window[100][n]-avg100), last_seen[n], n))[:15]
    top_pairs=[{'pair':list(p),'count':c} for p,c in pair_counts.most_common(12)]
    recent_numbers=set()
    for d in draws[:3]: recent_numbers.update(d['numbers'])
    return {
        'draws':draws,'freq':freq,'freq10':freq_by_window[10],'freq30':freq_by_window[30],'freq50':freq_by_window[50],'freq100':freq_by_window[100],
        'last_seen':last_seen,'hot':hot,'mid':mid,'cold':cold,'overdue':overdue,'top_pairs':top_pairs,'pair_counts':pair_counts,
        'end_counts':end_counts,'zone_counts':zone_counts,'odd_ratio':odd_total/(max(1,len(draws[:100])*6)),
        'sum_avg':round(sum(sums)/len(sums),1) if sums else 0,'recent_numbers':recent_numbers,
        'latest_round':draws[0]['round_no'] if draws else 1230
    }

def ac_value(combo):
    arr=sorted(combo)
    diffs={b-a for a,b in itertools.combinations(arr,2)}
    return max(0, len(diffs)-5)

def combo_score(combo, st):
    """0~100 AI 점수. 당첨 보장 점수가 아니라 통계 균형/분산 품질 점수입니다."""
    combo=sorted(parse_nums(combo));
    if len(combo)!=6: return 0
    f10,f30,f100=st['freq10'],st['freq30'],st['freq100']; last=st['last_seen']; pairs=st['pair_counts']
    total=sum(combo); odd=sum(n%2 for n in combo); even=6-odd
    zones=[sum(n<=15 for n in combo),sum(16<=n<=30 for n in combo),sum(n>=31 for n in combo)]
    cons=sum(1 for a,b in zip(combo,combo[1:]) if b-a==1)
    ends=len(set(n%10 for n in combo)); ac=ac_value(combo)
    # 기본 패턴 점수
    score=42.0
    score += {3:13,2:10,4:10,1:3,5:3,0:-8,6:-8}.get(odd,0)
    score += 13 if 105<=total<=175 else (8 if 95<=total<=190 else -10)
    score += 12 if max(zones)<=3 and min(zones)>=1 else (5 if max(zones)<=4 else -8)
    score += 7 if 5<=ac<=10 else (3 if 4<=ac<=11 else -4)
    score += 6 if ends>=5 else (3 if ends==4 else -4)
    score += 5 if cons<=1 else (-3 if cons==2 else -9)
    # 최근 흐름 점수: 과열수는 과하게 몰리지 않게, 중간/보강수를 같이 반영
    hot_hit=len(set(combo)&set(st['hot'][:10])); cold_hit=len(set(combo)&set(st['cold'][:10])); overdue_hit=len(set(combo)&set(st['overdue'][:10]))
    score += min(10, hot_hit*3.0) + min(7, cold_hit*2.1) + min(8, overdue_hit*2.2)
    if hot_hit>4: score -= 7
    if len(set(combo)&st['recent_numbers'])>=4: score -= 5
    # 동반출현은 1~3개 정도만 가산, 과도한 과거쌍 몰림은 감점
    pair_sum=sum(pairs.get(tuple(sorted((a,b))),0) for a,b in itertools.combinations(combo,2))
    strong_pairs=sum(1 for a,b in itertools.combinations(combo,2) if pairs.get(tuple(sorted((a,b))),0)>=4)
    score += min(10, pair_sum/4.0) + min(5, strong_pairs*1.5)
    if strong_pairs>5: score -= 5
    # 10/30/100 가중치 평균이 너무 한쪽으로 쏠리지 않게
    heat=sum(f10[n]*2.0 + f30[n]*1.1 + f100[n]*0.35 for n in combo)
    if 22 <= heat <= 55: score += 6
    elif heat > 70: score -= 8
    else: score += 2
    return round(max(35, min(99.7, score)), 1)

def tags_for_combo(combo, st):
    combo=sorted(combo); s=set(combo); tags=[]
    if len(s & set(st['hot'][:10]))>=2: tags.append('최근핵심')
    if len(s & set(st['overdue'][:10]))>=1: tags.append('미출현보강')
    if len(s & set(st['cold'][:10]))>=1: tags.append('저출현반등')
    if sum(n%2 for n in combo) in (2,3,4): tags.append('홀짝균형')
    zones=[sum(n<=15 for n in combo),sum(16<=n<=30 for n in combo),sum(n>=31 for n in combo)]
    if max(zones)<=3 and min(zones)>=1: tags.append('구간분산')
    if ac_value(combo) in range(5,11): tags.append('AC안정')
    pairs=st['pair_counts']
    if any(pairs.get(tuple(sorted((a,b))),0)>=4 for a,b in itertools.combinations(combo,2)): tags.append('동반출현')
    return tags[:5] or ['균형형']

def combo_detail(combo, st):
    combo=sorted(combo)
    odd=sum(n%2 for n in combo); zones=[sum(n<=15 for n in combo),sum(16<=n<=30 for n in combo),sum(n>=31 for n in combo)]
    pair_hits=[]
    for a,b in itertools.combinations(combo,2):
        cnt=st['pair_counts'].get(tuple(sorted((a,b))),0)
        if cnt>=3: pair_hits.append({'pair':[a,b], 'count':cnt})
    pair_hits=sorted(pair_hits, key=lambda x:-x['count'])[:3]
    return {'numbers':combo,'score':combo_score(combo,st),'tags':tags_for_combo(combo,st),'sum':sum(combo),'odd':odd,'even':6-odd,'zones':zones,'ac':ac_value(combo),'pair_hits':pair_hits}

def _weighted_pick(candidates, weights, k):
    picked=[]; pool=list(candidates); w=list(weights)
    for _ in range(min(k,len(pool))):
        total=sum(max(0.01,x) for x in w)
        r=random.random()*total; acc=0; idx=0
        for i,x in enumerate(w):
            acc += max(0.01,x)
            if acc>=r:
                idx=i; break
        picked.append(pool.pop(idx)); w.pop(idx)
    return picked

def make_premium_combos(count=10, fixed='', excluded='', mode='balanced'):
    st=latest_stats(120)
    fixed_set=set(parse_nums(fixed)); excluded_set=set(parse_nums(excluded))
    fixed_set={n for n in fixed_set if n not in excluded_set}
    if len(fixed_set)>6: fixed_set=set(sorted(fixed_set)[:6])
    pool=[n for n in range(1,46) if n not in excluded_set and n not in fixed_set]
    target=max(1,min(50,int(count or 10)))
    if len(pool)+len(fixed_set)<6:
        raise HTTPException(400, '고정수/제외수를 확인하세요. 선택 가능한 번호가 부족합니다.')
    past={tuple(d['numbers']) for d in st['draws']}
    f10,f30,f100,last=st['freq10'],st['freq30'],st['freq100'],st['last_seen']
    weights={}
    for n in pool:
        hot = f30[n]*1.4 + f10[n]*2.0 + f100[n]*0.35
        cold = max(0, 7-f100[n]) + min(8,last[n])*0.35
        mid = 6 - min(6, abs(f100[n] - (sum(f100.values())/45)))
        if mode=='aggressive': weights[n]=1 + hot*1.35 + cold*0.45 + mid*0.5
        elif mode=='conservative': weights[n]=1 + hot*0.65 + cold*1.25 + mid*0.9
        else: weights[n]=1 + hot*0.95 + cold*0.9 + mid*0.75
        # 직전 3회 과다 반영은 낮춤
        if n in st['recent_numbers']: weights[n]*=0.72
    candidates=[]; seen=set(); tries=0
    needed=max(target*75, 900)
    while len(candidates)<needed and tries<40000:
        tries+=1
        need=6-len(fixed_set)
        nums=set(fixed_set)
        picked=_weighted_pick(pool, [weights[n] for n in pool], need)
        nums.update(picked)
        arr=tuple(sorted(nums))
        if len(arr)!=6 or arr in seen or arr in past: continue
        odd=sum(n%2 for n in arr); total=sum(arr); zones=[sum(n<=15 for n in arr),sum(16<=n<=30 for n in arr),sum(n>=31 for n in arr)]
        cons=sum(1 for a,b in zip(arr,arr[1:]) if b-a==1)
        if odd not in (2,3,4): continue
        if not (90<=total<=195): continue
        if max(zones)>4 or 0 in zones: continue
        if cons>2: continue
        if len(set(n%10 for n in arr))<3: continue
        seen.add(arr)
        candidates.append((combo_score(arr,st), list(arr)))
    candidates=sorted(candidates, key=lambda x:(-x[0], x[1]))
    selected=[]
    for score, combo in candidates:
        # 최종 10개는 서로 너무 비슷하지 않게 분산
        if all(len(set(combo)&set(prev))<=4 for prev in selected):
            selected.append(combo)
        if len(selected)>=target: break
    if len(selected)<target:
        for score, combo in candidates:
            if combo not in selected:
                selected.append(combo)
            if len(selected)>=target: break
    details=[combo_detail(c, st) for c in selected[:target]]
    return selected[:target], details, st

def _engine_summary(details, st):
    if not details: return {'avg_score':0,'combo_count':0}
    scores=[d.get('score',0) for d in details]
    all_nums=[n for d in details for n in d.get('numbers',[])]
    freq=collections.Counter(all_nums)
    return {
        'avg_score':round(sum(scores)/len(scores),1),
        'max_score':max(scores),
        'combo_count':len(details),
        'latest_round':st.get('latest_round'),
        'hot':st.get('hot',[])[:8],
        'cold':st.get('cold',[])[:8],
        'overdue':st.get('overdue',[])[:8],
        'top_used':[n for n,_ in freq.most_common(8)],
        'top_pairs':st.get('top_pairs',[])[:5]
    }

def build_analysis_text(round_no, st, mode, fixed, excluded, details=None):
    """생성된 추천 조합과 실제 최근 100회 통계를 함께 사용해 매번 다른 4~5줄 분석을 만듭니다."""
    details=details or []
    engine=_engine_summary(details, st)
    top_scores=sorted(details, key=lambda x:-x.get('score',0))[:3]
    used_nums=engine.get('top_used',[])[:6]
    hot=[n for n in st.get('hot',[])[:10] if n in used_nums] or st.get('hot',[])[:4]
    overdue=[n for n in st.get('overdue',[])[:10] if n in used_nums] or st.get('overdue',[])[:4]
    cold=[n for n in st.get('cold',[])[:10] if n in used_nums] or st.get('cold',[])[:4]
    pair_text=', '.join([f"{p['pair'][0]}-{p['pair'][1]}({p['count']}회)" for p in st.get('top_pairs',[])[:3]]) or '동반출현 데이터 보강 중'
    best = top_scores[0] if top_scores else {}
    mode_name={'balanced':'균형형','conservative':'보강형','aggressive':'공격형'}.get(mode, mode or '균형형')
    variants=[
        f"{round_no}회차는 최근 10/30/100회 흐름을 합산해 {mode_name} 기준으로 재분석했습니다.",
        f"이번 생성은 직전 과열 흐름을 낮추고 최근 100회 누적 통계를 함께 반영했습니다.",
        f"{round_no}회차 분석은 출현빈도, 미출현 간격, 동반출현, AC값을 동시에 적용했습니다."
    ]
    line1=random.choice(variants)
    line2=random.choice([
        f"핵심 후보는 {', '.join(map(str,hot[:5]))}번이며, 미출현 보강 후보는 {', '.join(map(str,overdue[:5]))}번입니다.",
        f"최근 강한 흐름은 {', '.join(map(str,hot[:5]))}번, 저출현 반등 후보는 {', '.join(map(str,cold[:5]))}번 중심으로 잡았습니다.",
        f"번호 풀은 HOT {', '.join(map(str,hot[:4]))}번과 보강 {', '.join(map(str,overdue[:4]))}번을 섞어 구성했습니다."
    ])
    if best:
        line3=f"대표 조합은 합계 {best.get('sum')} / AC {best.get('ac')} / 홀짝 {best.get('odd')}:{best.get('even')} 기준이며 AI점수는 {best.get('score')}점입니다."
    else:
        line3=f"평균 AI점수는 {engine.get('avg_score',0)}점이며 과도한 중복과 구간 쏠림을 줄였습니다."
    line4=random.choice([
        f"동반출현 참고 흐름은 {pair_text}이며, 조합 간 중복을 낮춰 분산형으로 정리했습니다.",
        f"동반출현은 {pair_text} 흐름을 참고했고, 끝수와 구간이 한쪽으로 몰리지 않게 제한했습니다.",
        f"상위 동반 흐름({pair_text})은 참고만 하고, 동일 패턴 반복은 줄였습니다."
    ])
    line5=f"분석 갱신시각: {now()}"
    return '\n'.join([line1,line2,line3,line4,line5])

def build_sms(member_name, round_no, combos, analysis, details):
    name=member_name or '회원'
    best=sorted(details or [], key=lambda x:-x.get('score',0))[:1]
    best_line=''
    if best:
        b=best[0]
        best_line=f"대표 조합 포인트: 합계 {b.get('sum')} / AC {b.get('ac')} / 홀짝 {b.get('odd')}:{b.get('even')} / AI점수 {b.get('score')}점"
    return '\n'.join([
        f'안녕하세요 {name}님, BBLOTTO입니다.',
        f'{round_no}회차 추천번호와 이번 회차 분석을 안내드립니다.',
        '[추천번호]',
        *[f'{i+1}. '+', '.join(map(str,c)) for i,c in enumerate(combos)],
        '[분석요약]',
        analysis,
        best_line,
        '좋은 결과 있으시길 바랍니다.'
    ]).strip()


# RC3-7: 추천결과 화면/저장 안정화를 위한 상세정보 보강
def rc37_grade(score):
    try: s=float(score or 0)
    except Exception: s=0
    if s >= 97: return '1등'
    if s >= 94: return '2등'
    return '일반'

def rc37_star(score):
    try: s=float(score or 0)
    except Exception: s=0
    if s >= 95: return '★★★★★'
    if s >= 90: return '★★★★☆'
    if s >= 85: return '★★★★'
    return '★★★☆'

def rc37_enrich_details(combos, details):
    enriched=[]
    details=list(details or [])
    for i, combo in enumerate(combos or []):
        d=dict(details[i]) if i < len(details) and isinstance(details[i], dict) else {}
        score=d.get('score') or d.get('ai_score') or d.get('vip_score') or 0
        try: score=round(float(score),1)
        except Exception: score=0
        nums=[int(n) for n in combo]
        odd=sum(n%2 for n in nums)
        zones=[sum(n<=15 for n in nums), sum(16<=n<=30 for n in nums), sum(n>=31 for n in nums)]
        d.update({
            'rank': i+1,
            'score': score,
            'star': d.get('star') or rc37_star(score),
            'grade': d.get('grade') or rc37_grade(score),
            'sum': d.get('sum') or sum(nums),
            'odd': d.get('odd') if d.get('odd') is not None else odd,
            'even': d.get('even') if d.get('even') is not None else 6-odd,
            'zones': d.get('zones') or zones,
        })
        tags=list(d.get('tags') or d.get('reasons') or [])
        tags.append(d['grade'])
        tags.append(d['star'])
        d['tags']=list(dict.fromkeys(str(t) for t in tags if t))[:5]
        enriched.append(d)
    return enriched

def rc37_top3(combos, details):
    rows=[]
    for combo, d in sorted(zip(combos or [], details or []), key=lambda x: -float(x[1].get('score') or 0))[:3]:
        rows.append({'numbers': combo, 'score': d.get('score',0), 'star': d.get('star',''), 'grade': d.get('grade','STANDARD')})
    return rows


# RC3-8: 실사용 안정화 / 추천 포트폴리오 품질 보강
def rc38_overlap(a, b):
    try:
        return len(set(int(x) for x in a) & set(int(x) for x in b))
    except Exception:
        return 0

def rc38_portfolio_reorder(combos, details, max_overlap=3):
    """점수 순위만 따르지 않고 조합 간 중복을 줄여 실사용 추천 포트폴리오로 재정렬합니다."""
    combos = [list(map(int, c)) for c in (combos or [])]
    details = list(details or [])
    rows = []
    for i, combo in enumerate(combos):
        d = dict(details[i]) if i < len(details) and isinstance(details[i], dict) else {}
        try:
            score = float(d.get('score') or d.get('ai_score') or d.get('vip_score') or 0)
        except Exception:
            score = 0.0
        rows.append({'idx': i, 'combo': combo, 'detail': d, 'score': score})
    rows.sort(key=lambda x: (-x['score'], sum(x['combo']), x['combo']))
    selected, rest = [], []
    for row in rows:
        if not selected or all(rc38_overlap(row['combo'], s['combo']) <= max_overlap for s in selected):
            selected.append(row)
        else:
            rest.append(row)
    selected.extend(rest)
    new_combos = [r['combo'] for r in selected]
    new_details = [r['detail'] for r in selected]
    # 화면 순위와 실제 저장 순위를 일치시킵니다.
    for i, d in enumerate(new_details):
        d['rank'] = i + 1
    return new_combos, new_details

def rc38_generation_report(combos, details, safe_round, safe_mode):
    scores=[]
    for d in details or []:
        try: scores.append(float(d.get('score') or d.get('ai_score') or d.get('vip_score') or 0))
        except Exception: pass
    overlap_count=0
    max_overlap=0
    for i in range(len(combos or [])):
        for j in range(i+1, len(combos or [])):
            ov=rc38_overlap(combos[i], combos[j]); max_overlap=max(max_overlap, ov)
            if ov >= 4: overlap_count += 1
    flat=[int(n) for c in combos or [] for n in c]
    zone=[sum(1<=n<=15 for n in flat), sum(16<=n<=30 for n in flat), sum(31<=n<=45 for n in flat)]
    odd=sum(n%2 for n in flat)
    total=len(flat) or 1
    return {
        'rc_version': RC3_8_VERSION,
        'round_no': safe_round,
        'mode': safe_mode,
        'combo_count': len(combos or []),
        'avg_score': round(sum(scores)/len(scores), 1) if scores else 0,
        'max_score': round(max(scores), 1) if scores else 0,
        'min_score': round(min(scores), 1) if scores else 0,
        'max_overlap': max_overlap,
        'high_overlap_pairs': overlap_count,
        'zone_distribution': zone,
        'odd_even_total': {'odd': odd, 'even': total-odd},
        'quality_message': 'RC3-8 포트폴리오 보정: 고득점 조합을 우선하되 조합 간 중복과 구간 쏠림을 줄였습니다.'
    }

def rc38_db_health_snapshot():
    tables = ['admins','members','recommendations','sms_logs','winning_checks','admin_logs','login_logs','settings']
    out = {'version': RC3_8_VERSION, 'db_engine': DB_ENGINE, 'database_url_set': bool(DATABASE_URL), 'tables': {}, 'db_path': str(DB) if DB_ENGINE == 'sqlite' else 'postgresql'}
    with con() as c:
        for t in tables:
            try:
                out['tables'][t] = {'count': c.execute(f'SELECT COUNT(*) c FROM {t}').fetchone()['c'], 'columns': table_cols(c, t)}
            except Exception as e:
                out['tables'][t] = {'error': str(e)[:160], 'columns': []}
    return out

class LoginReq(BaseModel): username:str='admin'; password:str
class AdminReq(BaseModel): username:str; name:str='관리자'; password:str; role:str='전체권한'; memo:str=''
class AdminUpdateReq(BaseModel): name:str|None=None; password:str|None=None; role:str|None=None; memo:str|None=None; is_active:int|None=None
class MyAccountReq(BaseModel): name:str|None=None; phone:str|None=None; memo:str|None=None; current_password:str|None=None; new_password:str|None=None
class PasswordReq(BaseModel): password:str
class MemberReq(BaseModel): name:str; phone:str=''; grade:str='일반'; memo:str=''; status:str='활성'; priority:str='보통'; source:str='직접등록'; preferred_count:int=10; created_by:int|None=None; created_at:str|None=None; contract_months:int|None=12; contract_end_at:str|None=None
class MemberStatusReq(BaseModel): status:str; memo:str|None=None
class MemberMemoReq(BaseModel): memo:str=''
class MemberNoteReq(BaseModel): note:str; note_type:str='상담'
class MemberBulkStatusReq(BaseModel): member_ids:list[int]; status:str
class GenerateReq(BaseModel): member_id:int|None=None; round_no:int=1232; count:int=10; mode:str='balanced'; fixed:str=''; excluded:str=''; exclude:str='' # exclude는 기존 프론트 호환용
class SaveRecommendationReq(BaseModel): member_id:int|None=None; member_name:str=''; round_no:int; mode:str='balanced'; combos:list[list[int]]=[]; analysis:str=''; sms:str=''; details:list[dict]=[]; engine:dict={}
class SmsReq(BaseModel): member_id:int|None=None; member_name:str=''; phone:str=''; round_no:int; body:str; combos:list[list[int]]=[]; send_now:bool=False
class WinReq(BaseModel): round_no:int; win_numbers:list[int]=[]; bonus:int=0; combos:list[list[int]]=[]; member_id:int|None=None; member_name:str=''
class DrawReq(BaseModel): round_no:int; draw_date:str=''; numbers:list[int]; bonus:int
class AutoWinReq(BaseModel): round_no:int; winning:str|list[int]=''; bonus:int=0; draw_date:str=''
class SettingReq(BaseModel): key:str; value:str


@app.get('/api/rc6-10/status')
def rc6_10_status():
    return {'ok': True, 'version': 'RC6-10_SQL_PERCENT_STABLE', 'fix': 'postgres percent placeholder stable'}

@app.get('/api/health')
def health():
    return {'ok': True, 'app': APP_VERSION, 'phase': RC_VERSION, 'rc_version': RC_VERSION, 'time': now(), 'db_engine': DB_ENGINE, 'database_url_set': bool(DATABASE_URL), 'db_path': str(DB), 'persistent_dir': str(DB_DIR)}



@app.get('/api/rc5-12/status')
def rc5_12_status():
    """RC5-12: GitHub/Railway 배포 안정성 및 회원검색 상태 진단."""
    checks = []
    def add(name, ok, detail=''):
        checks.append({'name': name, 'ok': bool(ok), 'detail': detail})
    add('frontend_exists', FRONT.exists(), str(FRONT))
    add('database_dir_writable', os.access(str(DB_DIR), os.W_OK), str(DB_DIR))
    add('export_dir_writable', os.access(str(EXPORT_DIR), os.W_OK), str(EXPORT_DIR))
    add('db_file_ready', DB.exists(), str(DB))
    try:
        with con() as c:
            tables = {r['name'] for r in c.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()}
            add('members_table', 'members' in tables, '회원관리 테이블')
            add('admins_table', 'admins' in tables, '관리자 테이블')
            member_count = c.execute('SELECT COUNT(*) c FROM members').fetchone()['c'] if 'members' in tables else 0
            admin_count = c.execute('SELECT COUNT(*) c FROM admins').fetchone()['c'] if 'admins' in tables else 0
    except Exception as e:
        add('database_open', False, str(e))
        member_count = 0
        admin_count = 0
    return {
        'ok': all(x['ok'] for x in checks),
        'app': APP_VERSION,
        'rc_version': RC_VERSION,
        'time': now(),
        'counts': {'members': member_count, 'admins': admin_count},
        'checks': checks,
        'next': '회원검색은 /api/members?q=검색어 로 진단할 수 있습니다.'
    }


@app.get('/api/rc5-13/status')
def rc5_13_status():
    """RC5-13: GitHub/Railway 배포 전 최종 진단 및 핵심 테이블 상태 점검."""
    checks = []
    counts = {}
    def add(name, ok, detail=''):
        checks.append({'name': name, 'ok': bool(ok), 'detail': str(detail)})

    required_files = [
        ('start.py', BASE / 'start.py'),
        ('requirements.txt', BASE / 'requirements.txt'),
        ('runtime.txt', BASE / 'runtime.txt'),
        ('frontend/index.html', FRONT / 'index.html'),
        ('frontend/app.js', FRONT / 'app.js'),
        ('frontend/login.html', FRONT / 'login.html'),
        ('frontend/login.js', FRONT / 'login.js'),
    ]
    for name, path in required_files:
        add(f'file:{name}', path.exists(), path)

    add('database_dir_writable', DB_DIR.exists() and os.access(str(DB_DIR), os.W_OK), DB_DIR)
    add('export_dir_writable', EXPORT_DIR.exists() and os.access(str(EXPORT_DIR), os.W_OK), EXPORT_DIR)
    add('db_engine_ready', DB_ENGINE in ('sqlite', 'postgresql'), DB_ENGINE)

    required_tables = ['admins', 'members', 'recommendations', 'sms_logs', 'winning_checks', 'draws', 'settings', 'admin_logs']
    try:
        with con() as c:
            for table in required_tables:
                try:
                    row = c.execute(f'SELECT COUNT(*) c FROM {table}').fetchone()
                    counts[table] = int(row['c'] if hasattr(row, 'keys') else row[0])
                    add(f'table:{table}', True, f'{counts[table]} rows')
                except Exception as e:
                    counts[table] = 0
                    add(f'table:{table}', False, str(e)[:180])
    except Exception as e:
        add('database_connection', False, str(e)[:180])

    # GitHub에 올리면 안 되는 산출물/캐시가 남아있는지 간단 점검
    blocked = []
    for root, dirs, files in os.walk(BASE):
        rel_root = os.path.relpath(root, BASE)
        if '.git' in rel_root.split(os.sep):
            continue
        for d in list(dirs):
            if d == '__pycache__':
                blocked.append(os.path.join(rel_root, d))
        for f in files:
            if f.endswith(('.pyc', '.pyo', '.bak', '.tmp')):
                blocked.append(os.path.join(rel_root, f))
    add('github_clean_cache', len(blocked) == 0, ', '.join(blocked[:10]) if blocked else 'clean')

    return {
        'ok': all(x['ok'] for x in checks),
        'app': APP_VERSION,
        'rc_version': RC_VERSION,
        'time': now(),
        'db_engine': DB_ENGINE,
        'database_url_set': bool(DATABASE_URL),
        'paths': {'base': str(BASE), 'db_dir': str(DB_DIR), 'export_dir': str(EXPORT_DIR)},
        'counts': counts,
        'checks': checks,
        'message': 'RC5-13 배포 전 최종 점검입니다. ok가 true이면 GitHub/Railway 업로드 준비 상태입니다.'
    }


@app.get('/api/rc5-14/status')
def rc5_14_status():
    """RC5-14: GitHub/Railway 최종 업로드 전 파일/DB/환경 점검."""
    checks = []
    counts = {}
    def add(name, ok, detail=''):
        checks.append({'name': name, 'ok': bool(ok), 'detail': str(detail)[:300]})

    required_files = [
        ('start.py', BASE / 'start.py'),
        ('requirements.txt', BASE / 'requirements.txt'),
        ('runtime.txt', BASE / 'runtime.txt'),
        ('railway.json', BASE / 'railway.json'),
        ('Dockerfile', BASE / 'Dockerfile'),
        ('.env.example', BASE / '.env.example'),
        ('.gitignore', BASE / '.gitignore'),
        ('frontend/index.html', FRONT / 'index.html'),
        ('frontend/app.js', FRONT / 'app.js'),
        ('frontend/style.css', FRONT / 'style.css'),
    ]
    for name, path in required_files:
        add(f'file:{name}', path.exists(), path)

    add('database_dir_ready', DB_DIR.exists() and os.access(str(DB_DIR), os.W_OK), DB_DIR)
    add('export_dir_ready', EXPORT_DIR.exists() and os.access(str(EXPORT_DIR), os.W_OK), EXPORT_DIR)
    add('db_engine_valid', DB_ENGINE in ('sqlite', 'postgresql'), DB_ENGINE)

    required_tables = ['admins', 'members', 'recommendations', 'sms_logs', 'winning_checks', 'draws', 'settings', 'admin_logs']
    try:
        with con() as c:
            for table in required_tables:
                try:
                    row = c.execute(f'SELECT COUNT(*) c FROM {table}').fetchone()
                    counts[table] = int(row['c'] if hasattr(row, 'keys') else row[0])
                    add(f'table:{table}', True, f'{counts[table]} rows')
                except Exception as e:
                    counts[table] = 0
                    add(f'table:{table}', False, str(e))
    except Exception as e:
        add('database_connection', False, str(e))

    blocked = []
    secret_files = []
    backup_files = []
    for root, dirs, files in os.walk(BASE):
        rel_root = os.path.relpath(root, BASE)
        parts = set(rel_root.split(os.sep))
        if '.git' in parts:
            continue
        for d in list(dirs):
            if d in ('__pycache__', '.pytest_cache', '.mypy_cache'):
                blocked.append(os.path.join(rel_root, d))
        for f in files:
            p = os.path.join(rel_root, f)
            if f.endswith(('.pyc', '.pyo', '.tmp')):
                blocked.append(p)
            if f in ('.env', 'env.txt'):
                secret_files.append(p)
            if f.endswith(('.bak', '.backup', '.old')) or 'backup' in f.lower():
                backup_files.append(p)
    add('no_python_cache_files', len(blocked) == 0, ', '.join(blocked[:10]) if blocked else 'clean')
    add('no_secret_env_files', len(secret_files) == 0, ', '.join(secret_files[:10]) if secret_files else 'clean')
    add('no_backup_artifacts', len(backup_files) == 0, ', '.join(backup_files[:10]) if backup_files else 'clean')

    return {
        'ok': all(x['ok'] for x in checks),
        'app': APP_VERSION,
        'rc_version': RC_VERSION,
        'time': now(),
        'db_engine': DB_ENGINE,
        'database_url_set': bool(DATABASE_URL),
        'paths': {'base': str(BASE), 'db_dir': str(DB_DIR), 'export_dir': str(EXPORT_DIR)},
        'counts': counts,
        'checks': checks,
        'message': 'RC5-14 업로드 전 최종 점검입니다. ok가 true이면 GitHub/Railway 배포 준비 상태입니다.'
    }


@app.get('/api/rc5-15/status')
def rc5_15_status():
    """RC5-15: 배포 직전 GitHub/Railway 실행 안정성 점검."""
    checks = []
    counts = {}
    def add(name, ok, detail=''):
        checks.append({'name': name, 'ok': bool(ok), 'detail': str(detail)[:500]})

    required_files = [
        ('start.py', BASE / 'start.py'),
        ('requirements.txt', BASE / 'requirements.txt'),
        ('runtime.txt', BASE / 'runtime.txt'),
        ('railway.json', BASE / 'railway.json'),
        ('Procfile', BASE / 'Procfile'),
        ('Dockerfile', BASE / 'Dockerfile'),
        ('.env.example', BASE / '.env.example'),
        ('.gitignore', BASE / '.gitignore'),
        ('frontend/index.html', FRONT / 'index.html'),
        ('frontend/app.js', FRONT / 'app.js'),
        ('frontend/style.css', FRONT / 'style.css'),
    ]
    for name, path in required_files:
        add(f'file:{name}', path.exists(), path)

    # Railway/Render에서 가장 자주 나는 문제: PORT 미사용, 시작 명령 불일치, 로컬 주소 하드코딩
    try:
        start_text = (BASE / 'start.py').read_text(encoding='utf-8')
        add('start_uses_port_env', 'PORT' in start_text and '0.0.0.0' in start_text, 'start.py must bind 0.0.0.0:$PORT')
    except Exception as e:
        add('start_uses_port_env', False, str(e))

    try:
        railway_text = (BASE / 'railway.json').read_text(encoding='utf-8')
        add('railway_start_command_ready', 'python start.py' in railway_text or 'start.py' in railway_text, railway_text[:200])
    except Exception as e:
        add('railway_start_command_ready', False, str(e))

    try:
        app_js = (FRONT / 'app.js').read_text(encoding='utf-8')
        login_js = (FRONT / 'login.js').read_text(encoding='utf-8') if (FRONT / 'login.js').exists() else ''
        hardcoded = []
        for token in ('localhost:', '127.0.0.1:', 'http://localhost', 'http://127.0.0.1'):
            if token in app_js or token in login_js:
                hardcoded.append(token)
        add('frontend_no_localhost_api', not hardcoded, ', '.join(hardcoded) if hardcoded else 'clean')
    except Exception as e:
        add('frontend_no_localhost_api', False, str(e))

    add('database_dir_ready', DB_DIR.exists() and os.access(str(DB_DIR), os.W_OK), DB_DIR)
    add('export_dir_ready', EXPORT_DIR.exists() and os.access(str(EXPORT_DIR), os.W_OK), EXPORT_DIR)
    add('db_engine_valid', DB_ENGINE in ('sqlite', 'postgresql'), DB_ENGINE)

    required_tables = ['admins', 'members', 'recommendations', 'sms_logs', 'winning_checks', 'draws', 'settings', 'admin_logs']
    try:
        with con() as c:
            for table in required_tables:
                try:
                    row = c.execute(f'SELECT COUNT(*) c FROM {table}').fetchone()
                    counts[table] = int(row['c'] if hasattr(row, 'keys') else row[0])
                    add(f'table:{table}', True, f'{counts[table]} rows')
                except Exception as e:
                    counts[table] = 0
                    add(f'table:{table}', False, str(e))
    except Exception as e:
        add('database_connection', False, str(e))

    blocked = []
    secret_files = []
    backup_files = []
    large_files = []
    for root, dirs, files in os.walk(BASE):
        rel_root = os.path.relpath(root, BASE)
        parts = set(rel_root.split(os.sep))
        if '.git' in parts:
            continue
        for d in list(dirs):
            if d in ('__pycache__', '.pytest_cache', '.mypy_cache'):
                blocked.append(os.path.join(rel_root, d))
        for f in files:
            p = os.path.join(rel_root, f)
            fp = Path(root) / f
            if f.endswith(('.pyc', '.pyo', '.tmp')):
                blocked.append(p)
            if f in ('.env', 'env.txt'):
                secret_files.append(p)
            if f.endswith(('.bak', '.backup', '.old')) or 'backup' in f.lower():
                backup_files.append(p)
            try:
                if fp.stat().st_size > 20 * 1024 * 1024:
                    large_files.append(p)
            except Exception:
                pass
    add('no_python_cache_files', len(blocked) == 0, ', '.join(blocked[:10]) if blocked else 'clean')
    add('no_secret_env_files', len(secret_files) == 0, ', '.join(secret_files[:10]) if secret_files else 'clean')
    add('no_backup_artifacts', len(backup_files) == 0, ', '.join(backup_files[:10]) if backup_files else 'clean')
    add('no_large_artifacts_over_20mb', len(large_files) == 0, ', '.join(large_files[:10]) if large_files else 'clean')

    return {
        'ok': all(x['ok'] for x in checks),
        'app': APP_VERSION,
        'rc_version': RC_VERSION,
        'time': now(),
        'db_engine': DB_ENGINE,
        'database_url_set': bool(DATABASE_URL),
        'paths': {'base': str(BASE), 'db_dir': str(DB_DIR), 'export_dir': str(EXPORT_DIR)},
        'counts': counts,
        'checks': checks,
        'message': 'RC5-15 배포 직전 안정성 점검입니다. ok가 true이면 GitHub/Railway 업로드 준비 상태입니다.'
    }


@app.get('/api/rc5-16/status')
def rc5_16_status():
    """RC5-16: GitHub 보안/정리 상태 점검."""
    checks = []
    def add(name, ok, detail=''):
        checks.append({'name': name, 'ok': bool(ok), 'detail': str(detail)[:500]})

    for name in ['.gitignore', '.env.example', 'README.md', 'requirements.txt', 'start.py', 'Procfile', 'runtime.txt']:
        add(f'file:{name}', (BASE / name).exists(), BASE / name)

    cache_items, secret_items, db_items, backup_items = [], [], [], []
    for root, dirs, files in os.walk(BASE):
        rel_root = os.path.relpath(root, BASE)
        if '.git' in set(rel_root.split(os.sep)):
            continue
        for d in dirs:
            if d == '__pycache__':
                cache_items.append(os.path.join(rel_root, d))
        for f in files:
            rel = os.path.join(rel_root, f)
            low = f.lower()
            if low.endswith(('.pyc', '.pyo', '.pyd')):
                cache_items.append(rel)
            if f == '.env' or f.startswith('.env.') and f != '.env.example':
                secret_items.append(rel)
            if low.endswith(('.db', '.sqlite', '.sqlite3')):
                db_items.append(rel)
            if low.endswith(('.bak', '.backup', '.old')) or 'backup' in low:
                backup_items.append(rel)

    add('no_python_cache', not cache_items, ', '.join(cache_items[:10]) if cache_items else 'clean')
    add('no_env_secret_files', not secret_items, ', '.join(secret_items[:10]) if secret_items else 'clean')
    add('no_database_files_in_repo', not db_items, ', '.join(db_items[:10]) if db_items else 'clean')
    add('no_backup_artifacts', not backup_items, ', '.join(backup_items[:10]) if backup_items else 'clean')

    try:
        app_text = (BASE / 'backend' / 'app.py').read_text(encoding='utf-8')
        add('no_admin1234_hardcode', 'admin1234' not in app_text, 'hardcoded admin password removed')
        add('env_admin_password_supported', 'BBLOTTO_ADMIN_PASSWORD' in app_text, 'BBLOTTO_ADMIN_PASSWORD')
    except Exception as e:
        add('source_check', False, str(e))

    return {
        'ok': all(x['ok'] for x in checks),
        'app': APP_VERSION,
        'rc_version': RC_VERSION,
        'time': now(),
        'checks': checks,
        'message': 'RC5-16 GitHub 보안/정리 점검입니다. ok가 true이면 업로드 안전 상태입니다.'
    }

@app.get('/api/persistence_status')
def persistence_status(authorization: str|None = Header(default=None)):
    require_admin(authorization)
    with con() as c:
        counts = {
            'members': c.execute('SELECT COUNT(*) c FROM members').fetchone()['c'],
            'recommendations': c.execute('SELECT COUNT(*) c FROM recommendations').fetchone()['c'],
            'sms_logs': c.execute('SELECT COUNT(*) c FROM sms_logs').fetchone()['c'],
            'winning_checks': c.execute('SELECT COUNT(*) c FROM winning_checks').fetchone()['c'],
            'draws': c.execute('SELECT COUNT(*) c FROM draws').fetchone()['c'],
        }
    return {
        'ok': True,
        'db_engine': DB_ENGINE,
        'database_url_set': bool(DATABASE_URL),
        'db_path': str(DB),
        'db_dir': str(DB_DIR),
        'export_dir': str(EXPORT_DIR),
        'db_exists': DB.exists(),
        'db_size_bytes': DB.stat().st_size if DB.exists() else 0,
        'using_render_disk': str(DB).startswith('/data/'),
        'counts': counts,
        'message': 'BBLOTTO_DB_DIR 또는 /data Render Disk를 사용하면 재접속/재배포 후에도 데이터가 유지됩니다.'
    }


RC3_VERSION = 'RC3_3_MEMBER_DB_POSTGRES'

RC3_MIGRATION_TABLES = [
    'admins', 'sessions', 'admin_logs', 'login_logs', 'backup_history', 'members',
    'recommendations', 'sms_logs', 'winning_checks', 'draws', 'settings',
    'consultations', 'sms_templates', 'engine_runs', 'dashboard_snapshots'
]

def _sqlite_source_path():
    return Path(os.getenv('BBLOTTO_SQLITE_SOURCE_PATH', str(BASE / 'database' / 'bblotto_v34.db')))

def _table_exists_sqlite(path: Path, table: str) -> bool:
    if not path.exists():
        return False
    try:
        sc = sqlite3.connect(path)
        row = sc.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table,)).fetchone()
        sc.close()
        return bool(row)
    except Exception:
        return False

def _pg_set_sequence(c, table: str):
    if DB_ENGINE != 'postgresql' or not re.match(r'^[A-Za-z_][A-Za-z0-9_]*$', table):
        return
    try:
        c.execute('SAVEPOINT rc3_seq')
        c.execute(f"SELECT setval(pg_get_serial_sequence('{table}', 'id'), GREATEST(COALESCE((SELECT MAX(id) FROM {table}),0),1), true)")
        c.execute('RELEASE SAVEPOINT rc3_seq')
    except Exception:
        try:
            c.execute('ROLLBACK TO SAVEPOINT rc3_seq')
            c.execute('RELEASE SAVEPOINT rc3_seq')
        except Exception:
            pass

def rc3_migrate_sqlite_to_current_db(source_path: Path | None = None):
    """기존 SQLite DB 데이터를 현재 DB(PostgreSQL 권장)로 1회성 복사합니다."""
    source_path = source_path or _sqlite_source_path()
    if DB_ENGINE != 'postgresql':
        return {'ok': False, 'version': RC3_VERSION, 'message': '현재는 PostgreSQL 모드가 아닙니다.', 'db_engine': DB_ENGINE}
    if not source_path.exists():
        return {'ok': False, 'version': RC3_VERSION, 'message': '이전 SQLite DB 파일을 찾을 수 없습니다.', 'source_path': str(source_path)}

    migrated = []
    skipped = []
    errors = []
    sc = sqlite3.connect(source_path)
    sc.row_factory = sqlite3.Row
    with con() as c:
        for table in RC3_MIGRATION_TABLES:
            if not _table_exists_sqlite(source_path, table):
                skipped.append({'table': table, 'reason': 'source table missing'})
                continue
            try:
                source_cols = [r[1] for r in sc.execute(f'PRAGMA table_info({table})').fetchall()]
                target_cols = table_cols(c, table)
                cols = [x for x in source_cols if x in target_cols]
                if not cols:
                    skipped.append({'table': table, 'reason': 'no common columns'})
                    continue
                rows = sc.execute(f'SELECT {", ".join(cols)} FROM {table}').fetchall()
                if not rows:
                    migrated.append({'table': table, 'rows': 0})
                    continue
                placeholders = ','.join(['?'] * len(cols))
                col_sql = ','.join(cols)
                sql = f'INSERT INTO {table}({col_sql}) VALUES({placeholders}) ON CONFLICT DO NOTHING'
                count = 0
                for row in rows:
                    vals = [row[col] for col in cols]
                    c.execute(sql, vals)
                    count += 1
                _pg_set_sequence(c, table)
                migrated.append({'table': table, 'rows': count})
            except Exception as e:
                errors.append({'table': table, 'error': str(e)[:300]})
                try: c.rollback()
                except Exception: pass
        c.commit()
    sc.close()
    return {'ok': len(errors) == 0, 'version': RC3_VERSION, 'source_path': str(source_path), 'migrated': migrated, 'skipped': skipped, 'errors': errors}

@app.get('/api/rc3/database/status')
def rc3_database_status():
    counts = {}
    warnings = []
    try:
        with con() as c:
            for t in ['admins','members','recommendations','winning_checks','draws','settings']:
                try:
                    counts[t] = c.execute(f'SELECT COUNT(*) c FROM {t}').fetchone()['c']
                except Exception as e:
                    counts[t] = f'error: {str(e)[:120]}'
    except Exception as e:
        warnings.append('DB 연결 실패: ' + str(e)[:200])
    if DB_ENGINE != 'postgresql':
        warnings.append('현재 web 서비스에 DATABASE_URL 또는 Postgres 변수 참조가 연결되지 않았습니다.')
    return {
        'ok': DB_ENGINE == 'postgresql' and not warnings,
        'version': RC3_VERSION,
        'db_engine': DB_ENGINE,
        'database_url_set': bool(DATABASE_URL),
        'sqlite_source_path': str(_sqlite_source_path()),
        'sqlite_source_exists': _sqlite_source_path().exists(),
        'counts': counts,
        'warnings': warnings,
        'message': 'Railway에서는 web 서비스 Variables에 DATABASE_URL=${{Postgres.DATABASE_URL}} 를 연결해야 PostgreSQL을 사용합니다.'
    }

@app.post('/api/rc3/migrate/sqlite-to-postgres')
def rc3_migrate_sqlite_to_postgres(authorization: str|None = Header(default=None)):
    admin = require_admin(authorization)
    result = rc3_migrate_sqlite_to_current_db()
    try:
        log_action(admin, 'RC3_MIGRATE_SQLITE_TO_POSTGRES', json.dumps(result, ensure_ascii=False)[:500])
    except Exception:
        pass
    return result


@app.post('/api/rc3/member-db/ensure')
def rc3_member_db_ensure(authorization: str|None = Header(default=None)):
    admin = require_admin(authorization)
    init_db()
    try:
        log_action(admin, 'RC3_3_ENSURE_MEMBER_DB', '회원/관리자 PostgreSQL 스키마 점검 및 보강')
    except Exception:
        pass
    return rc3_member_db_status(authorization)

@app.get('/api/rc3/member-db/status')
def rc3_member_db_status(authorization: str|None = Header(default=None)):
    require_admin(authorization)
    tables = ['admins','sessions','login_logs','admin_logs','members','recommendations','sms_logs','winning_checks','settings']
    counts = {}
    columns = {}
    with con() as c:
        for t in tables:
            try:
                counts[t] = c.execute(f'SELECT COUNT(*) c FROM {t}').fetchone()['c']
                columns[t] = table_cols(c, t)
            except Exception as e:
                counts[t] = 'error: ' + str(e)[:160]
                columns[t] = []
        admin_roles = c.execute("SELECT COALESCE(role,'전체권한') label, COUNT(*) c FROM admins GROUP BY COALESCE(role,'전체권한')").fetchall()
        member_status = c.execute("SELECT COALESCE(status,'활성') label, COUNT(*) c FROM members GROUP BY COALESCE(status,'활성')").fetchall()
    required = {
        'admins': ['id','username','name','password_hash','is_active','role','last_login_at','last_ip'],
        'members': ['id','name','phone','grade','status','priority','source','created_by','created_at','updated_at'],
        'login_logs': ['id','admin_id','username','success','ip','user_agent','message','created_at'],
        'admin_logs': ['id','admin_id','username','action','detail','ip','created_at']
    }
    missing = {t:[col for col in req if col not in columns.get(t, [])] for t, req in required.items()}
    return {
        'ok': DB_ENGINE == 'postgresql' and not any(missing.values()),
        'version': RC3_VERSION,
        'db_engine': DB_ENGINE,
        'database_url_set': bool(DATABASE_URL),
        'counts': counts,
        'missing_columns': missing,
        'admin_roles': {r['label']: r['c'] for r in admin_roles},
        'member_status': {r['label']: r['c'] for r in member_status},
        'message': 'RC3-3 회원관리 DB 상태입니다. ok가 false이면 /api/rc3/member-db/ensure 를 1회 실행하세요.'
    }

@app.get('/api/rc3/member-db/login-logs')
def rc3_member_db_login_logs(limit:int=100, authorization: str|None = Header(default=None)):
    require_admin(authorization)
    limit=max(1, min(int(limit or 100), 500))
    with con() as c:
        rows=c.execute('SELECT id,admin_id,username,success,ip,user_agent,message,created_at FROM login_logs ORDER BY id DESC LIMIT ?', (limit,)).fetchall()
    return {'ok': True, 'items': [dict(r) for r in rows]}

@app.get('/api/version')
def version():
    return {'app': 'BBLOTTO PRO', 'version': 'V2 STABLE', 'phase': RC_VERSION, 'rc_version': RC_VERSION, 'features': ['server_foundation','members','recommendations','stats100','top3','score_grade','recommendation_history','admin_logs','db_health','cloud_deploy','backup_restore_guard','admin_audit','db_standardization','draw_auto_fetch_fallback','official_cache','ai_engine_v1_0','pair_triple_analysis','reason_based_scoring','member_linked_recommendations','member_linked_win_check','orphan_recommendation_repair','member_detail_message_history','member_detail_winning_history','member_detail_recommendation_hidden'], 'time': now()}

@app.get('/api/rc3-8/health')
def rc38_health(authorization: str|None = Header(default=None)):
    require_admin(authorization)
    snap = rc38_db_health_snapshot()
    required = {'admins':['id','username','password_hash'], 'members':['id','name','status'], 'recommendations':['id','member_id','round_no','numbers','created_at'], 'admin_logs':['id','action','created_at']}
    missing = {t:[col for col in cols if col not in snap['tables'].get(t,{}).get('columns',[])] for t, cols in required.items()}
    snap['ok'] = not any(missing.values())
    snap['missing_required_columns'] = missing
    snap['message'] = 'RC3-8 상태 점검입니다. ok=true이면 핵심 테이블과 컬럼이 준비된 상태입니다.'
    return snap

@app.get('/api/rc3-8/recommendation-summary')
def rc38_recommendation_summary(limit:int=20, authorization: str|None = Header(default=None)):
    require_admin(authorization)
    limit=max(1, min(int(limit or 20), 100))
    with con() as c:
        recent=c.execute('SELECT id,member_id,member_name,round_no,mode,count,avg_score,created_at FROM recommendations ORDER BY id DESC LIMIT ?', (limit,)).fetchall()
        by_round=c.execute('SELECT round_no, COUNT(*) c, AVG(avg_score) avg_score FROM recommendations GROUP BY round_no ORDER BY round_no DESC LIMIT 20').fetchall()
        by_member=c.execute('SELECT COALESCE(member_name,"미지정") member_name, COUNT(*) c, MAX(created_at) latest FROM recommendations GROUP BY COALESCE(member_name,"미지정") ORDER BY c DESC, latest DESC LIMIT 20').fetchall()
    return {'ok': True, 'version': RC3_8_VERSION, 'recent':[dict(r) for r in recent], 'by_round':[dict(r) for r in by_round], 'by_member':[dict(r) for r in by_member]}



@app.get('/api/rc6-7/status')
def rc67_status(authorization: str|None = Header(default=None)):
    require_admin(authorization)
    checks=[]
    with con() as c:
        for table in ['admins','members','recommendations','sms_logs','draws','settings']:
            try:
                cnt=c.execute(f'SELECT COUNT(*) c FROM {table}').fetchone()['c']
                checks.append({'table':table,'ok':True,'count':cnt})
            except Exception as e:
                checks.append({'table':table,'ok':False,'error':str(e)[:200]})
    return {'ok': all(x.get('ok') for x in checks), 'version':'RC6-10_SQL_PERCENT_STABLE', 'db_engine':DB_ENGINE, 'checks':checks, 'time':now()}


@app.get('/')
def login_page(): return FileResponse(FRONT/'login.html')

@app.get('/dashboard')
def dashboard_page(): return FileResponse(FRONT/'index.html')


@app.get('/style.css')
def style_css():
    return FileResponse(FRONT/'style.css', media_type='text/css', headers={'Cache-Control':'no-store, max-age=0'})

@app.get('/app.js')
def app_js():
    return FileResponse(FRONT/'app.js', media_type='application/javascript', headers={'Cache-Control':'no-store, max-age=0'})



@app.get('/api/ui-health')
def ui_health():
    return {'ok': True, 'version': 'STABLE-CORE-SAME-NUMBER-SAVE', 'event_owner': 'app.js', 'fallback_file': None, 'single_event_owner': True}

@app.get('/login.js')
def login_js():
    return FileResponse(FRONT/'login.js', media_type='application/javascript')



@app.post('/api/login')
def login(req:LoginReq, request:Request):
    username = str(req.username or '').strip()[:80]
    if not username or not req.password or len(str(req.password)) > 256:
        raise HTTPException(400, '아이디와 비밀번호를 확인해주세요.')
    _rc11_check_login_limit(request, username)
    with con() as c:
        admin = c.execute('SELECT * FROM admins WHERE username=? AND is_active=1', (username,)).fetchone()
        stored_hash = admin['password_hash'] if admin else 'pbkdf2_sha256$260000$00000000000000000000000000000000$' + ('0' * 64)
        password_ok = verify_password(req.password, stored_hash)
        if not admin or not password_ok:
            _rc11_record_login_failure(request, username)
            ip = _rc113_client_ip(request)
            c.execute('INSERT INTO admin_logs(admin_id,username,action,detail,ip,created_at) VALUES(?,?,?,?,?,?)', (None, username, 'LOGIN_FAILED', '로그인 실패', ip, now()))
            c.commit()
            log_login_event(0, username, 0, '로그인 실패', request)
            raise HTTPException(401, '아이디 또는 비밀번호가 맞지 않습니다.')
        _rc11_clear_login_failures(request, username)
        # 구형 또는 평문 해시는 로그인 성공 시 자동으로 RC11 방식으로 올립니다.
        if password_needs_rehash(admin['password_hash']):
            c.execute('UPDATE admins SET password_hash=? WHERE id=?', (hash_password(req.password), admin['id']))
        # 만료 세션과 오래된 중복 세션을 정리합니다.
        c.execute("DELETE FROM sessions WHERE datetime(expires_at) <= datetime('now','localtime')")
        old_sessions = c.execute('SELECT token FROM sessions WHERE admin_id=? ORDER BY created_at DESC', (admin['id'],)).fetchall()
        for old in old_sessions[2:]:
            c.execute('DELETE FROM sessions WHERE token=?', (old['token'],))
        timeout_minutes = 600
        row_timeout = c.execute('SELECT value FROM settings WHERE key=?', ('session_timeout_minutes',)).fetchone()
        if row_timeout:
            try: timeout_minutes = max(10, min(1440, int(row_timeout['value'])))
            except Exception: timeout_minutes = 600
        token=secrets.token_urlsafe(32); exp=(datetime.datetime.now()+datetime.timedelta(minutes=timeout_minutes)).strftime('%Y-%m-%d %H:%M:%S')
        ip = _rc113_client_ip(request)
        ua = request.headers.get('user-agent','')[:240]
        c.execute('INSERT INTO sessions(token,admin_id,created_at,expires_at,last_seen_at,ip,user_agent) VALUES(?,?,?,?,?,?,?)', (token,admin['id'],now(),exp,now(),ip,ua))
        c.execute('UPDATE admins SET last_login_at=?, last_ip=? WHERE id=?', (now(), ip, admin['id']))
        c.commit()
    log_action(dict(admin), 'LOGIN', '관리자 로그인', request)
    log_login_event(admin['id'], admin['username'], 1, '로그인 성공', request)
    return {'token':token,'admin':{'id':admin['id'],'username':admin['username'],'name':admin['name'],'role':admin['role'] if 'role' in admin.keys() else '전체권한','is_super_admin':is_super_admin(admin)},'expires_at':exp,'expires_in_seconds':timeout_minutes*60,'session_timeout_minutes':timeout_minutes}

@app.post('/api/logout')
def logout(request:Request, authorization: str|None = Header(default=None)):
    admin=require_admin(authorization); token=authorization.split(' ',1)[1].strip()
    with con() as c: c.execute('DELETE FROM sessions WHERE token=?',(token,)); c.commit()
    log_action(admin,'LOGOUT','관리자 로그아웃',request); return {'ok':True}

@app.get('/api/me')
def me(authorization: str|None = Header(default=None)):
    a=require_admin(authorization)
    
    left_seconds = 0
    try:
        exp_dt = datetime.datetime.strptime(a.get('expires_at',''), '%Y-%m-%d %H:%M:%S')
        left_seconds = max(0, int((exp_dt - datetime.datetime.now()).total_seconds()))
    except Exception:
        left_seconds = 0
    return {'id':a['id'],'username':a['username'],'name':a['name'],'phone':a.get('phone',''),'memo':a.get('memo',''),'role':a.get('role','전체권한'),'is_super_admin':is_super_admin(a),'expires_at':a.get('expires_at',''),'expires_in_seconds':left_seconds,'last_seen_at':a.get('last_seen_at',''),'last_login_at':a.get('last_login_at',''),'last_ip':a.get('last_ip','')}


@app.put('/api/me')
def update_my_account(req:MyAccountReq, request:Request, authorization: str|None = Header(default=None)):
    admin=require_admin(authorization)
    fields=[]; vals=[]
    if req.name is not None:
        fields.append('name=?'); vals.append(req.name.strip() or '관리자')
    if req.phone is not None:
        fields.append('phone=?'); vals.append(req.phone.strip())
    if req.memo is not None:
        fields.append('memo=?'); vals.append(req.memo.strip())
    if req.new_password:
        validate_password_strength(req.new_password)
        with con() as c:
            row=c.execute('SELECT password_hash FROM admins WHERE id=?', (admin['id'],)).fetchone()
        if not row or not verify_password(req.current_password or '', row['password_hash']):
            raise HTTPException(400,'현재 비밀번호가 맞지 않습니다.')
        fields.append('password_hash=?'); vals.append(hash_password(req.new_password))
    if not fields:
        return {'ok':True,'changed':0}
    fields.append('updated_at=?'); vals.append(now()); vals.append(admin['id'])
    with con() as c:
        c.execute('UPDATE admins SET '+', '.join(fields)+' WHERE id=?', vals)
        c.commit()
    action='UPDATE_MY_PASSWORD' if req.new_password and len(fields)<=2 else ('UPDATE_MY_ACCOUNT_WITH_PASSWORD' if req.new_password else 'UPDATE_MY_ACCOUNT')
    log_action(admin, action, '본인 계정 정보 수정', request)
    return {'ok':True,'changed':1}

@app.get('/api/admins')
def admins(authorization: str|None = Header(default=None)):
    require_admin(authorization)
    with con() as c: rows=c.execute('SELECT id,username,name,phone,role,memo,is_active,created_at,updated_at,last_login_at,last_ip FROM admins ORDER BY id').fetchall()
    return [dict(r) for r in rows]

@app.post('/api/admins')
def create_admin(req:AdminReq, request:Request, authorization: str|None = Header(default=None)):
    admin=require_admin(authorization)
    require_super_admin(admin)
    validate_password_strength(req.password)
    username = validate_admin_username(req.username)
    try:
        with con() as c:
            cur = c.execute('INSERT INTO admins(username,name,password_hash,is_active,created_at,updated_at,role,memo) VALUES(?,?,?,?,?,?,?,?)', (username,req.name.strip(),hash_password(req.password),1,now(),now(),req.role.strip() or '전체권한',req.memo.strip()))
            c.commit()
            if getattr(cur, 'rowcount', 1) == 0:
                raise HTTPException(400,'이미 존재하는 관리자 아이디입니다.')
        log_action(admin,'CREATE_ADMIN',f'관리자 생성: {username}',request); return {'ok':True}
    except sqlite3.IntegrityError:
        raise HTTPException(400,'이미 존재하는 관리자 아이디입니다.')
    except Exception as e:
        if 'duplicate' in str(e).lower() or 'unique' in str(e).lower():
            raise HTTPException(400,'이미 존재하는 관리자 아이디입니다.')
        raise

@app.delete('/api/admins/{admin_id}')
def delete_admin(admin_id:int, request:Request, authorization: str|None = Header(default=None)):
    admin=require_admin(authorization)
    require_super_admin(admin)
    if admin_id == admin['id']:
        raise HTTPException(400,'현재 로그인한 본인 계정은 삭제할 수 없습니다.')
    with con() as c:
        target=c.execute('SELECT id,username,is_active FROM admins WHERE id=?',(admin_id,)).fetchone()
        if not target:
            raise HTTPException(404,'관리자를 찾을 수 없습니다.')
        active_count=c.execute('SELECT COUNT(*) c FROM admins WHERE is_active=1').fetchone()['c']
        if int(target['is_active'] or 0)==1 and active_count <= 1:
            raise HTTPException(400,'마지막 활성 관리자는 삭제할 수 없습니다.')
        c.execute('DELETE FROM sessions WHERE admin_id=?',(admin_id,))
        c.execute('DELETE FROM admins WHERE id=?',(admin_id,))
        c.commit()
    log_action(admin,'DELETE_ADMIN',f'관리자 삭제: {target["username"]} / ID {admin_id}',request)
    return {'ok':True,'deleted':admin_id}


@app.get('/api/admins/{admin_id}')
def admin_detail(admin_id:int, authorization: str|None = Header(default=None)):
    admin=require_admin(authorization)
    if not is_super_admin(admin) and int(admin_id) != int(admin['id']):
        raise HTTPException(403, '일반 관리자는 본인 계정만 조회할 수 있습니다.')
    with con() as c:
        row=c.execute('SELECT id,username,name,phone,role,memo,is_active,created_at,updated_at,last_login_at,last_ip FROM admins WHERE id=?',(admin_id,)).fetchone()
        if not row: raise HTTPException(404,'관리자를 찾을 수 없습니다.')
        recent=c.execute('SELECT action,detail,ip,created_at FROM admin_logs WHERE admin_id=? ORDER BY id DESC LIMIT 30',(admin_id,)).fetchall()
    data=dict(row); data['recent_logs']=[dict(r) for r in recent]
    return data

@app.put('/api/admins/{admin_id}')
def update_admin(admin_id:int, req:AdminUpdateReq, request:Request, authorization: str|None = Header(default=None)):
    admin=require_admin(authorization)
    is_self = int(admin_id) == int(admin['id'])
    super_user = is_super_admin(admin)

    # 일반 관리자는 본인 비밀번호만 변경 가능
    if not super_user:
        if not is_self:
            raise HTTPException(403, '일반 관리자는 다른 관리자를 수정할 수 없습니다.')
        if req.name is not None or req.role is not None or req.memo is not None or req.is_active is not None:
            raise HTTPException(403, '일반 관리자는 본인 비밀번호만 변경할 수 있습니다.')
        if not req.password:
            raise HTTPException(400, '변경할 비밀번호를 입력하세요.')

    fields=[]; vals=[]
    if super_user:
        if req.name is not None:
            fields.append('name=?'); vals.append(req.name.strip() or '관리자')
        if req.role is not None:
            fields.append('role=?'); vals.append(req.role.strip() or '전체권한')
        if req.memo is not None:
            fields.append('memo=?'); vals.append(req.memo.strip())
        if req.is_active is not None:
            if is_self and int(req.is_active) == 0:
                raise HTTPException(400,'현재 로그인한 본인은 비활성화할 수 없습니다.')
            fields.append('is_active=?'); vals.append(1 if int(req.is_active) else 0)

    if req.password:
        validate_password_strength(req.password)
        fields.append('password_hash=?'); vals.append(hash_password(req.password))

    if not fields: return {'ok':True, 'changed':0}
    fields.append('updated_at=?'); vals.append(now()); vals.append(admin_id)
    with con() as c:
        target = c.execute('SELECT id,username,is_active FROM admins WHERE id=?', (admin_id,)).fetchone()
        if not target:
            raise HTTPException(404, '관리자를 찾을 수 없습니다.')
        if req.is_active is not None and int(req.is_active)==0 and int(target['is_active'] or 0)==1:
            active_count=c.execute('SELECT COUNT(*) c FROM admins WHERE is_active=1').fetchone()['c']
            if active_count <= 1:
                raise HTTPException(400,'마지막 활성 관리자는 비활성화할 수 없습니다.')
        c.execute('UPDATE admins SET '+', '.join(fields)+' WHERE id=?', vals)
        if req.is_active is not None and int(req.is_active)==0:
            c.execute('DELETE FROM sessions WHERE admin_id=?', (admin_id,))
        c.commit()
    log_action(admin,'UPDATE_ADMIN',f'관리자 수정 ID {admin_id}',request)
    return {'ok':True, 'changed':1}

@app.post('/api/admins/{admin_id}/activate')
def activate_admin(admin_id:int, request:Request, authorization: str|None = Header(default=None)):
    admin=require_admin(authorization)
    require_super_admin(admin)
    with con() as c:
        c.execute('UPDATE admins SET is_active=1, updated_at=? WHERE id=?', (now(), admin_id)); c.commit()
    log_action(admin,'ACTIVATE_ADMIN',f'관리자 활성화 ID {admin_id}',request)
    return {'ok':True}

@app.post('/api/sessions/cleanup')
def cleanup_sessions(request:Request, authorization: str|None = Header(default=None)):
    admin=require_admin(authorization); require_super_admin(admin)
    with con() as c:
        cur=c.execute('DELETE FROM sessions WHERE expires_at<?', (now(),)); deleted=cur.rowcount; c.commit()
    log_action(admin,'CLEANUP_SESSIONS',f'만료 세션 정리 {deleted}건',request)
    return {'ok':True,'deleted':deleted}

@app.get('/api/security_status')
def security_status(authorization: str|None = Header(default=None)):
    admin=require_admin(authorization); require_super_admin(admin)
    with con() as c:
        active_sessions = c.execute('SELECT COUNT(*) c FROM sessions WHERE expires_at>=?', (now(),)).fetchone()['c']
        failed_today = c.execute('SELECT COUNT(*) c FROM admin_logs WHERE action=? AND created_at LIKE ?', ('LOGIN_FAILED', datetime.datetime.now().strftime('%Y-%m-%d')+'%')).fetchone()['c']
        timeout = c.execute('SELECT value FROM settings WHERE key=?', ('session_timeout_minutes',)).fetchone()
        warn = c.execute('SELECT value FROM settings WHERE key=?', ('auto_logout_warning_minutes',)).fetchone()
    return {'ok':True,'active_sessions':active_sessions,'failed_login_today':failed_today,'session_timeout_minutes':int((timeout or {'value':600})['value'] or 600),'auto_logout_warning_minutes':int((warn or {'value':5})['value'] or 5),'is_super_admin':is_super_admin(admin),'password_hash':'PBKDF2-SHA256/260000','login_limit':'7회/15분','security_headers':True,'origin_check':True,'request_size_limit_mb':2}

@app.get('/api/sessions')
def list_sessions(authorization: str|None = Header(default=None)):
    admin=require_admin(authorization)
    require_super_admin(admin)
    with con() as c:
        rows=c.execute('''SELECT s.token,a.username,a.name,s.created_at,s.last_seen_at,s.expires_at,s.ip,s.user_agent
                          FROM sessions s JOIN admins a ON a.id=s.admin_id
                          WHERE s.expires_at>=? ORDER BY s.last_seen_at DESC, s.created_at DESC''', (now(),)).fetchall()
    out=[]
    for r in rows:
        d=dict(r); d['token_tail']=str(d.pop('token'))[-8:]; out.append(d)
    return out

@app.delete('/api/sessions/{token_tail}')
def revoke_session(token_tail:str, request:Request, authorization: str|None = Header(default=None)):
    admin=require_admin(authorization)
    require_super_admin(admin)
    suffix = re.sub(r'[^A-Za-z0-9_-]', '', str(token_tail or ''))[-8:]
    if len(suffix) != 8:
        raise HTTPException(400, '잘못된 세션 식별자입니다.')
    with con() as c:
        matches = c.execute('SELECT token FROM sessions').fetchall()
        tokens = [str(r['token']) for r in matches if str(r['token']).endswith(suffix)]
        if len(tokens) > 1:
            raise HTTPException(409, '세션 식별자가 중복됩니다. 목록을 새로고침해주세요.')
        deleted = 0
        if tokens:
            cur = c.execute('DELETE FROM sessions WHERE token=?', (tokens[0],)); deleted = cur.rowcount
        c.commit()
    log_action(admin,'REVOKE_SESSION',f'세션 종료: *{suffix}',request)
    return {'ok':True,'deleted':deleted}


@app.get('/api/dashboard')
def dashboard(authorization: str|None = Header(default=None)):
    require_admin(authorization)
    with con() as c:
        members=c.execute('SELECT COUNT(*) c FROM members').fetchone()['c']
        admins_count=c.execute('SELECT COUNT(*) c FROM admins WHERE is_active=1').fetchone()['c']
        recs=c.execute('SELECT COUNT(*) c FROM recommendations').fetchone()['c']
        checks=c.execute('SELECT COUNT(*) c, COALESCE(SUM(prize),0) prize, COALESCE(SUM(cost),0) cost, COALESCE(SUM(profit),0) profit FROM winning_checks').fetchone()
        latest=c.execute('SELECT * FROM draws ORDER BY round_no DESC LIMIT 1').fetchone()
    roi=round((checks['profit']/checks['cost']*100),2) if checks['cost'] else 0
    return {'members':members,'admins':admins_count,'recommendations':recs,'checks':checks['c'],'total_prize':checks['prize'],'total_cost':checks['cost'],'total_profit':checks['profit'],'roi':roi,'latest_draw':dict(latest) if latest else None}

@app.get('/api/dashboard_summary')
def dashboard_summary(authorization: str|None = Header(default=None)):
    require_admin(authorization)
    with con() as c:
        members=c.execute('SELECT COUNT(*) c FROM members').fetchone()['c']
        active_members=c.execute('SELECT COUNT(*) c FROM members WHERE COALESCE(status,"활성")="활성"').fetchone()['c']
        vip_members=c.execute('SELECT COUNT(*) c FROM members WHERE grade IN ("1등","2등","VIP","다이아","프리미엄")').fetchone()['c']
        high_priority=c.execute('SELECT COUNT(*) c FROM members WHERE COALESCE(priority,"보통") IN ("높음","최우선")').fetchone()['c']
        recs=c.execute('SELECT COUNT(*) c FROM recommendations').fetchone()['c']
        sms=c.execute('SELECT COUNT(*) c FROM sms_logs').fetchone()['c']
        latest=c.execute('SELECT round_no FROM draws ORDER BY round_no DESC LIMIT 1').fetchone()
        recent=c.execute('SELECT id,member_id,member_name,round_no,mode,count,created_at FROM recommendations ORDER BY id DESC LIMIT 12').fetchall()
        grade_rows=c.execute('SELECT COALESCE(grade,"일반") label, COUNT(*) c FROM members GROUP BY COALESCE(grade,"일반")').fetchall()
        status_rows=c.execute('SELECT COALESCE(status,"활성") label, COUNT(*) c FROM members GROUP BY COALESCE(status,"활성")').fetchall()
    return {'members':members,'active_members':active_members,'vip_members':vip_members,'high_priority_members':high_priority,'recommendations':recs,'sms':sms,'latest_round': latest['round_no'] if latest else None,'recent_recommendations':[dict(r) for r in recent],'grade_counts':{r['label']:r['c'] for r in grade_rows},'status_counts':{r['label']:r['c'] for r in status_rows}}

@app.get('/api/settings')
def get_settings(authorization: str|None = Header(default=None)):
    require_admin(authorization)
    with con() as c:
        rows=c.execute('SELECT key,value,updated_at FROM settings').fetchall()
    return {r['key']:{'value':clean_template_text(r['value']) if r['key']=='sms_template' else r['value'],'updated_at':r['updated_at']} for r in rows}

@app.get('/api/template')
def get_template(authorization: str|None = Header(default=None)):
    require_admin(authorization)
    with con() as c:
        r=c.execute('SELECT value,updated_at FROM settings WHERE key=?', ('sms_template',)).fetchone()
    return {'body': clean_template_text(r['value'] if r else ''), 'updated_at': r['updated_at'] if r else ''}

@app.post('/api/template')
def save_template(req: dict, request: Request, authorization: str|None = Header(default=None)):
    admin=require_admin(authorization)
    body=clean_template_text(req.get('body') or req.get('value') or '')
    with con() as c:
        c.execute('INSERT OR REPLACE INTO settings(key,value,updated_at) VALUES(?,?,?)', ('sms_template', body, now()))
        c.commit()
    log_action(admin,'SAVE_TEMPLATE','회원 안내 문구 템플릿 저장',request)
    return {'ok': True, 'body': body}

@app.post('/api/settings')
def save_setting(req:SettingReq, request:Request, authorization: str|None = Header(default=None)):
    admin=require_admin(authorization)
    allowed={'sms_template','sms_provider','sms_sender','sms_api_url','sms_api_key','session_timeout_minutes','auto_logout_warning_minutes'}
    if req.key not in allowed:
        raise HTTPException(400,'허용되지 않은 설정입니다.')
    with con() as c:
        c.execute('INSERT OR REPLACE INTO settings(key,value,updated_at) VALUES(?,?,?)',(req.key,clean_template_text(req.value) if req.key=='sms_template' else req.value,now()))
        c.commit()
    log_action(admin,'SAVE_SETTING',f'설정 저장: {req.key}',request)
    return {'ok':True,'key':req.key}

@app.get('/api/admin-stats')
def admin_stats(authorization: str|None = Header(default=None)):
    admin=require_admin(authorization); require_super_admin(admin)
    with con() as c:
        rows=c.execute('''SELECT a.id,a.username,a.name,a.last_login_at,
          SUM(CASE WHEN l.action='LOGIN' THEN 1 ELSE 0 END) login_count,
          SUM(CASE WHEN l.action='CREATE_MEMBER' THEN 1 ELSE 0 END) member_created,
          SUM(CASE WHEN l.action='GENERATE' THEN 1 ELSE 0 END) generated_count,
          COUNT(l.id) total_actions
          FROM admins a LEFT JOIN admin_logs l ON a.id=l.admin_id GROUP BY a.id ORDER BY a.id''').fetchall()
    return [dict(r) for r in rows]

@app.get('/api/admin-logs')
def admin_logs(limit:int=100, authorization: str|None = Header(default=None)):
    admin=require_admin(authorization); require_super_admin(admin)
    with con() as c: rows=c.execute('SELECT * FROM admin_logs ORDER BY id DESC LIMIT ?', (limit,)).fetchall()
    return [dict(r) for r in rows]

@app.get('/api/admin-overview')
def admin_overview(authorization: str|None = Header(default=None)):
    admin=require_admin(authorization); require_super_admin(admin)
    with con() as c:
        active_admins = c.execute('SELECT COUNT(*) c FROM admins WHERE is_active=1').fetchone()['c']
        active_sessions = c.execute('SELECT COUNT(*) c FROM sessions WHERE expires_at>=?', (now(),)).fetchone()['c']
        today = datetime.datetime.now().strftime('%Y-%m-%d')
        today_logins = c.execute('SELECT COUNT(*) c FROM admin_logs WHERE action=? AND created_at LIKE ?', ('LOGIN', today+'%')).fetchone()['c']
        today_actions = c.execute('SELECT COUNT(*) c FROM admin_logs WHERE created_at LIKE ?', (today+'%',)).fetchone()['c']
        backups = c.execute('SELECT COUNT(*) c, COALESCE(SUM(size_bytes),0) size FROM backup_history').fetchone()
        latest_backup = c.execute('SELECT * FROM backup_history ORDER BY id DESC LIMIT 1').fetchone()
    return {'active_admins':active_admins,'active_sessions':active_sessions,'today_logins':today_logins,'today_actions':today_actions,'backup_count':backups['c'],'backup_size':backups['size'],'latest_backup':dict(latest_backup) if latest_backup else None}

@app.get('/api/backups')
def backup_list(authorization: str|None = Header(default=None)):
    admin=require_admin(authorization); require_super_admin(admin)
    known = set()
    rows = []
    with con() as c:
        dbrows = c.execute('SELECT * FROM backup_history ORDER BY id DESC LIMIT 100').fetchall()
        for r in dbrows:
            d = dict(r); known.add(d.get('filename')); rows.append(d)
    for pattern in ('BBLOTTO*_BACKUP_*.db', 'BBLOTTO*_BACKUP_*.json', 'BBLOTTO_RC3_BACKUP_*.json'):
        for f in sorted(EXPORT_DIR.glob(pattern), key=lambda x: x.stat().st_mtime, reverse=True):
            if f.name not in known:
                rows.append({'id':None,'filename':f.name,'reason':'file','size_bytes':f.stat().st_size,'created_by':None,'created_at':datetime.datetime.fromtimestamp(f.stat().st_mtime).strftime('%Y-%m-%d %H:%M:%S'), 'format': f.suffix.lstrip('.')})
    return rows[:100]

@app.post('/api/backups/create')
def backup_create(request:Request, authorization: str|None = Header(default=None)):
    admin=require_admin(authorization); require_super_admin(admin)
    b=create_db_backup('manual', admin)
    log_action(admin, 'CREATE_BACKUP', f'DB 백업 생성: {b["filename"]}', request)
    return b

@app.post('/api/backups/restore/{filename}')
def backup_restore(filename:str, request:Request, authorization: str|None = Header(default=None)):
    admin=require_admin(authorization); require_super_admin(admin)
    safe = Path(filename).name
    path = EXPORT_DIR / safe
    result = _restore_json_backup(path, admin, request)
    return result

@app.get('/api/backups/validate/{filename}')
def backup_validate(filename:str, authorization: str|None = Header(default=None)):
    admin=require_admin(authorization); require_super_admin(admin)
    safe = Path(filename).name
    path = EXPORT_DIR / safe
    data = _validate_backup_json(path)
    tables = data.get('tables') or {}
    return {'ok': True, 'filename': safe, 'engine': data.get('engine'), 'created_at': data.get('created_at'), 'table_counts': {k: len(v or []) for k, v in tables.items()}}

@app.post('/api/backups/cleanup')
def backup_cleanup(keep:int=20, authorization: str|None = Header(default=None)):
    admin=require_admin(authorization); require_super_admin(admin)
    keep = max(3, min(int(keep or 20), 100))
    files = []
    for pattern in ('BBLOTTO*_BACKUP_*.db', 'BBLOTTO*_BACKUP_*.json', 'BBLOTTO_RC3_BACKUP_*.json'):
        files.extend(EXPORT_DIR.glob(pattern))
    files = sorted(set(files), key=lambda x: x.stat().st_mtime, reverse=True)
    removed=[]
    for f in files[keep:]:
        try:
            removed.append(f.name); f.unlink()
        except Exception:
            pass
    return {'ok': True, 'keep': keep, 'removed': removed, 'remaining': len(files)-len(removed)}

@app.get('/api/rc3-4/status')
def rc3_4_status(authorization: str|None = Header(default=None)):
    admin=require_admin(authorization)
    with con() as c:
        hist = c.execute('SELECT COUNT(*) c, COALESCE(SUM(size_bytes),0) size FROM backup_history').fetchone()
        latest = c.execute('SELECT * FROM backup_history ORDER BY id DESC LIMIT 1').fetchone()
    return {'ok': True, 'version': 'RC3-4_BACKUP_RESTORE', 'engine': DB_ENGINE, 'backup_dir': str(EXPORT_DIR), 'backup_count': hist['c'] if hist else 0, 'backup_size': hist['size'] if hist else 0, 'latest_backup': dict(latest) if latest else None, 'supports': ['json_backup', 'json_restore', 'sqlite_db_download', 'postgresql_export']}

@app.get('/api/backups/download/{filename}')
def backup_download(filename:str, token: str|None=None, authorization: str|None = Header(default=None)):
    require_admin_any(authorization, token)
    safe = Path(filename).name
    path = EXPORT_DIR / safe
    if not path.exists() or path.suffix.lower() not in ('.db', '.json'):
        raise HTTPException(404, '백업 파일을 찾을 수 없습니다.')
    media = 'application/json' if path.suffix.lower() == '.json' else 'application/octet-stream'
    return FileResponse(path, media_type=media, filename=safe)


@app.get('/api/members')
def list_members(q:str='', status:str='', grade:str='', priority:str='', sort:str='priority', limit:int=5000, authorization: str|None = Header(default=None)):
    admin=require_admin(authorization)
    wh=[]; args=[]
    scope_sql, scope_args = member_scope_condition(admin, 'm')
    if scope_sql:
        wh.append(scope_sql); args += scope_args
    if q:
        q_norm = re.sub(r'[\s\-_.()\[\]{}+~`\'"·,/:;]', '', str(q or '').lower())
        q_digits = ''.join(ch for ch in str(q or '') if ch.isdigit())
        wh.append('''(
            LOWER(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(COALESCE(m.name,''), ' ', ''), '-', ''), '.', ''), '(', ''), ')', '')) LIKE ?
            OR REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(COALESCE(m.phone,''), '-', ''), ' ', ''), '.', ''), '(', ''), ')', '') LIKE ?
            OR LOWER(COALESCE(m.phone,'')) LIKE ?
            OR LOWER(COALESCE(m.grade,'')) LIKE ?
            OR LOWER(COALESCE(m.status,'')) LIKE ?
            OR LOWER(COALESCE(m.memo,'')) LIKE ?
            OR LOWER(COALESCE(m.source,'')) LIKE ?
            OR LOWER(COALESCE(m.priority,'')) LIKE ?
            OR LOWER(COALESCE(a.name,'')) LIKE ?
            OR LOWER(COALESCE(a.username,'')) LIKE ?
        )''')
        like = f"%{str(q or '').lower()}%"
        norm_like = f"%{q_norm}%"
        digits_like = "%" + q_digits + "%"
        args += [norm_like, digits_like if q_digits else norm_like, like, like, like, like, like, like, like, like]
    if status:
        wh.append('COALESCE(m.status, "활성")=?')
        args.append(status)
    if grade:
        wh.append('COALESCE(m.grade, "일반")=?')
        args.append(grade)
    if priority:
        wh.append('COALESCE(m.priority, "보통")=?')
        args.append(priority)
    sort_map={
        'priority':'CASE COALESCE(m.priority,"보통") WHEN "최우선" THEN 0 WHEN "높음" THEN 1 WHEN "보통" THEN 2 ELSE 3 END, m.id DESC',
        'recent':'m.id DESC',
        'oldest':'m.id ASC',
        'name':'m.name COLLATE NOCASE ASC, m.id DESC',
        'updated':'COALESCE(m.updated_at,m.created_at,"") DESC, m.id DESC',
        'status':'COALESCE(m.status,"활성") ASC, m.id DESC'
    }
    order=sort_map.get(sort, sort_map['priority'])
    limit=max(1, min(int(limit or 5000), 5000))
    # PostgreSQL 호환 안정화:
    # SQL 문자열 안에 LIKE '%super%' 같은 % 리터럴을 직접 쓰면 psycopg2가
    # %s 파라미터로 오인해서 IndexError: tuple index out of range가 발생할 수 있다.
    # 그래서 대표/일반 관리자 판별 패턴도 모두 바인딩 파라미터로 처리한다.
    super_case_args = [
        '%대표관리자%', '%최고관리자%', '%super%', '%owner%', '%대표관리자%', '%최고관리자%'
    ]
    sql="""
        SELECT m.*,
               COALESCE(a.name, a.username, '미지정') AS registered_by_name,
               COALESCE(a.username, '') AS registered_by_username,
               COALESCE(a.role, '') AS registered_by_role,
               CASE
                   WHEN LOWER(COALESCE(a.username,''))='admin'
                     OR REPLACE(LOWER(COALESCE(a.role,'')), ' ', '') LIKE ?
                     OR REPLACE(LOWER(COALESCE(a.role,'')), ' ', '') LIKE ?
                     OR LOWER(COALESCE(a.role,'')) LIKE ?
                     OR LOWER(COALESCE(a.role,'')) LIKE ?
                     OR REPLACE(LOWER(COALESCE(a.name,'')), ' ', '') LIKE ?
                     OR REPLACE(LOWER(COALESCE(a.name,'')), ' ', '') LIKE ?
                   THEN 1 ELSE 0 END AS registered_by_super_admin
        FROM members m
        LEFT JOIN admins a ON a.id = COALESCE(m.created_by,0)
    """ + (' WHERE ' + ' AND '.join(wh) if wh else '') + f' ORDER BY {order} LIMIT ?'
    final_args = super_case_args + args + [limit]
    with con() as c:
        rows=c.execute(sql, final_args).fetchall()
    return [dict(r) for r in rows]

@app.get('/api/members_summary')
def members_summary(authorization: str|None = Header(default=None)):
    admin=require_admin(authorization)
    scope_sql, scope_args = member_scope_condition(admin)
    where = (' WHERE ' + scope_sql) if scope_sql else ''
    with con() as c:
        total=c.execute('SELECT COUNT(*) c FROM members' + where, scope_args).fetchone()['c']
        grade=c.execute('SELECT COALESCE(grade,"일반") label, COUNT(*) c FROM members' + where + ' GROUP BY COALESCE(grade,"일반")', scope_args).fetchall()
        status=c.execute('SELECT COALESCE(status,"활성") label, COUNT(*) c FROM members' + where + ' GROUP BY COALESCE(status,"활성")', scope_args).fetchall()
        priority=c.execute('SELECT COALESCE(priority,"보통") label, COUNT(*) c FROM members' + where + ' GROUP BY COALESCE(priority,"보통")', scope_args).fetchall()
        no_contact=c.execute('SELECT COUNT(*) c FROM members' + (where + ' AND ' if where else ' WHERE ') + 'COALESCE(last_contact_at,"")=""', scope_args).fetchone()['c']
    return {'total':total,'grade':{r['label']:r['c'] for r in grade},'status':{r['label']:r['c'] for r in status},'priority':{r['label']:r['c'] for r in priority},'no_contact':no_contact,'is_super_admin':is_super_admin(admin)}


@app.get('/api/members_manage_overview')
def members_manage_overview(authorization: str|None = Header(default=None)):
    admin=require_admin(authorization)
    scope_sql, scope_args = member_scope_condition(admin)
    where = (' WHERE ' + scope_sql) if scope_sql else ''
    with con() as c:
        total=c.execute('SELECT COUNT(*) c FROM members' + where, scope_args).fetchone()['c']
        active=c.execute('SELECT COUNT(*) c FROM members' + (where + ' AND ' if where else ' WHERE ') + 'COALESCE(status,"활성")="활성"', scope_args).fetchone()['c']
        paused=c.execute('SELECT COUNT(*) c FROM members' + (where + ' AND ' if where else ' WHERE ') + 'COALESCE(status,"활성") IN ("휴면","정지")', scope_args).fetchone()['c']
        closed=c.execute('SELECT COUNT(*) c FROM members' + (where + ' AND ' if where else ' WHERE ') + 'COALESCE(status,"활성") IN ("종료","탈퇴")', scope_args).fetchone()['c']
        recent=c.execute('SELECT id,name,phone,grade,status,priority,created_at,updated_at FROM members' + where + ' ORDER BY id DESC LIMIT 10', scope_args).fetchall()
        status_rows=c.execute('SELECT COALESCE(status,"활성") label, COUNT(*) c FROM members' + where + ' GROUP BY COALESCE(status,"활성")', scope_args).fetchall()
    return {'total':total,'active':active,'paused':paused,'closed':closed,'status_counts':{r['label']:r['c'] for r in status_rows},'recent':[dict(r) for r in recent],'is_super_admin':is_super_admin(admin)}


@app.post('/api/members/{member_id}/status')
def change_member_status(member_id:int, req:MemberStatusReq, request:Request, authorization: str|None = Header(default=None)):
    admin=require_admin(authorization)
    allowed={'활성','상담중','휴면','정지','종료','탈퇴'}
    if req.status not in allowed:
        raise HTTPException(400, '허용되지 않은 회원 상태입니다.')
    with con() as c:
        m=assert_member_access(c, admin, member_id)
        m=c.execute('SELECT id,name,status,memo FROM members WHERE id=?',(member_id,)).fetchone()
        memo = m['memo'] if req.memo is None else req.memo
        c.execute('UPDATE members SET status=?, memo=?, updated_at=? WHERE id=?',(req.status,memo,now(),member_id))
        c.commit()
    log_action(admin,'CHANGE_MEMBER_STATUS',f'회원 상태 변경: {m["name"]} / {m["status"] or "활성"} -> {req.status}',request)
    return {'ok':True,'id':member_id,'status':req.status}

@app.post('/api/members/bulk_status')
def bulk_member_status(req:MemberBulkStatusReq, request:Request, authorization: str|None = Header(default=None)):
    admin=require_admin(authorization)
    allowed={'활성','상담중','휴면','정지','종료','탈퇴'}
    if req.status not in allowed:
        raise HTTPException(400, '허용되지 않은 회원 상태입니다.')
    ids=[int(x) for x in req.member_ids if int(x)>0]
    if not ids: raise HTTPException(400,'변경할 회원을 선택하세요.')
    placeholders=','.join(['?']*len(ids))
    with con() as c:
        scope_sql, scope_args = member_scope_condition(admin)
        scope_tail = (' AND ' + scope_sql) if scope_sql else ''
        rows=c.execute(f'SELECT id,name,status FROM members WHERE id IN ({placeholders}){scope_tail}', [*ids, *scope_args]).fetchall()
        allowed_ids=[int(r['id']) for r in rows]
        if allowed_ids:
            ph=','.join(['?']*len(allowed_ids))
            c.execute(f'UPDATE members SET status=?, updated_at=? WHERE id IN ({ph})', [req.status, now(), *allowed_ids])
        c.commit()
    log_action(admin,'BULK_MEMBER_STATUS',f'회원 {len(rows)}명 상태 일괄 변경 -> {req.status}',request)
    return {'ok':True,'changed':len(rows),'status':req.status}

@app.post('/api/members')
def add_member(req:MemberReq, request:Request, authorization: str|None = Header(default=None)):
    admin=require_admin(authorization)
    created_at = normalize_date_text(req.created_at, datetime.datetime.now().strftime('%Y-%m-%d')) if is_super_admin(admin) else datetime.datetime.now().strftime('%Y-%m-%d')
    contract_months = int(req.contract_months or 12)
    if contract_months not in (6, 12, 24, 36):
        contract_months = 12
    contract_end_at = add_months_date(created_at, contract_months)
    owner_id = int(req.created_by or admin['id']) if is_super_admin(admin) else int(admin['id'])
    with con() as c:
        if owner_id:
            owner = c.execute('SELECT id FROM admins WHERE id=?', (owner_id,)).fetchone()
            if not owner:
                raise HTTPException(400, '등록 관리자를 찾을 수 없습니다.')
        preferred_count=max(1, min(int(req.preferred_count or 10), 100))
        cur=c.execute('INSERT INTO members(name,phone,grade,memo,status,priority,source,preferred_count,created_by,created_at,contract_months,contract_end_at,updated_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)',(req.name,req.phone,req.grade,req.memo,req.status,req.priority,req.source,preferred_count,owner_id,created_at,contract_months,contract_end_at,now()))
        c.commit(); mid=cur.lastrowid
    log_action(admin,'CREATE_MEMBER',f'회원 등록: {req.name}',request); return {'id':mid}

@app.delete('/api/members/{member_id}')
def del_member(member_id:int, request:Request, authorization: str|None = Header(default=None)):
    admin=require_admin(authorization)
    with con() as c:
        assert_member_access(c, admin, member_id)
        c.execute('DELETE FROM members WHERE id=?',(member_id,)); c.commit()
    log_action(admin,'DELETE_MEMBER',f'회원 삭제 ID {member_id}',request); return {'ok':True}

@app.put('/api/members/{member_id}')
def update_member(member_id:int, req:MemberReq, request:Request, authorization: str|None = Header(default=None)):
    admin=require_admin(authorization)
    with con() as c:
        assert_member_access(c, admin, member_id)
        preferred_count=max(1, min(int(req.preferred_count or 10), 100))
        current_member = c.execute('SELECT created_at, COALESCE(contract_months,12) AS contract_months, created_by FROM members WHERE id=?', (member_id,)).fetchone()
        if not current_member:
            raise HTTPException(404, '회원 정보를 찾을 수 없습니다.')

        # V3.0.0 STABLE 실제 수정:
        # - 등록일(created_at), 계약기간(contract_months)은 모든 관리자가 수정 가능
        # - 등록 관리자(created_by) 변경만 대표관리자만 가능
        # - 계약만료일(contract_end_at)은 등록일 + 계약기간으로 서버에서 항상 재계산
        base_created = normalize_date_text(req.created_at, '') or normalize_date_text(current_member['created_at'], datetime.datetime.now().strftime('%Y-%m-%d'))
        contract_months = int(req.contract_months or current_member['contract_months'] or 12)
        if contract_months not in (6, 12, 24, 36):
            contract_months = 12
        contract_end_at = add_months_date(base_created, contract_months)

        fields=[
            'name=?','phone=?','grade=?','memo=?','status=?','priority=?','source=?',
            'preferred_count=?','created_at=?','contract_months=?','contract_end_at=?','updated_at=?'
        ]
        vals=[
            req.name, req.phone, req.grade, req.memo, req.status, req.priority, req.source,
            preferred_count, base_created, contract_months, contract_end_at, now()
        ]

        if is_super_admin(admin):
            owner_id = int(req.created_by or 0)
            if owner_id:
                owner = c.execute('SELECT id FROM admins WHERE id=?', (owner_id,)).fetchone()
                if not owner:
                    raise HTTPException(400, '등록 관리자를 찾을 수 없습니다.')
                fields.append('created_by=?'); vals.append(owner_id)

        vals.append(member_id)
        c.execute('UPDATE members SET '+', '.join(fields)+' WHERE id=?', vals)
        c.commit()

        saved = c.execute('SELECT id,created_at,contract_months,contract_end_at,created_by FROM members WHERE id=?', (member_id,)).fetchone()
    log_action(admin,'UPDATE_MEMBER',f'회원 수정 ID {member_id}: {req.name}',request)
    return {'ok':True, 'member': dict(saved) if saved else {'id': member_id}}

@app.get('/api/members/{member_id}/detail')
def member_detail(member_id:int, authorization: str|None = Header(default=None)):
    admin=require_admin(authorization)
    with con() as c:
        assert_member_access(c, admin, member_id)
        m=c.execute('SELECT * FROM members WHERE id=?',(member_id,)).fetchone()
        recs=c.execute('SELECT id,round_no,mode,count,numbers,analysis,avg_score,created_at FROM recommendations WHERE member_id=? ORDER BY id DESC LIMIT 50',(member_id,)).fetchall()
        sms=c.execute('SELECT id,member_id,round_no,body,status,created_at FROM sms_logs WHERE member_id=? ORDER BY id DESC LIMIT 50',(member_id,)).fetchall()
        wins=c.execute('SELECT round_no,target_numbers,win_numbers,bonus,match_count,bonus_match,rank,prize,cost,profit,roi,created_at FROM winning_checks WHERE member_id=? ORDER BY id DESC LIMIT 50',(member_id,)).fetchall()
        notes=c.execute('SELECT id,note,note_type,created_by_name,created_at FROM member_notes WHERE member_id=? ORDER BY id DESC LIMIT 50',(member_id,)).fetchall()
    rec_list=[]
    for r in recs:
        d=dict(r)
        try: d['numbers']=json.loads(d.get('numbers') or '[]')
        except Exception: d['numbers']=[]
        rec_list.append(d)
    win_list=[]
    for r in wins:
        d=dict(r)
        for key in ('target_numbers','win_numbers'):
            try: d[key]=json.loads(d.get(key) or '[]')
            except Exception: d[key]=[]
        win_list.append(d)
    total_prize=sum(w.get('prize') or 0 for w in win_list)
    total_cost=sum(w.get('cost') or 0 for w in win_list)
    total_profit=sum(w.get('profit') or 0 for w in win_list)
    ranks=collections.Counter([w.get('rank','낙첨') for w in win_list])
    best='없음'
    order={'1등':1,'2등':2,'3등':3,'4등':4,'5등':5,'낙첨':9}
    ranked=[w.get('rank','낙첨') for w in win_list]
    if ranked: best=sorted(ranked, key=lambda x:order.get(x,9))[0]
    return {
        'member':dict(m),
        'recommendations':rec_list,
        'sms_logs':[dict(r) for r in sms],
        'winning_checks':win_list,
        'notes':[dict(r) for r in notes],
        'summary':{
            'recommendations':len(rec_list),'sms':len(sms),'checks':len(win_list),'notes':len(notes),'best_rank':best,
            'rank_counts':dict(ranks),'total_prize':total_prize,'total_cost':total_cost,'total_profit':total_profit,
            'roi':round((total_profit/total_cost*100),2) if total_cost else 0,
            'latest_recommendation': rec_list[0]['created_at'] if rec_list else '',
            'latest_sms': dict(sms[0])['created_at'] if sms else '',
            'latest_note': dict(notes[0])['created_at'] if notes else ''
        }
    }

@app.put('/api/members/{member_id}/memo')
def update_member_memo(member_id:int, req:MemberMemoReq, request:Request, authorization: str|None = Header(default=None)):
    admin=require_admin(authorization)
    with con() as c:
        m=assert_member_access(c, admin, member_id)
        c.execute('UPDATE members SET memo=?, updated_at=? WHERE id=?',(req.memo, now(), member_id))
        c.commit()
    log_action(admin,'UPDATE_MEMBER_MEMO',f'회원 메모 수정: {m["name"]}',request)
    return {'ok':True,'id':member_id}

@app.post('/api/members/{member_id}/notes')
def add_member_note(member_id:int, req:MemberNoteReq, request:Request, authorization: str|None = Header(default=None)):
    admin=require_admin(authorization)
    note=(req.note or '').strip()
    if not note: raise HTTPException(400,'상담 내용을 입력하세요.')
    note_type=(req.note_type or '상담').strip()[:20]
    with con() as c:
        m=assert_member_access(c, admin, member_id)
        c.execute('INSERT INTO member_notes(member_id,note,note_type,created_by,created_by_name,created_at) VALUES(?,?,?,?,?,?)',(member_id,note,note_type,admin['id'],admin.get('name') or admin.get('username') or '',now()))
        c.execute('UPDATE members SET last_contact_at=?, updated_at=? WHERE id=?',(now(),now(),member_id))
        c.commit()
    log_action(admin,'CREATE_MEMBER_NOTE',f'상담 이력 추가: {m["name"]}',request)
    return {'ok':True,'member_id':member_id}


def rc312_resolve_member(c, member_id=None, member_name=''):
    """RC3-12: 추천번호/당첨확인이 회원과 끊기지 않도록 회원 ID/이름을 표준화합니다."""
    mid = None
    name = (member_name or '').strip()
    if member_id:
        m = c.execute('SELECT id,name FROM members WHERE id=?', (member_id,)).fetchone()
        if m:
            mid = int(m['id'])
            name = m['name'] or name
    if not mid and name:
        m = c.execute('SELECT id,name FROM members WHERE name=? ORDER BY id DESC LIMIT 1', (name,)).fetchone()
        if m:
            mid = int(m['id'])
            name = m['name'] or name
    return mid, name


@app.get('/api/rc3-12/member-link-status')
def rc312_member_link_status(authorization: str|None = Header(default=None)):
    require_admin(authorization)
    with con() as c:
        members = c.execute('SELECT COUNT(*) c FROM members').fetchone()['c']
        recs = c.execute('SELECT COUNT(*) c FROM recommendations').fetchone()['c']
        linked = c.execute('SELECT COUNT(*) c FROM recommendations WHERE COALESCE(member_id,0)>0').fetchone()['c']
        orphan = c.execute('SELECT COUNT(*) c FROM recommendations WHERE COALESCE(member_id,0)=0').fetchone()['c']
        latest_orphans = c.execute('SELECT id,round_no,member_name,created_at FROM recommendations WHERE COALESCE(member_id,0)=0 ORDER BY id DESC LIMIT 20').fetchall()
    return {'ok': True, 'version': 'RC3-12', 'members': members, 'recommendations': recs, 'linked_recommendations': linked, 'orphan_recommendations': orphan, 'latest_orphans': [dict(r) for r in latest_orphans], 'message': '회원 선택 없이 생성된 기존 추천이력은 공통 추천으로 표시됩니다. 특정 회원으로 연결하려면 /api/rc3-12/link-orphan-recommendations 를 사용하세요.'}


@app.post('/api/rc3-12/link-orphan-recommendations')
def rc312_link_orphan_recommendations(req: dict, request:Request, authorization: str|None = Header(default=None)):
    admin = require_admin(authorization)
    member_id = int((req or {}).get('member_id') or 0)
    round_no = int((req or {}).get('round_no') or 0)
    if member_id <= 0:
        raise HTTPException(400, '연결할 회원을 선택하세요.')
    with con() as c:
        m = c.execute('SELECT id,name FROM members WHERE id=?', (member_id,)).fetchone()
        if not m:
            raise HTTPException(404, '회원을 찾을 수 없습니다.')
        params = [member_id, m['name'], now()]
        where = 'COALESCE(member_id,0)=0'
        if round_no > 0:
            where += ' AND round_no=?'
            params.append(round_no)
        cur = c.execute(f'UPDATE recommendations SET member_id=?, member_name=?, created_at=COALESCE(NULLIF(created_at,""), ?) WHERE {where}', params)
        # 이미 당첨확인이 되어 있는 공통 추천 결과도 같은 회차 기준으로 회원명을 보정합니다.
        wc_params = [member_id, m['name']]
        wc_where = 'COALESCE(member_id,0)=0'
        if round_no > 0:
            wc_where += ' AND round_no=?'
            wc_params.append(round_no)
        c.execute(f'UPDATE winning_checks SET member_id=?, member_name=? WHERE {wc_where}', wc_params)
        c.commit()
        changed = cur.rowcount if cur.rowcount is not None else 0
    log_action(admin, 'RC3_12_LINK_ORPHAN_RECOMMENDATIONS', f'{round_no or "전체"}회차 미연결 추천이력 {changed}건을 {m["name"]} 회원으로 연결', request)
    return {'ok': True, 'version': 'RC3-12', 'updated': changed, 'member_id': member_id, 'member_name': m['name'], 'round_no': round_no or None}

@app.post('/api/generate')
def generate(req:GenerateReq, request:Request, authorization: str|None = Header(default=None)):
    admin=require_admin(authorization)
    member_name=''
    member_grade='일반'
    member_id=req.member_id
    with con() as c:
        member_id, member_name = rc312_resolve_member(c, member_id, '')
        try:
            if member_id:
                _mg = c.execute('SELECT grade FROM members WHERE id=?', (member_id,)).fetchone()
                member_grade = rc45_grade_label(_mg['grade'] if _mg else '일반')
        except Exception:
            member_grade = '일반'
    # RC3-12: 회원을 선택하지 않은 추천은 기존 호환을 위해 허용하지만,
    # 프론트에서는 회원 선택을 안내하여 이후 당첨확인에서 '회원 선택 없음'이 나오지 않도록 합니다.
    excluded_value = req.excluded or req.exclude or ''
    safe_count=max(1, min(50, int(req.count or 10)))
    safe_round=max(1, int(req.round_no or 1))
    safe_mode=req.mode or 'balanced'
    combos, details, st = make_premium_combos(safe_count, req.fixed, excluded_value, safe_mode, member_grade, member_id=member_id)
    # RC7-1: 회원별 AI 엔진 V2 문구/번호 분산용 회원 시드 정보
    try:
        st['member_id'] = member_id or 0
        st['member_name'] = member_name or ''
        st['member_grade'] = member_grade
    except Exception:
        pass
    details = rc37_enrich_details(combos, details)
    combos, details = rc38_portfolio_reorder(combos, details)
    details = rc37_enrich_details(combos, details)
    analysis=clean_template_text(build_analysis_text(safe_round, st, safe_mode, req.fixed, excluded_value, details))
    sms=clean_template_text(build_sms(member_name, safe_round, combos, analysis, details))
    engine=_engine_summary(details, st)
    engine['phase']='RC11.2'
    engine['member_grade']=member_grade
    engine['grade_strength']=rc45_grade_strength_text(member_grade)
    engine['engine_label']=_rc729_engine_name(member_grade)
    engine['rc38_report']=rc38_generation_report(combos, details, safe_round, safe_mode)
    engine['top3']=rc37_top3(combos, details)
    engine['quality_guide']=f'{member_grade} 관리 기준 · 1회차부터 최신 회차까지의 기록과 실제 조합 특징을 반영한 설명형 추천'
    # RC8.18: 번호 생성 단계에서는 DB에 저장하지 않습니다.
    # 추천번호 저장/보낸문자 저장을 명시적으로 실행한 경우에만 recommendations에 등록됩니다.
    log_action(admin,'GENERATE_PREVIEW_RC8_18',f'{safe_round}회차 {len(combos)}조합 미리보기 생성 · 저장 안 함',request)
    return {'id':None,'saved':False,'round_no':safe_round,'round':safe_round,'sets':combos,'combos':combos,'details':details,'top3':engine.get('top3',[]),'engine':engine,'analysis':analysis,'sms':sms,'member_id':member_id,'member_name':member_name,'member_notice':sms,'quality_guide':engine.get('quality_guide')}


@app.post('/api/recommendations/save')
def save_recommendation(req:SaveRecommendationReq, request:Request, authorization: str|None = Header(default=None)):
    admin=require_admin(authorization)
    combos=[parse_nums(c) for c in (req.combos or [])]
    combos=[c for c in combos if len(c)==6]
    if not combos:
        raise HTTPException(400,'저장할 추천번호가 없습니다.')
    member_id=req.member_id
    member_name=(req.member_name or '').strip()
    member_grade='일반'
    with con() as c:
        member_id, resolved_name = rc312_resolve_member(c, member_id, member_name)
        member_name = resolved_name or member_name
        if member_id:
            m=c.execute('SELECT grade FROM members WHERE id=?',(member_id,)).fetchone()
            member_grade=rc45_grade_label(m['grade'] if m else '일반')
        rec_cols=table_cols(c,'recommendations')
        for col, ddl in {
            'member_id':'INTEGER','member_name':'TEXT DEFAULT ""','round_no':'INTEGER DEFAULT 0',
            'mode':'TEXT DEFAULT "balanced"','count':'INTEGER DEFAULT 0','numbers':'TEXT DEFAULT "[]"','analysis':'TEXT DEFAULT ""','sms':'TEXT DEFAULT ""',
            'created_by':'INTEGER DEFAULT 0','created_at':'TEXT DEFAULT ""',
            'avg_score':'REAL DEFAULT 0','grade':'TEXT DEFAULT "일반"','engine_json':'TEXT DEFAULT "{}"','details_json':'TEXT DEFAULT "[]"','explicit_saved':'INTEGER DEFAULT 0'
        }.items():
            if col not in rec_cols:
                c.execute(f'ALTER TABLE recommendations ADD COLUMN {col} {ddl}')
        rec_cols=table_cols(c,'recommendations')
        engine=req.engine or {}
        data={
            'member_id':member_id,'member_name':member_name,'round_no':max(1,int(req.round_no or 1)),
            'mode':req.mode or 'balanced','count':len(combos),'numbers':json.dumps(combos,ensure_ascii=False),
            'analysis':clean_template_text(req.analysis or ''),'sms':clean_template_text(req.sms or ''),
            'created_by':admin['id'],'created_at':now(),'avg_score':float(engine.get('avg_score') or 0),
            'grade':member_grade,'engine_json':json.dumps(engine,ensure_ascii=False),'details_json':json.dumps(req.details or [],ensure_ascii=False),'explicit_saved':1
        }
        cols=[k for k in data if k in rec_cols]
        cur=c.execute('INSERT INTO recommendations('+','.join(cols)+') VALUES('+','.join(['?']*len(cols))+')',[data[k] for k in cols])
        rid=cur.lastrowid
        c.commit()
    log_action(admin,'SAVE_RECOMMENDATION_RC8_18',f'{data["round_no"]}회차 {len(combos)}조합 저장 · {member_name or "회원 미선택"}',request)
    return {'ok':True,'saved':True,'id':rid,'member_id':member_id,'member_name':member_name,'round_no':data['round_no'],'count':len(combos)}


@app.get('/api/recommendations')
def recommendations(authorization: str|None = Header(default=None)):
    require_admin(authorization)
    with con() as c:
        rows=c.execute('SELECT id,member_id,member_name,round_no,mode,count,analysis,sms,created_by,created_at FROM recommendations ORDER BY id DESC LIMIT 200').fetchall()
    return [dict(r) for r in rows]

@app.get('/api/recommendations/{rec_id}')
def recommendation_detail(rec_id:int, authorization: str|None = Header(default=None)):
    require_admin(authorization)
    with con() as c:
        r=c.execute('SELECT * FROM recommendations WHERE id=?',(rec_id,)).fetchone()
    if not r: raise HTTPException(404,'추천번호를 찾을 수 없습니다.')
    d=dict(r); d['numbers']=json.loads(d.get('numbers') or '[]')
    st=latest_stats(120)
    try:
        d['details']=json.loads(d.get('details_json') or '[]')
    except Exception:
        d['details']=[]
    if not d['details']:
        d['details']=[combo_detail(c,st) for c in d['numbers']]
    d['details']=rc37_enrich_details(d.get('numbers') or [], d.get('details') or [])
    d['top3']=rc37_top3(d.get('numbers') or [], d.get('details') or [])
    return d


def get_setting_value(key, default=''):
    try:
        with con() as c:
            r = c.execute('SELECT value FROM settings WHERE key=?', (key,)).fetchone()
        return r['value'] if r else default
    except Exception:
        return default

def normalize_phone(phone):
    return re.sub(r'[^0-9]', '', phone or '')

def send_sms_provider(phone, body):
    """V34 문자 발송 게이트웨이.
    - 기본값 mock: 실제 과금/발송 없이 성공 처리하고 발송 이력만 남김
    - http: sms_api_url/sms_api_key/sms_sender 설정이 있으면 외부 문자 API에 POST 준비
      범용 JSON {to, from, text, api_key} 형태라 실제 업체 스펙에 맞게 URL만 연결/수정 가능
    """
    phone = normalize_phone(phone)
    provider = (get_setting_value('sms_provider', 'mock') or 'mock').lower()
    if not phone:
        return {'status':'failed','provider':provider,'message':'수신번호가 없습니다.'}
    if not body.strip():
        return {'status':'failed','provider':provider,'message':'문자 내용이 없습니다.'}
    if provider in ('mock','test','demo',''):
        return {'status':'sent_mock','provider':'mock','message':'모의 발송 완료: 실제 문자는 발송되지 않았고 이력만 저장되었습니다.'}
    if provider == 'http':
        url = get_setting_value('sms_api_url', '').strip()
        api_key = get_setting_value('sms_api_key', '').strip()
        sender = normalize_phone(get_setting_value('sms_sender', ''))
        if not url or not api_key:
            return {'status':'failed','provider':'http','message':'문자 API URL 또는 API KEY가 설정되지 않았습니다.'}
        payload = json.dumps({'to':phone,'from':sender,'text':body,'api_key':api_key}, ensure_ascii=False).encode('utf-8')
        req = urllib.request.Request(url, data=payload, headers={'Content-Type':'application/json'}, method='POST')
        try:
            with urllib.request.urlopen(req, timeout=8) as resp:
                txt = resp.read().decode('utf-8', 'ignore')[:500]
            return {'status':'sent','provider':'http','message':txt or 'API 발송 요청 완료'}
        except Exception as e:
            return {'status':'failed','provider':'http','message':str(e)}
    return {'status':'failed','provider':provider,'message':'지원하지 않는 문자 provider입니다. mock 또는 http를 사용하세요.'}

@app.post('/api/sms')
def save_sms(req:SmsReq, request:Request, authorization: str|None = Header(default=None)):
    admin=require_admin(authorization)
    name=req.member_name
    phone=req.phone
    if req.member_id:
        with con() as c:
            m=c.execute('SELECT name,phone FROM members WHERE id=?',(req.member_id,)).fetchone()
            if m:
                name=m['name'] if m['name'] else name
                phone=m['phone'] if m['phone'] else phone
    result = send_sms_provider(phone, req.body) if req.send_now else {'status':'saved','provider':get_setting_value('sms_provider','mock'),'message':'발송하지 않고 이력만 저장했습니다.'}
    sent_at = now() if result['status'] in ('sent','sent_mock') else ''
    with con() as c:
        cur=c.execute('INSERT INTO sms_logs(member_id,member_name,phone,round_no,body,combos,provider,status,result_message,sent_at,created_by,created_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)',(req.member_id,name,normalize_phone(phone),req.round_no,req.body,json.dumps(req.combos,ensure_ascii=False),result.get('provider','mock'),result.get('status','saved'),result.get('message',''),sent_at,admin['id'],now()))
        if req.member_id:
            c.execute('UPDATE members SET last_contact_at=?, updated_at=? WHERE id=?', (now(), now(), req.member_id))
        c.commit(); sid=cur.lastrowid
    action = 'SEND_SMS' if req.send_now else 'SAVE_SMS'
    log_action(admin,action,f'{req.round_no}회차 문자 {result.get("status")}',request)
    return {'id':sid, **result}

@app.post('/api/sms_log')
def save_sms_log_alias(req:SmsReq, request:Request, authorization: str|None = Header(default=None)):
    return save_sms(req, request, authorization)

@app.get('/api/sms')
def sms_logs(authorization: str|None = Header(default=None)):
    require_admin(authorization)
    with con() as c: rows=c.execute('SELECT * FROM sms_logs ORDER BY id DESC LIMIT 200').fetchall()
    return [dict(r) for r in rows]

@app.delete('/api/sms/{sms_id}')
def delete_sms_log(sms_id:int, request:Request, authorization: str|None = Header(default=None)):
    admin=require_admin(authorization)
    with con() as c:
        row=c.execute('SELECT id, member_id, member_name, round_no FROM sms_logs WHERE id=?',(sms_id,)).fetchone()
        if not row:
            raise HTTPException(404,'문구 이력을 찾을 수 없습니다.')
        if row['member_id']:
            assert_member_access(c, admin, row['member_id'])
        c.execute('DELETE FROM sms_logs WHERE id=?',(sms_id,))
        c.commit()
    log_action(admin,'DELETE_SMS_LOG',f"문구 이력 삭제 ID {sms_id} / {row['round_no'] or '-'}회 / {row['member_name'] or ''}",request)
    return {'ok':True,'deleted_id':sms_id}

@app.post('/api/win-check')
def win_check(req:WinReq, request:Request, authorization: str|None = Header(default=None)):
    admin=require_admin(authorization)
    wins=parse_nums(req.win_numbers)
    if not wins:
        d=get_draw(req.round_no)
        if not d: raise HTTPException(400,'당첨번호를 입력하거나 회차를 먼저 저장하세요.')
        wins=d['numbers']; req.bonus=d['bonus']
    results=[rank_result(c,wins,req.bonus) for c in req.combos]
    with con() as c:
        for r in results:
            c.execute('INSERT INTO winning_checks(member_id,member_name,round_no,target_numbers,win_numbers,bonus,match_count,bonus_match,rank,prize,cost,profit,roi,created_by,created_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)',(req.member_id,req.member_name,req.round_no,json.dumps(r['combo']),json.dumps(wins),req.bonus,r['match_count'],int(r['bonus_match']),r['rank'],r['prize'],r['cost'],r['profit'],r['roi'],admin['id'],now()))
        c.commit()
    summary={'count':len(results),'prize':sum(r['prize'] for r in results),'cost':sum(r['cost'] for r in results),'profit':sum(r['profit'] for r in results)}
    summary['roi']=round((summary['profit']/summary['cost']*100),2) if summary['cost'] else 0
    log_action(admin,'WIN_CHECK',f'{req.round_no}회차 수익률 계산',request); return {'wins':wins,'bonus':req.bonus,'results':results,'summary':summary}

@app.get('/api/win-checks')
def win_checks(authorization: str|None = Header(default=None)):
    require_admin(authorization)
    with con() as c: rows=c.execute('SELECT * FROM winning_checks ORDER BY id DESC LIMIT 300').fetchall()
    return [dict(r) for r in rows]

@app.get('/api/draws')
def draws(limit:int=100, authorization: str|None = Header(default=None)):
    require_admin(authorization)
    limit=max(1, min(200, int(limit or 100)))
    with con() as c: rows=c.execute('SELECT * FROM draws ORDER BY round_no DESC LIMIT ?', (limit,)).fetchall()
    return [{'round_no':r['round_no'],'draw_date':r['draw_date'],'numbers':parse_nums(r['numbers']),'bonus':r['bonus']} for r in rows]



@app.get('/api/draws/search')
def search_draw(round_no:int, authorization: str|None = Header(default=None)):
    """V3.0.0 STABLE: 1회차부터 추첨 완료 최신 회차까지 회차별 당첨번호 조회.
    DB에 없으면 동행복권 공개 조회를 시도하고, 성공 시 DB에 저장합니다.
    """
    require_admin(authorization)
    try:
        r = int(round_no or 0)
    except Exception:
        raise HTTPException(400, '회차는 숫자로 입력하세요.')
    expected, completed = _rc315_expected_round_and_completed()
    if r < 1:
        raise HTTPException(400, '1회차 이상만 조회할 수 있습니다.')
    if r > completed:
        return {
            'ok': False, 'round_no': r, 'expected_round': expected, 'completed_round': completed,
            'status': 'future', 'draw_date': draw_date_for_round(r), 'numbers': [], 'bonus': 0,
            'message': f'{r}회는 아직 추첨 완료 전입니다. 현재 조회 가능한 최신 완료 회차는 {completed}회입니다.'
        }
    draw = get_draw(r)
    source = 'db'
    if not draw:
        fetched = fetch_official_lotto(r)
        if fetched:
            saved = save_draw_if_missing(fetched)
            draw = saved or fetched
            source = draw.get('source', 'official')
    if draw and len(draw.get('numbers') or []) == 6:
        return {
            'ok': True, 'round_no': r, 'expected_round': expected, 'completed_round': completed,
            'status': 'saved', 'draw_date': draw.get('draw_date') or draw_date_for_round(r),
            'numbers': parse_nums(draw.get('numbers')), 'bonus': int(draw.get('bonus') or 0),
            'source': source, 'message': f'{r}회 당첨번호를 조회했습니다.'
        }
    return {
        'ok': False, 'round_no': r, 'expected_round': expected, 'completed_round': completed,
        'status': 'missing', 'draw_date': draw_date_for_round(r), 'numbers': [], 'bonus': 0,
        'message': f'{r}회 당첨번호가 DB에 없고 자동 조회도 실패했습니다. 인터넷 연결 또는 동행복권 조회 차단 여부를 확인하세요.'
    }

@app.get('/api/draws/next')
def next_draw_round(authorization: str|None = Header(default=None)):
    require_admin(authorization)
    expected = expected_lotto_round()
    check = resolve_draw_for_check(expected, allow_fetch=True)
    _, completed_round = _rc315_expected_round_and_completed()
    with con() as c:
        _rc315_clean_future_draws_in_conn(c)
        c.commit()
        latest=c.execute('SELECT * FROM draws WHERE round_no<=? ORDER BY round_no DESC LIMIT 1', (completed_round or 999999,)).fetchone()
    latest_obj = None
    latest_round = None
    if latest:
        latest_round = int(latest['round_no'])
        latest_obj = {'round_no':latest['round_no'], 'draw_date':latest['draw_date'], 'numbers':parse_nums(latest['numbers']), 'bonus':latest['bonus']}
    next_round = max(int(expected), int(latest_round or 0) + 1) if latest_round != expected else expected + 1
    return {
        'latest_round': latest_round,
        'expected_round': int(expected),
        'next_round': int(next_round),
        'check_round': int(check['round_no']),
        'draw_date': check.get('draw_date') or draw_date_for_round(expected),
        'draw_status': check.get('status'),
        'can_check': bool(check.get('can_check')),
        'latest': latest_obj,
        'current': {'round_no':check['round_no'], 'draw_date':check.get('draw_date',''), 'numbers':check.get('numbers',[]), 'bonus':check.get('bonus',0)} if check.get('numbers') else None,
        'check': check,
        'message': check.get('message','')
    }

@app.get('/api/draws/check-auto')
def check_draw_auto(round_no:int|None=None, authorization: str|None = Header(default=None)):
    require_admin(authorization)
    return resolve_draw_for_check(round_no or expected_lotto_round(), allow_fetch=True)

@app.post('/api/draws/fetch-official')
def fetch_draw_official_api(req: dict, request:Request, authorization: str|None = Header(default=None)):
    """RC3-10: 운영자가 회차를 강제로 공식/캐시 조회 후 DB에 저장할 수 있는 복구 API."""
    admin = require_admin(authorization)
    r = int((req or {}).get('round_no') or expected_lotto_round())
    fetched = fetch_official_lotto(r)
    if not fetched:
        raise HTTPException(404, f'{r}회 당첨번호를 공식 API/보조 캐시에서 찾지 못했습니다. 수동 입력이 필요합니다.')
    saved = save_draw_if_missing(fetched)
    log_action(admin, 'RC3_10_FETCH_DRAW', f'{r}회 당첨번호 자동조회/저장 source={saved.get("source")}', request)
    return {'ok': True, 'version': RC3_10_VERSION, 'draw': saved, 'message': f'{r}회 당첨번호를 저장했습니다.'}

@app.post('/api/draws')
def save_draw(req:DrawReq, request:Request, authorization: str|None = Header(default=None)):
    admin=require_admin(authorization)
    r, nums, bonus, expected, completed = _rc315_validate_draw_payload(req.round_no, req.numbers, req.bonus, allow_completed_only=True)
    with con() as c:
        _rc315_clean_future_draws_in_conn(c)
        c.execute('INSERT OR REPLACE INTO draws(round_no,draw_date,numbers,bonus,source,updated_at) VALUES(?,?,?,?,?,?)',(r,req.draw_date or draw_date_for_round(r),json.dumps(nums),bonus,'manual',now()))
        c.commit()
    log_action(admin,'SAVE_DRAW',f'{r}회차 당첨번호 저장 · 완료가능최신 {completed}회',request); return {'ok':True, 'round_no':r, 'numbers':nums, 'bonus':bonus, 'completed_round':completed}


def _auto_check_round(admin, req:AutoWinReq, request:Request):
    if not int(req.round_no or 0):
        req.round_no = expected_lotto_round()
    wins=parse_nums(req.winning)
    resolved = None
    # PHASE20: 당첨번호를 입력하지 않아도 회차 기준으로 자동 조회/저장 후 확인합니다.
    if len(wins)!=6 or not int(req.bonus or 0):
        resolved = resolve_draw_for_check(req.round_no, allow_fetch=True)
        if resolved.get('numbers') and resolved.get('bonus'):
            wins = parse_nums(resolved.get('numbers'))
            req.bonus = int(resolved.get('bonus') or 0)
            req.draw_date = req.draw_date or resolved.get('draw_date','')
        else:
            raise HTTPException(400, resolved.get('message') or '당첨번호가 아직 자동 확인되지 않았습니다.')
    if len(wins)!=6:
        raise HTTPException(400,'당첨번호 6개를 입력하세요.')
    if not (1 <= int(req.bonus) <= 45):
        raise HTTPException(400,'보너스 번호를 입력하세요.')
    if int(req.bonus) in wins:
        raise HTTPException(400,'보너스 번호는 당첨번호 6개와 달라야 합니다.')
    # RC3-15: 당첨확인 저장 회차는 반드시 실제 추첨 완료 회차까지만 허용합니다.
    req.round_no, wins, req.bonus, expected_round, completed_round = _rc315_validate_draw_payload(req.round_no, wins, req.bonus, allow_completed_only=True)
    with con() as c:
        _rc315_clean_future_draws_in_conn(c)
        c.execute('INSERT OR REPLACE INTO draws(round_no,draw_date,numbers,bonus,source,updated_at) VALUES(?,?,?,?,?,?)', (req.round_no, req.draw_date or draw_date_for_round(req.round_no), json.dumps(wins), int(req.bonus), (resolved.get('source') if resolved else 'manual_auto'), now()))
        # RC4-4 개선: 공동추천/회원 미지정 추천은 당첨확인 결과에서 제외하고, 회원별로 묶어서 확인합니다.
        rec_where = 'r.round_no=? AND COALESCE(r.member_id,0)>0 AND COALESCE(r.explicit_saved,0)=1'
        rec_args = [req.round_no]
        if not is_super_admin(admin):
            rec_where += ' AND COALESCE(m.created_by,0)=?'
            rec_args.append(int(admin.get('id') or 0))
        recs = [dict(r) for r in c.execute(f'''
            SELECT r.id,
                   r.member_id,
                   COALESCE(NULLIF(r.member_name,''), m.name, '회원명 미확인') AS member_name,
                   r.round_no,
                   r.numbers
            FROM recommendations r
            INNER JOIN members m ON m.id = r.member_id
            WHERE {rec_where}
            ORDER BY m.name ASC, r.id DESC
        ''', rec_args).fetchall()]

        # RC8.19: 기존 버전에서 '복사저장/보낸문자 저장/문구이력 저장'으로 남긴 번호도
        # 당첨확인 대상으로 복구합니다. sms_logs.combos가 비어 있지 않은 기록만 사용하며,
        # 추천번호 저장 기록과 같은 조합은 회원별로 중복 제거합니다.
        sms_where = "s.round_no=? AND COALESCE(s.member_id,0)>0 AND COALESCE(s.combos,'') NOT IN ('','[]')"
        sms_args = [req.round_no]
        if not is_super_admin(admin):
            sms_where += ' AND COALESCE(m.created_by,0)=?'
            sms_args.append(int(admin.get('id') or 0))
        sms_rows = c.execute(f'''
            SELECT s.id, s.member_id,
                   COALESCE(NULLIF(s.member_name,''), m.name, '회원명 미확인') AS member_name,
                   s.round_no, s.combos AS numbers
            FROM sms_logs s
            INNER JOIN members m ON m.id=s.member_id
            WHERE {sms_where}
            ORDER BY m.name ASC, s.id DESC
        ''', sms_args).fetchall()
        recs.extend({
            'id': -int(r['id'] or 0), 'member_id': r['member_id'],
            'member_name': r['member_name'], 'round_no': r['round_no'],
            'numbers': r['numbers']
        } for r in sms_rows)

        deduped_recs=[]
        seen_saved=set()
        for rec in recs:
            try:
                parsed=[parse_nums(x) for x in json.loads(rec.get('numbers') or '[]')]
                parsed=[x for x in parsed if len(x)==6]
            except Exception:
                parsed=[]
            if not parsed:
                continue
            canonical=json.dumps(parsed, ensure_ascii=False, separators=(',',':'))
            key=(int(rec.get('member_id') or 0), canonical)
            if key in seen_saved:
                continue
            seen_saved.add(key)
            rec['numbers']=json.dumps(parsed, ensure_ascii=False)
            deduped_recs.append(rec)
        recs=deduped_recs

        # 과거 '생성만 한 조합'으로 만들어진 당첨확인 결과를 같은 회차/회원 기준으로 정리합니다.
        explicit_member_ids=sorted({int(r['member_id'] or 0) for r in recs if int(r['member_id'] or 0)>0})
        if explicit_member_ids:
            marks=','.join(['?']*len(explicit_member_ids))
            c.execute(f'DELETE FROM winning_checks WHERE round_no=? AND member_id IN ({marks})', [req.round_no, *explicit_member_ids])
        checked=[]
        member_map={}
        rank_order={'1등':1,'2등':2,'3등':3,'4등':4,'5등':5,'낙첨':9}
        for rec in recs:
            try:
                combos=json.loads(rec['numbers'] or '[]')
            except Exception:
                combos=[]
            mid=int(rec['member_id'] or 0)
            mname=rec['member_name'] or '회원명 미확인'
            group=member_map.setdefault(mid, {
                'member_id': mid,
                'member_name': mname,
                'recommendation_count': 0,
                'total_combos': 0,
                'hit_count': 0,
                'lose_count': 0,
                'total_prize': 0,
                'best_rank': '낙첨',
                'best_prize': 0,
                'combos': []
            })
            if combos:
                group['recommendation_count'] += 1
            for idx, combo in enumerate(combos, start=1):
                r=rank_result(combo, wins, int(req.bonus))
                target=json.dumps(r['combo'], ensure_ascii=False)
                old=c.execute('SELECT id FROM winning_checks WHERE round_no=? AND COALESCE(member_id,0)=COALESCE(?,0) AND target_numbers=? AND win_numbers=? AND bonus=?', (req.round_no, rec['member_id'], target, json.dumps(wins), int(req.bonus))).fetchone()
                if old:
                    c.execute('UPDATE winning_checks SET member_name=?, match_count=?, bonus_match=?, rank=?, prize=?, cost=?, profit=?, roi=?, created_by=?, created_at=? WHERE id=?', (mname, r['match_count'], int(r['bonus_match']), r['rank'], r['prize'], r['cost'], r['profit'], r['roi'], admin['id'], now(), old['id']))
                else:
                    c.execute('INSERT INTO winning_checks(member_id,member_name,round_no,target_numbers,win_numbers,bonus,match_count,bonus_match,rank,prize,cost,profit,roi,created_by,created_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)', (rec['member_id'], mname, req.round_no, target, json.dumps(wins), int(req.bonus), r['match_count'], int(r['bonus_match']), r['rank'], r['prize'], r['cost'], r['profit'], r['roi'], admin['id'], now()))
                item={'member_id':mid, 'member_name':mname, 'recommendation_id':rec['id'], 'combo_index':idx, **r}
                checked.append(item)
                group['total_combos'] += 1
                group['total_prize'] += int(r.get('prize') or 0)
                if r.get('rank') and r.get('rank') != '낙첨':
                    group['hit_count'] += 1
                else:
                    group['lose_count'] += 1
                if rank_order.get(r.get('rank','낙첨'),9) < rank_order.get(group['best_rank'],9):
                    group['best_rank'] = r.get('rank') or '낙첨'
                    group['best_prize'] = int(r.get('prize') or 0)
                group['combos'].append(item)
        c.commit()
    member_results=list(member_map.values())
    member_results.sort(key=lambda x: (rank_order.get(x.get('best_rank','낙첨'),9), -int(x.get('total_prize') or 0), x.get('member_name') or ''))
    summary={'members':len(member_results), 'recommendations':len(recs), 'checked_combos':len(checked), 'hit_members':sum(1 for m in member_results if m.get('hit_count',0)>0), 'hit_combos':sum(1 for x in checked if x.get('rank')!='낙첨'), 'lose_combos':sum(1 for x in checked if x.get('rank')=='낙첨'), 'prize':sum(x['prize'] for x in checked), 'cost':sum(x['cost'] for x in checked), 'profit':sum(x['profit'] for x in checked)}
    summary['roi']=round((summary['profit']/summary['cost']*100),2) if summary['cost'] else 0
    log_action(admin,'AUTO_WIN_CHECK',f'{req.round_no}회차 회원별 자동 당첨확인 {len(member_results)}명/{len(checked)}조합',request)
    return {'ok':True, 'round_no':req.round_no, 'wins':wins, 'bonus':int(req.bonus), 'draw_date':req.draw_date, 'auto_resolved': bool(resolved), 'summary':summary, 'member_results':member_results, 'results':checked[:300]}

@app.post('/api/check_winning')
def check_winning_alias(req:AutoWinReq, request:Request, authorization: str|None = Header(default=None)):
    admin=require_admin(authorization)
    return _auto_check_round(admin, req, request)

@app.post('/api/win-check-auto')
def win_check_auto(req:AutoWinReq, request:Request, authorization: str|None = Header(default=None)):
    admin=require_admin(authorization)
    return _auto_check_round(admin, req, request)

@app.get('/api/stats')
def api_stats(limit:int=100, authorization: str|None = Header(default=None)):
    require_admin(authorization)
    st=latest_stats(limit)
    draws=st.get('draws',[])
    sums=[sum(d['numbers']) for d in draws]
    nums=[]
    for d in draws: nums.extend(d['numbers'])
    sections=[sum(1 for n in nums if n<=15),sum(1 for n in nums if 16<=n<=30),sum(1 for n in nums if n>=31)]
    odd=sum(1 for n in nums if n%2); even=len(nums)-odd
    pair_counter=collections.Counter()
    for d in draws:
        for a,b in itertools.combinations(d['numbers'],2):
            pair_counter[(a,b)]+=1
    return {
        'count': int(st.get('draw_count') or len(draws)),
        'display_count': len(draws),
        'limit': limit,
        'latest': draws[0] if draws else None,
        'recent_draws': draws if int(limit) <= 0 else draws[:max(100, int(limit))],
        'hot': st.get('hot', []),
        'cold': st.get('cold', []),
        'missing20': st.get('overdue', []),
        'odd': odd,
        'even': even,
        'sections': sections,
        'sum_min': min(sums) if sums else 0,
        'sum_avg': round(sum(sums)/len(sums),1) if sums else 0,
        'sum_max': max(sums) if sums else 0,
        'top_pairs': [{'pair':list(k),'count':v} for k,v in pair_counter.most_common(15)] or st.get('top_pairs', []),
        'freq': st.get('freq',{}),
        'freq100': st.get('freq100',{}),
        'engine_version': st.get('engine_version'),
        'analysis_confirm': st.get('analysis_confirm'),
        'round_range': st.get('round_range', []),
        'latest_round': st.get('latest_round', 0),
        'target_round': st.get('target_round', st.get('latest_round', 0)),
        'is_full_history': bool(st.get('is_full_history')),
        'missing_rounds_count': int(st.get('missing_rounds_count') or 0),
        'expected_count': int(st.get('expected_count') or 0),
        'actual_count': int(st.get('actual_count') or 0),
    }

def csv_response(filename, headers, rows):
    bio=io.StringIO()
    bio.write('\ufeff')
    w=csv.writer(bio); w.writerow(headers); w.writerows(rows)
    data=bio.getvalue().encode('utf-8-sig')
    return StreamingResponse(io.BytesIO(data), media_type='text/csv; charset=utf-8', headers={'Content-Disposition':f'attachment; filename={filename}'})

@app.get('/api/export/members_csv')
def export_members_csv(token: str|None=None, authorization: str|None = Header(default=None)):
    require_admin_any(authorization, token)
    with con() as c:
        rows=c.execute("SELECT m.id,m.name,m.phone,m.grade,m.status,COALESCE(m.priority,'보통') priority,COALESCE(m.source,'직접등록') source,m.last_contact_at,m.memo,m.created_at,COALESCE(m.contract_end_at,'') contract_end_at,COALESCE(m.contract_months,12) contract_months,COALESCE(a.name,a.username,'미지정') registered_by_name FROM members m LEFT JOIN admins a ON a.id=COALESCE(m.created_by,0) ORDER BY m.id DESC").fetchall()
    return csv_response('BBLOTTO_members.csv', ['ID','이름','연락처','등급','상태','우선순위','유입경로','최근연락','메모','등록일','계약만료일','등록관리자'], [[r['id'],r['name'],r['phone'],r['grade'],r['status'],r['priority'],r['source'],r['last_contact_at'],r['memo'],r['created_at'],r['contract_end_at'],r['registered_by_name']] for r in rows])

@app.get('/api/export/recommendations_csv')
def export_recommendations_csv(token: str|None=None, authorization: str|None = Header(default=None)):
    require_admin_any(authorization, token)
    with con() as c: rows=c.execute('SELECT id,member_name,round_no,mode,count,numbers,analysis,created_at FROM recommendations ORDER BY id DESC').fetchall()
    return csv_response('BBLOTTO_recommendations.csv', ['ID','회원','회차','모드','조합수','추천번호','분석','생성일'], [[r['id'],r['member_name'],r['round_no'],r['mode'],r['count'],r['numbers'],r['analysis'],r['created_at']] for r in rows])

@app.get('/api/export/winning_csv')
def export_winning_csv(token: str|None=None, authorization: str|None = Header(default=None)):
    require_admin_any(authorization, token)
    with con() as c: rows=c.execute('SELECT round_no,member_name,target_numbers,win_numbers,bonus,rank,prize,cost,profit,roi,created_at FROM winning_checks ORDER BY id DESC').fetchall()
    return csv_response('BBLOTTO_winning_checks.csv', ['회차','회원','추천번호','당첨번호','보너스','등수','당첨금','구매금','수익','수익률','확인일'], [[r['round_no'],r['member_name'],r['target_numbers'],r['win_numbers'],r['bonus'],r['rank'],r['prize'],r['cost'],r['profit'],r['roi'],r['created_at']] for r in rows])

@app.get('/api/backup_db')
def backup_db(token: str|None=None, authorization: str|None = Header(default=None)):
    admin = require_admin_any(authorization, token)
    if DB_ENGINE == 'postgresql':
        b = create_db_backup('download', admin)
        return FileResponse(EXPORT_DIR / b['filename'], media_type='application/json', filename=b['filename'])
    return FileResponse(DB, media_type='application/octet-stream', filename='BBLOTTO_lotto_backup.db')

@app.get('/api/export/excel')
def export_excel(token: str|None=None, authorization: str|None = Header(default=None)):
    require_admin_any(authorization, token)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except Exception:
        raise HTTPException(500,'openpyxl 설치가 필요합니다. pip install -r requirements.txt 를 실행하세요.')

    auth_header = authorization or ('Bearer '+token if token else None)
    data=dashboard(auth_header)
    wb=Workbook()

    header_fill = PatternFill('solid', fgColor='111111')
    gold_fill = PatternFill('solid', fgColor='D4AF37')
    gold_font = Font(bold=True, color='111111')
    white_font = Font(bold=True, color='FFFFFF')
    thin = Side(style='thin', color='D4AF37')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_sheet(ws):
        ws.freeze_panes = 'A2'
        for cell in ws[1]:
            cell.font = white_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = Border(bottom=Side(style='hair', color='555555'))
                cell.alignment = Alignment(vertical='top', wrap_text=True)
        for col in ws.columns:
            width = min(55, max(12, max(len(str(c.value or '')) for c in col) + 2))
            ws.column_dimensions[get_column_letter(col[0].column)].width = width

    ws=wb.active; ws.title='대시보드'
    ws.append(['항목','값'])
    dashboard_rows=[
        ('총 회원', data.get('members',0)),('관리자',data.get('admins',0)),('추천 생성',data.get('recommendations',0)),('문자 이력',data.get('sms',0)),('당첨 확인',data.get('checks',0)),('총 당첨금',data.get('total_prize',0)),('총 구매금',data.get('total_cost',0)),('총 수익',data.get('total_profit',0)),('수익률',str(data.get('roi',0))+'%'),('최근 회차',data.get('latest_round','-')),('생성일',now())]
    for r in dashboard_rows: ws.append(list(r))
    ws['A1'].fill=gold_fill; ws['B1'].fill=gold_fill; ws['A1'].font=gold_font; ws['B1'].font=gold_font
    style_sheet(ws)

    ws2=wb.create_sheet('회원목록'); ws2.append(['ID','이름','연락처','등급','상태','우선순위','유입경로','최근연락','메모','등록일'])
    with con() as c:
        members=c.execute("SELECT id,name,phone,grade,status,COALESCE(priority,'보통') priority,COALESCE(source,'직접등록') source,last_contact_at,memo,created_at FROM members ORDER BY id DESC").fetchall()
        recs=c.execute('SELECT id,member_name,round_no,mode,count,numbers,analysis,created_at FROM recommendations ORDER BY id DESC LIMIT 500').fetchall()
        wins=c.execute('SELECT round_no,member_name,target_numbers,win_numbers,bonus,rank,prize,cost,profit,roi,created_at FROM winning_checks ORDER BY id DESC LIMIT 500').fetchall()
        sms=c.execute('SELECT member_name,round_no,body,created_at FROM sms_logs ORDER BY id DESC LIMIT 500').fetchall()
        draws=c.execute('SELECT round_no,draw_date,numbers,bonus,source,updated_at FROM draws ORDER BY round_no DESC LIMIT 200').fetchall()
    for m in members: ws2.append([m['id'],m['name'],m['phone'],m['grade'],m['status'],m['priority'],m['source'],m['last_contact_at'],m['memo'],m['created_at']])
    style_sheet(ws2)

    ws3=wb.create_sheet('추천번호'); ws3.append(['ID','회원','회차','모드','조합수','추천번호','분석','생성일'])
    for r in recs: ws3.append([r['id'],r['member_name'],r['round_no'],r['mode'],r['count'],r['numbers'],r['analysis'],r['created_at']])
    style_sheet(ws3)

    ws4=wb.create_sheet('당첨확인_수익률'); ws4.append(['회차','회원','번호','당첨번호','보너스','등수','당첨금','구매금','수익','수익률%','확인일'])
    for w in wins: ws4.append([w['round_no'],w['member_name'],w['target_numbers'],w['win_numbers'],w['bonus'],w['rank'],w['prize'],w['cost'],w['profit'],w['roi'],w['created_at']])
    style_sheet(ws4)

    ws5=wb.create_sheet('문자이력'); ws5.append(['회원','회차','문자내용','저장일'])
    for x in sms: ws5.append([x['member_name'],x['round_no'],x['body'],x['created_at']])
    style_sheet(ws5)

    ws6=wb.create_sheet('당첨번호DB'); ws6.append(['회차','추첨일','당첨번호','보너스','출처','수정일'])
    for d in draws: ws6.append([d['round_no'],d['draw_date'],d['numbers'],d['bonus'],d['source'],d['updated_at']])
    style_sheet(ws6)

    bio=io.BytesIO(); wb.save(bio); bio.seek(0)
    return StreamingResponse(bio, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={'Content-Disposition':'attachment; filename=BBLOTTO_PRO_V34_FINAL_REPORT.xlsx'})


@app.get('/api/export/report_txt')
def export_report_txt(token: str|None=None, authorization: str|None = Header(default=None)):
    require_admin_any(authorization, token)
    auth_header = authorization or ('Bearer '+token if token else None)
    data=dashboard(auth_header)
    lines=[
        'BBLOTTO PRO V34 FINAL REPORT',
        f'생성일: {now()}',
        f'총 회원: {data.get("members",0)}',
        f'추천 생성: {data.get("recommendations",0)}',
        f'문자 이력: {data.get("sms",0)}',
        f'당첨 확인: {data.get("checks",0)}',
        f'총 당첨금: {data.get("total_prize",0)}',
        f'총 구매금: {data.get("total_cost",0)}',
        f'총 수익: {data.get("total_profit",0)}',
        f'수익률: {data.get("roi",0)}%',
    ]
    raw='\n'.join(lines).encode('utf-8-sig')
    return StreamingResponse(io.BytesIO(raw), media_type='text/plain; charset=utf-8', headers={'Content-Disposition':'attachment; filename=BBLOTTO_PRO_V34_FINAL_REPORT.txt'})

@app.get('/api/export/final_bundle')
def export_final_bundle(token: str|None=None, authorization: str|None = Header(default=None)):
    require_admin_any(authorization, token)
    import zipfile
    bio=io.BytesIO()
    auth_header = authorization or ('Bearer '+token if token else None)
    data=dashboard(auth_header)
    with zipfile.ZipFile(bio,'w',zipfile.ZIP_DEFLATED) as z:
        z.writestr('README_FINAL_PHASE5.txt','BBLOTTO PRO V34 FINAL PHASE5\n마감/최적화/백업 번들입니다.\n')
        z.writestr('dashboard_summary.json',json.dumps(data,ensure_ascii=False,indent=2))
        with con() as c:
            members=c.execute("SELECT id,name,phone,grade,status,COALESCE(priority,'보통') priority,COALESCE(source,'직접등록') source,last_contact_at,memo,created_at FROM members ORDER BY id DESC").fetchall()
            recs=c.execute('SELECT id,member_name,round_no,mode,count,numbers,analysis,created_at FROM recommendations ORDER BY id DESC').fetchall()
            wins=c.execute('SELECT * FROM winning_checks ORDER BY id DESC').fetchall()
        z.writestr('members.json',json.dumps([dict(x) for x in members],ensure_ascii=False,indent=2))
        z.writestr('recommendations.json',json.dumps([dict(x) for x in recs],ensure_ascii=False,indent=2))
        z.writestr('winning_checks.json',json.dumps([dict(x) for x in wins],ensure_ascii=False,indent=2))
        if DB.exists(): z.write(DB, 'database/lotto.db')
    bio.seek(0)
    return StreamingResponse(bio, media_type='application/zip', headers={'Content-Disposition':'attachment; filename=BBLOTTO_PRO_V34_FINAL_EXPORT_BUNDLE.zip'})

@app.get('/api/export/pdf')
def export_pdf(token: str|None=None, authorization: str|None = Header(default=None)):
    require_admin_any(authorization, token)
    # 외부 라이브러리 없이 간단한 PDF 리포트 생성(영문/숫자 중심, 한글 상세는 엑셀 사용 권장)
    auth_header = authorization or ('Bearer '+token if token else None)
    data=dashboard(auth_header)
    lines=['BBLOTTO PRO V34 REPORT', f'Generated: {now()}', '', f'Members: {data["members"]}', f'Admins: {data["admins"]}', f'Recommendations: {data["recommendations"]}', f'Win Checks: {data["checks"]}', f'Total Prize: {data["total_prize"]}', f'Total Cost: {data["total_cost"]}', f'Total Profit: {data["total_profit"]}', f'ROI: {data["roi"]}%']
    content='BT /F1 14 Tf 50 780 Td '
    escaped=[]
    for i,line in enumerate(lines):
        safe=line.replace('\\','\\\\').replace('(','\\(').replace(')','\\)')
        escaped.append(f'({safe}) Tj 0 -24 Td')
    stream=(content+' '.join(escaped)+' ET').encode('latin-1','ignore')
    objs=[]
    objs.append(b'1 0 obj << /Type /Catalog /Pages 2 0 R >> endobj')
    objs.append(b'2 0 obj << /Type /Pages /Kids [3 0 R] /Count 1 >> endobj')
    objs.append(b'3 0 obj << /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] /Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >> endobj')
    objs.append(b'4 0 obj << /Type /Font /Subtype /Type1 /BaseFont /Helvetica >> endobj')
    objs.append(b'5 0 obj << /Length '+str(len(stream)).encode()+b' >> stream\n'+stream+b'\nendstream endobj')
    pdf=b'%PDF-1.4\n'; offsets=[]
    for o in objs: offsets.append(len(pdf)); pdf+=o+b'\n'
    xref=len(pdf); pdf+=f'xref\n0 {len(objs)+1}\n0000000000 65535 f \n'.encode()
    for off in offsets: pdf+=f'{off:010d} 00000 n \n'.encode()
    pdf+=f'trailer << /Root 1 0 R /Size {len(objs)+1} >>\nstartxref\n{xref}\n%%EOF'.encode()
    return StreamingResponse(io.BytesIO(pdf), media_type='application/pdf', headers={'Content-Disposition':'attachment; filename=BBLOTTO_PRO_V34_REPORT.pdf'})


class TextPdfReq(BaseModel):
    text:str=''

@app.post('/api/export_pdf')
def export_text_pdf(req:TextPdfReq, authorization: str|None = Header(default=None), x_token: str|None = Header(default=None)):
    require_admin_any(authorization, x_token)
    # 간단 PDF: 한글은 환경에 따라 제한될 수 있어 텍스트 파일에 가까운 PDF 구조로 저장합니다.
    lines=(req.text or 'BBLOTTO PRO V34').splitlines()[:80]
    safe_lines=[]
    for line in lines:
        # PDF 기본 폰트는 한글 표시가 제한되므로 영문/숫자 중심으로 안전 처리
        safe=line.encode('latin-1','ignore').decode('latin-1') or 'BBLOTTO'
        safe=safe.replace('\\','\\\\').replace('(','\\(').replace(')','\\)')
        safe_lines.append(safe)
    content='BT /F1 12 Tf 45 800 Td ' + ' '.join([f'({line}) Tj 0 -18 Td' for line in safe_lines]) + ' ET'
    stream=content.encode('latin-1','ignore')
    objs=[b'1 0 obj << /Type /Catalog /Pages 2 0 R >> endobj', b'2 0 obj << /Type /Pages /Kids [3 0 R] /Count 1 >> endobj', b'3 0 obj << /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] /Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >> endobj', b'4 0 obj << /Type /Font /Subtype /Type1 /BaseFont /Helvetica >> endobj', b'5 0 obj << /Length '+str(len(stream)).encode()+b' >> stream\n'+stream+b'\nendstream endobj']
    pdf=b'%PDF-1.4\n'; offsets=[]
    for o in objs:
        offsets.append(len(pdf)); pdf+=o+b'\n'
    xref=len(pdf); pdf+=f'xref\n0 {len(objs)+1}\n0000000000 65535 f \n'.encode()
    for off in offsets: pdf+=f'{off:010d} 00000 n \n'.encode()
    pdf+=f'trailer << /Root 1 0 R /Size {len(objs)+1} >>\nstartxref\n{xref}\n%%EOF'.encode()
    return StreamingResponse(io.BytesIO(pdf), media_type='application/pdf', headers={'Content-Disposition':'attachment; filename=BBLOTTO_RECOMMENDATION.pdf'})

# -----------------------------------------------------------------------------
# V34 FINAL - AI ENGINE PHASE2 OVERRIDES
# 기존 API/DB를 건드리지 않고 추천엔진 함수만 후단에서 재정의합니다.
# 반영: 최근 10/30/50/100회 가중치, HOT/COLD, 미출현, 동반출현, AC값,
#      합계/홀짝/구간/끝수/연속수 필터, 엔진 요약 리턴.
# -----------------------------------------------------------------------------

def _draw_rows_for_ai(limit=120):
    with con() as c:
        rows = c.execute('SELECT * FROM draws ORDER BY round_no DESC LIMIT ?', (int(limit),)).fetchall()
    draws=[]
    for r in rows:
        nums=parse_nums(r['numbers'])
        if len(nums)==6:
            draws.append({'round_no':r['round_no'], 'draw_date':r['draw_date'], 'numbers':nums, 'bonus':r['bonus']})
    return draws

def _ai_ac_value(nums):
    nums=sorted(nums)
    diffs=set()
    for i,a in enumerate(nums):
        for b in nums[i+1:]:
            diffs.add(abs(b-a))
    return max(0, len(diffs)-5)

def _ai_consecutive(nums):
    nums=sorted(nums)
    return sum(1 for a,b in zip(nums, nums[1:]) if b-a==1)

def _ai_sections(nums):
    return [sum(1<=n<=15 for n in nums), sum(16<=n<=30 for n in nums), sum(31<=n<=45 for n in nums)]

def _ai_pair_key(a,b):
    return tuple(sorted((int(a), int(b))))

def latest_stats(limit=120):
    """AI Phase2 통계팩. 기존 함수명 유지로 충돌 방지."""
    draws=_draw_rows_for_ai(limit)
    freq_windows={10:{n:0 for n in range(1,46)}, 30:{n:0 for n in range(1,46)}, 50:{n:0 for n in range(1,46)}, 100:{n:0 for n in range(1,46)}}
    last_seen={n:999 for n in range(1,46)}
    pair_counter=collections.Counter()
    end_counts={i:0 for i in range(10)}
    zone_counts={'1~15':0,'16~30':0,'31~45':0}
    sums=[]; acs=[]; odd_total=0
    for idx,d in enumerate(draws):
        nums=d['numbers']
        if idx < 30:
            sums.append(sum(nums)); acs.append(_ai_ac_value(nums)); odd_total += sum(n%2 for n in nums)
            for n in nums:
                end_counts[n%10]+=1
                if n<=15: zone_counts['1~15']+=1
                elif n<=30: zone_counts['16~30']+=1
                else: zone_counts['31~45']+=1
            for a,b in itertools.combinations(nums,2):
                pair_counter[_ai_pair_key(a,b)] += 1
        for w in freq_windows:
            if idx < w:
                for n in nums:
                    freq_windows[w][n]+=1
        for n in nums:
            if last_seen[n] == 999:
                last_seen[n]=idx
    weighted_score={}
    for n in range(1,46):
        # 최신 회차일수록 강하게 반영하되 100회 장기흐름도 유지
        score=(freq_windows[10][n]*2.40 + freq_windows[30][n]*1.45 + freq_windows[50][n]*1.05 + freq_windows[100][n]*0.72)
        # 직전회차 번호 과다 반복 방지
        if last_seen[n] == 0: score -= 1.4
        # 오래 안 나온 번호는 반등 후보로 보정
        if last_seen[n] >= 12: score += 1.2
        if last_seen[n] >= 20: score += 1.0
        weighted_score[n]=round(score,3)
    hot=sorted(range(1,46), key=lambda n:(-weighted_score[n], last_seen[n], n))[:15]
    cold=sorted(range(1,46), key=lambda n:(freq_windows[30][n], -last_seen[n], n))[:15]
    overdue=sorted(range(1,46), key=lambda n:(-last_seen[n], freq_windows[100][n], n))[:15]
    mid=sorted(range(1,46), key=lambda n:(abs(weighted_score[n]-(sum(weighted_score.values())/45)), n))[:15]
    pair_top=[{'pair':list(k),'count':v} for k,v in pair_counter.most_common(20)]
    return {
        'draws':draws,
        'freq':freq_windows[100],
        'freq10':freq_windows[10],
        'freq30':freq_windows[30],
        'freq50':freq_windows[50],
        'freq100':freq_windows[100],
        'weighted_score':weighted_score,
        'last_seen':last_seen,
        'hot':hot,
        'cold':cold,
        'mid':mid,
        'overdue':overdue,
        'pair_counter':pair_counter,
        'pair_top':pair_top,
        'end_counts':end_counts,
        'zone_counts':zone_counts,
        'odd_ratio':odd_total/(max(1,len(draws[:30])*6)),
        'avg_sum30':round(sum(sums)/len(sums),1) if sums else 0,
        'avg_ac30':round(sum(acs)/len(acs),1) if acs else 0,
        'latest_round':draws[0]['round_no'] if draws else 1230,
    }

def _ai_weight_pool(st, mode='balanced'):
    weights={}
    hot=set(st['hot'][:12]); cold=set(st['cold'][:12]); overdue=set(st['overdue'][:12])
    avg=(sum(st['weighted_score'].values())/45) if st.get('weighted_score') else 1
    for n in range(1,46):
        w=1.0 + max(0, st['weighted_score'].get(n,0))/max(1,avg)*0.55
        if n in hot: w += 1.55
        if n in overdue: w += 1.05
        if n in cold: w += 0.55
        # 극단 구간 보정: 완전히 배제하지 않고 모드별로 가중치 조절
        if mode == 'conservative':
            if 11 <= n <= 35: w += 0.65
            if n in hot: w += 0.35
        elif mode == 'aggressive':
            if n in overdue or n in cold: w += 0.75
            if n <= 10 or n >= 36: w += 0.30
        else:
            if 16 <= n <= 40: w += 0.30
        weights[n]=round(max(0.15,w),4)
    return weights

def _weighted_choice_from_list(pool, weights):
    usable=[n for n in pool if n in weights]
    if not usable:
        return None
    total=sum(weights[n] for n in usable)
    r=random.random()*total
    for n in usable:
        r-=weights[n]
        if r <= 0:
            return n
    return usable[-1]

def _combo_pair_score(combo, st):
    pc=st.get('pair_counter') or collections.Counter()
    return sum(pc.get(_ai_pair_key(a,b),0) for a,b in itertools.combinations(combo,2))

def combo_score(combo, st):
    """AI Phase2 최종 점수. 기존 함수명 유지."""
    combo=sorted(parse_nums(combo));
    if len(combo)!=6: return 0
    freq=st.get('freq100') or st.get('freq') or {n:0 for n in range(1,46)}
    hot_hit=len(set(combo)&set(st['hot'][:10])); cold_hit=len(set(combo)&set(st['cold'][:10])); overdue_hit=len(set(combo)&set(st['overdue'][:10]))
    odd=sum(n%2 for n in combo); total=sum(combo); zones=_ai_sections(combo)
    cons=_ai_consecutive(combo); ends=len(set(n%10 for n in combo)); ac=_ai_ac_value(combo); pair_score=_combo_pair_score(combo, st)
    score=62
    if odd in (2,3,4): score+=10
    else: score-=12
    if max(zones)<=3 and min(zones)>=1: score+=10
    elif max(zones)==4: score+=2
    else: score-=10
    if 95<=total<=180: score+=9
    elif 85<=total<=195: score+=3
    else: score-=12
    if 6<=ac<=10: score+=9
    elif 5<=ac<=12: score+=4
    else: score-=8
    if cons<=1: score+=6
    elif cons==2: score+=1
    else: score-=9
    if ends>=5: score+=5
    elif ends==4: score+=2
    else: score-=6
    if 1<=hot_hit<=3: score+=5
    elif hot_hit>=4: score-=2
    if 1<=cold_hit<=3: score+=4
    if 1<=overdue_hit<=3: score+=5
    if pair_score>=5: score+=4
    elif pair_score>=2: score+=2
    score += min(5, sum(freq.get(n,0) for n in combo)//8)
    return max(0, min(99, int(round(score))))

def tags_for_combo(combo, st):
    combo=sorted(parse_nums(combo)); tags=[]
    if len(set(combo)&set(st['hot'][:10]))>=2: tags.append('HOT 흐름 반영')
    if len(set(combo)&set(st['overdue'][:10]))>=1: tags.append('미출현 보강')
    if len(set(combo)&set(st['cold'][:10]))>=1: tags.append('COLD 변동 후보')
    odd=sum(n%2 for n in combo)
    if odd in (2,3,4): tags.append('홀짝 균형')
    zones=_ai_sections(combo)
    if max(zones)<=3 and min(zones)>=1: tags.append('구간 분산')
    if 6 <= _ai_ac_value(combo) <= 10: tags.append('AC 적정')
    if len(set(n%10 for n in combo))>=5: tags.append('끝수 분산')
    if _combo_pair_score(combo, st)>=2: tags.append('동반출현 반영')
    return tags[:5] or ['균형형 조합']

def _combo_detail(combo, st):
    c=sorted(combo)
    return {
        'numbers':c,
        'score':combo_score(c,st),
        'tags':tags_for_combo(c,st),
        'sum':sum(c),
        'odd':sum(n%2 for n in c),
        'even':6-sum(n%2 for n in c),
        'zones':_ai_sections(c),
        'ac':_ai_ac_value(c),
        'consecutive':_ai_consecutive(c),
        'end_unique':len(set(n%10 for n in c)),
        'hot_hit':len(set(c)&set(st['hot'][:10])),
        'cold_hit':len(set(c)&set(st['cold'][:10])),
        'overdue_hit':len(set(c)&set(st['overdue'][:10])),
        'pair_score':_combo_pair_score(c,st),
    }

def _engine_summary(details, st):
    if not details:
        return {'avg_score':0,'max_score':0,'min_score':0,'filters':['생성 데이터 없음']}
    avg_score=round(sum(d['score'] for d in details)/len(details),1)
    sums=[d['sum'] for d in details]
    acs=[d['ac'] for d in details]
    pair_avg=round(sum(d['pair_score'] for d in details)/len(details),1)
    odd=sum(d['odd'] for d in details); even=sum(d['even'] for d in details)
    sections=[sum(d['zones'][i] for d in details) for i in range(3)]
    return {
        'avg_score':avg_score,
        'max_score':max(d['score'] for d in details),
        'min_score':min(d['score'] for d in details),
        'avg_sum':round(sum(sums)/len(sums),1),
        'avg_ac':round(sum(acs)/len(acs),1),
        'pair_avg':pair_avg,
        'odd_even':{'odd':odd,'even':even},
        'sections':sections,
        'hot':st['hot'][:8],
        'cold':st['cold'][:8],
        'overdue':st['overdue'][:8],
        'pair_top':st.get('pair_top',[])[:5],
        'filters':['홀짝 2:4~4:2','구간 몰림 제한','합계 90~190 우선','AC 5~12 우선','연속수 2쌍 이하','끝수 중복 제한','동반출현 보정'],
    }

def make_premium_combos(count=10, fixed='', excluded='', mode='balanced'):
    """AI Phase2 추천번호 생성. 기존 함수명 유지."""
    st=latest_stats(120)
    target=max(1,min(50,int(count or 10)))
    fixed_set=set(parse_nums(fixed)); excluded_set=set(parse_nums(excluded)) - fixed_set
    pool=[n for n in range(1,46) if n not in excluded_set and n not in fixed_set]
    if len(pool)+len(fixed_set)<6:
        return make_combos(target, fixed, excluded, mode), [], st
    base_weights=_ai_weight_pool(st, mode)
    for n in excluded_set:
        base_weights.pop(n, None)
    candidates=[]; seen=set(); guard=0
    buckets={
        'hot':[n for n in st['hot'] if n in pool],
        'mid':[n for n in st['mid'] if n in pool],
        'cold':[n for n in st['cold'] if n in pool],
        'overdue':[n for n in st['overdue'] if n in pool],
        'pool':pool,
    }
    while len(candidates) < target*12 and guard < 26000:
        guard += 1
        combo=set(fixed_set)
        # 모드별 후보 시드 구성
        if mode == 'aggressive':
            plan=['hot','overdue','cold','pool','pool','pool']
        elif mode == 'conservative':
            plan=['hot','mid','mid','pool','pool','pool']
        else:
            plan=['hot','overdue','mid','cold','pool','pool']
        for name in plan:
            if len(combo)>=6: break
            usable=[n for n in buckets.get(name, pool) if n not in combo]
            pick=_weighted_choice_from_list(usable, base_weights)
            if pick: combo.add(pick)
        while len(combo)<6:
            pick=_weighted_choice_from_list([n for n in pool if n not in combo], base_weights)
            if pick is None: break
            combo.add(pick)
        arr=tuple(sorted(combo))
        if len(arr)!=6 or arr in seen: continue
        seen.add(arr)
        odd=sum(n%2 for n in arr); total=sum(arr); zones=_ai_sections(arr); cons=_ai_consecutive(arr); ac=_ai_ac_value(arr); ends=len(set(n%10 for n in arr))
        # 필터는 너무 빡빡하지 않게 하되 품질 하한 확보
        if odd not in (2,3,4): continue
        if max(zones)>4 or min(zones)==0: continue
        if not (85<=total<=195): continue
        if cons>2: continue
        if not (4<=ac<=12): continue
        if ends<4: continue
        candidates.append((combo_score(arr,st),list(arr)))
    candidates=sorted(candidates, key=lambda x:(-x[0], x[1]))
    combos=[c for _,c in candidates[:target]]
    if len(combos)<target:
        fallback=make_combos(target-len(combos), fixed, excluded, mode)
        for f in fallback:
            if f not in combos:
                combos.append(f)
            if len(combos)>=target: break
    details=[_combo_detail(c,st) for c in combos[:target]]
    return combos[:target], details, st

def build_analysis_text(round_no, st, mode, fixed, excluded):
    """AI Phase2 분석 문장: 짧게 3~5줄, 번호 흐름 기반."""
    pack=_load_analysis_pack(); used=set(); cond=pack.get('conditional',{}) or {}
    lines=[]
    target=random.choice([3,4,4,5])
    # 통계 기반 문장 먼저 구성
    hot_txt=', '.join(map(str, st.get('hot',[])[:3]))
    overdue_txt=', '.join(map(str, st.get('overdue',[])[:3]))
    if random.random()<0.55:
        lines.append(f'최근 흐름에서는 핵심 번호 {hot_txt}번대의 반영 가치가 높게 나타났습니다.')
    else:
        lines.append(_pick(pack.get('start'), used) or '이번 회차는 최근 흐름을 기준으로 균형 있게 분석했습니다.')
    if len(lines)<target and overdue_txt:
        lines.append(f'장기 미출현 후보 {overdue_txt}번을 보조 흐름으로 함께 검토했습니다.')
    mode_key=mode if mode in ('balanced','conservative','aggressive') else 'balanced'
    candidates=[]; candidates += cond.get(mode_key, [])
    if fixed: candidates += cond.get('fixed', [])
    if excluded: candidates += cond.get('excluded', [])
    if candidates and len(lines)<target:
        line=_pick(candidates, used)
        if line: lines.append(line)
    for cat in random.sample(['flow','core','pattern','strategy','ending'],5):
        if len(lines)>=target: break
        line=_pick(pack.get(cat), used)
        if line: lines.append(line)
    if len(lines)<3:
        lines.append('홀짝, 구간, 끝수, 합계, AC값을 함께 고려해 조합을 정리했습니다.')
    return '\n'.join([x for x in lines if x][:5])

@app.get('/api/ai-engine/summary')
def ai_engine_summary(authorization: str|None = Header(default=None)):
    require_admin(authorization)
    st=latest_stats(120)
    return {
        'latest_round':st.get('latest_round'),
        'hot':st['hot'][:12],
        'cold':st['cold'][:12],
        'overdue':st['overdue'][:12],
        'avg_sum30':st.get('avg_sum30'),
        'avg_ac30':st.get('avg_ac30'),
        'pair_top':st.get('pair_top',[])[:10],
        'end_counts':st.get('end_counts'),
        'zone_counts':st.get('zone_counts'),
    }

# =========================================================
# BBLOTTO PRO V34 - PRIORITY 1 VIP AI ENGINE FINAL PATCH
# 기존 API/프론트 호출명은 그대로 유지하고, 엔진 로직만 최종 우선 적용합니다.
# =========================================================

V34_AI_ENGINE_VERSION = 'V34_PRIORITY1_VIP_AI_FINAL'

def _v34_draws(limit=200):
    """draws 테이블 구조(numbers JSON)을 기준으로 최근 회차를 안전하게 읽습니다."""
    with con() as c:
        rows = c.execute('SELECT * FROM draws ORDER BY round_no DESC LIMIT ?', (int(limit),)).fetchall()
    draws=[]
    for r in rows:
        nums=parse_nums(r['numbers'] if 'numbers' in r.keys() else '')
        if len(nums)==6:
            draws.append({'round_no':r['round_no'], 'draw_date':r['draw_date'], 'numbers':nums, 'bonus':r['bonus']})
    return draws

def _v34_ac(nums):
    nums=sorted(parse_nums(nums)); diffs=set()
    for i,a in enumerate(nums):
        for b in nums[i+1:]: diffs.add(abs(b-a))
    return max(0, len(diffs)-5)

def _v34_cons(nums):
    nums=sorted(parse_nums(nums)); return sum(1 for a,b in zip(nums,nums[1:]) if b-a==1)

def _v34_zones(nums):
    nums=parse_nums(nums); return [sum(1<=n<=15 for n in nums), sum(16<=n<=30 for n in nums), sum(31<=n<=45 for n in nums)]

def _v34_similarity(a,b):
    return len(set(a)&set(b))

def _v34_latest_stats(limit=200):
    draws=_v34_draws(limit)
    windows=[10,30,50,100,200]
    freq={w:{n:0 for n in range(1,46)} for w in windows}
    last_seen={n:999 for n in range(1,46)}
    pair_counter=collections.Counter(); triple_counter=collections.Counter()
    end_counts={i:0 for i in range(10)}; zone_counts={'1~15':0,'16~30':0,'31~45':0}
    sums=[]; acs=[]; odd_total=0
    latest_nums=set(draws[0]['numbers']) if draws else set()
    for idx,d in enumerate(draws):
        nums=d['numbers']
        for w in windows:
            if idx<w:
                for n in nums: freq[w][n]+=1
        for n in nums:
            if last_seen[n]==999: last_seen[n]=idx
        if idx<50:
            for a,b in itertools.combinations(nums,2): pair_counter[tuple(sorted((a,b)))] += 1
            for t in itertools.combinations(nums,3): triple_counter[tuple(sorted(t))] += 1
        if idx<30:
            sums.append(sum(nums)); acs.append(_v34_ac(nums)); odd_total += sum(n%2 for n in nums)
            for n in nums:
                end_counts[n%10]+=1
                if n<=15: zone_counts['1~15']+=1
                elif n<=30: zone_counts['16~30']+=1
                else: zone_counts['31~45']+=1
    weighted={}
    for n in range(1,46):
        score=freq[10][n]*2.80 + freq[30][n]*1.55 + freq[50][n]*1.10 + freq[100][n]*0.75 + freq[200][n]*0.35
        # 직전 회차 번호는 완전 제외하지 않고 과반영만 방지
        if n in latest_nums: score -= 1.75
        # 미출현/저출현 반등 후보 보정
        if last_seen[n] >= 10: score += 0.75
        if last_seen[n] >= 18: score += 1.15
        if last_seen[n] >= 25: score += 1.30
        weighted[n]=round(score,4)
    avg=sum(weighted.values())/45 if weighted else 1
    hot=sorted(range(1,46), key=lambda n:(-weighted[n], last_seen[n], n))[:18]
    cold=sorted(range(1,46), key=lambda n:(freq[30][n], -last_seen[n], n))[:18]
    overdue=sorted(range(1,46), key=lambda n:(-last_seen[n], freq[100][n], n))[:18]
    mid=sorted(range(1,46), key=lambda n:(abs(weighted[n]-avg), n))[:18]
    return {
        'engine_version':V34_AI_ENGINE_VERSION, 'draws':draws, 'latest_round':draws[0]['round_no'] if draws else 0,
        'freq':freq[100], 'freq10':freq[10], 'freq30':freq[30], 'freq50':freq[50], 'freq100':freq[100], 'freq200':freq[200],
        'weighted_score':weighted, 'last_seen':last_seen, 'hot':hot, 'cold':cold, 'mid':mid, 'overdue':overdue,
        'pair_counter':pair_counter, 'triple_counter':triple_counter,
        'pair_top':[{'pair':list(k),'count':v} for k,v in pair_counter.most_common(20)],
        'end_counts':end_counts, 'zone_counts':zone_counts,
        'odd_ratio':round(odd_total/max(1,len(draws[:30])*6),3),
        'avg_sum30':round(sum(sums)/len(sums),1) if sums else 0,
        'avg_ac30':round(sum(acs)/len(acs),1) if acs else 0,
    }

def latest_stats(limit=200):
    return _v34_latest_stats(limit)

def _v34_weights(st, mode='balanced'):
    avg=sum(st['weighted_score'].values())/45 if st.get('weighted_score') else 1
    hot=set(st['hot'][:12]); cold=set(st['cold'][:12]); overdue=set(st['overdue'][:12]); mid=set(st['mid'][:12])
    weights={}
    for n in range(1,46):
        w=1.0 + max(0, st['weighted_score'].get(n,0))/max(1,avg)*0.70
        if n in hot: w += 1.65
        if n in overdue: w += 1.25
        if n in cold: w += 0.70
        if n in mid: w += 0.35
        if mode=='conservative':
            if 11<=n<=35: w += 0.75
            if n in hot or n in mid: w += 0.35
        elif mode=='aggressive':
            if n in overdue or n in cold: w += 0.85
            if n<=10 or n>=36: w += 0.35
        else:
            if 8<=n<=40: w += 0.30
        weights[n]=max(0.1, round(w,4))
    return weights

def _v34_choice(pool, weights):
    pool=[n for n in pool if n in weights]
    if not pool: return None
    total=sum(weights[n] for n in pool); r=random.random()*total
    for n in pool:
        r-=weights[n]
        if r<=0: return n
    return pool[-1]

def _v34_pair_score(combo, st):
    pc=st.get('pair_counter') or collections.Counter()
    return sum(pc.get(tuple(sorted((a,b))),0) for a,b in itertools.combinations(combo,2))

def _v34_triple_penalty(combo, st):
    tc=st.get('triple_counter') or collections.Counter()
    # 최근 50회에서 동일 3수 묶음이 너무 강하면 과적합 방지로 감점
    return sum(1 for t in itertools.combinations(combo,3) if tc.get(tuple(sorted(t)),0) >= 2)

def combo_score(combo, st):
    c=sorted(parse_nums(combo))
    if len(c)!=6: return 0
    odd=sum(n%2 for n in c); total=sum(c); zones=_v34_zones(c); ac=_v34_ac(c); cons=_v34_cons(c); ends=len(set(n%10 for n in c))
    hot_hit=len(set(c)&set(st['hot'][:10])); cold_hit=len(set(c)&set(st['cold'][:10])); overdue_hit=len(set(c)&set(st['overdue'][:10]))
    pair_score=_v34_pair_score(c,st); triple_penalty=_v34_triple_penalty(c,st)
    score=60
    score += 12 if odd in (2,3,4) else -14
    if max(zones)<=3 and min(zones)>=1: score+=12
    elif max(zones)==4 and min(zones)>=1: score+=4
    else: score-=13
    if 95<=total<=180: score+=10
    elif 85<=total<=195: score+=4
    else: score-=14
    if 6<=ac<=10: score+=10
    elif 5<=ac<=12: score+=4
    else: score-=9
    if cons<=1: score+=7
    elif cons==2: score+=1
    else: score-=10
    score += 5 if ends>=5 else (2 if ends==4 else -7)
    if 1<=hot_hit<=3: score+=6
    elif hot_hit>=4: score-=3
    if 1<=cold_hit<=3: score+=4
    if 1<=overdue_hit<=3: score+=5
    if 3<=pair_score<=14: score+=5
    elif pair_score>18: score-=4
    score -= triple_penalty*3
    score += min(5, sum((st.get('freq100') or st.get('freq')).get(n,0) for n in c)//8)
    return max(0, min(99, int(round(score))))

def tags_for_combo(combo, st):
    c=sorted(parse_nums(combo)); tags=[]
    if len(set(c)&set(st['hot'][:10]))>=2: tags.append('핵심수 반영')
    if len(set(c)&set(st['overdue'][:10]))>=1: tags.append('미출현 보강')
    if len(set(c)&set(st['cold'][:10]))>=1: tags.append('저출현 후보')
    if sum(n%2 for n in c) in (2,3,4): tags.append('홀짝 균형')
    if max(_v34_zones(c))<=3 and min(_v34_zones(c))>=1: tags.append('구간 분산')
    if 6<=_v34_ac(c)<=10: tags.append('AC 적정')
    if len(set(n%10 for n in c))>=5: tags.append('끝수 분산')
    if 3<=_v34_pair_score(c,st)<=14: tags.append('동반출현 반영')
    return tags[:5] or ['VIP 균형 조합']

def _combo_detail(combo, st):
    c=sorted(parse_nums(combo))
    return {'numbers':c,'score':combo_score(c,st),'tags':tags_for_combo(c,st),'sum':sum(c),'odd':sum(n%2 for n in c),'even':6-sum(n%2 for n in c),'zones':_v34_zones(c),'ac':_v34_ac(c),'consecutive':_v34_cons(c),'end_unique':len(set(n%10 for n in c)),'hot_hit':len(set(c)&set(st['hot'][:10])),'cold_hit':len(set(c)&set(st['cold'][:10])),'overdue_hit':len(set(c)&set(st['overdue'][:10])),'pair_score':_v34_pair_score(c,st)}

def _engine_summary(details, st):
    if not details:
        return {'engine_version':V34_AI_ENGINE_VERSION,'avg_score':0,'filters':['생성 데이터 없음']}
    return {
        'engine_version':V34_AI_ENGINE_VERSION,
        'avg_score':round(sum(d['score'] for d in details)/len(details),1),
        'max_score':max(d['score'] for d in details), 'min_score':min(d['score'] for d in details),
        'avg_sum':round(sum(d['sum'] for d in details)/len(details),1),
        'avg_ac':round(sum(d['ac'] for d in details)/len(details),1),
        'pair_avg':round(sum(d['pair_score'] for d in details)/len(details),1),
        'odd_even':{'odd':sum(d['odd'] for d in details),'even':sum(d['even'] for d in details)},
        'sections':[sum(d['zones'][i] for d in details) for i in range(3)],
        'hot':st['hot'][:8], 'cold':st['cold'][:8], 'overdue':st['overdue'][:8], 'pair_top':st.get('pair_top',[])[:5],
        'filters':['최근 10/30/50/100/200회 가중치','동반출현 보정','과거 3수 과적합 감점','홀짝 2:4~4:2','구간 몰림 제한','합계 85~195','AC 4~12','연속수 2쌍 이하','끝수 4개 이상','조합간 유사도 제한'],
    }

def make_premium_combos(count=10, fixed='', excluded='', mode='balanced'):
    st=latest_stats(200); target=max(1,min(50,int(count or 10)))
    fixed_set=set(parse_nums(fixed)); excluded_set=set(parse_nums(excluded))-fixed_set
    if len(fixed_set)>6: fixed_set=set(sorted(fixed_set)[:6])
    pool=[n for n in range(1,46) if n not in excluded_set and n not in fixed_set]
    if len(pool)+len(fixed_set)<6:
        combos=make_combos(target, fixed, excluded, mode)
        return combos, [_combo_detail(c,st) for c in combos], st
    weights=_v34_weights(st, mode)
    for n in excluded_set: weights.pop(n,None)
    buckets={
        'hot':[n for n in st['hot'] if n in pool], 'mid':[n for n in st['mid'] if n in pool],
        'cold':[n for n in st['cold'] if n in pool], 'overdue':[n for n in st['overdue'] if n in pool], 'pool':pool,
    }
    plans = {
        'aggressive':['hot','overdue','cold','pool','pool','pool'],
        'conservative':['hot','mid','mid','pool','pool','pool'],
        'balanced':['hot','overdue','mid','cold','pool','pool'],
    }
    plan=plans.get(mode, plans['balanced'])
    candidates=[]; seen=set(); guard=0
    while len(candidates)<target*16 and guard<42000:
        guard+=1; combo=set(fixed_set)
        for name in plan:
            if len(combo)>=6: break
            pick=_v34_choice([n for n in buckets.get(name,pool) if n not in combo], weights)
            if pick: combo.add(pick)
        while len(combo)<6:
            pick=_v34_choice([n for n in pool if n not in combo], weights)
            if pick is None: break
            combo.add(pick)
        arr=tuple(sorted(combo))
        if len(arr)!=6 or arr in seen: continue
        seen.add(arr)
        odd=sum(n%2 for n in arr); total=sum(arr); zones=_v34_zones(arr); ac=_v34_ac(arr); cons=_v34_cons(arr); ends=len(set(n%10 for n in arr))
        if odd not in (2,3,4): continue
        if max(zones)>4 or min(zones)==0: continue
        if not (85<=total<=195): continue
        if not (4<=ac<=12): continue
        if cons>2 or ends<4: continue
        score=combo_score(arr,st)
        if score<78: continue
        candidates.append((score,list(arr)))
    candidates=sorted(candidates, key=lambda x:(-x[0], x[1]))
    combos=[]
    for score,c in candidates:
        # 회원 배포용으로 너무 비슷한 조합이 반복되지 않게 제한
        if all(_v34_similarity(c, old)<=3 for old in combos): combos.append(c)
        if len(combos)>=target: break
    if len(combos)<target:
        for _,c in candidates:
            if c not in combos: combos.append(c)
            if len(combos)>=target: break
    if len(combos)<target:
        for f in make_combos(target-len(combos), fixed, excluded, mode):
            if f not in combos: combos.append(f)
            if len(combos)>=target: break
    details=[_combo_detail(c,st) for c in combos[:target]]
    return combos[:target], details, st

def build_analysis_text(round_no, st, mode, fixed, excluded):
    hot=', '.join(map(str,st.get('hot',[])[:4])); overdue=', '.join(map(str,st.get('overdue',[])[:4])); cold=', '.join(map(str,st.get('cold',[])[:4]))
    mode_txt={'balanced':'균형형','conservative':'안정형','aggressive':'공격형'}.get(mode,'균형형')
    lines=[
        f'{round_no}회차는 V34 VIP AI 엔진 기준 {mode_txt}으로 분석했습니다.',
        f'최근 가중 핵심수는 {hot}번, 미출현 보강 후보는 {overdue}번 흐름이 강합니다.',
        f'저출현 변동 후보 {cold}번을 함께 반영해 과도한 HOT 편중을 줄였습니다.',
        '홀짝, 구간, 합계, AC값, 끝수, 연속수, 동반출현을 종합 필터링했습니다.'
    ]
    if fixed: lines.append(f'고정수 {fixed}번은 우선 반영하고 나머지 번호만 AI로 보정했습니다.')
    if excluded: lines.append(f'제외수 {excluded}번은 추천 후보에서 제외했습니다.')
    return '\n'.join(lines[:5])


# ===== V40 UPGRADE1 FINAL OVERRIDES: keep V34 UI, stabilize backend engine =====
def latest_stats(limit=100):
    """V40 Phase1 통계 엔진: 최근 10/30/50/100회 흐름, 미출현, 동반출현, 끝수/구간/홀짝/합계를 계산합니다."""
    with con() as c:
        rows = c.execute('SELECT * FROM draws ORDER BY round_no DESC LIMIT ?', (max(10, int(limit or 100)),)).fetchall()
    draws=[]
    for r in rows:
        nums=parse_nums(r['numbers'])
        if len(nums)==6:
            draws.append({'round_no':int(r['round_no']), 'draw_date':r['draw_date'] or '', 'numbers':nums, 'bonus':int(r['bonus'] or 0)})
    if not draws:
        draws=[{'round_no':r,'draw_date':d,'numbers':n,'bonus':b} for r,d,n,b in DEFAULT_DRAWS if len(n)==6]
    windows={10:draws[:10], 30:draws[:30], 50:draws[:50], 100:draws[:100]}
    freq_by_window={}
    for w,ds in windows.items():
        f={n:0 for n in range(1,46)}
        for d in ds:
            for n in d['numbers']:
                f[n]+=1
        freq_by_window[w]=f
    freq=freq_by_window[100]
    last_seen={n:999 for n in range(1,46)}
    for idx,d in enumerate(draws[:100]):
        for n in d['numbers']:
            if last_seen[n] == 999:
                last_seen[n]=idx
    pair_counts=collections.Counter()
    for d in draws[:100]:
        for a,b in itertools.combinations(d['numbers'],2):
            pair_counts[tuple(sorted((a,b)))] += 1
    end_counts={i:0 for i in range(10)}
    zone_counts={'1~15':0,'16~30':0,'31~45':0}
    odd_total=0; sums=[]
    for d in draws[:100]:
        sums.append(sum(d['numbers']))
        for n in d['numbers']:
            end_counts[n%10]+=1
            odd_total += n%2
            if n<=15: zone_counts['1~15']+=1
            elif n<=30: zone_counts['16~30']+=1
            else: zone_counts['31~45']+=1
    avg100=sum(freq.values())/45 if freq else 0
    hot=sorted(range(1,46), key=lambda n:(-(freq_by_window[30][n]*1.7 + freq_by_window[100][n]*0.7 + freq_by_window[10][n]*2.0), last_seen[n], n))[:12]
    cold=sorted(range(1,46), key=lambda n:(freq_by_window[100][n], -last_seen[n], n))[:12]
    overdue=sorted(range(1,46), key=lambda n:(last_seen[n], -freq_by_window[100][n], n), reverse=True)[:12]
    mid=sorted(range(1,46), key=lambda n:(abs(freq_by_window[100][n]-avg100), last_seen[n], n))[:15]
    top_pairs=[{'pair':list(p),'count':c} for p,c in pair_counts.most_common(12)]
    recent_numbers=set()
    for d in draws[:3]: recent_numbers.update(d['numbers'])
    return {
        'draws':draws,'freq':freq,'freq10':freq_by_window[10],'freq30':freq_by_window[30],'freq50':freq_by_window[50],'freq100':freq_by_window[100],
        'last_seen':last_seen,'hot':hot,'mid':mid,'cold':cold,'overdue':overdue,'top_pairs':top_pairs,'pair_counts':pair_counts,
        'end_counts':end_counts,'zone_counts':zone_counts,'odd_ratio':odd_total/(max(1,len(draws[:100])*6)),
        'sum_avg':round(sum(sums)/len(sums),1) if sums else 0,'recent_numbers':recent_numbers,
        'latest_round':draws[0]['round_no'] if draws else 1230
    }

def ac_value(combo):
    arr=sorted(combo)
    diffs={b-a for a,b in itertools.combinations(arr,2)}
    return max(0, len(diffs)-5)

def combo_score(combo, st):
    """0~100 AI 점수. 당첨 보장 점수가 아니라 통계 균형/분산 품질 점수입니다."""
    combo=sorted(parse_nums(combo));
    if len(combo)!=6: return 0
    f10,f30,f100=st['freq10'],st['freq30'],st['freq100']; last=st['last_seen']; pairs=st['pair_counts']
    total=sum(combo); odd=sum(n%2 for n in combo); even=6-odd
    zones=[sum(n<=15 for n in combo),sum(16<=n<=30 for n in combo),sum(n>=31 for n in combo)]
    cons=sum(1 for a,b in zip(combo,combo[1:]) if b-a==1)
    ends=len(set(n%10 for n in combo)); ac=ac_value(combo)
    # 기본 패턴 점수
    score=42.0
    score += {3:13,2:10,4:10,1:3,5:3,0:-8,6:-8}.get(odd,0)
    score += 13 if 105<=total<=175 else (8 if 95<=total<=190 else -10)
    score += 12 if max(zones)<=3 and min(zones)>=1 else (5 if max(zones)<=4 else -8)
    score += 7 if 5<=ac<=10 else (3 if 4<=ac<=11 else -4)
    score += 6 if ends>=5 else (3 if ends==4 else -4)
    score += 5 if cons<=1 else (-3 if cons==2 else -9)
    # 최근 흐름 점수: 과열수는 과하게 몰리지 않게, 중간/보강수를 같이 반영
    hot_hit=len(set(combo)&set(st['hot'][:10])); cold_hit=len(set(combo)&set(st['cold'][:10])); overdue_hit=len(set(combo)&set(st['overdue'][:10]))
    score += min(10, hot_hit*3.0) + min(7, cold_hit*2.1) + min(8, overdue_hit*2.2)
    if hot_hit>4: score -= 7
    if len(set(combo)&st['recent_numbers'])>=4: score -= 5
    # 동반출현은 1~3개 정도만 가산, 과도한 과거쌍 몰림은 감점
    pair_sum=sum(pairs.get(tuple(sorted((a,b))),0) for a,b in itertools.combinations(combo,2))
    strong_pairs=sum(1 for a,b in itertools.combinations(combo,2) if pairs.get(tuple(sorted((a,b))),0)>=4)
    score += min(10, pair_sum/4.0) + min(5, strong_pairs*1.5)
    if strong_pairs>5: score -= 5
    # 10/30/100 가중치 평균이 너무 한쪽으로 쏠리지 않게
    heat=sum(f10[n]*2.0 + f30[n]*1.1 + f100[n]*0.35 for n in combo)
    if 22 <= heat <= 55: score += 6
    elif heat > 70: score -= 8
    else: score += 2
    return round(max(35, min(99.7, score)), 1)

def tags_for_combo(combo, st):
    combo=sorted(combo); s=set(combo); tags=[]
    if len(s & set(st['hot'][:10]))>=2: tags.append('최근핵심')
    if len(s & set(st['overdue'][:10]))>=1: tags.append('미출현보강')
    if len(s & set(st['cold'][:10]))>=1: tags.append('저출현반등')
    if sum(n%2 for n in combo) in (2,3,4): tags.append('홀짝균형')
    zones=[sum(n<=15 for n in combo),sum(16<=n<=30 for n in combo),sum(n>=31 for n in combo)]
    if max(zones)<=3 and min(zones)>=1: tags.append('구간분산')
    if ac_value(combo) in range(5,11): tags.append('AC안정')
    pairs=st['pair_counts']
    if any(pairs.get(tuple(sorted((a,b))),0)>=4 for a,b in itertools.combinations(combo,2)): tags.append('동반출현')
    return tags[:5] or ['균형형']

def combo_detail(combo, st):
    combo=sorted(combo)
    odd=sum(n%2 for n in combo); zones=[sum(n<=15 for n in combo),sum(16<=n<=30 for n in combo),sum(n>=31 for n in combo)]
    pair_hits=[]
    for a,b in itertools.combinations(combo,2):
        cnt=st['pair_counts'].get(tuple(sorted((a,b))),0)
        if cnt>=3: pair_hits.append({'pair':[a,b], 'count':cnt})
    pair_hits=sorted(pair_hits, key=lambda x:-x['count'])[:3]
    return {'numbers':combo,'score':combo_score(combo,st),'tags':tags_for_combo(combo,st),'sum':sum(combo),'odd':odd,'even':6-odd,'zones':zones,'ac':ac_value(combo),'pair_hits':pair_hits}

def _weighted_pick(candidates, weights, k):
    picked=[]; pool=list(candidates); w=list(weights)
    for _ in range(min(k,len(pool))):
        total=sum(max(0.01,x) for x in w)
        r=random.random()*total; acc=0; idx=0
        for i,x in enumerate(w):
            acc += max(0.01,x)
            if acc>=r:
                idx=i; break
        picked.append(pool.pop(idx)); w.pop(idx)
    return picked

def make_premium_combos(count=10, fixed='', excluded='', mode='balanced'):
    st=latest_stats(120)
    fixed_set=set(parse_nums(fixed)); excluded_set=set(parse_nums(excluded))
    fixed_set={n for n in fixed_set if n not in excluded_set}
    if len(fixed_set)>6: fixed_set=set(sorted(fixed_set)[:6])
    pool=[n for n in range(1,46) if n not in excluded_set and n not in fixed_set]
    target=max(1,min(50,int(count or 10)))
    if len(pool)+len(fixed_set)<6:
        raise HTTPException(400, '고정수/제외수를 확인하세요. 선택 가능한 번호가 부족합니다.')
    past={tuple(d['numbers']) for d in st['draws']}
    f10,f30,f100,last=st['freq10'],st['freq30'],st['freq100'],st['last_seen']
    weights={}
    for n in pool:
        hot = f30[n]*1.4 + f10[n]*2.0 + f100[n]*0.35
        cold = max(0, 7-f100[n]) + min(8,last[n])*0.35
        mid = 6 - min(6, abs(f100[n] - (sum(f100.values())/45)))
        if mode=='aggressive': weights[n]=1 + hot*1.35 + cold*0.45 + mid*0.5
        elif mode=='conservative': weights[n]=1 + hot*0.65 + cold*1.25 + mid*0.9
        else: weights[n]=1 + hot*0.95 + cold*0.9 + mid*0.75
        # 직전 3회 과다 반영은 낮춤
        if n in st['recent_numbers']: weights[n]*=0.72
    candidates=[]; seen=set(); tries=0
    needed=max(target*75, 900)
    while len(candidates)<needed and tries<40000:
        tries+=1
        need=6-len(fixed_set)
        nums=set(fixed_set)
        picked=_weighted_pick(pool, [weights[n] for n in pool], need)
        nums.update(picked)
        arr=tuple(sorted(nums))
        if len(arr)!=6 or arr in seen or arr in past: continue
        odd=sum(n%2 for n in arr); total=sum(arr); zones=[sum(n<=15 for n in arr),sum(16<=n<=30 for n in arr),sum(n>=31 for n in arr)]
        cons=sum(1 for a,b in zip(arr,arr[1:]) if b-a==1)
        if odd not in (2,3,4): continue
        if not (90<=total<=195): continue
        if max(zones)>4 or 0 in zones: continue
        if cons>2: continue
        if len(set(n%10 for n in arr))<3: continue
        seen.add(arr)
        candidates.append((combo_score(arr,st), list(arr)))
    candidates=sorted(candidates, key=lambda x:(-x[0], x[1]))
    selected=[]
    for score, combo in candidates:
        # 최종 10개는 서로 너무 비슷하지 않게 분산
        if all(len(set(combo)&set(prev))<=4 for prev in selected):
            selected.append(combo)
        if len(selected)>=target: break
    if len(selected)<target:
        for score, combo in candidates:
            if combo not in selected:
                selected.append(combo)
            if len(selected)>=target: break
    details=[combo_detail(c, st) for c in selected[:target]]
    return selected[:target], details, st

def _engine_summary(details, st):
    if not details: return {'avg_score':0,'combo_count':0}
    scores=[d.get('score',0) for d in details]
    all_nums=[n for d in details for n in d.get('numbers',[])]
    freq=collections.Counter(all_nums)
    return {
        'avg_score':round(sum(scores)/len(scores),1),
        'max_score':max(scores),
        'combo_count':len(details),
        'latest_round':st.get('latest_round'),
        'hot':st.get('hot',[])[:8],
        'cold':st.get('cold',[])[:8],
        'overdue':st.get('overdue',[])[:8],
        'top_used':[n for n,_ in freq.most_common(8)],
        'top_pairs':st.get('top_pairs',[])[:5]
    }

def build_analysis_text(round_no, st, mode, fixed, excluded, details=None):
    """생성된 추천 조합과 실제 최근 100회 통계를 함께 사용해 매번 다른 4~5줄 분석을 만듭니다."""
    details=details or []
    engine=_engine_summary(details, st)
    top_scores=sorted(details, key=lambda x:-x.get('score',0))[:3]
    used_nums=engine.get('top_used',[])[:6]
    hot=[n for n in st.get('hot',[])[:10] if n in used_nums] or st.get('hot',[])[:4]
    overdue=[n for n in st.get('overdue',[])[:10] if n in used_nums] or st.get('overdue',[])[:4]
    cold=[n for n in st.get('cold',[])[:10] if n in used_nums] or st.get('cold',[])[:4]
    pair_text=', '.join([f"{p['pair'][0]}-{p['pair'][1]}({p['count']}회)" for p in st.get('top_pairs',[])[:3]]) or '동반출현 데이터 보강 중'
    best = top_scores[0] if top_scores else {}
    mode_name={'balanced':'균형형','conservative':'보강형','aggressive':'공격형'}.get(mode, mode or '균형형')
    variants=[
        f"{round_no}회차는 최근 10/30/100회 흐름을 합산해 {mode_name} 기준으로 재분석했습니다.",
        f"이번 생성은 직전 과열 흐름을 낮추고 최근 100회 누적 통계를 함께 반영했습니다.",
        f"{round_no}회차 분석은 출현빈도, 미출현 간격, 동반출현, AC값을 동시에 적용했습니다."
    ]
    line1=random.choice(variants)
    line2=random.choice([
        f"핵심 후보는 {', '.join(map(str,hot[:5]))}번이며, 미출현 보강 후보는 {', '.join(map(str,overdue[:5]))}번입니다.",
        f"최근 강한 흐름은 {', '.join(map(str,hot[:5]))}번, 저출현 반등 후보는 {', '.join(map(str,cold[:5]))}번 중심으로 잡았습니다.",
        f"번호 풀은 HOT {', '.join(map(str,hot[:4]))}번과 보강 {', '.join(map(str,overdue[:4]))}번을 섞어 구성했습니다."
    ])
    if best:
        line3=f"대표 조합은 합계 {best.get('sum')} / AC {best.get('ac')} / 홀짝 {best.get('odd')}:{best.get('even')} 기준이며 AI점수는 {best.get('score')}점입니다."
    else:
        line3=f"평균 AI점수는 {engine.get('avg_score',0)}점이며 과도한 중복과 구간 쏠림을 줄였습니다."
    line4=random.choice([
        f"동반출현 참고 흐름은 {pair_text}이며, 조합 간 중복을 낮춰 분산형으로 정리했습니다.",
        f"동반출현은 {pair_text} 흐름을 참고했고, 끝수와 구간이 한쪽으로 몰리지 않게 제한했습니다.",
        f"상위 동반 흐름({pair_text})은 참고만 하고, 동일 패턴 반복은 줄였습니다."
    ])
    line5=f"분석 갱신시각: {now()}"
    return '\n'.join([line1,line2,line3,line4,line5])

def build_sms(member_name, round_no, combos, analysis, details):
    name=member_name or '회원'
    best=sorted(details or [], key=lambda x:-x.get('score',0))[:1]
    best_line=''
    if best:
        b=best[0]
        best_line=f"대표 조합 포인트: 합계 {b.get('sum')} / AC {b.get('ac')} / 홀짝 {b.get('odd')}:{b.get('even')} / AI점수 {b.get('score')}점"
    return '\n'.join([
        f'안녕하세요 {name}님, BBLOTTO입니다.',
        f'{round_no}회차 추천번호와 이번 회차 분석을 안내드립니다.',
        '[추천번호]',
        *[f'{i+1}. '+', '.join(map(str,c)) for i,c in enumerate(combos)],
        '[분석요약]',
        analysis,
        best_line,
        '좋은 결과 있으시길 바랍니다.'
    ]).strip()



# ===== V40 UPGRADE1 DIVERSITY OVERRIDE =====
def make_premium_combos(count=10, fixed='', excluded='', mode='balanced'):
    st=latest_stats(120)
    fixed_set=set(parse_nums(fixed)); excluded_set=set(parse_nums(excluded))
    fixed_set={n for n in fixed_set if n not in excluded_set}
    if len(fixed_set)>6: fixed_set=set(sorted(fixed_set)[:6])
    pool=[n for n in range(1,46) if n not in excluded_set and n not in fixed_set]
    target=max(1,min(50,int(count or 10)))
    if len(pool)+len(fixed_set)<6:
        raise HTTPException(400, '고정수/제외수를 확인하세요. 선택 가능한 번호가 부족합니다.')
    past={tuple(d['numbers']) for d in st['draws']}
    f10,f30,f100,last=st['freq10'],st['freq30'],st['freq100'],st['last_seen']
    avg100=sum(f100.values())/45
    weights={}
    for n in pool:
        hot = f30[n]*1.0 + f10[n]*1.7 + f100[n]*0.20
        cold = max(0, avg100-f100[n]) + min(10,last[n])*0.40
        mid = max(0, 6 - abs(f100[n]-avg100))
        if mode=='aggressive': w=1 + hot*1.15 + cold*0.55 + mid*0.45
        elif mode=='conservative': w=1 + hot*0.55 + cold*1.30 + mid*0.85
        else: w=1 + hot*0.75 + cold*1.05 + mid*0.70
        if n in st['recent_numbers']: w*=0.62
        weights[n]=max(0.2,w)
    candidates=[]; seen=set(); tries=0
    needed=max(target*140, 1600)
    while len(candidates)<needed and tries<60000:
        tries+=1
        nums=set(fixed_set)
        # 조합별 시드 다양화: HOT/COLD/MID/OVERDUE 비율을 랜덤으로 섞음
        buckets=[]
        if mode=='aggressive': buckets=[st['hot'][:14], st['overdue'][:14], pool, pool, pool, pool]
        elif mode=='conservative': buckets=[st['mid'][:16], st['cold'][:14], st['overdue'][:14], pool, pool, pool]
        else: buckets=[st['hot'][:12], st['mid'][:15], st['cold'][:12], st['overdue'][:12], pool, pool]
        random.shuffle(buckets)
        for bucket in buckets:
            if len(nums)>=6: break
            usable=[n for n in bucket if n in pool and n not in nums]
            if usable:
                nums.update(_weighted_pick(usable, [weights[n] for n in usable], 1))
        while len(nums)<6:
            usable=[n for n in pool if n not in nums]
            nums.update(_weighted_pick(usable, [weights[n] for n in usable], 1))
        arr=tuple(sorted(nums))
        if len(arr)!=6 or arr in seen or arr in past: continue
        odd=sum(n%2 for n in arr); total=sum(arr); zones=[sum(n<=15 for n in arr),sum(16<=n<=30 for n in arr),sum(n>=31 for n in arr)]
        cons=sum(1 for a,b in zip(arr,arr[1:]) if b-a==1)
        if odd not in (2,3,4): continue
        if not (90<=total<=195): continue
        if max(zones)>4 or 0 in zones: continue
        if cons>2: continue
        if len(set(n%10 for n in arr))<4: continue
        seen.add(arr)
        candidates.append((combo_score(arr,st), list(arr)))
    candidates=sorted(candidates, key=lambda x:(-x[0], x[1]))
    selected=[]; usage=collections.Counter()
    # 1차: 조합 간 겹침과 특정 숫자 과사용 제한
    for score, combo in candidates:
        if any(usage[n]>=3 for n in combo if n not in fixed_set):
            continue
        if all(len(set(combo)&set(prev))<=3 for prev in selected):
            selected.append(combo); usage.update(combo)
        if len(selected)>=target: break
    # 2차: 부족할 때 조건 완화
    if len(selected)<target:
        for score, combo in candidates:
            if combo in selected: continue
            if any(usage[n]>=4 for n in combo if n not in fixed_set):
                continue
            if all(len(set(combo)&set(prev))<=4 for prev in selected):
                selected.append(combo); usage.update(combo)
            if len(selected)>=target: break
    # 3차: 그래도 부족하면 순위대로 보충
    if len(selected)<target:
        for score, combo in candidates:
            if combo not in selected:
                selected.append(combo); usage.update(combo)
            if len(selected)>=target: break
    details=[combo_detail(c, st) for c in selected[:target]]
    return selected[:target], details, st

# ===== V40 UPGRADE1 FINAL DIVERSITY STRICT =====
def make_premium_combos(count=10, fixed='', excluded='', mode='balanced'):
    st=latest_stats(120)
    fixed_set=set(parse_nums(fixed)); excluded_set=set(parse_nums(excluded))
    fixed_set={n for n in fixed_set if n not in excluded_set}
    if len(fixed_set)>6: fixed_set=set(sorted(fixed_set)[:6])
    pool=[n for n in range(1,46) if n not in excluded_set and n not in fixed_set]
    target=max(1,min(50,int(count or 10)))
    if len(pool)+len(fixed_set)<6:
        raise HTTPException(400, '고정수/제외수를 확인하세요. 선택 가능한 번호가 부족합니다.')
    past={tuple(d['numbers']) for d in st['draws']}
    f10,f30,f100,last=st['freq10'],st['freq30'],st['freq100'],st['last_seen']
    avg100=sum(f100.values())/45
    weights={}
    for n in pool:
        hot = f30[n]*0.85 + f10[n]*1.40 + f100[n]*0.18
        cold = max(0, avg100-f100[n]) + min(10,last[n])*0.38
        mid = max(0, 6 - abs(f100[n]-avg100))
        if mode=='aggressive': w=1 + hot*1.05 + cold*0.65 + mid*0.45
        elif mode=='conservative': w=1 + hot*0.50 + cold*1.25 + mid*0.90
        else: w=1 + hot*0.70 + cold*1.00 + mid*0.80
        if n in st['recent_numbers']: w*=0.58
        weights[n]=max(0.2,w)
    candidates=[]; seen=set(); tries=0; needed=max(target*180, 2200)
    while len(candidates)<needed and tries<80000:
        tries+=1; nums=set(fixed_set)
        if mode=='aggressive': buckets=[st['hot'][:14], st['overdue'][:14], st['mid'][:15], pool, pool, pool]
        elif mode=='conservative': buckets=[st['mid'][:16], st['cold'][:14], st['overdue'][:14], pool, pool, pool]
        else: buckets=[st['hot'][:12], st['mid'][:15], st['cold'][:12], st['overdue'][:12], pool, pool]
        random.shuffle(buckets)
        for bucket in buckets:
            if len(nums)>=6: break
            usable=[n for n in bucket if n in pool and n not in nums]
            if usable: nums.update(_weighted_pick(usable, [weights[n] for n in usable], 1))
        while len(nums)<6:
            usable=[n for n in pool if n not in nums]
            nums.update(_weighted_pick(usable, [weights[n] for n in usable], 1))
        arr=tuple(sorted(nums))
        if len(arr)!=6 or arr in seen or arr in past: continue
        odd=sum(n%2 for n in arr); total=sum(arr); zones=[sum(n<=15 for n in arr),sum(16<=n<=30 for n in arr),sum(n>=31 for n in arr)]
        cons=sum(1 for a,b in zip(arr,arr[1:]) if b-a==1)
        if odd not in (2,3,4): continue
        if not (92<=total<=190): continue
        if max(zones)>4 or 0 in zones: continue
        if cons>1: continue
        if len(set(n%10 for n in arr))<4: continue
        seen.add(arr); candidates.append((combo_score(arr,st), list(arr)))
    candidates=sorted(candidates, key=lambda x:(-x[0], x[1]))
    selected=[]; usage=collections.Counter(); pair_usage=collections.Counter()
    for score, combo in candidates:
        pairs=[tuple(sorted(p)) for p in itertools.combinations(combo,2)]
        if any(usage[n]>=3 for n in combo if n not in fixed_set): continue
        if any(pair_usage[p]>=1 for p in pairs): continue
        if all(len(set(combo)&set(prev))<=3 for prev in selected):
            selected.append(combo); usage.update(combo); pair_usage.update(pairs)
        if len(selected)>=target: break
    if len(selected)<target:
        for score, combo in candidates:
            if combo in selected: continue
            pairs=[tuple(sorted(p)) for p in itertools.combinations(combo,2)]
            if any(usage[n]>=4 for n in combo if n not in fixed_set): continue
            if any(pair_usage[p]>=2 for p in pairs): continue
            if all(len(set(combo)&set(prev))<=4 for prev in selected):
                selected.append(combo); usage.update(combo); pair_usage.update(pairs)
            if len(selected)>=target: break
    if len(selected)<target:
        for score, combo in candidates:
            if combo not in selected:
                selected.append(combo)
            if len(selected)>=target: break
    details=[combo_detail(c, st) for c in selected[:target]]
    return selected[:target], details, st

# ===== V40 UPGRADE1 SCORE CALIBRATION =====
def combo_score(combo, st):
    combo=sorted(parse_nums(combo))
    if len(combo)!=6: return 0
    f10,f30,f100=st['freq10'],st['freq30'],st['freq100']; last=st['last_seen']; pairs=st['pair_counts']
    total=sum(combo); odd=sum(n%2 for n in combo)
    zones=[sum(n<=15 for n in combo),sum(16<=n<=30 for n in combo),sum(n>=31 for n in combo)]
    cons=sum(1 for a,b in zip(combo,combo[1:]) if b-a==1); ends=len(set(n%10 for n in combo)); ac=ac_value(combo)
    score=58.0
    score += {3:9,2:7,4:7,1:1,5:1,0:-7,6:-7}.get(odd,0)
    score += 9 if 105<=total<=175 else (5 if 92<=total<=190 else -8)
    score += 8 if max(zones)<=3 and min(zones)>=1 else (3 if max(zones)<=4 and min(zones)>=1 else -7)
    score += 6 if 5<=ac<=10 else (2 if 4<=ac<=12 else -4)
    score += 4 if ends>=5 else (2 if ends==4 else -3)
    score += 4 if cons==0 else (2 if cons==1 else -5)
    hot_hit=len(set(combo)&set(st['hot'][:10])); cold_hit=len(set(combo)&set(st['cold'][:10])); overdue_hit=len(set(combo)&set(st['overdue'][:10]))
    score += min(7, hot_hit*1.6) + min(6, cold_hit*1.5) + min(6, overdue_hit*1.5)
    if hot_hit>4: score -= 5
    pair_sum=sum(pairs.get(tuple(sorted((a,b))),0) for a,b in itertools.combinations(combo,2))
    strong_pairs=sum(1 for a,b in itertools.combinations(combo,2) if pairs.get(tuple(sorted((a,b))),0)>=4)
    score += min(6, pair_sum/8.0) + min(4, strong_pairs*1.0)
    heat=sum(f10[n]*1.6 + f30[n]*0.8 + f100[n]*0.18 for n in combo)
    if 16 <= heat <= 48: score += 4
    elif heat > 65: score -= 5
    # 작은 분산값을 넣어 모든 조합이 같은 점수로 보이지 않게 함
    score += ((sum(n*n for n in combo) % 13) - 6) * 0.15
    return round(max(60, min(96.8, score)), 1)

# ===== V40 UPGRADE1 SCORE CALIBRATION 2 =====
def combo_score(combo, st):
    combo=sorted(parse_nums(combo))
    if len(combo)!=6: return 0
    f10,f30,f100=st['freq10'],st['freq30'],st['freq100']; last=st['last_seen']; pairs=st['pair_counts']
    total=sum(combo); odd=sum(n%2 for n in combo)
    zones=[sum(n<=15 for n in combo),sum(16<=n<=30 for n in combo),sum(n>=31 for n in combo)]
    cons=sum(1 for a,b in zip(combo,combo[1:]) if b-a==1); ends=len(set(n%10 for n in combo)); ac=ac_value(combo)
    score=50.0
    score += {3:8,2:6,4:6,1:0,5:0,0:-8,6:-8}.get(odd,0)
    score += 8 if 110<=total<=170 else (4 if 95<=total<=190 else -8)
    score += 7 if max(zones)<=3 and min(zones)>=1 else (2 if max(zones)<=4 and min(zones)>=1 else -6)
    score += 5 if 5<=ac<=10 else (2 if 4<=ac<=12 else -4)
    score += 3 if ends>=5 else (1 if ends==4 else -3)
    score += 3 if cons==0 else (1 if cons==1 else -5)
    hot_hit=len(set(combo)&set(st['hot'][:10])); cold_hit=len(set(combo)&set(st['cold'][:10])); overdue_hit=len(set(combo)&set(st['overdue'][:10]))
    score += min(6, hot_hit*1.3) + min(5, cold_hit*1.2) + min(5, overdue_hit*1.2)
    pair_sum=sum(pairs.get(tuple(sorted((a,b))),0) for a,b in itertools.combinations(combo,2))
    strong_pairs=sum(1 for a,b in itertools.combinations(combo,2) if pairs.get(tuple(sorted((a,b))),0)>=4)
    score += min(5, pair_sum/10.0) + min(3, strong_pairs*0.8)
    heat=sum(f10[n]*1.5 + f30[n]*0.7 + f100[n]*0.15 for n in combo)
    if 15 <= heat <= 45: score += 3
    elif heat > 62: score -= 5
    # 번호별 분산 보정
    score += ((sum(n*n for n in combo) % 17) - 8) * 0.18
    return round(max(65, min(94.9, score)), 1)

# ===== PATCH 023: BBLOTTO AI ENGINE V1 INSTALL =====
# RC8-2: 기존 recommend_engine_v1 설치 훅 제거. 추천번호 생성은 하단 AI V4 make_premium_combos()만 사용합니다.


# ===== RC2 SPRINT 2-3: AI ENGINE V2 + DASHBOARD INSIGHT PATCH =====
SPRINT23_ENGINE_VERSION = 'BBLOTTO_PRO_V2_STABLE_RC3_12'

def ensure_sprint23_schema():
    """추천 엔진 실행 이력/대시보드 캐시용 테이블을 안전하게 준비합니다."""
    try:
        with con() as c:
            c.execute('CREATE TABLE IF NOT EXISTS engine_runs(id INTEGER PRIMARY KEY AUTOINCREMENT, recommendation_id INTEGER DEFAULT 0, round_no INTEGER DEFAULT 0, member_id INTEGER DEFAULT 0, mode TEXT DEFAULT "balanced", count INTEGER DEFAULT 0, candidate_count INTEGER DEFAULT 0, selected_count INTEGER DEFAULT 0, avg_score REAL DEFAULT 0, max_score REAL DEFAULT 0, min_score REAL DEFAULT 0, engine_version TEXT DEFAULT "", created_by INTEGER DEFAULT 0, created_at TEXT)')
            c.execute('CREATE TABLE IF NOT EXISTS dashboard_snapshots(id INTEGER PRIMARY KEY AUTOINCREMENT, snapshot_json TEXT DEFAULT "{}", created_at TEXT)')
            c.commit()
    except Exception:
        pass

ensure_sprint23_schema()

def _s23_weighted_number_profile(st, mode='balanced'):
    f10,f30,f50,f100=st.get('freq10',{}),st.get('freq30',{}),st.get('freq50',{}),st.get('freq100',{})
    last=st.get('last_seen',{})
    avg100=(sum(f100.values())/45) if f100 else 0
    recent=set(st.get('recent_numbers',set()) or set())
    profile={}
    for n in range(1,46):
        hot=f10.get(n,0)*2.15 + f30.get(n,0)*1.15 + f50.get(n,0)*0.72 + f100.get(n,0)*0.35
        cold=max(0, avg100-f100.get(n,0))*0.85 + min(30,last.get(n,999))*0.11
        balance=max(0, 5.5-abs(f100.get(n,0)-avg100))*0.55
        pair_boost=0
        for k,v in (st.get('pair_counts') or {}).items():
            try:
                if n in k: pair_boost += min(4, v)*0.03
            except Exception:
                pass
        if mode=='aggressive': score=hot*1.18 + cold*0.55 + balance*0.52 + pair_boost
        elif mode=='conservative': score=hot*0.62 + cold*1.25 + balance*0.88 + pair_boost
        else: score=hot*0.83 + cold*1.03 + balance*0.75 + pair_boost
        if n in recent:
            score*=0.70
        profile[n]=round(max(0.25, score+1),4)
    return profile

def s23_combo_score(combo, st, mode='balanced'):
    combo=sorted(parse_nums(combo))
    if len(combo)!=6: return 0
    f10,f30,f50,f100=st.get('freq10',{}),st.get('freq30',{}),st.get('freq50',{}),st.get('freq100',{})
    last=st.get('last_seen',{})
    pairs=st.get('pair_counts') or {}
    total=sum(combo); odd=sum(n%2 for n in combo)
    zones=[sum(n<=15 for n in combo),sum(16<=n<=30 for n in combo),sum(n>=31 for n in combo)]
    cons=sum(1 for a,b in zip(combo,combo[1:]) if b-a==1)
    ends=len(set(n%10 for n in combo)); ac=ac_value(combo)
    score=52.0
    score += {3:9.0,2:7.0,4:7.0,1:1.0,5:1.0,0:-8.0,6:-8.0}.get(odd,0)
    score += 9 if 105<=total<=175 else (5 if 92<=total<=190 else -9)
    score += 8 if max(zones)<=3 and min(zones)>=1 else (3 if max(zones)<=4 and min(zones)>=1 else -8)
    score += 6 if 5<=ac<=10 else (2 if 4<=ac<=12 else -5)
    score += 5 if ends>=5 else (2 if ends==4 else -4)
    score += 4 if cons==0 else (1.5 if cons==1 else -6)
    hot_hit=len(set(combo)&set(st.get('hot',[])[:12])); cold_hit=len(set(combo)&set(st.get('cold',[])[:12])); overdue_hit=len(set(combo)&set(st.get('overdue',[])[:12]))
    score += min(7,hot_hit*1.25)+min(6,cold_hit*1.15)+min(6,overdue_hit*1.15)
    if hot_hit>=5: score-=4
    pair_sum=sum(pairs.get(tuple(sorted((a,b))),0) for a,b in itertools.combinations(combo,2))
    score += min(6.5, pair_sum/9.0)
    heat=sum(f10.get(n,0)*1.8 + f30.get(n,0)*0.85 + f50.get(n,0)*0.42 + f100.get(n,0)*0.17 for n in combo)
    if 16<=heat<=52: score+=4
    elif heat>68: score-=5
    # V2: 최근 100회에서 너무 흔한 번호만 몰리는 조합 감점, 오래된 번호 보완 가점
    avg100=(sum(f100.values())/45) if f100 else 0
    above=sum(1 for n in combo if f100.get(n,0)>avg100+2)
    overdue=sum(1 for n in combo if last.get(n,0)>=10)
    if above>=5: score-=4
    if 1<=overdue<=3: score+=3
    # 동일 10번대 쏠림 완화
    decades=collections.Counter((n-1)//10 for n in combo)
    if max(decades.values())>=4: score-=5
    # 안정적이지만 같은 점수 반복 방지
    score += ((sum(n*n for n in combo) % 19)-9)*0.13
    return round(max(67.0, min(96.3, score)),1)

def make_premium_combos(count=10, fixed='', excluded='', mode='balanced'):
    """RC2 Sprint 2-3 V2 엔진: 후보 확장, 포트폴리오 분산, 점수 기반 선별."""
    st=latest_stats(120)
    fixed_set=set(parse_nums(fixed)); excluded_set=set(parse_nums(excluded))
    fixed_set={n for n in fixed_set if n not in excluded_set}
    if len(fixed_set)>6: fixed_set=set(sorted(fixed_set)[:6])
    pool=[n for n in range(1,46) if n not in excluded_set and n not in fixed_set]
    target=max(1,min(50,int(count or 10)))
    if len(pool)+len(fixed_set)<6:
        raise HTTPException(400, '고정수/제외수를 확인하세요. 선택 가능한 번호가 부족합니다.')
    profile=_s23_weighted_number_profile(st, mode)
    past={tuple(d['numbers']) for d in st.get('draws',[])}
    buckets={
        'hot':[n for n in st.get('hot',[])[:18] if n in pool],
        'cold':[n for n in st.get('cold',[])[:18] if n in pool],
        'overdue':[n for n in st.get('overdue',[])[:18] if n in pool],
        'mid':[n for n in st.get('mid',[])[:18] if n in pool],
        'all':pool[:]
    }
    if mode=='aggressive': plan=['hot','hot','mid','overdue','all','all']
    elif mode=='conservative': plan=['cold','overdue','mid','mid','all','all']
    else: plan=['hot','cold','overdue','mid','all','all']
    needed=max(2600,target*240)
    candidates=[]; seen=set(); attempts=0
    while len(candidates)<needed and attempts<95000:
        attempts+=1
        nums=set(fixed_set)
        p=plan[:]; random.shuffle(p)
        for b in p:
            if len(nums)>=6: break
            usable=[n for n in buckets.get(b, pool) if n not in nums]
            if usable:
                nums.update(_weighted_pick(usable, [profile[n] for n in usable], 1))
        while len(nums)<6:
            usable=[n for n in pool if n not in nums]
            nums.update(_weighted_pick(usable, [profile[n] for n in usable], 1))
        arr=tuple(sorted(nums))
        if len(arr)!=6 or arr in seen or arr in past: continue
        odd=sum(n%2 for n in arr); total=sum(arr); zones=[sum(n<=15 for n in arr),sum(16<=n<=30 for n in arr),sum(n>=31 for n in arr)]
        cons=sum(1 for a,b in zip(arr,arr[1:]) if b-a==1)
        if odd not in (2,3,4): continue
        if not (92<=total<=190): continue
        if max(zones)>4 or min(zones)==0: continue
        if cons>1: continue
        if len(set(n%10 for n in arr))<4: continue
        seen.add(arr)
        candidates.append((s23_combo_score(arr,st,mode), list(arr)))
    candidates=sorted(candidates,key=lambda x:(-x[0], x[1]))
    selected=[]; usage=collections.Counter(); pair_usage=collections.Counter()
    for score, combo in candidates:
        pairs=[tuple(sorted(p)) for p in itertools.combinations(combo,2)]
        if any(usage[n]>=3 for n in combo if n not in fixed_set): continue
        if any(pair_usage[p]>=1 for p in pairs): continue
        if all(len(set(combo)&set(prev))<=3 for prev in selected):
            selected.append(combo); usage.update(combo); pair_usage.update(pairs)
        if len(selected)>=target: break
    if len(selected)<target:
        for score, combo in candidates:
            if combo in selected: continue
            pairs=[tuple(sorted(p)) for p in itertools.combinations(combo,2)]
            if any(usage[n]>=4 for n in combo if n not in fixed_set): continue
            if any(pair_usage[p]>=2 for p in pairs): continue
            if all(len(set(combo)&set(prev))<=4 for prev in selected):
                selected.append(combo); usage.update(combo); pair_usage.update(pairs)
            if len(selected)>=target: break
    if len(selected)<target:
        for score, combo in candidates:
            if combo not in selected:
                selected.append(combo)
            if len(selected)>=target: break
    details=[]
    for ccc in selected[:target]:
        d=combo_detail(ccc, st)
        d['score']=s23_combo_score(ccc,st,mode)
        d['engine_version']=SPRINT23_ENGINE_VERSION
        d['v2_reason']='최근 10/30/50/100회 가중치, 미출현 보정, 동반출현, 포트폴리오 분산을 반영했습니다.'
        details.append(d)
    st['s23_candidates']=len(candidates)
    return selected[:target], details, st

def _engine_summary(details, st):
    scores=[float(d.get('score') or d.get('ai_score') or d.get('vip_score') or 0) for d in (details or []) if (d.get('score') or d.get('ai_score') or d.get('vip_score'))]
    return {
        'version': SPRINT23_ENGINE_VERSION,
        'engine_version': SPRINT23_ENGINE_VERSION,
        'avg_score': round(sum(scores)/len(scores),1) if scores else 0,
        'max_score': round(max(scores),1) if scores else 0,
        'min_score': round(min(scores),1) if scores else 0,
        'candidate_count': int(st.get('s23_candidates') or 0),
        'selected_count': len(details or []),
        'latest_round': st.get('latest_round') or 0,
        'v2_pipeline_report': {
            'pipeline':'최근 100회 통계 → 번호별 V2 가중치 → 후보 확장 → 패턴 필터 → 포트폴리오 분산 → 최종 선별',
            'stage1_candidates': int(st.get('s23_candidates') or 0),
            'stage2_filters':'홀짝/합계/구간/AC/끝수/연속수',
            'stage3_portfolio':'숫자 과사용·중복 페어 제한',
            'summary':'RC2 Sprint 2-3 V2 엔진으로 후보 수와 분산 품질을 함께 강화했습니다.'
        }
    }

@app.get('/api/engine/insights')
def engine_insights(authorization: str|None = Header(default=None)):
    require_admin(authorization)
    ensure_sprint23_schema()
    st=latest_stats(120)
    profile=_s23_weighted_number_profile(st, 'balanced')
    top_weighted=sorted(profile.keys(), key=lambda n:(-profile[n], n))[:12]
    with con() as c:
        recent_runs=c.execute('SELECT * FROM engine_runs ORDER BY id DESC LIMIT 20').fetchall()
        rec_rows=c.execute('SELECT avg_score,engine_json,created_at FROM recommendations ORDER BY id DESC LIMIT 30').fetchall()
    scores=[]
    for r in rec_rows:
        try:
            if r['avg_score']: scores.append(float(r['avg_score']))
            else:
                ej=json.loads(r['engine_json'] or '{}'); scores.append(float(ej.get('avg_score') or 0))
        except Exception: pass
    return {
        'engine_version':SPRINT23_ENGINE_VERSION,
        'latest_round':st.get('latest_round'),
        'hot':st.get('hot',[])[:12],
        'cold':st.get('cold',[])[:12],
        'overdue':st.get('overdue',[])[:12],
        'top_weighted':top_weighted,
        'avg_recent_score':round(sum(scores)/len(scores),1) if scores else 0,
        'runs':[dict(r) for r in recent_runs],
        'summary':'최근 100회 통계와 추천 생성 이력을 기준으로 엔진 상태를 표시합니다.'
    }

@app.get('/api/dashboard_v2')
def dashboard_v2(authorization: str|None = Header(default=None)):
    require_admin(authorization)
    ensure_sprint23_schema()
    base=dashboard_summary(authorization)
    with con() as c:
        today=now()[:10]
        today_recs=c.execute('SELECT COUNT(*) c FROM recommendations WHERE created_at LIKE ?', (today+'%',)).fetchone()['c']
        today_sms=c.execute('SELECT COUNT(*) c FROM sms_logs WHERE created_at LIKE ?', (today+'%',)).fetchone()['c']
        avg=c.execute('SELECT COALESCE(AVG(avg_score),0) a, COALESCE(MAX(avg_score),0) m FROM recommendations WHERE avg_score IS NOT NULL').fetchone()
        recent_runs=c.execute('SELECT * FROM engine_runs ORDER BY id DESC LIMIT 5').fetchall()
    base.update({
        'engine_version':SPRINT23_ENGINE_VERSION,
        'today_recommendations':today_recs,
        'today_sms':today_sms,
        'avg_ai_score':round(float(avg['a'] or 0),1),
        'max_ai_score':round(float(avg['m'] or 0),1),
        'recent_engine_runs':[dict(r) for r in recent_runs]
    })
    return base

# generate 라우트가 만든 recommendations 저장 직후 engine_runs에도 보조 기록되도록 DB insert를 후킹하지 않고,
# 추천 상세 조회 시 누락된 실행 이력을 보강하는 안전 API를 제공합니다.
@app.post('/api/engine/backfill_runs')
def backfill_engine_runs(authorization: str|None = Header(default=None)):
    admin=require_admin(authorization)
    ensure_sprint23_schema()
    inserted=0
    with con() as c:
        rows=c.execute('SELECT id,member_id,round_no,mode,count,avg_score,engine_json,created_by,created_at FROM recommendations ORDER BY id DESC LIMIT 300').fetchall()
        for r in rows:
            exists=c.execute('SELECT id FROM engine_runs WHERE recommendation_id=?',(r['id'],)).fetchone()
            if exists: continue
            try: ej=json.loads(r['engine_json'] or '{}')
            except Exception: ej={}
            c.execute('INSERT INTO engine_runs(recommendation_id,round_no,member_id,mode,count,candidate_count,selected_count,avg_score,max_score,min_score,engine_version,created_by,created_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)',(
                r['id'],r['round_no'] or 0,r['member_id'] or 0,r['mode'] or 'balanced',r['count'] or 0,ej.get('candidate_count') or 0,ej.get('selected_count') or r['count'] or 0,r['avg_score'] or ej.get('avg_score') or 0,ej.get('max_score') or 0,ej.get('min_score') or 0,ej.get('engine_version') or ej.get('version') or SPRINT23_ENGINE_VERSION,r['created_by'] or admin['id'],r['created_at'] or now()))
            inserted+=1
        c.commit()
    return {'ok':True,'inserted':inserted}




# ===== RC4-4: Admin / Stats / AI / Members / Auto Update =====
RC4_4_VERSION = 'V2_STABLE_RC4_4_ADMIN_DASHBOARD'

def _rc44_safe_count(c, sql, params=()):
    try:
        r = c.execute(sql, params).fetchone()
        return int((r['c'] if r and 'c' in r.keys() else r[0]) or 0)
    except Exception:
        return 0

def _rc44_rows(c, sql, params=(), limit=20):
    try:
        return [dict(r) for r in c.execute(sql, params).fetchall()[:limit]]
    except Exception:
        return []

def _rc44_today_like():
    return datetime.datetime.now().strftime('%Y-%m-%d') + '%'

@app.get('/api/rc4-4/admin-dashboard')
def rc44_admin_dashboard(authorization: str|None = Header(default=None)):
    admin=require_admin(authorization)
    today = _rc44_today_like()
    with con() as c:
        scope_sql, scope_args = member_scope_condition(admin, 'm')
        mem_where = (' WHERE ' + scope_sql.replace('m.', '')) if scope_sql else ''
        rec_join = ' LEFT JOIN members m ON m.id=r.member_id'
        rec_scope = (' AND ' + scope_sql) if scope_sql else ''
        rec_scope_where = (' WHERE ' + scope_sql) if scope_sql else ''
        total_members = _rc44_safe_count(c, 'SELECT COUNT(*) c FROM members' + mem_where, scope_args)
        active_members = _rc44_safe_count(c, "SELECT COUNT(*) c FROM members" + (mem_where + " AND " if mem_where else " WHERE ") + "COALESCE(status,'활성')='활성'", scope_args)
        vip_members = _rc44_safe_count(c, "SELECT COUNT(*) c FROM members" + (mem_where + " AND " if mem_where else " WHERE ") + "COALESCE(grade,'일반') IN ('1등','2등','VIP','다이아','프리미엄')", scope_args)
        priority_members = _rc44_safe_count(c, "SELECT COUNT(*) c FROM members" + (mem_where + " AND " if mem_where else " WHERE ") + "COALESCE(priority,'보통') IN ('높음','최우선')", scope_args)
        rec_total = _rc44_safe_count(c, 'SELECT COUNT(*) c FROM recommendations r' + rec_join + rec_scope_where, scope_args)
        rec_today = _rc44_safe_count(c, 'SELECT COUNT(*) c FROM recommendations r' + rec_join + ' WHERE r.created_at LIKE ?' + rec_scope, (today, *scope_args))
        sms_today = _rc44_safe_count(c, 'SELECT COUNT(*) c FROM sms_logs WHERE created_at LIKE ?', (today,))
        wins_today = _rc44_safe_count(c, "SELECT COUNT(*) c FROM winning_checks WHERE created_at LIKE ? AND COALESCE(rank,'낙첨')<>'낙첨'", (today,))
        login_today = _rc44_safe_count(c, "SELECT COUNT(*) c FROM admin_logs WHERE action IN ('LOGIN','LOGIN_SUCCESS') AND created_at LIKE ?", (today,))
        activity_today = _rc44_safe_count(c, 'SELECT COUNT(*) c FROM admin_logs WHERE created_at LIKE ?', (today,))
        latest_draw = c.execute('SELECT round_no,draw_date,numbers,bonus FROM draws ORDER BY round_no DESC LIMIT 1').fetchone()
        avg = c.execute('SELECT COALESCE(AVG(avg_score),0) a, COALESCE(MAX(avg_score),0) m FROM recommendations WHERE avg_score IS NOT NULL').fetchone()
        prize_row = c.execute('SELECT COALESCE(SUM(prize),0) prize, COUNT(*) c FROM winning_checks').fetchone()
        recent_members = _rc44_rows(c, 'SELECT id,name,phone,grade,status,priority,created_at FROM members' + mem_where + ' ORDER BY id DESC LIMIT 8', scope_args)
        recent_logs = _rc44_rows(c, 'SELECT username,action,detail,created_at FROM admin_logs ORDER BY id DESC LIMIT 10')
        recent_recs = _rc44_rows(c, 'SELECT r.id,r.member_id,r.member_name,r.round_no,r.mode,r.count,r.avg_score,r.created_at FROM recommendations r' + rec_join + rec_scope_where + ' ORDER BY r.id DESC LIMIT 10', scope_args)
        recent_wins = _rc44_rows(c, "SELECT w.member_name,w.round_no,w.rank,w.prize,w.created_at FROM winning_checks w LEFT JOIN members m ON m.id=w.member_id WHERE COALESCE(w.rank,'낙첨')<>'낙첨'" + rec_scope + ' ORDER BY w.id DESC LIMIT 8', scope_args)
    alerts=[]
    if not latest_draw:
        alerts.append({'type':'warning','message':'저장된 당첨번호가 없습니다. 자동 업데이트를 실행하세요.'})
    else:
        alerts.append({'type':'success','message':f"최근 저장 회차 {latest_draw['round_no']}회 기준으로 운영 중입니다."})
    if rec_today:
        alerts.append({'type':'info','message':f'오늘 추천번호 {rec_today}건 생성되었습니다.'})
    if wins_today:
        alerts.append({'type':'success','message':f'오늘 적중 결과 {wins_today}건이 확인되었습니다.'})
    return {
        'ok': True, 'version': RC4_4_VERSION,
        'kpi': {
            'total_members': total_members, 'active_members': active_members, 'vip_members': vip_members,
            'priority_members': priority_members, 'recommendations_total': rec_total, 'recommendations_today': rec_today,
            'sms_today': sms_today, 'wins_today': wins_today, 'login_today': login_today, 'activity_today': activity_today,
            'avg_ai_score': round(float(avg['a'] or 0),1) if avg else 0, 'max_ai_score': round(float(avg['m'] or 0),1) if avg else 0,
            'total_prize': int(prize_row['prize'] or 0) if prize_row else 0, 'checked_total': int(prize_row['c'] or 0) if prize_row else 0
        },
        'latest_draw': dict(latest_draw) if latest_draw else None,
        'recent_members': recent_members, 'recent_logs': recent_logs, 'recent_recommendations': recent_recs,
        'recent_wins': recent_wins, 'alerts': alerts
    }

@app.get('/api/rc4-4/ai-status')
def rc44_ai_status(authorization: str|None = Header(default=None)):
    require_admin(authorization)
    today = _rc44_today_like()
    with con() as c:
        by_grade = _rc44_rows(c, """SELECT COALESCE(m.grade,'일반') grade, COUNT(r.id) c, COALESCE(AVG(r.avg_score),0) avg_score
                                     FROM recommendations r LEFT JOIN members m ON m.id=r.member_id
                                     GROUP BY COALESCE(m.grade,'일반') ORDER BY c DESC""", limit=20)
        by_mode = _rc44_rows(c, "SELECT COALESCE(mode,'balanced') mode, COUNT(*) c, COALESCE(AVG(avg_score),0) avg_score FROM recommendations GROUP BY COALESCE(mode,'balanced') ORDER BY c DESC", limit=20)
        today_rows = _rc44_rows(c, "SELECT id,member_name,round_no,mode,count,avg_score,created_at FROM recommendations WHERE created_at LIKE ? ORDER BY id DESC LIMIT 20", (today,), limit=20)
        recent_runs = _rc44_rows(c, 'SELECT recommendation_id,round_no,mode,count,avg_score,engine_version,created_at FROM engine_runs ORDER BY id DESC LIMIT 20', limit=20)
    return {'ok': True, 'version': RC4_4_VERSION, 'by_grade': by_grade, 'by_mode': by_mode, 'today': today_rows, 'recent_runs': recent_runs}

@app.get('/api/rc4-4/member-dashboard')
def rc44_member_dashboard(authorization: str|None = Header(default=None)):
    require_admin(authorization)
    with con() as c:
        grade = _rc44_rows(c, "SELECT COALESCE(grade,'일반') label, COUNT(*) c FROM members GROUP BY COALESCE(grade,'일반') ORDER BY c DESC", limit=20)
        status = _rc44_rows(c, "SELECT COALESCE(status,'활성') label, COUNT(*) c FROM members GROUP BY COALESCE(status,'활성') ORDER BY c DESC", limit=20)
        priority = _rc44_rows(c, "SELECT COALESCE(priority,'보통') label, COUNT(*) c FROM members GROUP BY COALESCE(priority,'보통') ORDER BY c DESC", limit=20)
        top = _rc44_rows(c, """SELECT m.id,m.name,m.grade,m.status,m.priority,COUNT(r.id) rec_count,COALESCE(MAX(r.created_at),'') latest_recommendation
                              FROM members m LEFT JOIN recommendations r ON r.member_id=m.id
                              GROUP BY m.id,m.name,m.grade,m.status,m.priority
                              ORDER BY rec_count DESC, m.id DESC LIMIT 10""", limit=10)
    return {'ok': True, 'version': RC4_4_VERSION, 'grade': grade, 'status': status, 'priority': priority, 'top_members': top}

@app.post('/api/rc4-4/auto-update')
def rc44_auto_update(backfill:int=12, request:Request=None, authorization: str|None = Header(default=None)):
    admin = require_admin(authorization)
    result = {'ok': True, 'version': RC4_4_VERSION, 'steps': []}
    try:
        sync = _s3_sync_recent_draws(backfill)
        result['steps'].append({'name':'최신회차 동기화','ok':True,'data':sync})
    except Exception as e:
        result['steps'].append({'name':'최신회차 동기화','ok':False,'error':str(e)[:180]})
    try:
        st = latest_stats(100)
        result['steps'].append({'name':'최근 100회 통계 갱신','ok':True,'latest_round':st.get('latest_round'),'sample_size':len(st.get('draws',[]) or [])})
    except Exception as e:
        result['steps'].append({'name':'최근 100회 통계 갱신','ok':False,'error':str(e)[:180]})
    try:
        with con() as c:
            latest = c.execute('SELECT round_no FROM draws ORDER BY round_no DESC LIMIT 1').fetchone()
        if latest:
            auto = _auto_check_round(admin, AutoWinReq(round_no=int(latest['round_no'])), request)
            result['steps'].append({'name':'회원 적중 자동계산','ok':True,'data':auto})
        else:
            result['steps'].append({'name':'회원 적중 자동계산','ok':False,'error':'저장된 회차 없음'})
    except Exception as e:
        result['steps'].append({'name':'회원 적중 자동계산','ok':False,'error':str(e)[:180]})
    try:
        log_action(admin, 'RC4_4_AUTO_UPDATE', json.dumps(result, ensure_ascii=False), request)
    except Exception:
        pass
    result['success_count'] = sum(1 for s in result['steps'] if s.get('ok'))
    result['failed_count'] = sum(1 for s in result['steps'] if not s.get('ok'))
    return result


# ===== RC2 Sprint 3: draw automation / 100-round stats / mobile status =====
SPRINT3_VERSION = 'RC2-Sprint3-DrawStats-Mobile'

def _s3_row_to_draw(r):
    if not r:
        return None
    return {
        'round_no': int(r['round_no']),
        'draw_date': r['draw_date'] or draw_date_for_round(int(r['round_no'])),
        'numbers': parse_nums(r['numbers']),
        'bonus': int(r['bonus'] or 0),
        'source': r['source'] if 'source' in r.keys() else 'db',
        'updated_at': r['updated_at'] if 'updated_at' in r.keys() else ''
    }

def _s3_save_draw(draw, replace=True):
    if not draw or len(parse_nums(draw.get('numbers'))) != 6:
        return None
    with con() as c:
        if replace:
            c.execute('INSERT OR REPLACE INTO draws(round_no,draw_date,numbers,bonus,source,updated_at) VALUES(?,?,?,?,?,?)', (
                int(draw['round_no']), draw.get('draw_date') or draw_date_for_round(int(draw['round_no'])),
                json.dumps(sorted(parse_nums(draw.get('numbers'))), ensure_ascii=False), int(draw.get('bonus') or 0),
                draw.get('source') or 'official', now()
            ))
        else:
            c.execute('INSERT OR IGNORE INTO draws(round_no,draw_date,numbers,bonus,source,updated_at) VALUES(?,?,?,?,?,?)', (
                int(draw['round_no']), draw.get('draw_date') or draw_date_for_round(int(draw['round_no'])),
                json.dumps(sorted(parse_nums(draw.get('numbers'))), ensure_ascii=False), int(draw.get('bonus') or 0),
                draw.get('source') or 'official', now()
            ))
        c.commit()
    return draw

def _s3_sync_recent_draws(backfill=12):
    expected = expected_lotto_round()
    start = max(1, expected - max(1, min(int(backfill or 12), 60)) + 1)
    inserted=[]; skipped=[]; failed=[]
    for round_no in range(start, expected + 1):
        status = draw_status_for_round(round_no)
        existing = get_draw(round_no)
        if existing and len(existing.get('numbers') or []) == 6:
            skipped.append(round_no); continue
        if status in ('future','scheduled'):
            skipped.append(round_no); continue
        fetched = fetch_official_lotto(round_no)
        if fetched:
            _s3_save_draw(fetched, replace=True); inserted.append(round_no)
        else:
            failed.append(round_no)
    latest = None
    with con() as c:
        latest = c.execute('SELECT * FROM draws ORDER BY round_no DESC LIMIT 1').fetchone()
    return {
        'ok': True, 'version': SPRINT3_VERSION, 'expected_round': expected,
        'inserted_rounds': inserted, 'skipped_rounds': skipped, 'failed_rounds': failed,
        'latest_draw': _s3_row_to_draw(latest),
        'message': '동행복권 공개 데이터 기준으로 최근 회차 자동 동기화를 시도했습니다.'
    }

@app.post('/api/draws/sync')
def sprint3_sync_draws(backfill:int=12, authorization: str|None = Header(default=None)):
    admin=require_admin(authorization)
    result=_s3_sync_recent_draws(backfill)
    try: log_action(admin.get('username') or admin.get('name') or 'admin', 'DRAW_SYNC', json.dumps(result, ensure_ascii=False))
    except Exception: pass
    return result

@app.get('/api/draws/status_v2')
def sprint3_draw_status_v2(authorization: str|None = Header(default=None)):
    require_admin(authorization)
    expected = expected_lotto_round()
    current = resolve_draw_for_check(expected, allow_fetch=True)
    previous = resolve_draw_for_check(max(1, expected-1), allow_fetch=True)
    with con() as c:
        total = c.execute('SELECT COUNT(*) c FROM draws').fetchone()['c']
        latest = c.execute('SELECT * FROM draws ORDER BY round_no DESC LIMIT 1').fetchone()
    return {
        'version': SPRINT3_VERSION, 'expected_round': expected, 'total_draws': total,
        'current': current, 'previous': previous, 'latest_draw': _s3_row_to_draw(latest),
        'next_draw_date': draw_date_for_round(expected),
        'summary': '현재 관리 회차와 당첨번호 저장 상태를 자동으로 판정합니다.'
    }

@app.get('/api/stats/round100')
def sprint3_stats_round100(authorization: str|None = Header(default=None)):
    require_admin(authorization)
    st = latest_stats(100)
    draws = st.get('draws', [])[:100]
    trend=[]
    for d in draws[:20]:
        nums=d.get('numbers') or []
        trend.append({
            'round_no': d.get('round_no'), 'draw_date': d.get('draw_date'), 'sum': sum(nums),
            'odd': sum(n%2 for n in nums), 'even': 6-sum(n%2 for n in nums),
            'low': sum(n<=22 for n in nums), 'high': sum(n>=23 for n in nums),
            'numbers': nums, 'bonus': d.get('bonus')
        })
    freq100=st.get('freq100') or {}
    return {
        'version': SPRINT3_VERSION, 'latest_round': st.get('latest_round'), 'sample_size': len(draws),
        'hot': st.get('hot', [])[:15], 'cold': st.get('cold', [])[:15], 'overdue': st.get('overdue', [])[:15],
        'top_pairs': st.get('top_pairs', [])[:15], 'zone_counts': st.get('zone_counts'), 'end_counts': st.get('end_counts'),
        'sum_avg': st.get('sum_avg'), 'odd_ratio': st.get('odd_ratio'),
        'frequency_top': sorted([{'number':int(n),'count':int(c)} for n,c in freq100.items()], key=lambda x:(-x['count'], x['number']))[:20],
        'recent_trend': trend,
        'summary': '최근 100회 기준 빈도/미출현/동반출현/구간/끝수/합계 흐름을 통합 계산했습니다.'
    }

@app.get('/api/mobile/status')
def sprint3_mobile_status(authorization: str|None = Header(default=None)):
    require_admin(authorization)
    with con() as c:
        members=c.execute('SELECT COUNT(*) c FROM members').fetchone()['c']
        recs=c.execute('SELECT COUNT(*) c FROM recommendations').fetchone()['c']
        latest=c.execute('SELECT * FROM draws ORDER BY round_no DESC LIMIT 1').fetchone()
    return {
        'version': SPRINT3_VERSION, 'ok': True, 'members': members, 'recommendations': recs,
        'latest_draw': _s3_row_to_draw(latest),
        'pwa': {'manifest': False, 'service_worker': False, 'mobile_layout': True},
        'summary': '모바일 접속용 핵심 상태 점검이 정상입니다.'
    }

# ===== RC2 Sprint 4: operations hardening / health / backup cleanup =====
SPRINT4_VERSION = 'RC2-Sprint4-OpsHardening'

def _s4_table_count(conn, table):
    try:
        return int(conn.execute(f'SELECT COUNT(*) c FROM {table}').fetchone()['c'])
    except Exception:
        return 0

def _s4_latest_backup():
    latest = None
    try:
        with con() as c:
            latest = c.execute('SELECT * FROM backup_history ORDER BY id DESC LIMIT 1').fetchone()
            if latest:
                return dict(latest)
    except Exception:
        pass
    try:
        files = sorted(EXPORT_DIR.glob('*.db'), key=lambda p: p.stat().st_mtime, reverse=True)
        if files:
            f = files[0]
            return {'filename': f.name, 'reason': 'file', 'size_bytes': f.stat().st_size, 'created_at': datetime.datetime.fromtimestamp(f.stat().st_mtime).strftime('%Y-%m-%d %H:%M:%S')}
    except Exception:
        pass
    return None

def _s4_disk_status():
    try:
        usage = shutil.disk_usage(str(DB_DIR))
        return {'total': usage.total, 'used': usage.used, 'free': usage.free, 'free_mb': round(usage.free/1024/1024, 2)}
    except Exception as e:
        return {'error': str(e)}

def _s4_required_files():
    required = ['backend/app.py','frontend/index.html','frontend/app.js','frontend/login.html','frontend/login.js','requirements.txt','Dockerfile','Procfile']
    rows=[]
    for rel in required:
        p = BASE / rel
        rows.append({'path': rel, 'exists': p.exists(), 'size_bytes': p.stat().st_size if p.exists() else 0})
    return rows

@app.get('/api/ops/health')
def sprint4_ops_health(authorization: str|None = Header(default=None)):
    admin = require_admin(authorization)
    counts = {}
    table_status = []
    required_tables = ['admins','sessions','admin_logs','members','recommendations','draws','settings','backup_history','sms_logs','winning_checks']
    with con() as c:
        for t in required_tables:
            try:
                counts[t] = _s4_table_count(c, t)
                table_status.append({'table': t, 'ok': True, 'count': counts[t]})
            except Exception as e:
                table_status.append({'table': t, 'ok': False, 'error': str(e)})
        try:
            recent_errors = c.execute("SELECT action,detail,created_at FROM admin_logs WHERE action LIKE '%ERROR%' ORDER BY id DESC LIMIT 10").fetchall()
        except Exception:
            recent_errors = []
    db_exists = DB.exists()
    response = {
        'ok': True,
        'version': SPRINT4_VERSION,
        'checked_by': admin.get('username'),
        'time': now(),
        'db': {'engine': DB_ENGINE, 'path': str(DB), 'exists': db_exists, 'size_bytes': DB.stat().st_size if db_exists else 0, 'dir': str(DB_DIR)},
        'disk': _s4_disk_status(),
        'counts': counts,
        'tables': table_status,
        'latest_backup': _s4_latest_backup(),
        'required_files': _s4_required_files(),
        'recent_errors': [dict(r) for r in recent_errors],
        'summary': '운영에 필요한 DB, 백업, 필수 파일, 저장공간 상태를 통합 점검했습니다.'
    }
    return response

@app.post('/api/ops/backup/create')
def sprint4_ops_backup_create(request:Request, authorization: str|None = Header(default=None)):
    admin = require_admin(authorization)
    b = create_db_backup('sprint4_manual', admin)
    try: log_action(admin, 'OPS_BACKUP_CREATE', f'Sprint4 운영 백업 생성: {b.get("filename")}', request)
    except Exception: pass
    return {'ok': True, 'version': SPRINT4_VERSION, 'backup': b}

@app.post('/api/ops/backups/cleanup')
def sprint4_ops_backups_cleanup(keep:int=10, authorization: str|None = Header(default=None)):
    admin = require_admin(authorization); require_super_admin(admin)
    keep = max(3, min(int(keep or 10), 50))
    files = sorted(EXPORT_DIR.glob('*.db'), key=lambda p: p.stat().st_mtime, reverse=True)
    removed=[]
    for f in files[keep:]:
        try:
            removed.append({'filename': f.name, 'size_bytes': f.stat().st_size})
            f.unlink()
        except Exception:
            pass
    return {'ok': True, 'version': SPRINT4_VERSION, 'keep': keep, 'removed': removed, 'remaining': len(files)-len(removed)}

@app.get('/api/ops/audit/recent')
def sprint4_ops_audit_recent(limit:int=100, authorization: str|None = Header(default=None)):
    admin = require_admin(authorization)
    limit = max(10, min(int(limit or 100), 500))
    with con() as c:
        logs = c.execute('SELECT id,username,action,detail,ip,created_at FROM admin_logs ORDER BY id DESC LIMIT ?', (limit,)).fetchall()
    return {'ok': True, 'version': SPRINT4_VERSION, 'logs': [dict(r) for r in logs]}

@app.get('/api/ops/validation')
def sprint4_ops_validation(authorization: str|None = Header(default=None)):
    require_admin(authorization)
    files = _s4_required_files()
    missing_files = [f['path'] for f in files if not f['exists']]
    health = sprint4_ops_health(authorization)
    failed_tables = [t['table'] for t in health.get('tables', []) if not t.get('ok')]
    warnings=[]
    if not health['latest_backup']:
        warnings.append('백업 파일이 아직 없습니다. /api/ops/backup/create 로 백업을 생성하세요.')
    if health.get('disk', {}).get('free_mb', 9999) < 200:
        warnings.append('저장공간이 200MB 미만입니다. 백업 파일 정리가 필요합니다.')
    return {
        'ok': not missing_files and not failed_tables,
        'version': SPRINT4_VERSION,
        'missing_files': missing_files,
        'failed_tables': failed_tables,
        'warnings': warnings,
        'summary': 'Sprint 4 최종 운영 검증 결과입니다.'
    }


# === RC2 Sprint 5: Release / Deployment Readiness ===
SPRINT5_VERSION = 'RC2-Sprint5-Release-Ready'

def _s5_env_status():
    keys = ['PORT','DATABASE_URL','BBLOTTO_DB_DIR','BBLOTTO_EXPORT_DIR','BBLOTTO_SECRET_KEY','ADMIN_USERNAME']
    rows=[]
    for k in keys:
        v=os.getenv(k,'')
        rows.append({'key': k, 'set': bool(str(v).strip()), 'masked': ('***' if str(v).strip() else '')})
    return rows

def _s5_deploy_files():
    required = ['Dockerfile','Procfile','railway.json','requirements.txt','runtime.txt','frontend/index.html','backend/app.py','.gitignore','.dockerignore']
    optional = ['render.yaml','docker-compose.yml','deploy/ubuntu_run.sh','deploy/oracle_cloud_setup.sh','.env.example','START_HERE_DEPLOY.md']
    out=[]
    for rel in required:
        f=BASE/rel
        out.append({'path': rel, 'required': True, 'exists': f.exists(), 'size_bytes': f.stat().st_size if f.exists() else 0})
    for rel in optional:
        f=BASE/rel
        out.append({'path': rel, 'required': False, 'exists': f.exists(), 'size_bytes': f.stat().st_size if f.exists() else 0})
    return out

def _s5_requirements_check():
    req = BASE/'requirements.txt'
    needed = ['fastapi','uvicorn','python-multipart']
    if DB_ENGINE == 'postgresql' or os.getenv('DATABASE_URL'):
        needed.append('psycopg2-binary')
    content = req.read_text(encoding='utf-8', errors='ignore').lower() if req.exists() else ''
    return [{'package': n, 'ok': n.lower() in content} for n in needed]

@app.get('/api/release/readiness')
def sprint5_release_readiness(authorization: str|None = Header(default=None)):
    admin = require_admin(authorization)
    files = _s5_deploy_files()
    missing_required = [x['path'] for x in files if x['required'] and not x['exists']]
    reqs = _s5_requirements_check()
    missing_packages = [x['package'] for x in reqs if not x['ok']]
    health = {'ok': False}
    try:
        health = sprint4_ops_health(authorization)
    except Exception as e:
        health = {'ok': False, 'error': str(e)}
    warnings=[]
    if DB_ENGINE == 'sqlite' and not os.getenv('BBLOTTO_DB_DIR'):
        warnings.append('클라우드 배포에서는 BBLOTTO_DB_DIR 또는 DATABASE_URL 설정을 권장합니다.')
    if not os.getenv('BBLOTTO_SECRET_KEY'):
        warnings.append('운영 환경에서는 BBLOTTO_SECRET_KEY를 설정하는 것이 좋습니다.')
    return {
        'ok': not missing_required and not missing_packages and bool(health.get('ok', False)),
        'version': SPRINT5_VERSION,
        'checked_by': admin.get('username'),
        'time': now(),
        'deployment': {'platform_ready': True, 'recommended_start': 'uvicorn backend.app:app --host 0.0.0.0 --port $PORT'},
        'files': files,
        'missing_required_files': missing_required,
        'requirements': reqs,
        'missing_packages': missing_packages,
        'environment': _s5_env_status(),
        'health_summary': {'ok': health.get('ok'), 'db_engine': DB_ENGINE, 'db_dir': str(DB_DIR)},
        'warnings': warnings,
        'summary': 'Sprint 5 출시/배포 준비 상태 점검 결과입니다.'
    }

@app.get('/api/release/checklist')
def sprint5_release_checklist(authorization: str|None = Header(default=None)):
    require_admin(authorization)
    return {
        'ok': True,
        'version': SPRINT5_VERSION,
        'items': [
            {'step': 1, 'title': 'GitHub 업로드', 'detail': 'ZIP 압축 해제 후 내용물을 저장소 최상위에 업로드'},
            {'step': 2, 'title': '환경변수 설정', 'detail': 'PORT는 플랫폼 자동값 사용, 운영 DB는 DATABASE_URL 또는 BBLOTTO_DB_DIR 지정'},
            {'step': 3, 'title': '배포 시작 명령', 'detail': 'uvicorn backend.app:app --host 0.0.0.0 --port $PORT'},
            {'step': 4, 'title': '상태 확인', 'detail': '/api/health 접속 후 정상 응답 확인'},
            {'step': 5, 'title': '관리자 로그인', 'detail': '관리자 로그인 후 /api/release/readiness 로 최종 점검'},
            {'step': 6, 'title': '백업 생성', 'detail': '운영 전 /api/ops/backup/create 실행 권장'}
        ]
    }


# === RC2 Sprint 6: Release Candidate Stabilization ===
SPRINT6_VERSION = 'RC2-Sprint6-Release-Candidate-Stable'


def _s6_table_count(table):
    try:
        with con() as c:
            row = c.execute(f'SELECT COUNT(*) AS cnt FROM {table}').fetchone()
            return {'table': table, 'ok': True, 'count': int(row['cnt'] if hasattr(row, 'keys') and 'cnt' in row.keys() else row[0])}
    except Exception as e:
        return {'table': table, 'ok': False, 'error': str(e)}


def _s6_file_fingerprint(rel):
    f = BASE / rel
    if not f.exists():
        return {'path': rel, 'exists': False, 'size_bytes': 0}
    h = hashlib.sha256()
    with f.open('rb') as fp:
        for chunk in iter(lambda: fp.read(1024 * 1024), b''):
            h.update(chunk)
    return {'path': rel, 'exists': True, 'size_bytes': f.stat().st_size, 'sha256_12': h.hexdigest()[:12]}


def _s6_safe_backup_name(label='rc2_sprint6'):
    ts = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    return EXPORT_DIR / f'{label}_{ts}.db'


@app.get('/api/rc/version')
def sprint6_rc_version():
    return {
        'ok': True,
        'version': SPRINT6_VERSION,
        'base': 'RC2 Sprint 5',
        'time': now(),
        'db_engine': DB_ENGINE
    }


@app.get('/api/rc/db-safety')
def sprint6_db_safety(authorization: str|None = Header(default=None)):
    require_admin(authorization)
    tables = ['admins','members','draws','admin_logs','settings']
    counts = [_s6_table_count(t) for t in tables]
    warnings = []
    if DB_ENGINE == 'sqlite' and not str(DB_DIR).startswith('/data') and not os.getenv('BBLOTTO_DB_DIR'):
        warnings.append('클라우드 운영에서는 SQLite 기본 경로가 재배포 때 초기화될 수 있습니다. DATABASE_URL 또는 BBLOTTO_DB_DIR 설정을 권장합니다.')
    if not DB.exists() and DB_ENGINE == 'sqlite':
        warnings.append('SQLite DB 파일이 아직 생성되지 않았습니다. 최초 실행 후 다시 확인하세요.')
    return {
        'ok': all(x.get('ok') for x in counts),
        'version': SPRINT6_VERSION,
        'engine': DB_ENGINE,
        'db_dir': str(DB_DIR),
        'db_file': str(DB),
        'tables': counts,
        'warnings': warnings,
        'summary': 'DB 초기화 위험과 핵심 테이블 상태를 점검했습니다.'
    }


@app.post('/api/rc/safe-backup')
def sprint6_safe_backup(authorization: str|None = Header(default=None)):
    admin = require_admin(authorization)
    if DB_ENGINE != 'sqlite':
        return {'ok': False, 'version': SPRINT6_VERSION, 'message': 'PostgreSQL 사용 중에는 플랫폼 DB 백업 기능을 사용하세요.'}
    if not DB.exists():
        raise HTTPException(status_code=404, detail='백업할 SQLite DB 파일이 없습니다.')
    target = _s6_safe_backup_name()
    shutil.copy2(DB, target)
    try:
        log_admin(admin.get('username','admin'), 'rc_safe_backup', f'backup={target.name}', '')
    except Exception:
        pass
    return {'ok': True, 'version': SPRINT6_VERSION, 'backup_file': target.name, 'size_bytes': target.stat().st_size}


@app.get('/api/rc/smoke-test')
def sprint6_smoke_test(authorization: str|None = Header(default=None)):
    admin = require_admin(authorization)
    results = []
    checks = [
        ('health', lambda: health()),
        ('ops_health', lambda: sprint4_ops_health(authorization)),
        ('release_readiness', lambda: sprint5_release_readiness(authorization)),
        ('db_safety', lambda: sprint6_db_safety(authorization)),
    ]
    for name, fn in checks:
        try:
            data = fn()
            results.append({'name': name, 'ok': bool(data.get('ok', False)), 'message': data.get('summary','OK')})
        except Exception as e:
            results.append({'name': name, 'ok': False, 'message': str(e)})
    files = [_s6_file_fingerprint(x) for x in ['backend/app.py','frontend/index.html','frontend/app.js','requirements.txt','Procfile','Dockerfile','START_HERE_DEPLOY.md']]
    ok = all(r['ok'] for r in results) and all(f['exists'] for f in files)
    try:
        log_admin(admin.get('username','admin'), 'rc_smoke_test', f'ok={ok}', '')
    except Exception:
        pass
    return {
        'ok': ok,
        'version': SPRINT6_VERSION,
        'checked_by': admin.get('username'),
        'results': results,
        'files': files,
        'summary': 'RC2 Sprint 6 출시 후보 스모크 테스트 결과입니다.'
    }


@app.get('/api/rc/error-policy')
def sprint6_error_policy():
    return {
        'ok': True,
        'version': SPRINT6_VERSION,
        'policy': {
            'success_shape': {'ok': True, 'data': 'endpoint-specific'},
            'error_shape': {'ok': False, 'error': {'type': '...', 'message': '...'}, 'path': '/api/...', 'time': 'YYYY-MM-DD HH:MM:SS'},
            'admin_required': '관리자 API는 Authorization 헤더가 필요합니다.',
            'safe_message': '예상치 못한 오류는 상세 내부정보 대신 안전한 문구로 반환합니다.'
        }
    }


# === RC3-9: DB / 관리자 로그 / 백업 안정화 ===
# 운영 중 가장 많이 문제가 생기는 데이터 유지, 관리자 기록, 백업 상태를 한 화면에서 점검하기 위한 API입니다.

RC3_9_REQUIRED_TABLES = ['admins','members','recommendations','sms_logs','winning_checks','admin_logs','login_logs','backup_history','settings','draws']


def _rc39_table_count(table: str):
    try:
        with con() as c:
            row = c.execute(f'SELECT COUNT(*) c FROM {table}').fetchone()
            return {'table': table, 'ok': True, 'count': int(row['c'] if row else 0)}
    except Exception as e:
        return {'table': table, 'ok': False, 'count': 0, 'error': str(e)}


def _rc39_latest_backup():
    try:
        with con() as c:
            row = c.execute('SELECT * FROM backup_history ORDER BY id DESC LIMIT 1').fetchone()
            return dict(row) if row else None
    except Exception:
        return None


def _rc39_backup_files(limit: int = 20):
    out = []
    for pattern in ('BBLOTTO*_BACKUP_*.json','BBLOTTO*_BACKUP_*.db','BBLOTTO_RC3_BACKUP_*.json','rc2_sprint6_*.db'):
        for f in EXPORT_DIR.glob(pattern):
            try:
                out.append({'filename': f.name, 'size_bytes': f.stat().st_size, 'modified_at': datetime.datetime.fromtimestamp(f.stat().st_mtime).strftime('%Y-%m-%d %H:%M:%S'), 'format': f.suffix.lstrip('.')})
            except Exception:
                pass
    out.sort(key=lambda x: x.get('modified_at',''), reverse=True)
    return out[:max(1, min(int(limit or 20), 100))]


def _rc39_database_candidates():
    candidates = []
    for rel in ['database/bblotto_v34.db', 'database/lotto.db']:
        f = BASE / rel
        candidates.append({'path': rel, 'exists': f.exists(), 'size_bytes': f.stat().st_size if f.exists() else 0, 'primary': str(f.resolve()) == str(DB.resolve()) if f.exists() else False})
    return candidates


@app.get('/api/rc3-9/status')
def rc39_status(authorization: str|None = Header(default=None)):
    admin = require_admin(authorization)
    counts = [_rc39_table_count(t) for t in RC3_9_REQUIRED_TABLES]
    warnings = []
    if DB_ENGINE == 'sqlite' and not os.getenv('BBLOTTO_DB_DIR') and not str(DB_DIR).startswith('/data'):
        warnings.append('Railway 운영에서는 PostgreSQL DATABASE_URL 또는 영구 저장소 DB 경로 설정을 권장합니다.')
    if not _rc39_latest_backup():
        warnings.append('최근 백업 기록이 없습니다. /api/rc3-9/backup-create 실행을 권장합니다.')
    failed = [x for x in counts if not x.get('ok')]
    return {
        'ok': len(failed) == 0,
        'version': RC3_9_VERSION,
        'checked_by': admin.get('username'),
        'time': now(),
        'db_engine': DB_ENGINE,
        'db_path': str(DB) if DB_ENGINE == 'sqlite' else 'postgresql',
        'export_dir': str(EXPORT_DIR),
        'tables': counts,
        'failed_tables': failed,
        'latest_backup': _rc39_latest_backup(),
        'backup_files': _rc39_backup_files(10),
        'database_candidates': _rc39_database_candidates(),
        'warnings': warnings,
        'summary': 'RC3-9 DB/백업/운영 로그 안정화 점검 결과입니다.'
    }


@app.post('/api/rc3-9/backup-create')
def rc39_backup_create(request: Request, authorization: str|None = Header(default=None)):
    admin = require_admin(authorization)
    result = create_db_backup('rc3_9_manual', admin)
    log_action(admin, 'RC3_9_BACKUP_CREATE', f'RC3-9 수동 백업 생성: {result.get("filename","")}', request)
    return {'ok': True, 'version': RC3_9_VERSION, 'backup': result}


@app.get('/api/rc3-9/admin-audit')
def rc39_admin_audit(limit: int = 100, authorization: str|None = Header(default=None)):
    require_admin(authorization)
    limit = max(10, min(int(limit or 100), 300))
    with con() as c:
        recent = c.execute('SELECT id,username,action,detail,ip,created_at FROM admin_logs ORDER BY id DESC LIMIT ?', (limit,)).fetchall()
        by_action = c.execute('SELECT action, COUNT(*) c, MAX(created_at) latest FROM admin_logs GROUP BY action ORDER BY c DESC LIMIT 30').fetchall()
        logins = c.execute('SELECT username, success, message, ip, created_at FROM login_logs ORDER BY id DESC LIMIT ?', (limit,)).fetchall()
        failed_today = c.execute('SELECT COUNT(*) c FROM login_logs WHERE success=0 AND created_at LIKE ?', (datetime.datetime.now().strftime('%Y-%m-%d')+'%',)).fetchone()['c']
    return {
        'ok': True,
        'version': RC3_9_VERSION,
        'recent_logs': [dict(r) for r in recent],
        'action_counts': [dict(r) for r in by_action],
        'login_logs': [dict(r) for r in logins],
        'failed_login_today': failed_today
    }


@app.get('/api/rc3-9/recommendation-audit')
def rc39_recommendation_audit(limit: int = 100, authorization: str|None = Header(default=None)):
    require_admin(authorization)
    limit = max(10, min(int(limit or 100), 300))
    with con() as c:
        recent = c.execute('SELECT id,member_id,member_name,round_no,mode,count,avg_score,grade,created_at FROM recommendations ORDER BY id DESC LIMIT ?', (limit,)).fetchall()
        by_round = c.execute('SELECT round_no, COUNT(*) c, COALESCE(AVG(avg_score),0) avg_score, MAX(created_at) latest FROM recommendations GROUP BY round_no ORDER BY round_no DESC LIMIT 30').fetchall()
        by_grade = c.execute('SELECT COALESCE(grade,"미분류") grade, COUNT(*) c FROM recommendations GROUP BY COALESCE(grade,"미분류") ORDER BY c DESC').fetchall()
        today = datetime.datetime.now().strftime('%Y-%m-%d')
        today_count = c.execute('SELECT COUNT(*) c FROM recommendations WHERE created_at LIKE ?', (today+'%',)).fetchone()['c']
    return {
        'ok': True,
        'version': RC3_9_VERSION,
        'today_recommendations': today_count,
        'recent': [dict(r) for r in recent],
        'by_round': [dict(r) for r in by_round],
        'by_grade': [dict(r) for r in by_grade]
    }


@app.get('/api/rc3-9/db-standard')
def rc39_db_standard(authorization: str|None = Header(default=None)):
    require_admin(authorization)
    return {
        'ok': True,
        'version': RC3_9_VERSION,
        'primary_database': str(DB) if DB_ENGINE == 'sqlite' else 'DATABASE_URL(PostgreSQL)',
        'db_engine': DB_ENGINE,
        'rule': '운영 기준 DB는 bblotto_v34.db 또는 PostgreSQL DATABASE_URL 하나로 통일합니다. lotto.db는 과거 데이터/참조용으로만 유지합니다.',
        'candidates': _rc39_database_candidates(),
        'recommendation': 'Railway 운영에서는 PostgreSQL 연결을 유지하거나, SQLite를 쓸 경우 BBLOTTO_DB_DIR를 영구 볼륨 경로로 설정하세요.'
    }


# === RC3-10: 당첨번호 자동조회 안정화 ===
@app.get('/api/rc3-10/status')
def rc310_status(round_no:int|None=None, authorization: str|None = Header(default=None)):
    require_admin(authorization)
    r = int(round_no or expected_lotto_round())
    saved = get_draw(r)
    checked = resolve_draw_for_check(r, allow_fetch=True)
    return {
        'ok': True,
        'version': RC3_10_VERSION,
        'round_no': r,
        'draw_status': draw_status_for_round(r),
        'saved': saved,
        'resolved': checked,
        'fallback_available': r in OFFICIAL_DRAW_FALLBACKS,
        'summary': 'RC3-10 당첨번호 자동조회/보조캐시/수동입력 연동 상태입니다.'
    }



# === RC3-12: 회원 연동 당첨확인 안정화 ===
@app.get('/api/rc3-12/status')
def rc312_status(authorization: str|None = Header(default=None)):
    require_admin(authorization)
    with con() as c:
        members = c.execute('SELECT COUNT(*) c FROM members').fetchone()['c']
        recs = c.execute('SELECT COUNT(*) c FROM recommendations').fetchone()['c']
        linked = c.execute('SELECT COUNT(*) c FROM recommendations WHERE COALESCE(member_id,0)>0').fetchone()['c']
        checks = c.execute('SELECT COUNT(*) c FROM winning_checks').fetchone()['c']
    return {'ok': True, 'version': 'RC3-12', 'members': members, 'recommendations': recs, 'linked_recommendations': linked, 'winning_checks': checks, 'summary': '회원 선택 → 추천번호 생성 → 추천이력 저장 → 회원별 당첨확인 흐름을 보강했습니다.'}

# === RC3-11: BBLOTTO AI Engine V1.0 추천번호 고도화 ===
AI_ENGINE_V1_VERSION = 'BBLOTTO_AI_ENGINE_V1_0_RC3_11'


def _ai_v1_window_stats(st, window=300):
    draws = list(st.get('draws') or [])[:max(10, int(window or 300))]
    freq = {n: 0 for n in range(1, 46)}
    pair_counts = collections.Counter()
    triple_counts = collections.Counter()
    last_seen = {n: 999 for n in range(1, 46)}
    for idx, d in enumerate(draws):
        nums = sorted(parse_nums(d.get('numbers') or d.get('nums') or []))
        if len(nums) != 6:
            continue
        for n in nums:
            freq[n] += 1
            if last_seen[n] == 999:
                last_seen[n] = idx
        for p in itertools.combinations(nums, 2):
            pair_counts[tuple(sorted(p))] += 1
        for t in itertools.combinations(nums, 3):
            triple_counts[tuple(sorted(t))] += 1
    avg = sum(freq.values()) / 45 if freq else 0
    hot = sorted(range(1, 46), key=lambda n: (-(freq[n]), last_seen[n], n))[:15]
    cold = sorted(range(1, 46), key=lambda n: (freq[n], -last_seen[n], n))[:15]
    overdue = sorted(range(1, 46), key=lambda n: (-last_seen[n], freq[n], n))[:15]
    mid = sorted(range(1, 46), key=lambda n: (abs(freq[n] - avg), last_seen[n], n))[:15]
    return {
        'draws_used': len(draws), 'freq300': freq, 'pair300': pair_counts, 'triple300': triple_counts,
        'last_seen300': last_seen, 'hot300': hot, 'cold300': cold, 'overdue300': overdue, 'mid300': mid,
        'avg300': avg,
    }


def _ai_v1_profile(st, mode='balanced'):
    ext = _ai_v1_window_stats(st, 300)
    f10 = st.get('freq10') or {}; f30 = st.get('freq30') or {}; f50 = st.get('freq50') or {}; f100 = st.get('freq100') or {}
    f300 = ext['freq300']; last = ext['last_seen300']; avg = ext['avg300'] or 1
    recent = set(st.get('recent_numbers') or [])
    weights = {}
    for n in range(1, 46):
        hot_flow = f10.get(n, 0) * 2.2 + f30.get(n, 0) * 1.35 + f50.get(n, 0) * 0.72 + f100.get(n, 0) * 0.38 + f300.get(n, 0) * 0.16
        cold_flow = max(0, avg - f300.get(n, 0)) * 1.25 + min(18, last.get(n, 999)) * 0.23
        mid_flow = max(0, 9 - abs(f100.get(n, 0) - (sum(f100.values()) / 45 if f100 else avg))) * 0.42
        band = 0.9 if 11 <= n <= 35 else 0.55
        if mode == 'aggressive':
            w = 1.0 + hot_flow * 1.15 + cold_flow * 0.62 + mid_flow * 0.45 + band
        elif mode == 'conservative':
            w = 1.0 + hot_flow * 0.72 + cold_flow * 1.12 + mid_flow * 0.85 + band
        else:
            w = 1.0 + hot_flow * 0.92 + cold_flow * 0.88 + mid_flow * 0.72 + band
        if n in recent:
            w *= 0.68  # 직전 3회 과다 반복 방지
        weights[n] = max(0.2, round(w, 4))
    return weights, ext


def _ai_v1_signature(combo):
    combo = sorted(parse_nums(combo))
    odd = sum(n % 2 for n in combo)
    zones = [sum(n <= 15 for n in combo), sum(16 <= n <= 30 for n in combo), sum(n >= 31 for n in combo)]
    decades = collections.Counter((n - 1) // 10 for n in combo)
    end_dup = 6 - len(set(n % 10 for n in combo))
    cons = sum(1 for a, b in zip(combo, combo[1:]) if b - a == 1)
    return {'odd': odd, 'even': 6 - odd, 'zones': zones, 'sum': sum(combo), 'ac': ac_value(combo), 'decade_max': max(decades.values() or [0]), 'end_dup': end_dup, 'cons': cons}


def _ai_v1_combo_score(combo, st, mode='balanced', ext=None, profile=None):
    combo = sorted(parse_nums(combo))
    if len(combo) != 6:
        return 0.0
    ext = ext or _ai_v1_window_stats(st, 300)
    profile = profile or _ai_v1_profile(st, mode)[0]
    sig = _ai_v1_signature(combo)
    pairs = st.get('pair_counts') or collections.Counter()
    pair300 = ext.get('pair300') or collections.Counter()
    triple300 = ext.get('triple300') or collections.Counter()
    score = 46.0
    score += {3: 14, 2: 11, 4: 11, 1: 2, 5: 2}.get(sig['odd'], -9)
    score += 14 if 105 <= sig['sum'] <= 175 else (8 if 95 <= sig['sum'] <= 190 else -12)
    score += 13 if max(sig['zones']) <= 3 and min(sig['zones']) >= 1 else (4 if max(sig['zones']) <= 4 and min(sig['zones']) >= 1 else -12)
    score += 8 if 6 <= sig['ac'] <= 10 else (3 if 5 <= sig['ac'] <= 11 else -7)
    score += 7 if sig['end_dup'] <= 1 else (3 if sig['end_dup'] == 2 else -6)
    score += 5 if sig['cons'] == 0 else (2 if sig['cons'] == 1 else -8)
    score += 4 if sig['decade_max'] <= 2 else (-5 if sig['decade_max'] >= 4 else 1)
    # 번호별 가중치: 높을수록 좋지만 한쪽으로만 몰리면 감점
    wsum = sum(profile.get(n, 1) for n in combo)
    score += min(10, wsum / 13)
    hot = set(ext.get('hot300', [])[:12]); cold = set(ext.get('cold300', [])[:12]); overdue = set(ext.get('overdue300', [])[:12]); mid = set(ext.get('mid300', [])[:12])
    hot_hit = len(set(combo) & hot); cold_hit = len(set(combo) & cold); overdue_hit = len(set(combo) & overdue); mid_hit = len(set(combo) & mid)
    score += min(7, hot_hit * 1.8) + min(5.5, cold_hit * 1.55) + min(5.5, overdue_hit * 1.55) + min(3.5, mid_hit * 0.8)
    if hot_hit >= 5:
        score -= 5
    if len(set(combo) & set(st.get('recent_numbers') or [])) >= 4:
        score -= 5
    pair_score = sum((pairs.get(tuple(sorted(p)), 0) * 0.62 + pair300.get(tuple(sorted(p)), 0) * 0.38) for p in itertools.combinations(combo, 2))
    strong_pairs = sum(1 for p in itertools.combinations(combo, 2) if (pairs.get(tuple(sorted(p)), 0) + pair300.get(tuple(sorted(p)), 0)) >= 4)
    triple_score = sum(triple300.get(tuple(sorted(t)), 0) for t in itertools.combinations(combo, 3))
    score += min(8, pair_score / 8.5) + min(5, strong_pairs * 1.1) + min(4.5, triple_score / 3.5)
    if strong_pairs >= 8:
        score -= 4
    if triple_score >= 10:
        score -= 3
    # 작은 난수 대신 결정적 흔들림으로 동점 완화
    score += ((sum(n * n for n in combo) + sig['sum']) % 23 - 11) * 0.08
    return round(max(64.0, min(98.8, score)), 1)


def _ai_v1_reasons(combo, st, ext=None):
    combo = sorted(parse_nums(combo)); ext = ext or _ai_v1_window_stats(st, 300); s = set(combo)
    sig = _ai_v1_signature(combo)
    reasons = []
    if len(s & set(ext.get('hot300', [])[:12])) >= 2:
        reasons.append('최근/장기 핵심수 반영')
    if len(s & set(ext.get('overdue300', [])[:12])) >= 1:
        reasons.append('미출현 보강수 포함')
    if len(s & set(ext.get('cold300', [])[:12])) >= 1:
        reasons.append('저출현 반등 후보 포함')
    if sig['odd'] in (2, 3, 4):
        reasons.append(f'홀짝 {sig["odd"]}:{sig["even"]} 균형')
    if max(sig['zones']) <= 3 and min(sig['zones']) >= 1:
        reasons.append('저·중·고 구간 분산')
    if 6 <= sig['ac'] <= 10:
        reasons.append(f'AC값 {sig["ac"]} 안정권')
    if sig['end_dup'] <= 1:
        reasons.append('끝수 중복 최소화')
    if sig['cons'] <= 1:
        reasons.append('연속수 과다 배제')
    pair_hits = []
    for a, b in itertools.combinations(combo, 2):
        c = (st.get('pair_counts') or {}).get(tuple(sorted((a, b))), 0) + (ext.get('pair300') or {}).get(tuple(sorted((a, b))), 0)
        if c >= 3:
            pair_hits.append({'pair': [a, b], 'count': int(c)})
    return reasons[:6] or ['균형형 후보'], sorted(pair_hits, key=lambda x: -x['count'])[:3]


def _ai_v1_detail(combo, st, mode, ext=None, profile=None):
    ext = ext or _ai_v1_window_stats(st, 300); profile = profile or _ai_v1_profile(st, mode)[0]
    combo = sorted(parse_nums(combo)); sig = _ai_v1_signature(combo); reasons, pair_hits = _ai_v1_reasons(combo, st, ext)
    score = _ai_v1_combo_score(combo, st, mode, ext, profile)
    return {
        'numbers': combo, 'score': score, 'ai_score': score, 'vip_score': score,
        'grade': 'VIP' if score >= 94 else 'PREMIUM' if score >= 88 else 'STANDARD',
        'star': '★★★★★' if score >= 94 else '★★★★☆' if score >= 88 else '★★★★',
        'tags': reasons[:4], 'reasons': reasons, 'reason_text': ' · '.join(reasons[:4]),
        'sum': sig['sum'], 'odd': sig['odd'], 'even': sig['even'], 'zones': sig['zones'], 'ac': sig['ac'],
        'end_digit_dup': sig['end_dup'], 'consecutive': sig['cons'], 'pair_hits': pair_hits,
        'engine_version': AI_ENGINE_V1_VERSION,
        'v2_reason': '최근 10/30/50/100/300회 가중치, 페어·트리플, AC값, 끝수, 구간, 포트폴리오 중복 제한을 종합했습니다.'
    }


def make_premium_combos(count=10, fixed='', excluded='', mode='balanced'):
    """RC3-11 / BBLOTTO AI Engine V1.0: 후보 대량 생성 → 점수화 → 포트폴리오 분산 → 근거 표시."""
    st = latest_stats(300)
    fixed_set = set(parse_nums(fixed)); excluded_set = set(parse_nums(excluded))
    fixed_set = {n for n in fixed_set if n not in excluded_set}
    if len(fixed_set) > 6:
        fixed_set = set(sorted(fixed_set)[:6])
    pool = [n for n in range(1, 46) if n not in excluded_set and n not in fixed_set]
    target = max(1, min(50, int(count or 10)))
    if len(pool) + len(fixed_set) < 6:
        raise HTTPException(400, '고정수/제외수를 확인하세요. 선택 가능한 번호가 부족합니다.')
    profile, ext = _ai_v1_profile(st, mode)
    past = {tuple(d['numbers']) for d in st.get('draws', []) if len(d.get('numbers', [])) == 6}
    buckets = {
        'hot': [n for n in ext['hot300'][:18] if n in pool],
        'cold': [n for n in ext['cold300'][:18] if n in pool],
        'overdue': [n for n in ext['overdue300'][:18] if n in pool],
        'mid': [n for n in ext['mid300'][:18] if n in pool],
        'all': pool[:],
    }
    if mode == 'aggressive':
        plans = [['hot','hot','mid','overdue','all','all'], ['hot','mid','all','all','overdue','cold']]
    elif mode == 'conservative':
        plans = [['cold','overdue','mid','mid','all','all'], ['mid','cold','all','overdue','all','hot']]
    else:
        plans = [['hot','cold','overdue','mid','all','all'], ['hot','mid','cold','overdue','all','all']]
    candidates = []
    seen = set()
    tries = 0
    needed = max(3800, target * 320)
    while len(candidates) < needed and tries < max(85000, needed * 26):
        tries += 1
        nums = set(fixed_set)
        plan = random.choice(plans)[:]
        random.shuffle(plan)
        for b in plan:
            if len(nums) >= 6:
                break
            usable = [n for n in buckets.get(b, pool) if n not in nums]
            if usable:
                nums.update(_weighted_pick(usable, [profile[n] for n in usable], 1))
        while len(nums) < 6:
            usable = [n for n in pool if n not in nums]
            nums.update(_weighted_pick(usable, [profile[n] for n in usable], 1))
        arr = tuple(sorted(nums))
        if len(arr) != 6 or arr in seen or arr in past:
            continue
        sig = _ai_v1_signature(arr)
        if sig['odd'] not in (2, 3, 4):
            continue
        if not (92 <= sig['sum'] <= 190):
            continue
        if max(sig['zones']) > 3 or min(sig['zones']) == 0:
            continue
        if sig['cons'] > 1:
            continue
        if sig['end_dup'] > 2:
            continue
        if not (5 <= sig['ac'] <= 11):
            continue
        seen.add(arr)
        candidates.append((_ai_v1_combo_score(arr, st, mode, ext, profile), list(arr)))
    candidates.sort(key=lambda x: (-x[0], x[1]))
    selected = []
    usage = collections.Counter()
    pair_usage = collections.Counter()
    for score, combo in candidates:
        pairs = [tuple(sorted(p)) for p in itertools.combinations(combo, 2)]
        if any(usage[n] >= 3 for n in combo if n not in fixed_set):
            continue
        if any(pair_usage[p] >= 1 for p in pairs):
            continue
        if all(len(set(combo) & set(prev)) <= 3 for prev in selected):
            selected.append(combo); usage.update(combo); pair_usage.update(pairs)
        if len(selected) >= target:
            break
    if len(selected) < target:
        for score, combo in candidates:
            if combo in selected:
                continue
            pairs = [tuple(sorted(p)) for p in itertools.combinations(combo, 2)]
            if any(usage[n] >= 4 for n in combo if n not in fixed_set):
                continue
            if any(pair_usage[p] >= 2 for p in pairs):
                continue
            if all(len(set(combo) & set(prev)) <= 4 for prev in selected):
                selected.append(combo); usage.update(combo); pair_usage.update(pairs)
            if len(selected) >= target:
                break
    if len(selected) < target:
        for score, combo in candidates:
            if combo not in selected:
                selected.append(combo)
            if len(selected) >= target:
                break
    details = [_ai_v1_detail(c, st, mode, ext, profile) for c in selected[:target]]
    st.update(ext)
    st['ai_v1_candidates'] = len(candidates)
    st['ai_v1_attempts'] = tries
    st['engine_version'] = AI_ENGINE_V1_VERSION
    return selected[:target], details, st


def _engine_summary(details, st):
    scores = [float(d.get('score') or d.get('ai_score') or d.get('vip_score') or 0) for d in (details or []) if (d.get('score') or d.get('ai_score') or d.get('vip_score'))]
    return {
        'version': AI_ENGINE_V1_VERSION,
        'engine_version': AI_ENGINE_V1_VERSION,
        'avg_score': round(sum(scores) / len(scores), 1) if scores else 0,
        'max_score': round(max(scores), 1) if scores else 0,
        'min_score': round(min(scores), 1) if scores else 0,
        'candidate_count': int(st.get('ai_v1_candidates') or st.get('s23_candidates') or 0),
        'selected_count': len(details or []),
        'latest_round': st.get('latest_round') or 0,
        'ai_engine_v1_report': {
            'pipeline': '최근 10/30/50/100/300회 통계 → 번호별 가중치 → 페어·트리플 분석 → 후보 대량 생성 → 패턴 필터 → 포트폴리오 분산 → TOP3 선정',
            'draws_used': int(st.get('draws_used') or len(st.get('draws') or [])),
            'candidate_count': int(st.get('ai_v1_candidates') or 0),
            'attempts': int(st.get('ai_v1_attempts') or 0),
            'filters': '홀짝 2:4~4:2 / 합계 92~190 / 구간 1개 이상 / AC 5~11 / 끝수·연속수 제한',
            'portfolio': '동일 번호 과사용, 동일 페어 반복, 조합 간 4~5개 이상 중복을 제한합니다.',
            'summary': 'BBLOTTO AI Engine V1.0은 단순 랜덤이 아니라 후보 조합을 대량 생성한 뒤 통계 근거와 분산 품질로 최종 선별합니다.'
        },
        'v2_pipeline_report': {
            'pipeline': '최근 10/30/50/100/300회 통계 → 가중치 → 후보 생성 → 점수화 → 분산 선별',
            'stage1_candidates': int(st.get('ai_v1_candidates') or 0),
            'stage2_filters': '페어/트리플/AC/끝수/연속수/구간 필터',
            'stage3_portfolio': '번호·페어 과사용 제한 및 TOP3 우선 표시',
            'summary': 'RC3-11 AI Engine V1.0 적용 완료'
        }
    }


def build_analysis_text(round_no, st, mode, fixed, excluded, details=None):
    details = details or []
    engine = _engine_summary(details, st)
    best = sorted(details, key=lambda x: -float(x.get('score') or 0))[:3]
    top_nums = []
    for d in best:
        for n in d.get('numbers', []):
            if n not in top_nums:
                top_nums.append(n)
    avg = engine.get('avg_score', 0)
    hot = (st.get('hot300') or st.get('hot') or [])[:6]
    overdue = (st.get('overdue300') or st.get('overdue') or [])[:6]
    mode_name = {'balanced': '균형형', 'aggressive': '공격형', 'conservative': '보수형'}.get(mode, mode or '균형형')
    lines = [
        f'{round_no}회차는 {mode_name} 기준으로 최근 10/30/50/100/300회 흐름을 함께 반영했습니다.',
        f'핵심 흐름은 {", ".join(map(str, hot)) if hot else "자동 분석"} / 보강 흐름은 {", ".join(map(str, overdue)) if overdue else "분산 보강"} 중심입니다.',
        f'TOP3 후보 핵심 번호는 {", ".join(map(str, top_nums[:8])) if top_nums else "생성 결과 기준 자동 산출"}이며 평균 AI 점수는 {avg}점입니다.',
        '페어·트리플 출현, AC값, 끝수, 구간, 홀짝, 조합 간 중복을 동시에 제한했습니다.',
        ''
    ]
    return '\n'.join(lines)


@app.get('/api/rc3-11/engine-status')
def rc311_engine_status(authorization: str|None = Header(default=None)):
    require_admin(authorization)
    st = latest_stats(300)
    profile, ext = _ai_v1_profile(st, 'balanced')
    top_weighted = sorted(profile.keys(), key=lambda n: (-profile[n], n))[:12]
    sample_combos, sample_details, sample_stats = make_premium_combos(5, '', '', 'balanced')
    return {
        'ok': True,
        'version': AI_ENGINE_V1_VERSION,
        'latest_round': st.get('latest_round'),
        'draws_used': ext.get('draws_used'),
        'top_weighted': top_weighted,
        'hot300': ext.get('hot300', [])[:12],
        'cold300': ext.get('cold300', [])[:12],
        'overdue300': ext.get('overdue300', [])[:12],
        'sample': {'combos': sample_combos, 'details': sample_details, 'engine': _engine_summary(sample_details, sample_stats)},
        'summary': 'RC3-11 BBLOTTO AI Engine V1.0 상태 점검입니다.'
    }


# === RC3-14: 회원 상세 화면 정리 ===
@app.get('/api/rc3-14/status')
def rc314_status(authorization: str|None = Header(default=None)):
    admin = require_admin(authorization)
    return {
        'ok': True,
        'version': 'RC3-14',
        'summary': '회원 상세 페이지에서 추천이력 노출을 제거하고 문구이력/당첨이력 중심으로 정리했습니다.',
        'sections': ['member_profile', 'memo', 'sms_logs', 'winning_checks']
    }


# === RC3-15: 당첨번호 회차 무결성 점검/복구 ===
@app.get('/api/rc3-15/draw-integrity')
def rc315_draw_integrity(authorization: str|None = Header(default=None)):
    require_admin(authorization)
    expected, completed = _rc315_expected_round_and_completed()
    with con() as c:
        future_rows = c.execute('SELECT round_no,draw_date,numbers,bonus,source,updated_at FROM draws WHERE round_no>? ORDER BY round_no DESC', (completed,)).fetchall()
        latest_rows = c.execute('SELECT round_no,draw_date,numbers,bonus,source,updated_at FROM draws WHERE round_no<=? ORDER BY round_no DESC LIMIT 15', (completed or 999999,)).fetchall()
    return {
        'ok': len(future_rows) == 0,
        'version': 'RC3-15',
        'expected_round': expected,
        'completed_round': completed,
        'future_or_invalid_draws': [dict(r) for r in future_rows],
        'latest_valid_draws': [{'round_no':r['round_no'], 'draw_date':r['draw_date'], 'numbers':parse_nums(r['numbers']), 'bonus':r['bonus'], 'source':r['source'], 'updated_at':r['updated_at']} for r in latest_rows],
        'message': 'ok=false이면 /api/rc3-15/repair-draws 를 실행하면 추첨 전 회차 당첨번호가 정리됩니다.'
    }

@app.post('/api/rc3-15/repair-draws')
def rc315_repair_draws(request:Request, authorization: str|None = Header(default=None)):
    admin = require_admin(authorization)
    with con() as c:
        result = _rc315_clean_future_draws_in_conn(c)
        c.commit()
    log_action(admin, 'RC3_15_REPAIR_DRAWS', f'추첨 전/미래 회차 당첨번호 {result.get("removed",0)}건 정리', request)
    return {'ok': True, 'version': 'RC3-15', **result, 'message': f'잘못 저장된 추첨 전/미래 회차 {result.get("removed",0)}건을 정리했습니다.'}

# ===== RC4-2: Recommendation Engine Quality Upgrade =====
# 최종 실행 시점에서 /api/generate가 사용하는 추천 엔진을 RC4-2 버전으로 재정의합니다.
AI_ENGINE_RC42_VERSION = 'BBLOTTO_PRO_V2_RC4_2_AI_ENGINE'

def _rc42_gap_signature(combo):
    nums = sorted(parse_nums(combo))
    gaps = [nums[i] - nums[i-1] for i in range(1, len(nums))]
    return {
        'gaps': gaps,
        'min_gap': min(gaps) if gaps else 0,
        'max_gap': max(gaps) if gaps else 0,
        'small_gaps': sum(1 for g in gaps if g <= 2),
        'wide_gaps': sum(1 for g in gaps if g >= 15),
    }

def _rc42_end_digits(combo):
    digits = [int(n) % 10 for n in parse_nums(combo)]
    return digits, collections.Counter(digits)

def _rc42_pattern_key(combo):
    sig = _ai_v1_signature(combo)
    digits, dc = _rc42_end_digits(combo)
    decade = tuple(sum(1 for n in combo if lo <= n <= hi) for lo, hi in [(1,9),(10,19),(20,29),(30,39),(40,45)])
    return (sig['odd'], tuple(sig['zones']), decade, tuple(sorted(dc.values(), reverse=True)[:2]))

def _rc42_combo_ok(combo, past=None, fixed_set=None):
    combo = tuple(sorted(parse_nums(combo)))
    if len(combo) != 6 or len(set(combo)) != 6:
        return False, '중복 번호'
    if past and combo in past:
        return False, '과거 당첨 조합과 동일'
    sig = _ai_v1_signature(combo)
    gap = _rc42_gap_signature(combo)
    digits, dc = _rc42_end_digits(combo)
    if sig['odd'] not in (2, 3, 4):
        return False, '홀짝 불균형'
    if not (96 <= sig['sum'] <= 184):
        return False, '합계 범위 이탈'
    if max(sig['zones']) > 3 or min(sig['zones']) < 1:
        return False, '저중고 구간 불균형'
    if sig['cons'] > 1:
        return False, '연속수 과다'
    if sig['ac'] < 6 or sig['ac'] > 11:
        return False, 'AC값 범위 이탈'
    if max(dc.values() or [0]) > 2:
        return False, '끝수 중복 과다'
    if len(set(digits)) < 5:
        return False, '끝수 분산 부족'
    if gap['small_gaps'] > 2:
        return False, '근접 번호 과다'
    if gap['wide_gaps'] > 1 or gap['max_gap'] >= 20:
        return False, '번호 간격 과다'
    return True, '통과'

def _rc42_adjusted_score(combo, st, mode, ext, profile):
    score = float(_ai_v1_combo_score(combo, st, mode, ext, profile))
    sig = _ai_v1_signature(combo)
    gap = _rc42_gap_signature(combo)
    digits, dc = _rc42_end_digits(combo)
    if sig['odd'] == 3:
        score += 1.5
    if max(sig['zones']) == 2:
        score += 1.2
    if len(set(digits)) >= 6:
        score += 1.4
    elif len(set(digits)) == 5:
        score += 0.8
    if 118 <= sig['sum'] <= 166:
        score += 1.1
    if 7 <= sig['ac'] <= 10:
        score += 1.2
    if gap['small_gaps'] == 0:
        score += 0.8
    if gap['wide_gaps'] == 0:
        score += 0.6
    return round(max(70.0, min(99.4, score)), 1)

def _rc42_detail(combo, st, mode, ext=None, profile=None):
    ext = ext or _ai_v1_window_stats(st, 300)
    profile = profile or _ai_v1_profile(st, mode)[0]
    detail = _ai_v1_detail(combo, st, mode, ext, profile)
    combo = sorted(parse_nums(combo))
    sig = _ai_v1_signature(combo)
    gap = _rc42_gap_signature(combo)
    digits, dc = _rc42_end_digits(combo)
    score = _rc42_adjusted_score(combo, st, mode, ext, profile)
    tags = list(detail.get('tags') or [])
    for tag in ['RC4-2 고품질 필터', '끝수 5개 이상 분산', '포트폴리오 중복 제한']:
        if tag not in tags:
            tags.append(tag)
    detail.update({
        'score': score,
        'ai_score': score,
        'vip_score': score,
        'grade': 'VIP' if score >= 94 else 'PREMIUM' if score >= 89 else 'STANDARD',
        'star': '★★★★★' if score >= 94 else '★★★★☆' if score >= 89 else '★★★★',
        'tags': tags[:6],
        'reason_text': ' · '.join(tags[:5]),
        'gap_min': gap['min_gap'],
        'gap_max': gap['max_gap'],
        'end_digit_diversity': len(set(digits)),
        'end_digit_counts': dict(dc),
        'quality_rule': '홀짝·저중고·끝수·AC·합계·간격·연속수·중복패턴 RC4-2 필터 통과',
        'engine_version': AI_ENGINE_RC42_VERSION,
        'v2_reason': 'RC4-2는 후보를 대량 생성한 뒤 끝수 분산, 근접/광역 간격, 패턴 반복, 페어 반복, 조합 간 중복을 추가로 제한합니다.',
    })
    return detail

def make_premium_combos(count=10, fixed='', excluded='', mode='balanced'):
    """RC4-2: AI 후보 생성 + 고품질 패턴 필터 + 포트폴리오 중복 제한."""
    st = latest_stats(300)
    fixed_set = set(parse_nums(fixed)); excluded_set = set(parse_nums(excluded))
    fixed_set = {n for n in fixed_set if n not in excluded_set}
    if len(fixed_set) > 6:
        fixed_set = set(sorted(fixed_set)[:6])
    pool = [n for n in range(1, 46) if n not in excluded_set and n not in fixed_set]
    target = max(1, min(50, int(count or 10)))
    if len(pool) + len(fixed_set) < 6:
        raise HTTPException(400, '고정수/제외수를 확인하세요. 선택 가능한 번호가 부족합니다.')
    profile, ext = _ai_v1_profile(st, mode)
    past = {tuple(sorted(d['numbers'])) for d in st.get('draws', []) if len(d.get('numbers', [])) == 6}
    buckets = {
        'hot': [n for n in ext['hot300'][:20] if n in pool],
        'cold': [n for n in ext['cold300'][:20] if n in pool],
        'overdue': [n for n in ext['overdue300'][:20] if n in pool],
        'mid': [n for n in ext['mid300'][:24] if n in pool],
        'all': pool[:],
    }
    if mode == 'aggressive':
        plans = [['hot','hot','mid','overdue','all','cold'], ['hot','mid','all','overdue','all','all'], ['hot','cold','overdue','all','all','mid']]
    elif mode == 'conservative':
        plans = [['mid','mid','cold','overdue','all','all'], ['cold','overdue','mid','hot','all','all'], ['mid','cold','all','all','overdue','all']]
    else:
        plans = [['hot','cold','overdue','mid','all','all'], ['hot','mid','cold','overdue','all','all'], ['mid','hot','overdue','all','cold','all']]
    candidates = []
    seen = set()
    tries = 0
    needed = max(6200, target * 520)
    max_tries = max(125000, needed * 24)
    while len(candidates) < needed and tries < max_tries:
        tries += 1
        nums = set(fixed_set)
        plan = random.choice(plans)[:]
        random.shuffle(plan)
        for b in plan:
            if len(nums) >= 6:
                break
            usable = [n for n in buckets.get(b, pool) if n not in nums]
            if usable:
                nums.update(_weighted_pick(usable, [profile[n] for n in usable], 1))
        while len(nums) < 6:
            usable = [n for n in pool if n not in nums]
            nums.update(_weighted_pick(usable, [profile[n] for n in usable], 1))
        arr = tuple(sorted(nums))
        if arr in seen:
            continue
        ok, reason = _rc42_combo_ok(arr, past=past, fixed_set=fixed_set)
        if not ok:
            continue
        seen.add(arr)
        candidates.append((_rc42_adjusted_score(arr, st, mode, ext, profile), list(arr)))
    candidates.sort(key=lambda x: (-x[0], x[1]))
    selected = []
    usage = collections.Counter()
    pair_usage = collections.Counter()
    pattern_usage = collections.Counter()
    def can_add(combo, max_number_use, max_pair_use, max_overlap, max_pattern_use):
        pairs = [tuple(sorted(p)) for p in itertools.combinations(combo, 2)]
        pattern = _rc42_pattern_key(combo)
        if any(usage[n] >= max_number_use for n in combo if n not in fixed_set):
            return False
        if any(pair_usage[p] >= max_pair_use for p in pairs):
            return False
        if pattern_usage[pattern] >= max_pattern_use:
            return False
        if not all(len(set(combo) & set(prev)) <= max_overlap for prev in selected):
            return False
        return True
    def add_combo(combo):
        pairs = [tuple(sorted(p)) for p in itertools.combinations(combo, 2)]
        selected.append(combo)
        usage.update(combo)
        pair_usage.update(pairs)
        pattern_usage.update([_rc42_pattern_key(combo)])
    for score, combo in candidates:
        if can_add(combo, 3, 1, 3, 1):
            add_combo(combo)
        if len(selected) >= target:
            break
    if len(selected) < target:
        for score, combo in candidates:
            if combo in selected:
                continue
            if can_add(combo, 4, 2, 3, 2):
                add_combo(combo)
            if len(selected) >= target:
                break
    if len(selected) < target:
        for score, combo in candidates:
            if combo in selected:
                continue
            if can_add(combo, 5, 2, 4, 3):
                add_combo(combo)
            if len(selected) >= target:
                break
    if len(selected) < target:
        for score, combo in candidates:
            if combo not in selected:
                add_combo(combo)
            if len(selected) >= target:
                break
    details = [_rc42_detail(c, st, mode, ext, profile) for c in selected[:target]]
    details.sort(key=lambda x: -float(x.get('score') or 0))
    selected_sorted = [d['numbers'] for d in details]
    st.update(ext)
    st['ai_v1_candidates'] = len(candidates)
    st['ai_v1_attempts'] = tries
    st['engine_version'] = AI_ENGINE_RC42_VERSION
    st['rc42_portfolio'] = {
        'max_overlap': max((len(set(a) & set(b)) for i, a in enumerate(selected_sorted) for b in selected_sorted[i+1:]), default=0),
        'unique_patterns': len({_rc42_pattern_key(c) for c in selected_sorted}),
        'number_usage_max': max(usage.values()) if usage else 0,
    }
    return selected_sorted[:target], details[:target], st

def _engine_summary(details, st):
    scores = [float(d.get('score') or d.get('ai_score') or d.get('vip_score') or 0) for d in (details or []) if (d.get('score') or d.get('ai_score') or d.get('vip_score'))]
    return {
        'version': AI_ENGINE_RC42_VERSION,
        'engine_version': AI_ENGINE_RC42_VERSION,
        'phase': 'RC4-2',
        'avg_score': round(sum(scores) / len(scores), 1) if scores else 0,
        'max_score': round(max(scores), 1) if scores else 0,
        'min_score': round(min(scores), 1) if scores else 0,
        'candidate_count': int(st.get('ai_v1_candidates') or 0),
        'selected_count': len(details or []),
        'latest_round': st.get('latest_round') or 0,
        'rc42_report': {
            'pipeline': '최근 10/30/50/100/300회 통계 → 가중치 후보 생성 → RC4-2 품질 필터 → 점수화 → 포트폴리오 중복 제한',
            'filters': '홀짝 2:4~4:2 / 합계 96~184 / 저중고 각 1개 이상 / AC 6~11 / 끝수 5개 이상 / 연속수 1쌍 이하 / 과도한 간격 제한',
            'portfolio': st.get('rc42_portfolio') or {},
            'summary': 'RC4-2는 번호 품질, 끝수 분산, 반복 패턴 제거, 조합 간 중복 제한을 강화한 추천 엔진입니다.'
        },
        'ai_engine_v1_report': {
            'pipeline': 'RC4-2 품질 필터가 적용된 BBLOTTO AI Engine',
            'draws_used': int(st.get('draws_used') or len(st.get('draws') or [])),
            'candidate_count': int(st.get('ai_v1_candidates') or 0),
            'attempts': int(st.get('ai_v1_attempts') or 0),
            'filters': '홀짝·합계·구간·AC·끝수·간격·연속수·과거조합 제외',
            'portfolio': '동일 번호 과사용, 동일 페어 반복, 유사 패턴 반복, 조합 간 4개 이상 중복을 제한합니다.',
            'summary': 'BBLOTTO PRO RC4-2 추천 엔진 적용 완료'
        },
        'v2_pipeline_report': {
            'pipeline': '통계 가중치 → 후보 대량 생성 → 품질 필터 → 포트폴리오 선별',
            'stage1_candidates': int(st.get('ai_v1_candidates') or 0),
            'stage2_filters': '끝수/간격/AC/구간/홀짝/연속수 필터 강화',
            'stage3_portfolio': '번호·페어·패턴 중복 제한 및 TOP3 우선 표시',
            'summary': 'RC4-2 추천번호 AI 엔진 고도화 완료'
        }
    }

def build_analysis_text(round_no, st, mode, fixed, excluded, details=None):
    details = details or []
    engine = _engine_summary(details, st)
    best = sorted(details, key=lambda x: -float(x.get('score') or 0))[:3]
    top_nums = []
    for d in best:
        for n in d.get('numbers', []):
            if n not in top_nums:
                top_nums.append(n)
    avg = engine.get('avg_score', 0)
    hot = (st.get('hot300') or st.get('hot') or [])[:6]
    overdue = (st.get('overdue300') or st.get('overdue') or [])[:6]
    mode_name = {'balanced': '균형형', 'aggressive': '공격형', 'conservative': '보수형'}.get(mode, mode or '균형형')
    lines = [
        f'{round_no}회차는 RC4-2 {mode_name} 엔진으로 최근 10/30/50/100/300회 흐름을 반영했습니다.',
        f'핵심 흐름 {", ".join(map(str, hot)) if hot else "자동 분석"} / 보강 흐름 {", ".join(map(str, overdue)) if overdue else "분산 보강"} 기준입니다.',
        f'TOP3 핵심 후보 번호는 {", ".join(map(str, top_nums[:8])) if top_nums else "생성 결과 기준 자동 산출"}이며 평균 AI 점수는 {avg}점입니다.',
        '끝수 5개 이상 분산, 연속수 1쌍 이하, AC값 6~11, 조합 간 중복 제한을 적용했습니다.',
        ''
    ]
    return '\n'.join(lines)


# ===================== RC4-5 DEEP RECOMMEND ENGINE =====================
def rc45_grade_label(raw='일반'):
    v = str(raw or '일반').strip()
    alias = {
        'VIP':'1등', '다이아':'1등', '다이아몬드':'1등', '프리미엄':'2등',
        '1등관리':'1등', '1등 관리':'1등', '2등관리':'2등', '2등 관리':'2등',
        '일반관리':'일반', '일반 관리':'일반'
    }
    return alias.get(v, v if v in ('1등','2등','일반') else '일반')

def rc45_grade_strength_text(grade='일반'):
    g = rc45_grade_label(grade)
    if g == '1등':
        return '고강도 심층분석: 후보군 확대, 페어 반복 억제, 끝수·구간·AC·중복 제한을 가장 엄격하게 적용'
    if g == '2등':
        return '중강도 심층분석: 최근 흐름과 미출현 보강을 균형 반영하고 유사 조합을 제한'
    return '표준 심층분석: 안정형 균형 조합 중심으로 과한 패턴과 중복을 제한'

def _rc45_grade_params(grade='일반'):
    g = rc45_grade_label(grade)
    if g == '1등':
        return {'needed': 820, 'base': 9200, 'tries': 30, 'first': (2,1,3,1), 'second': (3,1,3,2), 'third': (4,2,3,2), 'score_boost': 1.8}
    if g == '2등':
        return {'needed': 660, 'base': 7600, 'tries': 26, 'first': (3,1,3,1), 'second': (4,2,3,2), 'third': (5,2,4,3), 'score_boost': 1.0}
    return {'needed': 520, 'base': 6200, 'tries': 24, 'first': (4,2,4,2), 'second': (5,2,4,3), 'third': (6,3,4,4), 'score_boost': 0.3}

def make_premium_combos(count=10, fixed='', excluded='', mode='balanced', member_grade='일반'):
    """RC4-5: 회원 등급(1등/2등/일반)별 추천 강도 차등 + 심층분석 후보 선별."""
    member_grade = rc45_grade_label(member_grade)
    gp = _rc45_grade_params(member_grade)
    st = latest_stats(300)
    fixed_set = set(parse_nums(fixed)); excluded_set = set(parse_nums(excluded))
    fixed_set = {n for n in fixed_set if n not in excluded_set}
    if len(fixed_set) > 6:
        fixed_set = set(sorted(fixed_set)[:6])
    pool = [n for n in range(1, 46) if n not in excluded_set and n not in fixed_set]
    target = max(1, min(50, int(count or 10)))
    if len(pool) + len(fixed_set) < 6:
        raise HTTPException(400, '고정수/제외수를 확인하세요. 선택 가능한 번호가 부족합니다.')
    profile, ext = _ai_v1_profile(st, mode)
    for n in range(1, 46):
        if n in (ext.get('hot300') or [])[:12]:
            profile[n] = profile.get(n, 1) + gp['score_boost']
        if n in (ext.get('overdue300') or [])[:12]:
            profile[n] = profile.get(n, 1) + gp['score_boost'] * 0.75
        if member_grade == '1등' and n in (ext.get('mid300') or [])[:16]:
            profile[n] = profile.get(n, 1) + 0.6
    past = {tuple(sorted(d['numbers'])) for d in st.get('draws', []) if len(d.get('numbers', [])) == 6}
    buckets = {
        'hot': [n for n in ext['hot300'][:22] if n in pool],
        'cold': [n for n in ext['cold300'][:22] if n in pool],
        'overdue': [n for n in ext['overdue300'][:22] if n in pool],
        'mid': [n for n in ext['mid300'][:26] if n in pool],
        'all': pool[:],
    }
    if member_grade == '1등':
        plans = [['hot','overdue','mid','cold','all','all'], ['hot','mid','overdue','all','cold','all'], ['mid','hot','overdue','cold','all','all']]
    elif member_grade == '2등':
        plans = [['hot','cold','overdue','mid','all','all'], ['mid','hot','overdue','all','cold','all'], ['cold','mid','hot','all','overdue','all']]
    elif mode == 'aggressive':
        plans = [['hot','hot','mid','overdue','all','cold'], ['hot','mid','all','overdue','all','all'], ['hot','cold','overdue','all','all','mid']]
    elif mode == 'conservative':
        plans = [['mid','mid','cold','overdue','all','all'], ['cold','overdue','mid','hot','all','all'], ['mid','cold','all','all','overdue','all']]
    else:
        plans = [['hot','cold','overdue','mid','all','all'], ['hot','mid','cold','overdue','all','all'], ['mid','hot','overdue','all','cold','all']]
    candidates = []
    seen = set()
    tries = 0
    needed = max(gp['base'], target * gp['needed'])
    max_tries = max(125000, needed * gp['tries'])
    while len(candidates) < needed and tries < max_tries:
        tries += 1
        nums = set(fixed_set)
        plan = random.choice(plans)[:]
        random.shuffle(plan)
        for b in plan:
            if len(nums) >= 6:
                break
            usable = [n for n in buckets.get(b, pool) if n not in nums]
            if usable:
                nums.update(_weighted_pick(usable, [profile[n] for n in usable], 1))
        while len(nums) < 6:
            usable = [n for n in pool if n not in nums]
            nums.update(_weighted_pick(usable, [profile[n] for n in usable], 1))
        arr = tuple(sorted(nums))
        if arr in seen:
            continue
        ok, reason = _rc42_combo_ok(arr, past=past, fixed_set=fixed_set)
        if not ok:
            continue
        seen.add(arr)
        score = _rc42_adjusted_score(arr, st, mode, ext, profile)
        spread_bonus = len({n % 10 for n in arr}) * (0.12 if member_grade == '1등' else 0.08)
        section_bonus = len({0 if n <= 15 else 1 if n <= 30 else 2 for n in arr}) * (0.18 if member_grade == '1등' else 0.10)
        candidates.append((score + spread_bonus + section_bonus, list(arr)))
    candidates.sort(key=lambda x: (-x[0], x[1]))
    selected = []
    usage = collections.Counter()
    pair_usage = collections.Counter()
    pattern_usage = collections.Counter()
    def can_add(combo, max_number_use, max_pair_use, max_overlap, max_pattern_use):
        pairs = [tuple(sorted(p)) for p in itertools.combinations(combo, 2)]
        pattern = _rc42_pattern_key(combo)
        if any(usage[n] >= max_number_use for n in combo if n not in fixed_set):
            return False
        if any(pair_usage[p] >= max_pair_use for p in pairs):
            return False
        if pattern_usage[pattern] >= max_pattern_use:
            return False
        if not all(len(set(combo) & set(prev)) <= max_overlap for prev in selected):
            return False
        return True
    def add_combo(combo):
        selected.append(combo)
        usage.update(combo)
        pair_usage.update([tuple(sorted(p)) for p in itertools.combinations(combo, 2)])
        pattern_usage.update([_rc42_pattern_key(combo)])
    for score, combo in candidates:
        if can_add(combo, *gp['first']):
            add_combo(combo)
        if len(selected) >= target:
            break
    if len(selected) < target:
        for score, combo in candidates:
            if combo in selected:
                continue
            if can_add(combo, *gp['second']):
                add_combo(combo)
            if len(selected) >= target:
                break
    if len(selected) < target:
        for score, combo in candidates:
            if combo in selected:
                continue
            if can_add(combo, *gp['third']):
                add_combo(combo)
            if len(selected) >= target:
                break
    if len(selected) < target:
        for score, combo in candidates:
            if combo not in selected:
                add_combo(combo)
            if len(selected) >= target:
                break
    details = [_rc42_detail(c, st, mode, ext, profile) for c in selected[:target]]
    for d in details:
        d['member_grade'] = member_grade
        d['grade_strength'] = rc45_grade_strength_text(member_grade)
        d['engine_version'] = 'RC4-5_DEEP_GRADE_ENGINE'
    details.sort(key=lambda x: -float(x.get('score') or 0))
    selected_sorted = [d['numbers'] for d in details]
    st.update(ext)
    st['ai_v1_candidates'] = len(candidates)
    st['ai_v1_attempts'] = tries
    st['engine_version'] = 'RC4-5_DEEP_GRADE_ENGINE'
    st['member_grade'] = member_grade
    st['grade_strength'] = rc45_grade_strength_text(member_grade)
    st['rc42_portfolio'] = {
        'max_overlap': max((len(set(a) & set(b)) for i, a in enumerate(selected_sorted) for b in selected_sorted[i+1:]), default=0),
        'unique_patterns': len({_rc42_pattern_key(c) for c in selected_sorted}),
        'number_usage_max': max(usage.values()) if usage else 0,
    }
    return selected_sorted[:target], details[:target], st

def _engine_summary(details, st):
    scores = [float(d.get('score') or d.get('ai_score') or d.get('vip_score') or 0) for d in (details or []) if (d.get('score') or d.get('ai_score') or d.get('vip_score'))]
    grade = rc45_grade_label(st.get('member_grade') or (details[0].get('member_grade') if details else '일반'))
    return {
        'version': 'RC4-5_DEEP_GRADE_ENGINE',
        'engine_version': 'RC4-5_DEEP_GRADE_ENGINE',
        'phase': 'RC4-5',
        'member_grade': grade,
        'grade_strength': rc45_grade_strength_text(grade),
        'avg_score': round(sum(scores) / len(scores), 1) if scores else 0,
        'max_score': round(max(scores), 1) if scores else 0,
        'min_score': round(min(scores), 1) if scores else 0,
        'candidate_count': int(st.get('ai_v1_candidates') or 0),
        'selected_count': len(details or []),
        'latest_round': st.get('latest_round') or 0,
        'rc45_report': {
            'pipeline': '최근 10/30/50/100/300회 통계 → 등급별 후보군 가중치 → 심층 품질 필터 → 포트폴리오 중복 제한 → 최종 점수 선별',
            'grade_policy': '회원 등급은 1등/2등/일반으로 운영하며 등급별 후보 수, 중복 제한, 조합 선별 강도를 다르게 적용합니다.',
            'filters': '홀짝·합계·저중고 구간·AC값·끝수·간격·연속수·과거 동일 조합·페어 반복 제한',
            'portfolio': st.get('rc42_portfolio') or {},
            'summary': '최근 흐름과 누적 통계를 함께 반영한 심층 추천 결과입니다.'
        }
    }

def build_analysis_text(round_no, st, mode, fixed, excluded, details=None):
    """회원용 3~5줄 분석 문구. 개발자/엔진명 문구는 노출하지 않고, 생성 결과에 따라 문장 조합이 달라지도록 구성합니다."""
    import hashlib
    details = details or []
    engine = _engine_summary(details, st)
    grade = engine.get('member_grade') or rc45_grade_label(st.get('member_grade'))
    best = sorted(details, key=lambda x: -float(x.get('score') or 0))[:3]
    top_nums = []
    for d in best:
        for n in d.get('numbers', []):
            if n not in top_nums:
                top_nums.append(n)
    hot = (st.get('hot300') or st.get('hot') or [])[:8]
    overdue = (st.get('overdue300') or st.get('overdue') or [])[:8]
    mode_name = {'balanced': '균형형', 'aggressive': '공격형', 'conservative': '안정형'}.get(mode, mode or '균형형')
    core_text = ', '.join(map(str, top_nums[:6])) if top_nums else '주요 후보군'
    hot_text = ', '.join(map(str, hot[:4])) if hot else '최근 흐름 번호'
    sub_text = ', '.join(map(str, overdue[:4])) if overdue else '보강 후보 번호'
    combos = [d.get('numbers', []) for d in details if d.get('numbers')]
    flat = [int(n) for c in combos for n in c if str(n).isdigit()]
    low = sum(1 for n in flat if 1 <= n <= 15); mid = sum(1 for n in flat if 16 <= n <= 30); high = sum(1 for n in flat if 31 <= n <= 45)
    odd = sum(1 for n in flat if n % 2 == 1); even = len(flat) - odd
    focus = []
    if mode == 'conservative': focus.append('안정형')
    elif mode == 'aggressive': focus.append('공격형')
    else: focus.append('균형형')
    if hot: focus.append('최근흐름')
    if overdue: focus.append('보강후보')
    if abs(odd-even) <= max(2, len(flat)//8): focus.append('홀짝균형')
    if max(low, mid, high) - min(low, mid, high) <= max(2, len(flat)//10): focus.append('구간분산')

    seed_src = f"{round_no}|{mode}|{grade}|{core_text}|{hot_text}|{sub_text}|{sum(flat)}|{len(details)}"
    seed = int(hashlib.sha256(seed_src.encode('utf-8')).hexdigest()[:12], 16)
    def pick(arr, salt=0):
        return arr[(seed + salt) % len(arr)]

    openers = {
        '균형형': [
            f'{round_no}회차는 최근 흐름과 누적 데이터를 함께 비교해 안정적인 분포의 조합으로 구성했습니다.',
            f'{round_no}회차는 특정 번호대에 치우치지 않도록 전체 흐름을 기준으로 조합을 선별했습니다.',
            f'이번 회차는 최근 당첨 흐름과 장기 통계를 함께 반영해 균형 중심으로 구성했습니다.',
            f'번호 분포와 최근 출현 흐름을 함께 검토해 {round_no}회차 추천 조합을 구성했습니다.'
        ],
        '안정형': [
            f'{round_no}회차는 과도한 변동보다 안정적인 번호 흐름을 우선해 조합을 선별했습니다.',
            f'이번 회차는 최근 흐름 안에서 무리한 편중을 줄이고 안정성을 높이는 방향으로 구성했습니다.',
            f'{round_no}회차는 누적 통계와 반복 패턴을 함께 살펴 안정적인 조합을 중심으로 선별했습니다.'
        ],
        '공격형': [
            f'{round_no}회차는 최근 흐름 변화가 큰 구간을 함께 반영해 적극적인 조합으로 구성했습니다.',
            f'이번 회차는 출현 가능성이 높아진 후보를 중심으로 변화를 준 조합을 선별했습니다.',
            f'{round_no}회차는 최근 강세 번호와 보강 후보를 함께 반영해 흐름 전환 가능성을 고려했습니다.'
        ]
    }
    middles = [
        f'주요 후보는 {core_text}이며, 최근 흐름 번호와 보강 후보를 함께 배분했습니다.',
        f'최근 흐름 번호({hot_text})와 보강 후보({sub_text})를 조합해 한쪽으로 쏠리지 않게 맞췄습니다.',
        f'핵심 후보군은 {core_text} 중심이며, 전체 조합 간 중복 가능성을 낮추는 방향으로 정리했습니다.',
        f'번호별 출현 흐름을 비교해 {hot_text} 계열과 {sub_text} 계열을 균형 있게 반영했습니다.',
        '최근 반복된 패턴은 일부만 반영하고, 새롭게 움직일 가능성이 있는 번호를 함께 보강했습니다.',
        '당첨 흐름이 강한 번호와 장기적으로 보강이 필요한 번호를 나누어 조합에 배치했습니다.'
    ]
    balances = [
        '홀짝 비율과 저·중·고 구간 분포를 함께 맞춰 전체적인 안정성을 높였습니다.',
        '끝수 흐름과 번호 간 간격을 함께 확인해 비슷한 형태의 조합이 반복되지 않도록 조정했습니다.',
        '연속수와 반복 패턴은 필요한 범위 안에서만 반영해 조합 간 차이를 살렸습니다.',
        '구간 분산과 번호 간 연결성을 함께 고려해 조합별 완성도를 높였습니다.',
        '최근 회차와 지나치게 유사한 구성은 줄이고, 조합별 다양성을 확보했습니다.',
        f'{mode_name} 기준에 맞춰 번호대, 끝수, 반복 흐름을 함께 점검했습니다.'
    ]
    closers = [
        '전체적으로 최근 데이터와 누적 통계를 함께 고려한 심층 추천 결과입니다.',
        '이번 추천은 안정성과 변화 가능성을 함께 반영한 구성입니다.',
        '단순 빈도보다 번호 간 균형과 최근 흐름을 함께 본 추천입니다.',
        '회원별 추천 이력과 중복 가능성까지 고려해 최종 조합을 정리했습니다.',
        '최근 흐름을 유지하면서도 새로운 출현 가능성을 함께 고려했습니다.'
    ]
    opener_pool = openers.get(focus[0], openers['균형형'])
    lines = [pick(opener_pool, 1), pick(middles, 7), pick(balances, 13)]
    if (seed % 3) != 0:
        lines.append(pick(closers, 19))
    # 3~4줄 유지, 중복 문장 제거
    clean = []
    for line in lines:
        if line and line not in clean:
            clean.append(line)
    return '\n'.join(clean[:4])
# ===================== /RC4-5 DEEP RECOMMEND ENGINE =====================

# ===================== RC5-5 RECOMMEND ENGINE UPGRADE =====================
# 추천번호 생성 품질 강화: 등급별 후보 강도, 최근 추천 중복 방지, 조합별 다양성 강화.
RC55_ENGINE_VERSION = 'RC7-29_GRADE_TIER_ENGINE'

def _rc55_recent_recommendation_combos(member_id=None, limit=180):
    """최근 생성 이력에서 추천 조합을 가져와 동일/유사 조합 반복을 줄입니다."""
    out = []
    try:
        with con() as c:
            cols = table_cols(c, 'recommendations')
            if 'numbers' not in cols:
                return out
            if member_id and 'member_id' in cols:
                rs = c.execute('SELECT numbers FROM recommendations WHERE member_id=? ORDER BY id DESC LIMIT ?', (member_id, limit)).fetchall()
            else:
                rs = c.execute('SELECT numbers FROM recommendations ORDER BY id DESC LIMIT ?', (limit,)).fetchall()
            for r in rs:
                try:
                    data = json.loads(r['numbers'] or '[]')
                except Exception:
                    data = []
                if isinstance(data, list):
                    for item in data:
                        nums = item.get('numbers') if isinstance(item, dict) else item
                        nums = parse_nums(nums)
                        if len(nums) == 6:
                            out.append(tuple(sorted(nums)))
                if len(out) >= limit:
                    break
    except Exception:
        return []
    return out[:limit]

def _rc55_too_close_to_history(combo, history, max_overlap=4):
    s = set(parse_nums(combo))
    for h in history or []:
        hs = set(parse_nums(h))
        if len(s & hs) > max_overlap:
            return True
    return False

def _rc55_combo_quality_bonus(combo, st, ext, usage=None, member_grade='일반'):
    nums = sorted(parse_nums(combo))
    sig = _ai_v1_signature(nums)
    digits = [n % 10 for n in nums]
    decade = [sum(1 for n in nums if lo <= n <= hi) for lo, hi in [(1,9),(10,19),(20,29),(30,39),(40,45)]]
    bonus = 0.0
    # 가장 안정적인 합계/AC/구간에 가산점
    if 115 <= sig.get('sum', 0) <= 170:
        bonus += 1.2
    if 7 <= sig.get('ac', 0) <= 10:
        bonus += 1.1
    if sig.get('odd') in (2, 3, 4):
        bonus += 0.8
    if len(set(digits)) >= 5:
        bonus += 0.9
    if max(decade or [0]) <= 2:
        bonus += 0.7
    # 최근 추천에서 과사용된 번호는 감점해 회원에게 비슷한 느낌이 반복되지 않게 함
    if usage:
        bonus -= sum(max(0, usage.get(n, 0) - 1) * 0.18 for n in nums)
    g = rc45_grade_label(member_grade)
    if g == '1등':
        bonus += 1.8
    elif g == '2등':
        bonus += 0.9
    return bonus


def _rc729_engine_name(grade='일반'):
    g = rc45_grade_label(grade)
    if g == '1등':
        return 'AI MASTER'
    if g == '2등':
        return 'AI PREMIUM'
    return 'AI BASIC'

def _rc729_grade_score(base_score, grade='일반'):
    """RC7-29: 등급별 점수 구간을 분리합니다.
    일반은 90~94점대, 2등은 94~97점대, 1등은 97~99점대로 표시해
    화면에서 등급별 엔진 차이가 확실히 보이도록 합니다.
    """
    try:
        base = float(base_score or 0)
    except Exception:
        base = 85.0
    # 기존 엔진 점수를 0~1 값으로 정규화하여 등급 구간 안에서 상대 품질을 유지
    norm = max(0.0, min(1.0, (base - 70.0) / 29.7))
    g = rc45_grade_label(grade)
    if g == '1등':
        lo, hi = 97.0, 99.2
    elif g == '2등':
        lo, hi = 94.0, 96.8
    else:
        lo, hi = 90.0, 93.8
    return round(lo + (hi - lo) * norm, 1)

def _rc729_grade_tags(grade='일반'):
    g = rc45_grade_label(grade)
    if g == '1등':
        return ['1등 전용', 'AI MASTER', '상위 후보 선별']
    if g == '2등':
        return ['2등 전용', 'AI PREMIUM', '강화 후보 선별']
    return ['일반 전용', 'AI BASIC', '기본 균형 선별']

def _rc55_grade_params(grade='일반'):
    # RC7-29: 1등/2등/일반 3단계만 사용하고 등급별 후보 선별 강도를 분리합니다.
    g = rc45_grade_label(grade)
    if g == '1등':
        return {
            'base': 2200, 'mult': 220, 'tries': 14,
            'first': (2, 1, 3, 1), 'second': (3, 1, 3, 2), 'third': (4, 2, 3, 3),
            'history_overlap': 4, 'hot_boost': 3.2, 'overdue_boost': 2.15, 'mid_boost': 1.25,
        }
    if g == '2등':
        return {
            'base': 1650, 'mult': 165, 'tries': 11,
            'first': (3, 1, 3, 1), 'second': (4, 2, 3, 2), 'third': (5, 2, 4, 3),
            'history_overlap': 4, 'hot_boost': 2.25, 'overdue_boost': 1.55, 'mid_boost': 0.8,
        }
    return {
        'base': 1000, 'mult': 120, 'tries': 9,
        'first': (4, 2, 4, 2), 'second': (5, 2, 4, 3), 'third': (6, 3, 4, 4),
        'history_overlap': 5, 'hot_boost': 1.15, 'overdue_boost': 0.75, 'mid_boost': 0.25,
    }

def make_premium_combos(count=10, fixed='', excluded='', mode='balanced', member_grade='일반', member_id=None):
    """RC5-5: 추천번호 생성 엔진 업그레이드.
    - 1등/2등/일반 등급별 후보군 강도 차등
    - 최근 추천 이력과 유사한 조합 제한
    - 번호/페어/패턴 중복 제한 강화
    - 조합 간 분산성과 안정성 동시 반영
    """
    member_grade = rc45_grade_label(member_grade)
    gp = _rc55_grade_params(member_grade)
    st = latest_stats(300)
    fixed_set = set(parse_nums(fixed)); excluded_set = set(parse_nums(excluded))
    fixed_set = {n for n in fixed_set if n not in excluded_set}
    if len(fixed_set) > 6:
        fixed_set = set(sorted(fixed_set)[:6])
    pool = [n for n in range(1, 46) if n not in excluded_set and n not in fixed_set]
    target = max(1, min(50, int(count or 10)))
    if len(pool) + len(fixed_set) < 6:
        raise HTTPException(400, '고정수/제외수를 확인하세요. 선택 가능한 번호가 부족합니다.')

    profile, ext = _ai_v1_profile(st, mode)
    hot = ext.get('hot300') or []
    cold = ext.get('cold300') or []
    overdue = ext.get('overdue300') or []
    mid = ext.get('mid300') or []
    # 등급별 가중치 강화
    for n in range(1, 46):
        if n in hot[:14]:
            profile[n] = profile.get(n, 1) + gp['hot_boost']
        if n in overdue[:14]:
            profile[n] = profile.get(n, 1) + gp['overdue_boost']
        if n in mid[:18]:
            profile[n] = profile.get(n, 1) + gp['mid_boost']
        if n in cold[:10] and member_grade == '1등':
            profile[n] = profile.get(n, 1) + 0.35

    past_draws = {tuple(sorted(d['numbers'])) for d in st.get('draws', []) if len(d.get('numbers', [])) == 6}
    recent_history = _rc55_recent_recommendation_combos(member_id=member_id, limit=220)
    recent_usage = collections.Counter(n for combo in recent_history[:80] for n in combo)

    buckets = {
        'hot': [n for n in hot[:24] if n in pool],
        'cold': [n for n in cold[:24] if n in pool],
        'overdue': [n for n in overdue[:24] if n in pool],
        'mid': [n for n in mid[:28] if n in pool],
        'low': [n for n in range(1, 16) if n in pool],
        'middle': [n for n in range(16, 31) if n in pool],
        'high': [n for n in range(31, 46) if n in pool],
        'all': pool[:],
    }
    # 다양한 계획을 섞어 매번 후보군 성격이 달라지도록 구성
    plans = [
        ['hot','cold','overdue','mid','all','all'],
        ['hot','mid','overdue','low','middle','high'],
        ['mid','hot','cold','overdue','all','all'],
        ['low','middle','high','hot','overdue','all'],
        ['cold','mid','hot','all','overdue','all'],
        ['overdue','hot','middle','high','low','all'],
    ]
    if member_grade == '1등':
        plans += [
            ['hot','overdue','mid','cold','low','high'],
            ['hot','hot','overdue','mid','middle','all'],
            ['overdue','mid','hot','cold','all','all'],
        ]
    elif member_grade == '2등':
        plans += [
            ['hot','cold','overdue','mid','middle','all'],
            ['mid','hot','overdue','low','high','all'],
        ]
    if mode == 'aggressive':
        plans += [['hot','hot','overdue','high','all','all'], ['hot','mid','high','overdue','all','all']]
    elif mode == 'conservative':
        plans += [['mid','mid','cold','middle','all','all'], ['cold','overdue','mid','low','middle','all']]

    candidates = []
    seen = set()
    tries = 0
    needed = max(gp['base'], target * gp['mult'])
    max_tries = max(9000, needed * gp['tries'])
    while len(candidates) < needed and tries < max_tries:
        tries += 1
        nums = set(fixed_set)
        plan = random.choice(plans)[:]
        random.shuffle(plan)
        for b in plan:
            if len(nums) >= 6:
                break
            usable = [n for n in buckets.get(b, pool) if n not in nums]
            if usable:
                nums.update(_weighted_pick(usable, [profile.get(n, 1) for n in usable], 1))
        while len(nums) < 6:
            usable = [n for n in pool if n not in nums]
            nums.update(_weighted_pick(usable, [profile.get(n, 1) for n in usable], 1))
        arr = tuple(sorted(nums))
        if arr in seen or arr in past_draws:
            continue
        if _rc55_too_close_to_history(arr, recent_history, max_overlap=gp['history_overlap']):
            continue
        ok, _reason = _rc42_combo_ok(arr, past=past_draws, fixed_set=fixed_set)
        if not ok:
            continue
        seen.add(arr)
        base_score = _rc42_adjusted_score(arr, st, mode, ext, profile)
        score = base_score + _rc55_combo_quality_bonus(arr, st, ext, recent_usage, member_grade)
        candidates.append((round(max(70.0, min(99.7, score)), 1), list(arr)))

    candidates.sort(key=lambda x: (-x[0], x[1]))
    selected = []
    usage = collections.Counter()
    pair_usage = collections.Counter()
    pattern_usage = collections.Counter()

    def can_add(combo, max_number_use, max_pair_use, max_overlap, max_pattern_use):
        pairs = [tuple(sorted(p)) for p in itertools.combinations(combo, 2)]
        pattern = _rc42_pattern_key(combo)
        if any(usage[n] >= max_number_use for n in combo if n not in fixed_set):
            return False
        if any(pair_usage[p] >= max_pair_use for p in pairs):
            return False
        if pattern_usage[pattern] >= max_pattern_use:
            return False
        if not all(len(set(combo) & set(prev)) <= max_overlap for prev in selected):
            return False
        return True

    def add_combo(combo):
        selected.append(combo)
        usage.update(combo)
        pair_usage.update([tuple(sorted(p)) for p in itertools.combinations(combo, 2)])
        pattern_usage.update([_rc42_pattern_key(combo)])

    for limits in (gp['first'], gp['second'], gp['third'], (6, 3, 4, 5)):
        if len(selected) >= target:
            break
        for score, combo in candidates:
            if combo in selected:
                continue
            if can_add(combo, *limits):
                add_combo(combo)
            if len(selected) >= target:
                break
    if len(selected) < target:
        for score, combo in candidates:
            if combo not in selected:
                add_combo(combo)
            if len(selected) >= target:
                break

    details = [_rc42_detail(c, st, mode, ext, profile) for c in selected[:target]]
    # 실제 최종 점수도 RC5-6 빠른 엔진 기준으로 재보정
    for d in details:
        nums = d.get('numbers') or []
        raw_score = float(d.get('score') or 0) + _rc55_combo_quality_bonus(nums, st, ext, recent_usage, member_grade)
        score = _rc729_grade_score(raw_score, member_grade)
        d['raw_score'] = round(max(70.0, min(99.7, raw_score)), 1)
        d['score'] = d['ai_score'] = d['vip_score'] = score
        d['member_grade'] = member_grade
        d['grade_strength'] = rc45_grade_strength_text(member_grade)
        d['engine_version'] = RC55_ENGINE_VERSION
        d['engine_label'] = _rc729_engine_name(member_grade)
        d['grade'] = '1등' if member_grade == '1등' else '2등' if member_grade == '2등' else '일반'
        tags = [t for t in (d.get('tags') or []) if not str(t).startswith('RC') and str(t) not in ('VIP','PREMIUM','STANDARD')]
        for tag in _rc729_grade_tags(member_grade) + ['최근 추천 중복 제한', '구간·끝수 분산', '조합 간 유사도 제한']:
            if tag not in tags:
                tags.append(tag)
        d['tags'] = tags[:6]
        d['reason_text'] = ' · '.join(tags[:5])
        d['quality_rule'] = '홀짝·구간·끝수·간격·AC·연속수·최근 추천 유사도까지 종합 검토'
    details.sort(key=lambda x: -float(x.get('score') or 0))
    selected_sorted = [d['numbers'] for d in details]

    st.update(ext)
    st['ai_v1_candidates'] = len(candidates)
    st['ai_v1_attempts'] = tries
    st['engine_version'] = RC55_ENGINE_VERSION
    st['member_grade'] = member_grade
    st['grade_strength'] = rc45_grade_strength_text(member_grade)
    st['rc55_history_checked'] = len(recent_history)
    st['rc42_portfolio'] = {
        'max_overlap': max((len(set(a) & set(b)) for i, a in enumerate(selected_sorted) for b in selected_sorted[i+1:]), default=0),
        'unique_patterns': len({_rc42_pattern_key(c) for c in selected_sorted}),
        'number_usage_max': max(usage.values()) if usage else 0,
        'recent_history_checked': len(recent_history),
    }
    return selected_sorted[:target], details[:target], st

def _engine_summary(details, st):
    scores = [float(d.get('score') or d.get('ai_score') or d.get('vip_score') or 0) for d in (details or []) if (d.get('score') or d.get('ai_score') or d.get('vip_score'))]
    grade = rc45_grade_label(st.get('member_grade') or (details[0].get('member_grade') if details else '일반'))
    return {
        'version': RC55_ENGINE_VERSION,
        'engine_version': RC55_ENGINE_VERSION,
        'phase': 'RC7-29',
        'member_grade': grade,
        'engine_label': _rc729_engine_name(grade),
        'grade_strength': rc45_grade_strength_text(grade),
        'avg_score': round(sum(scores) / len(scores), 1) if scores else 0,
        'max_score': round(max(scores), 1) if scores else 0,
        'min_score': round(min(scores), 1) if scores else 0,
        'candidate_count': int(st.get('ai_v1_candidates') or 0),
        'selected_count': len(details or []),
        'latest_round': st.get('latest_round') or 0,
        'rc55_report': {
            'grade_policy': '1등/2등/일반 3단계 전용 · 일반 90~94점대 / 2등 94~97점대 / 1등 97~99점대 구간 분리',
            'portfolio': st.get('rc42_portfolio') or {},
            'summary': '최근 추천 이력과 유사한 조합을 줄이고 분산형 후보를 우선 선별했습니다.'
        },
        # 화면 호환용 기존 키 유지
        'rc45_report': {
            'summary': '최근 흐름과 누적 통계를 함께 반영한 심층 추천 결과입니다.',
            'portfolio': st.get('rc42_portfolio') or {},
        }
    }
# ===================== /RC5-5 RECOMMEND ENGINE UPGRADE =====================


@app.get('/api/rc6-11/status')
def rc6_11_status(authorization: str|None = Header(default=None)):
    admin=require_admin(authorization)
    return {'ok': True, 'version': 'RC6-11_MEMBER_QUERY_REAL_FIX', 'engine': DB_ENGINE, 'admin': admin.get('username')}

# ===================== RC7-1: MEMBER PERSONALIZED AI ENGINE V2 =====================
# 회원별 시드(회원ID/이름/회차/조합)를 분석 문구에 반영하여
# 같은 회차 CSV에서도 회원별 추천번호·요약이 반복되지 않도록 보강합니다.
RC71_ENGINE_VERSION = 'RC7-1_MEMBER_PERSONALIZED_AI_ENGINE_V2'

def _rc71_seed(*parts):
    try:
        import hashlib
        raw = '|'.join(str(p or '') for p in parts)
        return int(hashlib.sha256(raw.encode('utf-8')).hexdigest()[:12], 16)
    except Exception:
        return 0

def build_analysis_text(round_no, st, mode, fixed, excluded, details=None):
    details = details or []
    engine = _engine_summary(details, st)
    member_id = st.get('member_id') or 0
    member_name = st.get('member_name') or ''
    grade = engine.get('member_grade') or rc45_grade_label(st.get('member_grade'))
    combos = [d.get('numbers', []) for d in details if d.get('numbers')]
    flat = [int(n) for c in combos for n in c if str(n).isdigit()]
    best = sorted(details, key=lambda x: -float(x.get('score') or 0))[:3]
    top_nums=[]
    for d in best:
        for n in d.get('numbers', []):
            if n not in top_nums:
                top_nums.append(n)
    hot = (st.get('hot300') or st.get('hot') or [])[:8]
    overdue = (st.get('overdue300') or st.get('overdue') or [])[:8]
    mode_name = {'balanced': '균형형', 'aggressive': '공격형', 'conservative': '안정형'}.get(mode, mode or '균형형')
    core_text = ', '.join(map(str, top_nums[:6])) if top_nums else '주요 후보군'
    hot_text = ', '.join(map(str, hot[:4])) if hot else '최근 흐름 번호'
    sub_text = ', '.join(map(str, overdue[:4])) if overdue else '보강 후보 번호'
    odd = sum(1 for n in flat if n % 2 == 1); even = len(flat) - odd
    low = sum(1 for n in flat if 1 <= n <= 15); mid = sum(1 for n in flat if 16 <= n <= 30); high = sum(1 for n in flat if 31 <= n <= 45)
    seed = _rc71_seed(round_no, member_id, member_name, mode, grade, sum(flat), core_text)
    def pick(arr, salt=0):
        return arr[(seed + salt) % len(arr)]
    openers = [
        f'{round_no}회차는 {mode_name} 기준으로 최근 흐름과 누적 데이터를 함께 비교했습니다.',
        f'{round_no}회차는 회원별 추천 이력과 번호 분포를 나누어 조합을 선별했습니다.',
        f'이번 회차는 최근 강세 구간과 보강 후보를 함께 반영해 맞춤형으로 구성했습니다.',
        f'{round_no}회차는 특정 번호대 편중을 줄이고 조합별 차이를 확보하는 방향으로 정리했습니다.',
        f'이번 추천은 최근 출현 흐름, 끝수, 구간, 홀짝 균형을 함께 점검했습니다.',
    ]
    middles = [
        f'핵심 후보는 {core_text} 중심이며, {hot_text} 흐름을 일부 반영했습니다.',
        f'최근 흐름 번호({hot_text})와 보강 후보({sub_text})를 나누어 배치했습니다.',
        f'조합 간 번호 중복을 줄이고 {core_text} 후보군의 분산도를 높였습니다.',
        f'장기 보강 후보({sub_text})를 함께 넣어 단순 고빈도 조합을 피했습니다.',
        '최근 반복된 패턴은 일부만 반영하고 새롭게 움직일 가능성이 있는 번호를 보강했습니다.',
    ]
    balances = [
        f'전체 홀짝 흐름은 홀수 {odd}개/짝수 {even}개 기준으로 검토했습니다.',
        f'저·중·고 구간 분포는 {low}/{mid}/{high} 흐름으로 맞춰 편중을 줄였습니다.',
        '끝수 반복과 연속수 과다 사용을 제한해 조합별 형태가 겹치지 않게 했습니다.',
        '번호 간 간격과 AC값을 함께 확인해 단순 나열식 조합을 줄였습니다.',
        '회원별 발송 조합이 서로 비슷하게 반복되지 않도록 분산 기준을 추가했습니다.',
    ]
    closers = [
        '단순 빈도보다 번호 간 균형과 최근 흐름을 함께 본 추천입니다.',
        '최근 데이터와 누적 통계를 함께 고려한 심층 추천 결과입니다.',
        '안정성과 변화 가능성을 동시에 반영한 구성입니다.',
    ]
    lines = [pick(openers,1), pick(middles,7), pick(balances,13), pick(closers,19)]
    clean=[]
    for line in lines:
        if line and line not in clean:
            clean.append(line)
    return '\n'.join(clean[:4])

@app.get('/api/rc7-1/status')
def rc7_1_status(authorization: str|None = Header(default=None)):
    admin=require_admin(authorization)
    return {'ok': True, 'version': RC71_ENGINE_VERSION, 'engine': DB_ENGINE, 'summary': '회원별 추천번호/분석요약 분산 엔진 적용', 'admin': admin.get('username')}
# ===================== /RC7-1 MEMBER PERSONALIZED AI ENGINE V2 =====================




# ===================== RC7-3 SMSGANDA REAL XLS EXPORT =====================
class SmsGandaRow(BaseModel):
    name: str = ''
    phone: str = ''
    seg1: str = ''
    seg2: str = ''
    seg3: str = ''
    seg4: str = ''

class SmsGandaXlsReq(BaseModel):
    rows: list[SmsGandaRow] = []
    scope: str = 'all'
    round_no: str | int | None = None

def _smsganda_clean_phone(value):
    return re.sub(r'[^0-9]', '', str(value or ''))

def _safe_download_name(name):
    return re.sub(r'[^0-9A-Za-z_\-가-힣]', '_', str(name or 'download'))[:120] or 'download'

def _smsganda_cell_text(value):
    """RC7-21 문자간다 최종 포맷.
    문자간다 엑셀 업로드가 CR/LF를 지우는 환경이 있어, 실제 미리보기에서 줄로 인식된
    Unicode LINE SEPARATOR(U+2028)를 저장합니다. 또한 이전 패치의 / 구분과 1) 형식을
    1. 형식의 한 줄 1조합으로 복구합니다.
    """
    text = str(value or '')
    text = text.replace('\r\n', '\n').replace('\r', '\n').replace('\\n', '\n')
    text = re.sub(r'\s*/\s*(?=\d{1,2}[\.\)]\s*)', '\n', text)
    text = re.sub(r'(\[추천번호\])\s*(?=\d{1,2}[\.\)]\s*)', r'\1\n', text)
    text = re.sub(r'([^\n])\s+(?=\d{1,2}[\.\)]\s*\d)', r'\1\n', text)
    text = re.sub(r'(^|\n)(\d{1,2})\)\s*', r'\1\2. ', text)
    text = re.sub(r'\n{3,}', '\n\n', text).strip()
    return text.replace('\n', '\u2028')

@app.post('/api/export/smsganda_xls')
def export_smsganda_real_xls(req: SmsGandaXlsReq, authorization: str|None = Header(default=None)):
    require_admin(authorization)
    try:
        import xlwt
    except Exception as e:
        raise HTTPException(status_code=500, detail='문자간다 XLS 생성 모듈(xlwt)이 설치되지 않았습니다. requirements.txt에 xlwt를 추가한 뒤 재배포하세요.') from e

    cleaned=[]
    seen=set()
    for row in (req.rows or []):
        name = str(row.name or '').strip()
        phone = _smsganda_clean_phone(row.phone)
        if not name or not phone:
            continue
        key=(name, phone)
        if key in seen:
            continue
        seen.add(key)
        cleaned.append((
            name,
            phone,
            str(getattr(row, 'seg1', '') or ''),
            str(getattr(row, 'seg2', '') or ''),
            str(getattr(row, 'seg3', '') or ''),
            str(getattr(row, 'seg4', '') or '')
        ))
    if not cleaned:
        raise HTTPException(status_code=400, detail='엑셀로 만들 회원 이름/전화번호가 없습니다.')

    # 문자간다 제공 샘플 형식 그대로 맞춥니다.
    # 샘플 1행: 이름, 휴대전화, [*1*], [*2*], [*3*], [*4*]
    # 실제 데이터는 2행부터 A열=이름, B열=휴대전화, C~F열=[*1*]~[*4*] 치환값으로 저장합니다.
    wb = xlwt.Workbook(encoding='cp949')
    ws = wb.add_sheet('Sheet1')

    header_style = xlwt.XFStyle()
    header_style.num_format_str = '@'
    header_font = xlwt.Font()
    header_font.bold = True
    header_style.font = header_font

    text_style = xlwt.XFStyle()
    text_style.num_format_str = '@'
    text_alignment = xlwt.Alignment()
    text_alignment.wrap = 1
    text_alignment.vert = xlwt.Alignment.VERT_TOP
    text_style.alignment = text_alignment

    headers = ['이름', '휴대전화', '[*1*]', '[*2*]', '[*3*]', '[*4*]']
    for col, header in enumerate(headers):
        ws.write(0, col, header, header_style)

    for idx, (name, phone, seg1, seg2, seg3, seg4) in enumerate(cleaned, start=1):
        ws.write(idx, 0, name, text_style)
        ws.write(idx, 1, phone, text_style)
        ws.write(idx, 2, _smsganda_cell_text(seg1), text_style)
        ws.write(idx, 3, _smsganda_cell_text(seg2), text_style)
        ws.write(idx, 4, _smsganda_cell_text(seg3), text_style)
        ws.write(idx, 5, _smsganda_cell_text(seg4), text_style)
        max_lines = max(1, *[_smsganda_cell_text(v).count('\n') + 1 for v in (seg1, seg2, seg3, seg4)])
        ws.row(idx).height_mismatch = True
        ws.row(idx).height = min(9000, max(360, max_lines * 300))

    widths = [16, 18, 34, 44, 44, 28]
    for col, width in enumerate(widths):
        ws.col(col).width = width * 256

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    scope_label = {'all':'전체회원','representative':'대표관리자회원','general':'일반관리자회원','selected':'선택회원'}.get(str(req.scope or 'all'), '회원')
    round_part = f'{req.round_no}회차_' if req.round_no else ''
    filename = _safe_download_name(f'BBLOTTO_{round_part}문자간다_주소록_{scope_label}.xls')
    # Starlette/HTTP headers are latin-1 encoded. Korean text in the plain filename= part
    # causes a server 500 error, so keep filename= ASCII and put the real Korean filename
    # only in RFC 5987 filename*=UTF-8''.
    ascii_filename = 'bblotto_smsganda_address.xls'
    quoted = urllib.parse.quote(filename)
    return StreamingResponse(
        bio,
        media_type='application/vnd.ms-excel',
        headers={'Content-Disposition': f"attachment; filename={ascii_filename}; filename*=UTF-8''{quoted}"}
    )



# ===================== RC7-5 SMSGANDA TXT CP949 EXPORT =====================
@app.post('/api/export/smsganda_txt')
def export_smsganda_txt(req: SmsGandaXlsReq, authorization: str|None = Header(default=None)):
    require_admin(authorization)
    cleaned=[]
    seen=set()
    for row in (req.rows or []):
        name = str(row.name or '').strip()
        phone = _smsganda_clean_phone(row.phone)
        if not name or not phone:
            continue
        key=(name, phone)
        if key in seen:
            continue
        seen.add(key)
        # 문자간다 텍스트 업로드용: 이름,전화번호 한 줄씩
        cleaned.append(f'{name},{phone}')
    if not cleaned:
        raise HTTPException(status_code=400, detail='TXT로 만들 회원 이름/전화번호가 없습니다.')

    # 문자간다 구형 업로드 화면 호환을 위해 CP949(ANSI) + CRLF로 저장합니다.
    text = '\r\n'.join(cleaned) + '\r\n'
    data = text.encode('cp949', errors='replace')
    bio = io.BytesIO(data)
    scope_label = {'all':'전체회원','representative':'대표관리자회원','general':'일반관리자회원','selected':'선택회원'}.get(str(req.scope or 'all'), '회원')
    round_part = f'{req.round_no}회차_' if req.round_no else ''
    filename = _safe_download_name(f'BBLOTTO_{round_part}문자간다_주소록_{scope_label}.txt')
    ascii_filename = 'bblotto_smsganda_address.txt'
    quoted = urllib.parse.quote(filename)
    return StreamingResponse(
        bio,
        media_type='text/plain; charset=cp949',
        headers={'Content-Disposition': f"attachment; filename={ascii_filename}; filename*=UTF-8''{quoted}"}
    )

@app.get('/api/rc7-5/status')
def rc7_5_status(authorization: str|None = Header(default=None)):
    admin=require_admin(authorization)
    return {'ok': True, 'version': 'RC7-5 SMSGANDA TXT CP949', 'engine': DB_ENGINE, 'summary': '문자간다 TXT ANSI/CP949 주소록 생성 기본 지원', 'admin': admin.get('username')}

@app.get('/api/rc7-6/status')
def rc7_6_status(authorization: str|None = Header(default=None)):
    admin=require_admin(authorization)
    return {'ok': True, 'version': 'RC7-6 SMSGANDA TEMPLATE XLS', 'engine': DB_ENGINE, 'summary': '문자간다 샘플 XLS 헤더 형식(이름/휴대전화/[*1*]~[*4*]) 적용', 'admin': admin.get('username')}

@app.get('/api/rc7-8/status')
def rc7_8_status(authorization: str|None = Header(default=None)):
    admin=require_admin(authorization)
    return {'ok': True, 'version': 'RC7-8 SMSGANDA SEND CENTER', 'engine': DB_ENGINE, 'summary': '문자간다 [*1*]~[*4*] 문구 분리/수정/미리보기/XLS 연동 적용', 'admin': admin.get('username')}
# ===================== /RC7-5 SMSGANDA TXT CP949 EXPORT =====================

@app.get('/api/rc7-4/status')
def rc7_4_status(authorization: str|None = Header(default=None)):
    admin=require_admin(authorization)
    return {'ok': True, 'version': 'RC7-4 SMSGANDA HEADER FIX', 'engine': DB_ENGINE, 'summary': '문자간다 XLS 다운로드 한글 파일명 헤더 오류 수정', 'admin': admin.get('username')}

@app.get('/api/rc7-3/status')
def rc7_3_status(authorization: str|None = Header(default=None)):
    admin=require_admin(authorization)
    return {'ok': True, 'version': 'RC7-3 SMSGANDA REAL XLS', 'engine': DB_ENGINE, 'summary': '문자간다 샘플 기준 Excel 97-2003 BIFF .xls 주소록 생성', 'admin': admin.get('username')}
# ===================== /RC7-3 SMSGANDA REAL XLS EXPORT =====================


# ===================== RC7-2 SMSGANDA XLS EXPORT =====================
@app.get('/api/rc7-2/status')
def rc7_2_status(authorization: str|None = Header(default=None)):
    admin=require_admin(authorization)
    return {'ok': True, 'version': 'RC7-2 SMSGANDA XLS', 'engine': DB_ENGINE, 'summary': '문자간다 A열 이름/B열 전화번호 XLS 업로드 파일 생성 지원', 'admin': admin.get('username')}
# ===================== /RC7-2 SMSGANDA XLS EXPORT =====================

# ===================== BBLOTTO AI V4 ENGINE FULL REPLACEMENT =====================
# 기존 API/DB/화면 구조는 유지하고 실제 추천번호 생성 엔진만 최종 오버라이드합니다.
BBLOTTO_AI_V4_ENGINE_VERSION = 'BBLOTTO_AI_V4_ENGINE_FULL_REPLACEMENT_RC8_1'


def _v4_grade_params(grade='일반'):
    g = rc45_grade_label(grade)
    if g == '1등':
        return {'candidates': 18000, 'tries': 240000, 'score_shift': 5.8, 'max_num_use': 3, 'max_pair_use': 1, 'max_overlap': 3, 'history_overlap': 3}
    if g == '2등':
        return {'candidates': 13000, 'tries': 180000, 'score_shift': 3.2, 'max_num_use': 4, 'max_pair_use': 1, 'max_overlap': 4, 'history_overlap': 4}
    return {'candidates': 8500, 'tries': 120000, 'score_shift': 0.0, 'max_num_use': 5, 'max_pair_use': 2, 'max_overlap': 4, 'history_overlap': 5}


def _v4_weight_profile(st, ext, mode='balanced', member_grade='일반'):
    profile, _ = _ai_v1_profile(st, mode)
    f10 = st.get('freq10') or {}
    f30 = st.get('freq30') or {}
    f50 = st.get('freq50') or {}
    f100 = st.get('freq100') or {}
    f300 = ext.get('freq300') or {}
    last = ext.get('last_seen300') or {}
    hot = set((ext.get('hot300') or [])[:16])
    cold = set((ext.get('cold300') or [])[:16])
    overdue = set((ext.get('overdue300') or [])[:18])
    mid = set((ext.get('mid300') or [])[:20])
    recent_numbers = set(st.get('recent_numbers') or [])
    g = rc45_grade_label(member_grade)
    for n in range(1, 46):
        trend = f10.get(n, 0) * 2.8 + f30.get(n, 0) * 1.45 + f50.get(n, 0) * 0.75 + f100.get(n, 0) * 0.35 + f300.get(n, 0) * 0.12
        gap = min(28, int(last.get(n, 999) if last.get(n, 999) != 999 else 28)) * 0.32
        band = 1.25 if 11 <= n <= 35 else 0.75
        w = 1.0 + trend + gap + band + profile.get(n, 1.0) * 0.45
        if n in hot: w += 3.2 if g == '1등' else 2.1 if g == '2등' else 1.2
        if n in overdue: w += 2.4 if g == '1등' else 1.5 if g == '2등' else 0.9
        if n in cold: w += 0.8 if g == '1등' else 0.45
        if n in mid: w += 1.05 if g != '일반' else 0.55
        if n in recent_numbers: w *= 0.72
        if mode == 'aggressive' and n >= 31: w += 0.9
        if mode == 'conservative' and 10 <= n <= 35: w += 0.9
        profile[n] = max(0.2, round(w, 4))
    return profile


def _v4_gap_score(combo, ext):
    last = ext.get('last_seen300') or {}
    gaps = [min(30, int(last.get(n, 999) if last.get(n, 999) != 999 else 30)) for n in combo]
    avg_gap = sum(gaps) / 6
    # 너무 최근 번호만 몰리거나 너무 장기 미출현만 몰리는 것을 방지
    score = 0.0
    if 3.0 <= avg_gap <= 16.0: score += 5.5
    elif 1.5 <= avg_gap <= 23.0: score += 2.5
    if sum(1 for g in gaps if g >= 12) in (1, 2, 3): score += 3.0
    if sum(1 for g in gaps if g <= 2) >= 4: score -= 4.0
    return score


def _v4_combo_score(combo, st, mode, ext, profile, member_grade='일반'):
    combo = sorted(parse_nums(combo))
    if len(combo) != 6:
        return 0.0
    sig = _ai_v1_signature(combo)
    base = _rc42_adjusted_score(combo, st, mode, ext, profile)
    pairs100 = st.get('pair_counts') or collections.Counter()
    pairs300 = ext.get('pair300') or collections.Counter()
    triples300 = ext.get('triple300') or collections.Counter()
    hot = set((ext.get('hot300') or [])[:14])
    cold = set((ext.get('cold300') or [])[:14])
    overdue = set((ext.get('overdue300') or [])[:16])
    mid = set((ext.get('mid300') or [])[:20])
    s = float(base)
    # 패턴 안정성
    s += 4.6 if sig['odd'] in (2, 3, 4) else -7.5
    s += 5.2 if 105 <= sig['sum'] <= 178 else 1.5 if 92 <= sig['sum'] <= 190 else -8.0
    s += 5.0 if max(sig['zones']) <= 3 and min(sig['zones']) >= 1 else -8.0
    s += 3.5 if 6 <= sig['ac'] <= 10 else 1.0 if 5 <= sig['ac'] <= 11 else -5.0
    s += 2.8 if sig['end_dup'] <= 1 else 0.6 if sig['end_dup'] == 2 else -4.5
    s += 2.4 if sig['cons'] <= 1 else -4.5
    # 흐름 배합
    hits = {'hot': len(set(combo) & hot), 'cold': len(set(combo) & cold), 'overdue': len(set(combo) & overdue), 'mid': len(set(combo) & mid)}
    s += min(5.0, hits['hot'] * 1.5) + min(4.5, hits['overdue'] * 1.35) + min(3.4, hits['cold'] * 1.0) + min(3.2, hits['mid'] * 0.65)
    if hits['hot'] >= 5 or hits['overdue'] >= 5:
        s -= 4.0
    s += _v4_gap_score(combo, ext)
    # 동반출현/트리플은 과다하지 않게 가산
    pair_score = 0.0
    strong_pairs = 0
    for p in itertools.combinations(combo, 2):
        key = tuple(sorted(p))
        pc = pairs100.get(key, 0) * 0.55 + pairs300.get(key, 0) * 0.45
        pair_score += pc
        if pc >= 3.5:
            strong_pairs += 1
    triple_score = sum(triples300.get(tuple(sorted(t)), 0) for t in itertools.combinations(combo, 3))
    s += min(7.0, pair_score / 8.0) + min(4.0, strong_pairs * 0.9) + min(3.5, triple_score / 4.0)
    if strong_pairs >= 8:
        s -= 3.5
    if triple_score >= 10:
        s -= 3.0
    s += _v4_grade_params(member_grade)['score_shift']
    s += ((sum(n * n for n in combo) + sum(combo) * 3) % 29 - 14) * 0.045
    return round(max(72.0, min(99.8, s)), 1)


def _v4_detail(combo, st, mode, ext, profile, member_grade='일반'):
    combo = sorted(parse_nums(combo))
    d = _rc42_detail(combo, st, mode, ext, profile)
    score = _v4_combo_score(combo, st, mode, ext, profile, member_grade)
    sig = _ai_v1_signature(combo)
    reasons, pair_hits = _ai_v1_reasons(combo, st, ext)
    extra = []
    if _v4_gap_score(combo, ext) >= 5: extra.append('출현간격 안정권')
    if len(pair_hits) >= 1: extra.append('동반출현 보정')
    if max(sig['zones']) <= 3 and min(sig['zones']) >= 1: extra.append('구간 분산 우수')
    tags = []
    for t in _rc729_grade_tags(member_grade) + reasons + extra + ['AI V4 후보대량선별']:
        if t and t not in tags:
            tags.append(t)
    d.update({
        'numbers': combo, 'score': score, 'ai_score': score, 'vip_score': score,
        'raw_score': score,
        'grade': rc45_grade_label(member_grade),
        'member_grade': rc45_grade_label(member_grade),
        'grade_strength': rc45_grade_strength_text(member_grade),
        'engine_version': BBLOTTO_AI_V4_ENGINE_VERSION,
        'engine_label': _rc729_engine_name(member_grade),
        'sum': sig['sum'], 'odd': sig['odd'], 'even': sig['even'], 'zones': sig['zones'], 'ac': sig['ac'],
        'end_digit_dup': sig['end_dup'], 'consecutive': sig['cons'],
        'tags': tags[:7], 'reasons': tags[:7], 'reason_text': ' · '.join(tags[:5]),
        'pair_hits': pair_hits,
        'quality_rule': '최근 10/30/50/100/300회, 출현간격, 동반출현, 트리플, 합계, 홀짝, 구간, 끝수, AC, 연속수, 이전 추천 유사도를 종합 검토',
    })
    return d


def make_premium_combos(count=10, fixed='', excluded='', mode='balanced', member_grade='일반', member_id=None):
    """BBLOTTO AI V4: 추천번호 생성 엔진 전면 교체 버전.
    기존 라우트와 DB 저장 방식은 유지하면서 후보 생성/점수/포트폴리오 선별만 새로 수행합니다.
    """
    member_grade = rc45_grade_label(member_grade)
    params = _v4_grade_params(member_grade)
    st = latest_stats(300)
    ext = _ai_v1_window_stats(st, 300)
    fixed_set = set(parse_nums(fixed)); excluded_set = set(parse_nums(excluded))
    fixed_set = {n for n in fixed_set if n not in excluded_set}
    if len(fixed_set) > 6:
        fixed_set = set(sorted(fixed_set)[:6])
    pool = [n for n in range(1, 46) if n not in excluded_set and n not in fixed_set]
    target = max(1, min(50, int(count or 10)))
    if len(pool) + len(fixed_set) < 6:
        raise HTTPException(400, '고정수/제외수를 확인하세요. 선택 가능한 번호가 부족합니다.')

    profile = _v4_weight_profile(st, ext, mode, member_grade)
    past_draws = {tuple(sorted(d.get('numbers') or [])) for d in st.get('draws', []) if len(d.get('numbers') or []) == 6}
    recent_history = _rc55_recent_recommendation_combos(member_id=member_id, limit=300)
    recent_usage = collections.Counter(n for combo in recent_history[:100] for n in combo)

    buckets = {
        'hot': [n for n in (ext.get('hot300') or [])[:25] if n in pool],
        'cold': [n for n in (ext.get('cold300') or [])[:25] if n in pool],
        'overdue': [n for n in (ext.get('overdue300') or [])[:25] if n in pool],
        'mid': [n for n in (ext.get('mid300') or [])[:30] if n in pool],
        'low': [n for n in range(1, 16) if n in pool],
        'middle': [n for n in range(16, 31) if n in pool],
        'high': [n for n in range(31, 46) if n in pool],
        'all': pool[:],
    }
    plans = [
        ['hot','overdue','mid','low','middle','high'],
        ['hot','cold','overdue','mid','all','all'],
        ['mid','hot','cold','low','middle','high'],
        ['overdue','hot','mid','all','all','all'],
        ['low','middle','high','hot','overdue','mid'],
        ['cold','mid','hot','overdue','all','all'],
    ]
    if member_grade == '1등':
        plans += [['hot','hot','overdue','mid','cold','all'], ['overdue','hot','mid','low','middle','high']]
    elif member_grade == '2등':
        plans += [['hot','overdue','mid','cold','all','all'], ['mid','hot','overdue','low','high','all']]
    if mode == 'aggressive':
        plans += [['hot','hot','high','overdue','all','all'], ['hot','mid','high','cold','all','all']]
    elif mode == 'conservative':
        plans += [['mid','mid','cold','overdue','low','middle'], ['cold','mid','middle','all','all','all']]

    candidates, seen = [], set()
    tries = 0
    while len(candidates) < params['candidates'] and tries < params['tries']:
        tries += 1
        nums = set(fixed_set)
        plan = random.choice(plans)[:]
        random.shuffle(plan)
        for name in plan:
            if len(nums) >= 6:
                break
            usable = [n for n in buckets.get(name, pool) if n not in nums]
            if usable:
                nums.update(_weighted_pick(usable, [profile.get(n, 1) for n in usable], 1))
        while len(nums) < 6:
            usable = [n for n in pool if n not in nums]
            if not usable:
                break
            nums.update(_weighted_pick(usable, [profile.get(n, 1) for n in usable], 1))
        arr = tuple(sorted(nums))
        if len(arr) != 6 or arr in seen or arr in past_draws:
            continue
        if _rc55_too_close_to_history(arr, recent_history, max_overlap=params['history_overlap']):
            continue
        ok, _why = _rc42_combo_ok(arr, past=past_draws, fixed_set=fixed_set)
        if not ok:
            continue
        sig = _ai_v1_signature(arr)
        if sig['odd'] not in (2, 3, 4): continue
        if not (92 <= sig['sum'] <= 190): continue
        if max(sig['zones']) > 3 or min(sig['zones']) == 0: continue
        if sig['cons'] > 1: continue
        if sig['end_dup'] > 2: continue
        if not (5 <= sig['ac'] <= 11): continue
        seen.add(arr)
        score = _v4_combo_score(arr, st, mode, ext, profile, member_grade)
        # 최근 추천에 많이 쓰인 번호는 살짝 감점해 매번 같은 번호 반복을 줄임
        score -= min(2.8, sum(recent_usage.get(n, 0) for n in arr) * 0.04)
        candidates.append((round(max(72.0, min(99.8, score)), 1), list(arr)))

    candidates.sort(key=lambda x: (-x[0], x[1]))
    selected = []
    usage = collections.Counter()
    pair_usage = collections.Counter()
    pattern_usage = collections.Counter()

    def pattern_key(combo):
        try:
            return _rc42_pattern_key(combo)
        except Exception:
            sig = _ai_v1_signature(combo)
            return (sig['odd'], tuple(sig['zones']), sig['sum'] // 10, sig['ac'])

    def can_add(combo, max_num_use, max_pair_use, max_overlap, max_pattern_use=3):
        pairs = [tuple(sorted(p)) for p in itertools.combinations(combo, 2)]
        pkey = pattern_key(combo)
        if any(usage[n] >= max_num_use for n in combo if n not in fixed_set): return False
        if any(pair_usage[p] >= max_pair_use for p in pairs): return False
        if pattern_usage[pkey] >= max_pattern_use: return False
        if not all(len(set(combo) & set(prev)) <= max_overlap for prev in selected): return False
        return True

    def add_combo(combo):
        selected.append(combo)
        usage.update(combo)
        pair_usage.update([tuple(sorted(p)) for p in itertools.combinations(combo, 2)])
        pattern_usage.update([pattern_key(combo)])

    limit_sets = [
        (params['max_num_use'], params['max_pair_use'], params['max_overlap'], 2),
        (params['max_num_use'] + 1, params['max_pair_use'] + 1, 4, 3),
        (6, 3, 4, 5),
    ]
    for limits in limit_sets:
        if len(selected) >= target: break
        for _score, combo in candidates:
            if combo in selected: continue
            if can_add(combo, *limits): add_combo(combo)
            if len(selected) >= target: break
    if len(selected) < target:
        for _score, combo in candidates:
            if combo not in selected:
                add_combo(combo)
            if len(selected) >= target:
                break
    if len(selected) < target:
        # 극단적인 고정/제외 조건에서도 화면이 멈추지 않도록 최종 안전 fallback
        fallback = make_combos(target - len(selected), ','.join(map(str, fixed_set)), ','.join(map(str, excluded_set)), mode)
        for combo in fallback:
            if combo not in selected:
                add_combo(combo)
            if len(selected) >= target:
                break

    details = [_v4_detail(c, st, mode, ext, profile, member_grade) for c in selected[:target]]
    details.sort(key=lambda x: -float(x.get('score') or 0))
    selected_sorted = [d['numbers'] for d in details]
    st.update(ext)
    st.update({
        'engine_version': BBLOTTO_AI_V4_ENGINE_VERSION,
        'member_grade': member_grade,
        'grade_strength': rc45_grade_strength_text(member_grade),
        'ai_v4_candidates': len(candidates),
        'ai_v4_attempts': tries,
        'ai_v1_candidates': len(candidates),
        'ai_v1_attempts': tries,
        'rc55_history_checked': len(recent_history),
        'rc42_portfolio': {
            'max_overlap': max((len(set(a) & set(b)) for i, a in enumerate(selected_sorted) for b in selected_sorted[i+1:]), default=0),
            'unique_patterns': len({pattern_key(c) for c in selected_sorted}),
            'number_usage_max': max(usage.values()) if usage else 0,
            'pair_usage_max': max(pair_usage.values()) if pair_usage else 0,
            'recent_history_checked': len(recent_history),
        }
    })
    return selected_sorted[:target], details[:target], st


def _engine_summary(details, st):
    scores = [float(d.get('score') or d.get('ai_score') or d.get('vip_score') or 0) for d in (details or []) if (d.get('score') or d.get('ai_score') or d.get('vip_score'))]
    grade = rc45_grade_label(st.get('member_grade') or (details[0].get('member_grade') if details else '일반'))
    return {
        'version': BBLOTTO_AI_V4_ENGINE_VERSION,
        'engine_version': BBLOTTO_AI_V4_ENGINE_VERSION,
        'phase': 'RC8-1',
        'member_grade': grade,
        'engine_label': _rc729_engine_name(grade),
        'grade_strength': rc45_grade_strength_text(grade),
        'avg_score': round(sum(scores) / len(scores), 1) if scores else 0,
        'max_score': round(max(scores), 1) if scores else 0,
        'min_score': round(min(scores), 1) if scores else 0,
        'candidate_count': int(st.get('ai_v4_candidates') or st.get('ai_v1_candidates') or 0),
        'selected_count': len(details or []),
        'latest_round': st.get('latest_round') or 0,
        'v4_report': {
            'summary': '추천번호 엔진을 AI V4로 전면 교체했습니다. 최근 10/30/50/100/300회 흐름, 출현간격, 동반출현, 트리플, AC값, 끝수, 구간, 연속수, 이전 추천 유사도를 종합 선별합니다.',
            'portfolio': st.get('rc42_portfolio') or {},
            'candidates': int(st.get('ai_v4_candidates') or 0),
            'attempts': int(st.get('ai_v4_attempts') or 0),
        },
        'rc55_report': {
            'grade_policy': '1등/2등/일반 3단계 전용 점수 구간은 유지하되, 후보 생성·점수화·포트폴리오 선별은 AI V4 기준으로 교체',
            'portfolio': st.get('rc42_portfolio') or {},
            'summary': '후보 대량 생성 후 번호/페어/패턴/최근 추천 중복을 줄여 최종 조합을 선별했습니다.'
        },
        'rc45_report': {
            'summary': '최근 흐름과 누적 통계를 함께 반영한 심층 추천 결과입니다.',
            'portfolio': st.get('rc42_portfolio') or {},
        }
    }
# ===================== /BBLOTTO AI V4 ENGINE FULL REPLACEMENT =====================

# RC8-1 점수 보정: 기존 고점수 함수 영향으로 모든 조합이 99점대에 몰리지 않도록 V4 자체 점수화로 재정의합니다.
def _v4_combo_score(combo, st, mode, ext, profile, member_grade='일반'):
    combo = sorted(parse_nums(combo))
    if len(combo) != 6:
        return 0.0
    sig = _ai_v1_signature(combo)
    pairs100 = st.get('pair_counts') or collections.Counter()
    pairs300 = ext.get('pair300') or collections.Counter()
    triples300 = ext.get('triple300') or collections.Counter()
    hot = set((ext.get('hot300') or [])[:14]); cold = set((ext.get('cold300') or [])[:14]); overdue = set((ext.get('overdue300') or [])[:16]); mid = set((ext.get('mid300') or [])[:20])
    s = 58.0
    s += {3: 9.0, 2: 7.0, 4: 7.0, 1: 1.0, 5: 1.0}.get(sig['odd'], -6.0)
    s += 9.0 if 105 <= sig['sum'] <= 178 else 4.0 if 92 <= sig['sum'] <= 190 else -8.0
    s += 8.0 if max(sig['zones']) <= 3 and min(sig['zones']) >= 1 else -9.0
    s += 5.5 if 6 <= sig['ac'] <= 10 else 2.0 if 5 <= sig['ac'] <= 11 else -5.0
    s += 4.0 if sig['end_dup'] <= 1 else 1.5 if sig['end_dup'] == 2 else -5.0
    s += 3.5 if sig['cons'] == 0 else 1.5 if sig['cons'] == 1 else -5.0
    wavg = sum(profile.get(n, 1) for n in combo) / 6.0
    s += min(9.0, wavg * 0.36)
    hits = {'hot': len(set(combo) & hot), 'cold': len(set(combo) & cold), 'overdue': len(set(combo) & overdue), 'mid': len(set(combo) & mid)}
    s += min(5.2, hits['hot'] * 1.45) + min(4.6, hits['overdue'] * 1.30) + min(3.2, hits['cold'] * 0.9) + min(3.0, hits['mid'] * 0.55)
    if hits['hot'] >= 5 or hits['overdue'] >= 5: s -= 4.0
    s += _v4_gap_score(combo, ext)
    pair_score = 0.0; strong_pairs = 0
    for p in itertools.combinations(combo, 2):
        key = tuple(sorted(p)); pc = pairs100.get(key, 0) * 0.55 + pairs300.get(key, 0) * 0.45
        pair_score += pc
        if pc >= 3.5: strong_pairs += 1
    triple_score = sum(triples300.get(tuple(sorted(t)), 0) for t in itertools.combinations(combo, 3))
    s += min(5.0, pair_score / 11.0) + min(3.0, strong_pairs * 0.65) + min(2.4, triple_score / 5.5)
    if strong_pairs >= 8: s -= 3.5
    if triple_score >= 10: s -= 3.0
    # 등급별 표시는 분리하되 일반/2등/1등 점수대가 완전히 같아 보이지 않도록 차등 보정
    g = rc45_grade_label(member_grade)
    if g == '1등': s += 4.2
    elif g == '2등': s += 2.0
    s += ((sum(n * n for n in combo) + sum(combo) * 3) % 29 - 14) * 0.055
    return round(max(72.0, min(99.2, s)), 1)

# RC8-1 점수 보정 2차: 점수 분포를 일반 88~94, 2등 92~96, 1등 95~99 근처로 자연스럽게 분산합니다.
def _v4_combo_score(combo, st, mode, ext, profile, member_grade='일반'):
    combo = sorted(parse_nums(combo))
    if len(combo) != 6:
        return 0.0
    sig = _ai_v1_signature(combo)
    pairs100 = st.get('pair_counts') or collections.Counter(); pairs300 = ext.get('pair300') or collections.Counter(); triples300 = ext.get('triple300') or collections.Counter()
    hot = set((ext.get('hot300') or [])[:14]); cold = set((ext.get('cold300') or [])[:14]); overdue = set((ext.get('overdue300') or [])[:16]); mid = set((ext.get('mid300') or [])[:20])
    s = 41.0
    s += {3: 7.5, 2: 5.8, 4: 5.8, 1: 0.5, 5: 0.5}.get(sig['odd'], -5.5)
    s += 7.0 if 105 <= sig['sum'] <= 178 else 3.0 if 92 <= sig['sum'] <= 190 else -7.0
    s += 6.8 if max(sig['zones']) <= 3 and min(sig['zones']) >= 1 else -8.0
    s += 4.8 if 6 <= sig['ac'] <= 10 else 1.6 if 5 <= sig['ac'] <= 11 else -4.5
    s += 3.2 if sig['end_dup'] <= 1 else 1.0 if sig['end_dup'] == 2 else -4.0
    s += 2.8 if sig['cons'] == 0 else 1.0 if sig['cons'] == 1 else -4.5
    wavg = sum(profile.get(n, 1) for n in combo) / 6.0
    s += min(7.0, wavg * 0.13)
    hits = {'hot': len(set(combo) & hot), 'cold': len(set(combo) & cold), 'overdue': len(set(combo) & overdue), 'mid': len(set(combo) & mid)}
    s += min(4.0, hits['hot'] * 1.15) + min(3.6, hits['overdue'] * 1.05) + min(2.8, hits['cold'] * 0.75) + min(2.5, hits['mid'] * 0.45)
    if hits['hot'] >= 5 or hits['overdue'] >= 5: s -= 3.5
    s += min(5.0, _v4_gap_score(combo, ext) * 0.65)
    pair_score = 0.0; strong_pairs = 0
    for p in itertools.combinations(combo, 2):
        key = tuple(sorted(p)); pc = pairs100.get(key, 0) * 0.55 + pairs300.get(key, 0) * 0.45
        pair_score += pc
        if pc >= 3.5: strong_pairs += 1
    triple_score = sum(triples300.get(tuple(sorted(t)), 0) for t in itertools.combinations(combo, 3))
    s += min(4.0, pair_score / 18.0) + min(2.3, strong_pairs * 0.45) + min(1.8, triple_score / 8.0)
    if strong_pairs >= 8: s -= 3.0
    if triple_score >= 10: s -= 2.5
    g = rc45_grade_label(member_grade)
    if g == '1등': s += 5.0
    elif g == '2등': s += 3.7
    s += ((sum(n * n for n in combo) + sum(combo) * 3) % 29 - 14) * 0.065
    return round(max(72.0, min(99.2, s)), 1)

# =========================================================
# BBLOTTO AI V6 DB FULL HISTORY CACHE FINAL PATCH
# - 분석 결과를 database/bblotto_v34.db 의 ai_analysis_cache 테이블에 저장
# - 1회차~1231회차 전체 보유 여부를 API로 확인 가능
# - 추천번호 생성은 최종적으로 ai_engine_v6만 사용
# =========================================================
try:
    from .ai_engine_v7 import make_premium_combos as make_premium_combos
    from .ai_engine_v7 import latest_stats as latest_stats
    from .ai_engine_v7 import get_analysis_cache as _ai_v6_get_analysis_cache
    BBLOTTO_AI_V6_ENGINE_VERSION = 'BBLOTTO_RC10_AUTO_FULL_HISTORY'

    @app.get('/api/ai-engine/v6-cache')
    def ai_engine_v6_cache(authorization: str|None = Header(default=None), force: int = 0, target_round: int|None = None):
        require_admin(authorization)
        cache = _ai_v6_get_analysis_cache(bool(force), target_round=target_round)
        return {
            'ok': True,
            'engine_version': cache.get('engine_version'),
            'cache_storage': cache.get('cache_storage'),
            'analysis_confirm': cache.get('analysis_confirm'),
            'draw_count': cache.get('draw_count'),
            'actual_count': cache.get('actual_count'),
            'expected_count': cache.get('expected_count'),
            'round_range': cache.get('round_range'),
            'latest_round': cache.get('latest_round'),
            'target_round': cache.get('target_round'),
            'is_full_history': cache.get('is_full_history'),
            'missing_rounds_count': cache.get('missing_rounds_count'),
            'missing_rounds_sample': cache.get('missing_rounds_sample'),
            'hot': cache.get('hot', [])[:12],
            'cold': cache.get('cold', [])[:12],
            'overdue': cache.get('overdue', [])[:12],
            'cache_used': True,
        }
except Exception as _v6_import_error:
    BBLOTTO_AI_V6_ENGINE_VERSION = 'BBLOTTO_AI_V6_IMPORT_FAILED'
    print('[BBLOTTO] AI V6 engine import failed:', _v6_import_error)

try:
    from .ai_engine_v7 import sync_official_full_history as _ai_v6_sync_official_full_history

    @app.post('/api/ai-engine/v6-sync-full')
    def ai_engine_v6_sync_full(authorization: str|None = Header(default=None), max_round: int|None = None):
        require_admin(authorization)
        return _ai_v6_sync_official_full_history(max_round=max_round)
except Exception as _v6_sync_import_error:
    print('[BBLOTTO] AI V6 sync endpoint failed:', _v6_sync_import_error)


# =========================================================
# BBLOTTO AI V6 관리자 화면용 전체 동기화/분석 API
# - 관리자 버튼에서 사용: POST /api/admin/ai-v6/full-sync
# - 주소 직접 입력 시 Not Found 대신 안내/실행 가능 여부 반환
# =========================================================
try:
    from .ai_engine_v7 import sync_official_full_history as _bb_v6_sync_full_ui
    from .ai_engine_v7 import get_analysis_cache as _bb_v6_cache_ui

    @app.post('/api/admin/ai-v6/full-sync')
    def admin_ai_v6_full_sync(authorization: str|None = Header(default=None), max_round: int|None = None):
        require_admin(authorization)
        sync_result = _bb_v6_sync_full_ui(max_round=max_round)
        cache = _bb_v6_cache_ui(True, target_round=max_round)
        return {
            'ok': True,
            'message': sync_result.get('message') or (f'1회차~{cache.get("target_round", max_round)}회차 전체 동기화/분석 저장 완료' if cache.get('is_full_history') else f'전체 분석 미완료: {cache.get("missing_rounds_count", 0)}개 누락'),
            'completed': bool(cache.get('is_full_history')),
            'sync_result': sync_result,
            'cache': {
                'engine_version': cache.get('engine_version'),
                'cache_storage': cache.get('cache_storage'),
                'analysis_confirm': cache.get('analysis_confirm'),
                'actual_count': cache.get('actual_count'),
                'expected_count': cache.get('expected_count'),
                'round_range': cache.get('round_range'),
                'latest_round': cache.get('latest_round'),
                'target_round': cache.get('target_round'),
                'is_full_history': cache.get('is_full_history'),
                'missing_rounds_count': cache.get('missing_rounds_count'),
                'missing_rounds_sample': cache.get('missing_rounds_sample'),
            }
        }


    from .ai_engine_v7 import sync_official_history_step as _bb_v6_sync_step_ui

    @app.post('/api/admin/ai-v6/full-sync-step')
    def admin_ai_v6_full_sync_step(authorization: str|None = Header(default=None), max_round: int|None = None, chunk_size: int = 25):
        require_admin(authorization)
        try:
            return _bb_v6_sync_step_ui(max_round=max_round, chunk_size=chunk_size)
        except Exception as exc:
            # 브라우저에는 원인을 알 수 없는 500 오류 대신 재시도 가능한 안내를 반환합니다.
            print('[BBLOTTO] AI V7 step sync failed:', repr(exc))
            return {
                'ok': False,
                'completed': False,
                'message': '회차 동기화 처리 중 오류가 발생했습니다. 잠시 후 다시 실행해주세요.',
                'error_type': type(exc).__name__,
                'retryable': True,
            }

    @app.get('/api/admin/ai-v6/cache-status')
    def admin_ai_v6_cache_status(authorization: str|None = Header(default=None), target_round: int|None = None):
        require_admin(authorization)
        cache = _bb_v6_cache_ui(False, target_round=target_round)
        return {
            'ok': True,
            'engine_version': cache.get('engine_version'),
            'cache_storage': cache.get('cache_storage'),
            'analysis_confirm': cache.get('analysis_confirm'),
            'actual_count': cache.get('actual_count'),
            'expected_count': cache.get('expected_count'),
            'round_range': cache.get('round_range'),
            'latest_round': cache.get('latest_round'),
            'target_round': cache.get('target_round'),
            'is_full_history': cache.get('is_full_history'),
            'missing_rounds_count': cache.get('missing_rounds_count'),
            'missing_rounds_sample': cache.get('missing_rounds_sample'),
        }

    @app.get('/admin/sync-full-history')
    def admin_sync_full_history_url_notice(authorization: str|None = Header(default=None), max_round: int|None = None):
        # 브라우저 주소창 직접 입력 시 기존처럼 Not Found가 나오지 않도록 안내한다.
        # 실제 실행은 로그인 후 관리자 화면 버튼 또는 Authorization 헤더가 있는 요청에서만 가능하다.
        if not authorization:
            return {
                'ok': False,
                'message': '주소창 직접 입력은 로그인 토큰이 없어 실행하지 않습니다. 관리자 화면 > AI 엔진 > 전체 회차 동기화 버튼을 눌러주세요.',
                'api': '/api/admin/ai-v6/full-sync',
                'target_round': max_round
            }
        require_admin(authorization)
        sync_result = _bb_v6_sync_full_ui(max_round=max_round)
        cache = _bb_v6_cache_ui(True, target_round=max_round)
        return {'ok': bool(cache.get('is_full_history')), 'message': sync_result.get('message') or ('전체 저장/분석 완료' if cache.get('is_full_history') else '전체 분석 미완료'), 'sync_result': sync_result, 'cache': cache}
except Exception as _bb_v6_ui_sync_error:
    print('[BBLOTTO] AI V6 admin UI sync endpoint failed:', _bb_v6_ui_sync_error)


# ===================== RC10.2 DYNAMIC MEMBER-FRIENDLY ANALYSIS =====================
def build_analysis_text(round_no, st, mode, fixed, excluded, details=None):
    """추천번호의 실제 특징을 바탕으로, 생성할 때마다 표현이 달라지는 쉬운 분석문을 만든다."""
    import random
    import secrets

    details = details or []
    combos = []
    for item in details:
        nums = item.get('numbers') or item.get('nums') or item.get('combo') or []
        try:
            nums = sorted({int(n) for n in nums if 1 <= int(n) <= 45})
        except Exception:
            nums = []
        if len(nums) == 6:
            combos.append(nums)

    flat = [n for combo in combos for n in combo]
    rng = random.Random(secrets.randbits(64))
    latest_round = int(st.get('latest_round') or st.get('target_round') or max(0, int(round_no or 1) - 1))
    hot = [int(x) for x in (st.get('hot300') or st.get('hot100') or st.get('hot') or [])[:10] if str(x).isdigit()]
    overdue = [int(x) for x in (st.get('overdue300') or st.get('overdue100') or st.get('overdue') or [])[:10] if str(x).isdigit()]

    if not flat:
        return '\n'.join(rng.sample([
            '최근 흐름과 전체 기록을 함께 살펴 번호가 한쪽에 몰리지 않도록 구성했습니다.',
            '자주 나온 번호와 한동안 쉬었던 번호를 섞어 이번 회차의 변화를 함께 살폈습니다.',
            '비슷한 모양의 조합이 반복되지 않도록 번호 간 간격과 구간을 고르게 맞췄습니다.',
            '이번 추천은 단순히 많이 나온 번호만 고르지 않고 전체적인 균형을 우선했습니다.',
        ], 3))

    low = sum(1 for n in flat if n <= 15)
    mid = sum(1 for n in flat if 16 <= n <= 30)
    high = len(flat) - low - mid
    odd = sum(1 for n in flat if n % 2)
    even = len(flat) - odd
    zone_counts = [('낮은 번호대', low), ('중간 번호대', mid), ('높은 번호대', high)]
    strongest_zone = max(zone_counts, key=lambda x: x[1])[0]
    most_common = []
    freq = {}
    for n in flat:
        freq[n] = freq.get(n, 0) + 1
    most_common = [n for n, _ in sorted(freq.items(), key=lambda x: (-x[1], x[0]))[:5]]
    hot_used = [n for n in hot if n in freq][:4]
    overdue_used = [n for n in overdue if n in freq][:4]
    consecutive_count = sum(1 for c in combos for a, b in zip(c, c[1:]) if b - a == 1)
    same_end_count = 0
    for c in combos:
        ends = [n % 10 for n in c]
        same_end_count += len(ends) - len(set(ends))

    mode_label = {'balanced': '균형을 우선한', 'conservative': '안정적인 흐름을 우선한', 'aggressive': '변화를 조금 더 반영한'}.get(mode, '균형을 우선한')
    opening_pool = [
        f'{latest_round}회차까지의 기록과 최근 흐름을 함께 살펴 이번 추천을 구성했습니다.',
        '이번 회차는 오래된 기록과 최근 움직임을 같이 비교해 번호를 골랐습니다.',
        '최근 결과만 보지 않고 전체 흐름까지 함께 확인해 추천번호를 정리했습니다.',
        f'1회차부터 {latest_round}회차까지의 흐름을 바탕으로 이번 조합을 새로 구성했습니다.',
        '이번 추천은 자주 나온 번호와 잠시 쉬었던 번호를 함께 비교해 만들었습니다.',
    ]
    zone_pool = [
        f'{strongest_zone}의 흐름이 비교적 눈에 띄어 다른 번호대와 함께 고르게 섞었습니다.',
        '낮은 번호, 중간 번호, 높은 번호가 한쪽에 몰리지 않도록 나누어 배치했습니다.',
        '특정 번호대만 반복되지 않도록 여러 구간을 섞어 조합의 폭을 넓혔습니다.',
        f'{mode_label} 방식으로 번호대를 나누어 전체적인 모양을 맞췄습니다.',
    ]
    flow_pool = []
    if hot_used:
        flow_pool += [
            f'최근 자주 보인 {", ".join(map(str, hot_used))}번을 중심 후보로 두되 다른 번호와 함께 분산했습니다.',
            f'{", ".join(map(str, hot_used))}번은 최근 흐름이 이어져 일부 조합에 나누어 반영했습니다.',
        ]
    if overdue_used:
        flow_pool += [
            f'한동안 쉬었던 {", ".join(map(str, overdue_used))}번도 일부 포함해 변화 가능성을 살폈습니다.',
            f'{", ".join(map(str, overdue_used))}번은 오랜 공백 뒤 다시 나올 가능성을 고려해 보강 후보로 넣었습니다.',
        ]
    if not flow_pool:
        flow_pool = [
            '최근에 자주 나온 번호와 한동안 쉬었던 번호를 함께 섞어 한쪽 흐름에 치우치지 않게 했습니다.',
            f'반복해서 사용된 중심 번호는 {", ".join(map(str, most_common))}번이며, 나머지 번호는 조합마다 다르게 배치했습니다.',
            '단순 출현 횟수보다 최근 움직임과 번호 간 조화를 함께 살폈습니다.',
        ]
    shape_pool = [
        '홀수와 짝수가 지나치게 한쪽으로 쏠리지 않도록 조합마다 균형을 맞췄습니다.',
        '조합끼리 같은 번호가 너무 많이 겹치지 않도록 서로 다른 구성을 우선했습니다.',
        '번호 간 간격이 너무 좁거나 넓은 조합은 줄이고 자연스러운 배열을 선택했습니다.',
        '끝자리가 같은 번호가 지나치게 반복되지 않도록 조합별로 나누어 배치했습니다.',
    ]
    if consecutive_count:
        shape_pool.append('연속번호는 일부 조합에만 넣어 최근 흐름을 반영하면서도 반복을 줄였습니다.')
    if same_end_count <= max(2, len(combos)):
        shape_pool.append('끝자리 분포가 비교적 고르게 나오도록 비슷한 끝수의 반복을 줄였습니다.')
    close_pool = [
        '전체적으로 비슷한 조합의 반복을 줄이고 각 조합의 차이를 살린 추천입니다.',
        '이번 결과는 한 가지 흐름에만 기대지 않고 여러 가능성을 나누어 담았습니다.',
        '번호가 한곳에 몰리지 않도록 조정해 보기 편하고 이해하기 쉬운 구성으로 정리했습니다.',
        '최근 흐름을 반영하되 과도한 편중은 줄이는 방향으로 최종 조합을 골랐습니다.',
    ]

    candidates = [rng.choice(opening_pool), rng.choice(zone_pool), rng.choice(flow_pool), rng.choice(shape_pool), rng.choice(close_pool)]
    result = []
    for line in candidates:
        if line not in result:
            result.append(line)
    # 화면이 지나치게 길어지지 않도록 4줄만 제공한다.
    return '\n'.join(result[:4])
# ===================== /RC10.2 DYNAMIC MEMBER-FRIENDLY ANALYSIS =====================


# ===== RC9 AI V7 integrity audit =====
@app.get('/api/ai-engine/rc9-audit')
def rc9_engine_audit(authorization: str|None = Header(default=None)):
    require_admin(authorization)
    from .ai_engine_v7 import rc9_audit
    return rc9_audit()


# ===================== RC11 EXPLAINABLE ANALYSIS OVERRIDE =====================
try:
    from .analysis_engine_rc11 import build_member_friendly_analysis as _rc11_build_analysis

    def build_analysis_text(round_no, st, mode, fixed, excluded, details=None):
        return _rc11_build_analysis(round_no, st, mode, fixed, excluded, details or [])
except Exception as _rc11_analysis_import_error:
    print('[BBLOTTO] RC11 analysis engine load failed:', repr(_rc11_analysis_import_error))
# ===================== /RC11 EXPLAINABLE ANALYSIS OVERRIDE =====================
