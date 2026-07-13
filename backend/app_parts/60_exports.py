# Extracted from legacy backend/app.py lines 3408-3921.
@router.get('/api/export/members_csv')
def export_members_csv(token: str|None=None, authorization: str|None = Header(default=None)):
    require_admin_any(authorization, token)
    with con() as c:
        rows=c.execute("SELECT m.id,m.name,m.phone,m.grade,m.status,COALESCE(m.priority,'보통') priority,COALESCE(m.source,'직접등록') source,m.last_contact_at,m.memo,m.created_at,COALESCE(m.contract_end_at,'') contract_end_at,COALESCE(m.contract_months,12) contract_months,COALESCE(a.name,a.username,'미지정') registered_by_name FROM members m LEFT JOIN admins a ON a.id=COALESCE(m.created_by,0) ORDER BY m.id DESC").fetchall()
    return csv_response('BBLOTTO_members.csv', ['ID','이름','연락처','등급','상태','우선순위','유입경로','최근연락','메모','등록일','계약만료일','등록관리자'], [[r['id'],r['name'],r['phone'],r['grade'],r['status'],r['priority'],r['source'],r['last_contact_at'],r['memo'],r['created_at'],r['contract_end_at'],r['registered_by_name']] for r in rows])

@router.get('/api/export/recommendations_csv')
def export_recommendations_csv(token: str|None=None, authorization: str|None = Header(default=None)):
    require_admin_any(authorization, token)
    with con() as c: rows=c.execute('SELECT id,member_name,round_no,mode,count,numbers,analysis,created_at FROM recommendations ORDER BY id DESC').fetchall()
    return csv_response('BBLOTTO_recommendations.csv', ['ID','회원','회차','모드','조합수','추천번호','분석','생성일'], [[r['id'],r['member_name'],r['round_no'],r['mode'],r['count'],r['numbers'],r['analysis'],r['created_at']] for r in rows])

@router.get('/api/export/winning_csv')
def export_winning_csv(token: str|None=None, authorization: str|None = Header(default=None)):
    require_admin_any(authorization, token)
    with con() as c: rows=c.execute('SELECT round_no,member_name,target_numbers,win_numbers,bonus,rank,prize,cost,profit,roi,created_at FROM winning_checks ORDER BY id DESC').fetchall()
    return csv_response('BBLOTTO_winning_checks.csv', ['회차','회원','추천번호','당첨번호','보너스','등수','당첨금','구매금','수익','수익률','확인일'], [[r['round_no'],r['member_name'],r['target_numbers'],r['win_numbers'],r['bonus'],r['rank'],r['prize'],r['cost'],r['profit'],r['roi'],r['created_at']] for r in rows])

@router.get('/api/backup_db')
def backup_db(token: str|None=None, authorization: str|None = Header(default=None)):
    admin = require_admin_any(authorization, token)
    if DB_ENGINE == 'postgresql':
        b = create_db_backup('download', admin)
        return FileResponse(EXPORT_DIR / b['filename'], media_type='application/json', filename=b['filename'])
    return FileResponse(DB, media_type='application/octet-stream', filename='BBLOTTO_lotto_backup.db')

@router.get('/api/export/excel')
def export_excel(token: str|None=None, authorization: str|None = Header(default=None)):
    require_admin_any(authorization, token)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except Exception:
        raise HTTPException(500,'openpyxl 설치가 필요합니다. pip install -r requirements.txt 를 실행하세요.')

    auth_header = authorization or ('Bearer '+token if token else None)
    data=dashboard(auth_header)
    wb=Workbook()

    header_fill = PatternFill('solid', fgColor='111111')
    gold_fill = PatternFill('solid', fgColor='D4AF37')
    gold_font = Font(bold=True, color='111111')
    white_font = Font(bold=True, color='FFFFFF')
    thin = Side(style='thin', color='D4AF37')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_sheet(ws):
        ws.freeze_panes = 'A2'
        for cell in ws[1]:
            cell.font = white_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = Border(bottom=Side(style='hair', color='555555'))
                cell.alignment = Alignment(vertical='top', wrap_text=True)
        for col in ws.columns:
            width = min(55, max(12, max(len(str(c.value or '')) for c in col) + 2))
            ws.column_dimensions[get_column_letter(col[0].column)].width = width

    ws=wb.active; ws.title='대시보드'
    ws.append(['항목','값'])
    dashboard_rows=[
        ('총 회원', data.get('members',0)),('관리자',data.get('admins',0)),('추천 생성',data.get('recommendations',0)),('문자 이력',data.get('sms',0)),('당첨 확인',data.get('checks',0)),('총 당첨금',data.get('total_prize',0)),('총 구매금',data.get('total_cost',0)),('총 수익',data.get('total_profit',0)),('수익률',str(data.get('roi',0))+'%'),('최근 회차',data.get('latest_round','-')),('생성일',now())]
    for r in dashboard_rows: ws.append(list(r))
    ws['A1'].fill=gold_fill; ws['B1'].fill=gold_fill; ws['A1'].font=gold_font; ws['B1'].font=gold_font
    style_sheet(ws)

    ws2=wb.create_sheet('회원목록'); ws2.append(['ID','이름','연락처','등급','상태','우선순위','유입경로','최근연락','메모','등록일'])
    with con() as c:
        members=c.execute("SELECT id,name,phone,grade,status,COALESCE(priority,'보통') priority,COALESCE(source,'직접등록') source,last_contact_at,memo,created_at FROM members ORDER BY id DESC").fetchall()
        recs=c.execute('SELECT id,member_name,round_no,mode,count,numbers,analysis,created_at FROM recommendations ORDER BY id DESC LIMIT 500').fetchall()
        wins=c.execute('SELECT round_no,member_name,target_numbers,win_numbers,bonus,rank,prize,cost,profit,roi,created_at FROM winning_checks ORDER BY id DESC LIMIT 500').fetchall()
        sms=c.execute('SELECT member_name,round_no,body,created_at FROM sms_logs ORDER BY id DESC LIMIT 500').fetchall()
        draws=c.execute('SELECT round_no,draw_date,numbers,bonus,source,updated_at FROM draws ORDER BY round_no DESC LIMIT 200').fetchall()
    for m in members: ws2.append([m['id'],m['name'],m['phone'],m['grade'],m['status'],m['priority'],m['source'],m['last_contact_at'],m['memo'],m['created_at']])
    style_sheet(ws2)

    ws3=wb.create_sheet('추천번호'); ws3.append(['ID','회원','회차','모드','조합수','추천번호','분석','생성일'])
    for r in recs: ws3.append([r['id'],r['member_name'],r['round_no'],r['mode'],r['count'],r['numbers'],r['analysis'],r['created_at']])
    style_sheet(ws3)

    ws4=wb.create_sheet('당첨확인_수익률'); ws4.append(['회차','회원','번호','당첨번호','보너스','등수','당첨금','구매금','수익','수익률%','확인일'])
    for w in wins: ws4.append([w['round_no'],w['member_name'],w['target_numbers'],w['win_numbers'],w['bonus'],w['rank'],w['prize'],w['cost'],w['profit'],w['roi'],w['created_at']])
    style_sheet(ws4)

    ws5=wb.create_sheet('문자이력'); ws5.append(['회원','회차','문자내용','저장일'])
    for x in sms: ws5.append([x['member_name'],x['round_no'],x['body'],x['created_at']])
    style_sheet(ws5)

    ws6=wb.create_sheet('당첨번호DB'); ws6.append(['회차','추첨일','당첨번호','보너스','출처','수정일'])
    for d in draws: ws6.append([d['round_no'],d['draw_date'],d['numbers'],d['bonus'],d['source'],d['updated_at']])
    style_sheet(ws6)

    bio=io.BytesIO(); wb.save(bio); bio.seek(0)
    return StreamingResponse(bio, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={'Content-Disposition':'attachment; filename=BBLOTTO_PRO_V34_FINAL_REPORT.xlsx'})


@router.get('/api/export/report_txt')
def export_report_txt(token: str|None=None, authorization: str|None = Header(default=None)):
    require_admin_any(authorization, token)
    auth_header = authorization or ('Bearer '+token if token else None)
    data=dashboard(auth_header)
    lines=[
        'BBLOTTO PRO V34 FINAL REPORT',
        f'생성일: {now()}',
        f'총 회원: {data.get("members",0)}',
        f'추천 생성: {data.get("recommendations",0)}',
        f'문자 이력: {data.get("sms",0)}',
        f'당첨 확인: {data.get("checks",0)}',
        f'총 당첨금: {data.get("total_prize",0)}',
        f'총 구매금: {data.get("total_cost",0)}',
        f'총 수익: {data.get("total_profit",0)}',
        f'수익률: {data.get("roi",0)}%',
    ]
    raw='\n'.join(lines).encode('utf-8-sig')
    return StreamingResponse(io.BytesIO(raw), media_type='text/plain; charset=utf-8', headers={'Content-Disposition':'attachment; filename=BBLOTTO_PRO_V34_FINAL_REPORT.txt'})

@router.get('/api/export/final_bundle')
def export_final_bundle(token: str|None=None, authorization: str|None = Header(default=None)):
    require_admin_any(authorization, token)
    import zipfile
    bio=io.BytesIO()
    auth_header = authorization or ('Bearer '+token if token else None)
    data=dashboard(auth_header)
    with zipfile.ZipFile(bio,'w',zipfile.ZIP_DEFLATED) as z:
        z.writestr('README_FINAL_PHASE5.txt','BBLOTTO PRO V34 FINAL PHASE5\n마감/최적화/백업 번들입니다.\n')
        z.writestr('dashboard_summary.json',json.dumps(data,ensure_ascii=False,indent=2))
        with con() as c:
            members=c.execute("SELECT id,name,phone,grade,status,COALESCE(priority,'보통') priority,COALESCE(source,'직접등록') source,last_contact_at,memo,created_at FROM members ORDER BY id DESC").fetchall()
            recs=c.execute('SELECT id,member_name,round_no,mode,count,numbers,analysis,created_at FROM recommendations ORDER BY id DESC').fetchall()
            wins=c.execute('SELECT * FROM winning_checks ORDER BY id DESC').fetchall()
        z.writestr('members.json',json.dumps([dict(x) for x in members],ensure_ascii=False,indent=2))
        z.writestr('recommendations.json',json.dumps([dict(x) for x in recs],ensure_ascii=False,indent=2))
        z.writestr('winning_checks.json',json.dumps([dict(x) for x in wins],ensure_ascii=False,indent=2))
        if DB.exists(): z.write(DB, 'database/bblotto_v34.db')
    bio.seek(0)
    return StreamingResponse(bio, media_type='application/zip', headers={'Content-Disposition':'attachment; filename=BBLOTTO_PRO_V34_FINAL_EXPORT_BUNDLE.zip'})

@router.get('/api/export/pdf')
def export_pdf(token: str|None=None, authorization: str|None = Header(default=None)):
    require_admin_any(authorization, token)
    # 외부 라이브러리 없이 간단한 PDF 리포트 생성(영문/숫자 중심, 한글 상세는 엑셀 사용 권장)
    auth_header = authorization or ('Bearer '+token if token else None)
    data=dashboard(auth_header)
    lines=['BBLOTTO PRO V34 REPORT', f'Generated: {now()}', '', f'Members: {data["members"]}', f'Admins: {data["admins"]}', f'Recommendations: {data["recommendations"]}', f'Win Checks: {data["checks"]}', f'Total Prize: {data["total_prize"]}', f'Total Cost: {data["total_cost"]}', f'Total Profit: {data["total_profit"]}', f'ROI: {data["roi"]}%']
    content='BT /F1 14 Tf 50 780 Td '
    escaped=[]
    for i,line in enumerate(lines):
        safe=line.replace('\\','\\\\').replace('(','\\(').replace(')','\\)')
        escaped.append(f'({safe}) Tj 0 -24 Td')
    stream=(content+' '.join(escaped)+' ET').encode('latin-1','ignore')
    objs=[]
    objs.append(b'1 0 obj << /Type /Catalog /Pages 2 0 R >> endobj')
    objs.append(b'2 0 obj << /Type /Pages /Kids [3 0 R] /Count 1 >> endobj')
    objs.append(b'3 0 obj << /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] /Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >> endobj')
    objs.append(b'4 0 obj << /Type /Font /Subtype /Type1 /BaseFont /Helvetica >> endobj')
    objs.append(b'5 0 obj << /Length '+str(len(stream)).encode()+b' >> stream\n'+stream+b'\nendstream endobj')
    pdf=b'%PDF-1.4\n'; offsets=[]
    for o in objs: offsets.append(len(pdf)); pdf+=o+b'\n'
    xref=len(pdf); pdf+=f'xref\n0 {len(objs)+1}\n0000000000 65535 f \n'.encode()
    for off in offsets: pdf+=f'{off:010d} 00000 n \n'.encode()
    pdf+=f'trailer << /Root 1 0 R /Size {len(objs)+1} >>\nstartxref\n{xref}\n%%EOF'.encode()
    return StreamingResponse(io.BytesIO(pdf), media_type='application/pdf', headers={'Content-Disposition':'attachment; filename=BBLOTTO_PRO_V34_REPORT.pdf'})


class TextPdfReq(BaseModel):
    text:str=''

@router.post('/api/export_pdf')
def export_text_pdf(req:TextPdfReq, authorization: str|None = Header(default=None), x_token: str|None = Header(default=None)):
    require_admin_any(authorization, x_token)
    # 간단 PDF: 한글은 환경에 따라 제한될 수 있어 텍스트 파일에 가까운 PDF 구조로 저장합니다.
    lines=(req.text or 'BBLOTTO PRO V34').splitlines()[:80]
    safe_lines=[]
    for line in lines:
        # PDF 기본 폰트는 한글 표시가 제한되므로 영문/숫자 중심으로 안전 처리
        safe=line.encode('latin-1','ignore').decode('latin-1') or 'BBLOTTO'
        safe=safe.replace('\\','\\\\').replace('(','\\(').replace(')','\\)')
        safe_lines.append(safe)
    content='BT /F1 12 Tf 45 800 Td ' + ' '.join([f'({line}) Tj 0 -18 Td' for line in safe_lines]) + ' ET'
    stream=content.encode('latin-1','ignore')
    objs=[b'1 0 obj << /Type /Catalog /Pages 2 0 R >> endobj', b'2 0 obj << /Type /Pages /Kids [3 0 R] /Count 1 >> endobj', b'3 0 obj << /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] /Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >> endobj', b'4 0 obj << /Type /Font /Subtype /Type1 /BaseFont /Helvetica >> endobj', b'5 0 obj << /Length '+str(len(stream)).encode()+b' >> stream\n'+stream+b'\nendstream endobj']
    pdf=b'%PDF-1.4\n'; offsets=[]
    for o in objs:
        offsets.append(len(pdf)); pdf+=o+b'\n'
    xref=len(pdf); pdf+=f'xref\n0 {len(objs)+1}\n0000000000 65535 f \n'.encode()
    for off in offsets: pdf+=f'{off:010d} 00000 n \n'.encode()
    pdf+=f'trailer << /Root 1 0 R /Size {len(objs)+1} >>\nstartxref\n{xref}\n%%EOF'.encode()
    return StreamingResponse(io.BytesIO(pdf), media_type='application/pdf', headers={'Content-Disposition':'attachment; filename=BBLOTTO_RECOMMENDATION.pdf'})

# -----------------------------------------------------------------------------
# V34 FINAL - AI ENGINE PHASE2 OVERRIDES
# 기존 API/DB를 건드리지 않고 추천엔진 함수만 후단에서 재정의합니다.
# 반영: 최근 10/30/50/100회 가중치, HOT/COLD, 미출현, 동반출현, AC값,
#      합계/홀짝/구간/끝수/연속수 필터, 엔진 요약 리턴.
# -----------------------------------------------------------------------------

def _draw_rows_for_ai(limit=120):
    with con() as c:
        rows = c.execute('SELECT * FROM draws ORDER BY round_no DESC LIMIT ?', (int(limit),)).fetchall()
    draws=[]
    for r in rows:
        nums=parse_nums(r['numbers'])
        if len(nums)==6:
            draws.append({'round_no':r['round_no'], 'draw_date':r['draw_date'], 'numbers':nums, 'bonus':r['bonus']})
    return draws

def _ai_ac_value(nums):
    nums=sorted(nums)
    diffs=set()
    for i,a in enumerate(nums):
        for b in nums[i+1:]:
            diffs.add(abs(b-a))
    return max(0, len(diffs)-5)

def _ai_consecutive(nums):
    nums=sorted(nums)
    return sum(1 for a,b in zip(nums, nums[1:]) if b-a==1)

def _ai_sections(nums):
    return [sum(1<=n<=15 for n in nums), sum(16<=n<=30 for n in nums), sum(31<=n<=45 for n in nums)]

def _ai_pair_key(a,b):
    return tuple(sorted((int(a), int(b))))

def latest_stats(limit=120):
    """AI Phase2 통계팩. 기존 함수명 유지로 충돌 방지."""
    draws=_draw_rows_for_ai(limit)
    freq_windows={10:{n:0 for n in range(1,46)}, 30:{n:0 for n in range(1,46)}, 50:{n:0 for n in range(1,46)}, 100:{n:0 for n in range(1,46)}}
    last_seen={n:999 for n in range(1,46)}
    pair_counter=collections.Counter()
    end_counts={i:0 for i in range(10)}
    zone_counts={'1~15':0,'16~30':0,'31~45':0}
    sums=[]; acs=[]; odd_total=0
    for idx,d in enumerate(draws):
        nums=d['numbers']
        if idx < 30:
            sums.append(sum(nums)); acs.append(_ai_ac_value(nums)); odd_total += sum(n%2 for n in nums)
            for n in nums:
                end_counts[n%10]+=1
                if n<=15: zone_counts['1~15']+=1
                elif n<=30: zone_counts['16~30']+=1
                else: zone_counts['31~45']+=1
            for a,b in itertools.combinations(nums,2):
                pair_counter[_ai_pair_key(a,b)] += 1
        for w in freq_windows:
            if idx < w:
                for n in nums:
                    freq_windows[w][n]+=1
        for n in nums:
            if last_seen[n] == 999:
                last_seen[n]=idx
    weighted_score={}
    for n in range(1,46):
        # 최신 회차일수록 강하게 반영하되 100회 장기흐름도 유지
        score=(freq_windows[10][n]*2.40 + freq_windows[30][n]*1.45 + freq_windows[50][n]*1.05 + freq_windows[100][n]*0.72)
        # 직전회차 번호 과다 반복 방지
        if last_seen[n] == 0: score -= 1.4
        # 오래 안 나온 번호는 반등 후보로 보정
        if last_seen[n] >= 12: score += 1.2
        if last_seen[n] >= 20: score += 1.0
        weighted_score[n]=round(score,3)
    hot=sorted(range(1,46), key=lambda n:(-weighted_score[n], last_seen[n], n))[:15]
    cold=sorted(range(1,46), key=lambda n:(freq_windows[30][n], -last_seen[n], n))[:15]
    overdue=sorted(range(1,46), key=lambda n:(-last_seen[n], freq_windows[100][n], n))[:15]
    mid=sorted(range(1,46), key=lambda n:(abs(weighted_score[n]-(sum(weighted_score.values())/45)), n))[:15]
    pair_top=[{'pair':list(k),'count':v} for k,v in pair_counter.most_common(20)]
    return {
        'draws':draws,
        'freq':freq_windows[100],
        'freq10':freq_windows[10],
        'freq30':freq_windows[30],
        'freq50':freq_windows[50],
        'freq100':freq_windows[100],
        'weighted_score':weighted_score,
        'last_seen':last_seen,
        'hot':hot,
        'cold':cold,
        'mid':mid,
        'overdue':overdue,
        'pair_counter':pair_counter,
        'pair_top':pair_top,
        'end_counts':end_counts,
        'zone_counts':zone_counts,
        'odd_ratio':odd_total/(max(1,len(draws[:30])*6)),
        'avg_sum30':round(sum(sums)/len(sums),1) if sums else 0,
        'avg_ac30':round(sum(acs)/len(acs),1) if acs else 0,
        'latest_round':draws[0]['round_no'] if draws else 1230,
    }

def _ai_weight_pool(st, mode='balanced'):
    weights={}
    hot=set(st['hot'][:12]); cold=set(st['cold'][:12]); overdue=set(st['overdue'][:12])
    avg=(sum(st['weighted_score'].values())/45) if st.get('weighted_score') else 1
    for n in range(1,46):
        w=1.0 + max(0, st['weighted_score'].get(n,0))/max(1,avg)*0.55
        if n in hot: w += 1.55
        if n in overdue: w += 1.05
        if n in cold: w += 0.55
        # 극단 구간 보정: 완전히 배제하지 않고 모드별로 가중치 조절
        if mode == 'conservative':
            if 11 <= n <= 35: w += 0.65
            if n in hot: w += 0.35
        elif mode == 'aggressive':
            if n in overdue or n in cold: w += 0.75
            if n <= 10 or n >= 36: w += 0.30
        else:
            if 16 <= n <= 40: w += 0.30
        weights[n]=round(max(0.15,w),4)
    return weights

def _weighted_choice_from_list(pool, weights):
    usable=[n for n in pool if n in weights]
    if not usable:
        return None
    total=sum(weights[n] for n in usable)
    r=random.random()*total
    for n in usable:
        r-=weights[n]
        if r <= 0:
            return n
    return usable[-1]

def _combo_pair_score(combo, st):
    pc=st.get('pair_counter') or collections.Counter()
    return sum(pc.get(_ai_pair_key(a,b),0) for a,b in itertools.combinations(combo,2))

def combo_score(combo, st):
    """AI Phase2 최종 점수. 기존 함수명 유지."""
    combo=sorted(parse_nums(combo));
    if len(combo)!=6: return 0
    freq=st.get('freq100') or st.get('freq') or {n:0 for n in range(1,46)}
    hot_hit=len(set(combo)&set(st['hot'][:10])); cold_hit=len(set(combo)&set(st['cold'][:10])); overdue_hit=len(set(combo)&set(st['overdue'][:10]))
    odd=sum(n%2 for n in combo); total=sum(combo); zones=_ai_sections(combo)
    cons=_ai_consecutive(combo); ends=len(set(n%10 for n in combo)); ac=_ai_ac_value(combo); pair_score=_combo_pair_score(combo, st)
    score=62
    if odd in (2,3,4): score+=10
    else: score-=12
    if max(zones)<=3 and min(zones)>=1: score+=10
    elif max(zones)==4: score+=2
    else: score-=10
    if 95<=total<=180: score+=9
    elif 85<=total<=195: score+=3
    else: score-=12
    if 6<=ac<=10: score+=9
    elif 5<=ac<=12: score+=4
    else: score-=8
    if cons<=1: score+=6
    elif cons==2: score+=1
    else: score-=9
    if ends>=5: score+=5
    elif ends==4: score+=2
    else: score-=6
    if 1<=hot_hit<=3: score+=5
    elif hot_hit>=4: score-=2
    if 1<=cold_hit<=3: score+=4
    if 1<=overdue_hit<=3: score+=5
    if pair_score>=5: score+=4
    elif pair_score>=2: score+=2
    score += min(5, sum(freq.get(n,0) for n in combo)//8)
    return max(0, min(99, int(round(score))))

def tags_for_combo(combo, st):
    combo=sorted(parse_nums(combo)); tags=[]
    if len(set(combo)&set(st['hot'][:10]))>=2: tags.append('HOT 흐름 반영')
    if len(set(combo)&set(st['overdue'][:10]))>=1: tags.append('미출현 보강')
    if len(set(combo)&set(st['cold'][:10]))>=1: tags.append('COLD 변동 후보')
    odd=sum(n%2 for n in combo)
    if odd in (2,3,4): tags.append('홀짝 균형')
    zones=_ai_sections(combo)
    if max(zones)<=3 and min(zones)>=1: tags.append('구간 분산')
    if 6 <= _ai_ac_value(combo) <= 10: tags.append('AC 적정')
    if len(set(n%10 for n in combo))>=5: tags.append('끝수 분산')
    if _combo_pair_score(combo, st)>=2: tags.append('동반출현 반영')
    return tags[:5] or ['균형형 조합']

def _combo_detail(combo, st):
    c=sorted(combo)
    return {
        'numbers':c,
        'score':combo_score(c,st),
        'tags':tags_for_combo(c,st),
        'sum':sum(c),
        'odd':sum(n%2 for n in c),
        'even':6-sum(n%2 for n in c),
        'zones':_ai_sections(c),
        'ac':_ai_ac_value(c),
        'consecutive':_ai_consecutive(c),
        'end_unique':len(set(n%10 for n in c)),
        'hot_hit':len(set(c)&set(st['hot'][:10])),
        'cold_hit':len(set(c)&set(st['cold'][:10])),
        'overdue_hit':len(set(c)&set(st['overdue'][:10])),
        'pair_score':_combo_pair_score(c,st),
    }

def _engine_summary(details, st):
    if not details:
        return {'avg_score':0,'max_score':0,'min_score':0,'filters':['생성 데이터 없음']}
    avg_score=round(sum(d['score'] for d in details)/len(details),1)
    sums=[d['sum'] for d in details]
    acs=[d['ac'] for d in details]
    pair_avg=round(sum(d['pair_score'] for d in details)/len(details),1)
    odd=sum(d['odd'] for d in details); even=sum(d['even'] for d in details)
    sections=[sum(d['zones'][i] for d in details) for i in range(3)]
    return {
        'avg_score':avg_score,
        'max_score':max(d['score'] for d in details),
        'min_score':min(d['score'] for d in details),
        'avg_sum':round(sum(sums)/len(sums),1),
        'avg_ac':round(sum(acs)/len(acs),1),
        'pair_avg':pair_avg,
        'odd_even':{'odd':odd,'even':even},
        'sections':sections,
        'hot':st['hot'][:8],
        'cold':st['cold'][:8],
        'overdue':st['overdue'][:8],
        'pair_top':st.get('pair_top',[])[:5],
        'filters':['홀짝 2:4~4:2','구간 몰림 제한','합계 90~190 우선','AC 5~12 우선','연속수 2쌍 이하','끝수 중복 제한','동반출현 보정'],
    }

def make_premium_combos(count=10, fixed='', excluded='', mode='balanced'):
    """AI Phase2 추천번호 생성. 기존 함수명 유지."""
    st=latest_stats(120)
    target=max(1,min(50,int(count or 10)))
    fixed_set=set(parse_nums(fixed)); excluded_set=set(parse_nums(excluded)) - fixed_set
    pool=[n for n in range(1,46) if n not in excluded_set and n not in fixed_set]
    if len(pool)+len(fixed_set)<6:
        return make_combos(target, fixed, excluded, mode), [], st
    base_weights=_ai_weight_pool(st, mode)
    for n in excluded_set:
        base_weights.pop(n, None)
    candidates=[]; seen=set(); guard=0
    buckets={
        'hot':[n for n in st['hot'] if n in pool],
        'mid':[n for n in st['mid'] if n in pool],
        'cold':[n for n in st['cold'] if n in pool],
        'overdue':[n for n in st['overdue'] if n in pool],
        'pool':pool,
    }
    while len(candidates) < target*12 and guard < 26000:
        guard += 1
        combo=set(fixed_set)
        # 모드별 후보 시드 구성
        if mode == 'aggressive':
            plan=['hot','overdue','cold','pool','pool','pool']
        elif mode == 'conservative':
            plan=['hot','mid','mid','pool','pool','pool']
        else:
            plan=['hot','overdue','mid','cold','pool','pool']
        for name in plan:
            if len(combo)>=6: break
            usable=[n for n in buckets.get(name, pool) if n not in combo]
            pick=_weighted_choice_from_list(usable, base_weights)
            if pick: combo.add(pick)
        while len(combo)<6:
            pick=_weighted_choice_from_list([n for n in pool if n not in combo], base_weights)
            if pick is None: break
            combo.add(pick)
        arr=tuple(sorted(combo))
        if len(arr)!=6 or arr in seen: continue
        seen.add(arr)
        odd=sum(n%2 for n in arr); total=sum(arr); zones=_ai_sections(arr); cons=_ai_consecutive(arr); ac=_ai_ac_value(arr); ends=len(set(n%10 for n in arr))
        # 필터는 너무 빡빡하지 않게 하되 품질 하한 확보
        if odd not in (2,3,4): continue
        if max(zones)>4 or min(zones)==0: continue
        if not (85<=total<=195): continue
        if cons>2: continue
        if not (4<=ac<=12): continue
        if ends<4: continue
        candidates.append((combo_score(arr,st),list(arr)))
    candidates=sorted(candidates, key=lambda x:(-x[0], x[1]))
    combos=[c for _,c in candidates[:target]]
    if len(combos)<target:
        fallback=make_combos(target-len(combos), fixed, excluded, mode)
        for f in fallback:
            if f not in combos:
                combos.append(f)
            if len(combos)>=target: break
    details=[_combo_detail(c,st) for c in combos[:target]]
    return combos[:target], details, st

def build_analysis_text(round_no, st, mode, fixed, excluded):
    """AI Phase2 분석 문장: 짧게 3~5줄, 번호 흐름 기반."""
    pack=_load_analysis_pack(); used=set(); cond=pack.get('conditional',{}) or {}
    lines=[]
    target=random.choice([3,4,4,5])
    # 통계 기반 문장 먼저 구성
    hot_txt=', '.join(map(str, st.get('hot',[])[:3]))
    overdue_txt=', '.join(map(str, st.get('overdue',[])[:3]))
    if random.random()<0.55:
        lines.append(f'최근 흐름에서는 핵심 번호 {hot_txt}번대의 반영 가치가 높게 나타났습니다.')
    else:
        lines.append(_pick(pack.get('start'), used) or '이번 회차는 최근 흐름을 기준으로 균형 있게 분석했습니다.')
    if len(lines)<target and overdue_txt:
        lines.append(f'장기 미출현 후보 {overdue_txt}번을 보조 흐름으로 함께 검토했습니다.')
    mode_key=mode if mode in ('balanced','conservative','aggressive') else 'balanced'
    candidates=[]; candidates += cond.get(mode_key, [])
    if fixed: candidates += cond.get('fixed', [])
    if excluded: candidates += cond.get('excluded', [])
    if candidates and len(lines)<target:
        line=_pick(candidates, used)
        if line: lines.append(line)
    for cat in random.sample(['flow','core','pattern','strategy','ending'],5):
        if len(lines)>=target: break
        line=_pick(pack.get(cat), used)
        if line: lines.append(line)
    if len(lines)<3:
        lines.append('홀짝, 구간, 끝수, 합계, AC값을 함께 고려해 조합을 정리했습니다.')
    return '\n'.join([x for x in lines if x][:5])

